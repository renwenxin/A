"""
涨停复制战法 — 过去一年回测

选股: 沪深主板 · 非ST · 近20日有过涨停 · 处于缩量回调企稳状态
买入: 回调后再次放量突破回调高点 + 量 > MAVOL180×1.2 → 次日开盘买入
      + 竞价确认（低开>3%跳过、缩量>50%跳过、开盘跌停跳过）
卖出:
  0. -5% 硬止损
  1. 涨停复制成功（再次涨停）→ 继续持有，断板卖出
  2. 移动止盈: 最高收盘价回落 > 5%
  3. 持仓 ≥ 5天兜底

三大核心形态:
  - N字反包: 涨停 → 缩量回调1-3天 → 再次放量突破 → 买入
  - 涨停双响炮: 两根涨停中间夹缩量小K线 → 第二根涨停次日开盘买
  - 缩量回踩不破: 涨停后缩量回踩MA5不破 → 再次放量启动 → 买入

输出: xlsx 文件（交易明细 / 统计汇总 / 按形态统计 / 按卖出原因 / 月度统计）
"""
import sys, os, json, struct, time as _time
from datetime import date, datetime, timedelta
from typing import Dict, List, Optional, Tuple
from collections import defaultdict

import numpy as np
import pandas as pd

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

from ashare_review.data.tdx_reader import TdxReader, RECORD_SIZE
from ashare_review.analysis.indicators import calc_ma
from ashare_review.utils.calendar import TradingCalendar

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ═══════════════════════════════════════════════════════════════════════════
# Constants
# ═══════════════════════════════════════════════════════════════════════════
FEE = 0.0015
SLIPPAGE = 0.002
TOTAL_COST = FEE + SLIPPAGE
MAX_HOLD_DAYS = 5
MAX_LOOKBACK_ZT = 20        # 往前找涨停的最大天数
MIN_PULLBACK_DAYS = 1       # 涨停后至少回调1天
MAX_PULLBACK_DAYS = 10      # 涨停后最多回调10天（超过则股性失效）
MAVOL_PERIOD = 180
MAVOL_MULTIPLIER = 1.2
VOL_EXPAND_RATIO = 1.5      # 相对于前一日放量倍数

DATA_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'data')
LIMIT_UP_POOL_FILE = os.path.join(DATA_DIR, 'limit_up_pool.json')
NAME_CACHE_FILE = os.path.join(DATA_DIR, 'stock_name_map.json')
OUTPUT_DIR = os.path.join(DATA_DIR, '..', 'analysis')

# ═══════════════════════════════════════════════════════════════════════════
# Excel 样式
# ═══════════════════════════════════════════════════════════════════════════
TITLE_FONT = Font(name='微软雅黑', size=14, bold=True, color='1F2937')
HEADER_FONT = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='374151', end_color='374151', fill_type='solid')
HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
CENTER = Alignment(horizontal='center', vertical='center')
THIN_BORDER = Border(
    left=Side(style='thin', color='D1D5DB'),
    right=Side(style='thin', color='D1D5DB'),
    top=Side(style='thin', color='D1D5DB'),
    bottom=Side(style='thin', color='D1D5DB'),
)
WIN_FILL = PatternFill(start_color='DCFCE7', end_color='DCFCE7', fill_type='solid')
LOSS_FILL = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
SUMMARY_LABEL_FONT = Font(name='微软雅黑', size=10, bold=True, color='374151')
SUMMARY_VAL_FONT = Font(name='Consolas', size=11, color='1F2937')
GREEN_FONT = Font(name='Consolas', size=14, bold=True, color='059669')
RED_FONT = Font(name='Consolas', size=14, bold=True, color='DC2626')


def _style_header(ws, row: int, ncols: int):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def _style_data(ws, start_row: int, end_row: int, ncols: int):
    for r in range(start_row, end_row + 1):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='center', vertical='center')


def _auto_width(ws, min_width: int = 8, max_width: int = 22):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            if cell.value:
                val = str(cell.value)
                length = sum(2 if ord(c) > 127 else 1 for c in val)
                max_len = max(max_len, length)
        ws.column_dimensions[col_letter].width = max(min_width, min(max_len + 3, max_width))


# ═══════════════════════════════════════════════════════════════════════════
# 涨停复制战法 回测引擎
# ═══════════════════════════════════════════════════════════════════════════

class ZTReplicaBacktest:
    """涨停复制战法历史回测

    核心逻辑：
    1. 找近20日内有涨停的股票
    2. 判断是否处于缩量回调企稳状态
    3. 检测三类复制买入信号（N字反包/双响炮/回踩不破）
    4. 次日开盘买入 + V3卖出规则
    """

    def __init__(self, only_double_cannon: bool = False):
        self.tdx = TdxReader()
        self.cal = TradingCalendar()
        self.only_double_cannon = only_double_cannon  # 仅做涨停双响炮（砍掉缩量回踩不破）
        self._name_map: Dict[str, str] = {}
        self._sector_map: Dict[str, str] = {}
        self._load_name_map()
        self._load_sector_map()

    def _load_name_map(self):
        if os.path.exists(NAME_CACHE_FILE):
            try:
                with open(NAME_CACHE_FILE, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                if isinstance(data, dict) and len(data) > 1000:
                    self._name_map = data
            except Exception:
                pass

    def _load_sector_map(self):
        """加载行业板块映射（用于板块共振检测）。"""
        imap_path = os.path.join(DATA_DIR, 'industry_map.json')
        if os.path.exists(imap_path):
            try:
                with open(imap_path, 'r', encoding='utf-8') as f:
                    self._sector_map = json.load(f)
            except Exception:
                pass

    def _get_name(self, code: str) -> str:
        return self._name_map.get(str(code).zfill(6), code)

    def _get_sector(self, code: str) -> str:
        """获取股票所属行业板块。"""
        return self._sector_map.get(str(code).zfill(6), '')

    @staticmethod
    def _limit_threshold(code: str) -> float:
        code = str(code).zfill(6)
        if code.startswith(('300', '301', '688')): return 0.199
        if code.startswith(('8', '4')): return 0.299
        return 0.095

    @staticmethod
    def _is_main_board(code: str) -> bool:
        code = str(code).zfill(6)
        return code.startswith(('60', '00', '001', '002'))

    # ─── 获取候选池 ───

    def get_universe(self) -> List[str]:
        """从缓存获取候选池股票代码列表（年涨停≥10 + 主板 + 非ST）。"""
        if os.path.exists(LIMIT_UP_POOL_FILE):
            try:
                with open(LIMIT_UP_POOL_FILE, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                pool = data.get('pool', [])
                if pool:
                    codes = [s['code'] for s in pool
                             if self._is_main_board(s['code'])
                             and 'ST' not in self._get_name(s['code'])]
                    print(f'从缓存加载候选池: {len(codes)} 只')
                    return codes
            except Exception:
                pass
        print('[WARN] 候选池缓存不存在，请先运行 sim_portfolio.py 构建候选池')
        return []

    # ─── 读取单只股票全量数据 ───

    def _read_stock_full(self, code: str) -> Optional[pd.DataFrame]:
        """读取股票全量日线 + 计算MA指标 + MAVOL180。"""
        market = 'sh' if str(code).startswith('6') else 'sz'
        if str(code).startswith(('8', '4')):
            market = 'bj'
        try:
            df = self.tdx.read_daily(code, market)
            if df is None or df.empty or len(df) < MAVOL_PERIOD + 60:
                return None
            if len(df) > 600:
                df = df.iloc[-600:].copy()
                df.reset_index(drop=True, inplace=True)
            df = calc_ma(df, [5, 10, 20, 60])
            df['mavol180'] = df['volume'].rolling(MAVOL_PERIOD).mean()
            return df
        except Exception:
            return None

    # ─── 找最近涨停日 ───

    def _find_recent_zt(self, df: pd.DataFrame, idx: int, code: str
                        ) -> Optional[Dict]:
        """在idx位置往前找最近一次涨停（20日内）。

        Returns:
            {zt_idx, zt_date, zt_close, zt_open, zt_high, zt_low, zt_vol,
             zt_change_pct, days_since_zt}
        """
        limit_pct = self._limit_threshold(code)
        lookback = min(MAX_LOOKBACK_ZT, idx)
        if lookback < 2:
            return None

        for j in range(idx - 1, max(idx - lookback - 1, 0), -1):
            prev_close = float(df['close'].iloc[j - 1]) if j > 0 else 0
            close_j = float(df['close'].iloc[j])
            if prev_close <= 0:
                continue
            chg = (close_j - prev_close) / prev_close
            if chg >= limit_pct:
                # 排除一字板（开盘即涨停，无交易机会）
                open_j = float(df['open'].iloc[j])
                if abs(open_j - close_j) / max(close_j, 0.01) < 0.005:
                    # 一字板 → 继续往前找
                    continue
                return {
                    'zt_idx': j,
                    'zt_date': df['trade_date'].iloc[j],
                    'zt_close': close_j,
                    'zt_open': open_j,
                    'zt_high': float(df['high'].iloc[j]),
                    'zt_low': float(df['low'].iloc[j]),
                    'zt_vol': float(df['volume'].iloc[j]),
                    'zt_change_pct': round(chg * 100, 1),
                    'days_since_zt': idx - j,
                }
        return None

    # ─── 判断回调状态 ───

    def _analyze_pullback(self, df: pd.DataFrame, zt_idx: int, current_idx: int
                          ) -> Optional[Dict]:
        """分析涨停后的回调状态。

        Returns:
            {pullback_days, pb_high, pb_low, pb_vol_max, pb_vol_min,
             is_shrinking, is_above_zt_low, is_above_ma5, is_above_ma10,
             zt_retrace_pct, pattern_type}
        """
        if current_idx - zt_idx < MIN_PULLBACK_DAYS:
            return None  # 涨停当天/次日不参与
        if current_idx - zt_idx > MAX_PULLBACK_DAYS:
            return None  # 回调太久，股性已失效

        pb_start = zt_idx + 1
        pb_end = current_idx
        pb_days = pb_end - pb_start + 1

        zt_low = float(df['low'].iloc[zt_idx])
        zt_close = float(df['close'].iloc[zt_idx])
        zt_vol = float(df['volume'].iloc[zt_idx])

        # 回调区间的最高价、最低价、成交量
        pb_high = max(float(df['high'].iloc[i]) for i in range(pb_start, pb_end + 1))
        pb_low = min(float(df['low'].iloc[i]) for i in range(pb_start, pb_end + 1))
        pb_volumes = [float(df['volume'].iloc[i]) for i in range(pb_start, pb_end + 1)]
        pb_vol_max = max(pb_volumes)
        pb_vol_min = min(pb_volumes)
        pb_vol_last = float(df['volume'].iloc[current_idx])

        # 缩量判断：回调区最大量 < 涨停日量 × 0.8
        is_shrinking = pb_vol_max < zt_vol * 0.8

        # 不破涨停日最低价
        is_above_zt_low = pb_low >= zt_low * 0.98

        # 是否在均线上方
        ma5 = float(df['ma5'].iloc[current_idx]) if 'ma5' in df.columns and not pd.isna(df['ma5'].iloc[current_idx]) else 0
        ma10 = float(df['ma10'].iloc[current_idx]) if 'ma10' in df.columns and not pd.isna(df['ma10'].iloc[current_idx]) else 0
        close_cur = float(df['close'].iloc[current_idx])
        is_above_ma5 = close_cur > ma5 if ma5 > 0 else False
        is_above_ma10 = close_cur > ma10 if ma10 > 0 else False

        # 回调深度（相对涨停日收盘）
        zt_retrace_pct = (pb_low - zt_close) / zt_close * 100

        # 形态分类
        if pb_days <= 3 and is_shrinking and is_above_zt_low:
            pattern_type = 'N字反包'
        elif pb_days <= 4 and pb_vol_max < zt_vol * 0.6:
            pattern_type = '涨停双响炮候补'
        elif is_shrinking and is_above_ma5:
            pattern_type = '缩量回踩不破'
        elif is_shrinking and is_above_ma10:
            pattern_type = '回踩10日线'
        else:
            pattern_type = '回调企稳'

        return {
            'pullback_days': pb_days,
            'pb_high': round(pb_high, 2),
            'pb_low': round(pb_low, 2),
            'pb_vol_max': int(pb_vol_max),
            'pb_vol_min': int(pb_vol_min),
            'pb_vol_last': int(pb_vol_last),
            'zt_vol': int(zt_vol),
            'is_shrinking': is_shrinking,
            'is_above_zt_low': is_above_zt_low,
            'is_above_ma5': is_above_ma5,
            'is_above_ma10': is_above_ma10,
            'zt_retrace_pct': round(zt_retrace_pct, 1),
            'pattern_type': pattern_type,
        }

    # ─── 检测买入信号 ───

    def _check_buy_signal(self, df: pd.DataFrame, idx: int, code: str,
                          zt_info: Dict, pb_info: Dict) -> Optional[Dict]:
        """检测三类涨停复制买入信号。

        Returns None if no signal, or dict with signal details.
        """
        close = float(df['close'].iloc[idx])
        vol = float(df['volume'].iloc[idx])
        mavol180 = float(df['mavol180'].iloc[idx])
        open_p = float(df['open'].iloc[idx])
        high = float(df['high'].iloc[idx])

        if pd.isna(mavol180) or mavol180 <= 0:
            return None

        pb_high = pb_info['pb_high']
        zt_close = zt_info['zt_close']
        zt_vol = zt_info['zt_vol']
        days_since = zt_info['days_since_zt']

        # ── 条件1: 收盘价突破回调区间最高价 ──
        break_pb_high = close > pb_high

        # ── 条件2: 放量 ──
        vol_ratio = vol / mavol180
        vol_expand = vol > mavol180 * MAVOL_MULTIPLIER

        # 相对前一日放量
        if idx >= 1:
            prev_vol = float(df['volume'].iloc[idx - 1])
            vol_vs_prev = vol / max(prev_vol, 1)
        else:
            vol_vs_prev = 1.0

        # ── 条件3: 收盘在MA5上方 ──
        ma5 = float(df['ma5'].iloc[idx]) if 'ma5' in df.columns and not pd.isna(df['ma5'].iloc[idx]) else 0
        above_ma5 = close > ma5 if ma5 > 0 else True

        # ── 条件4: 非一字板（今日开板有交易机会） ──
        limit_pct = self._limit_threshold(code)
        is_yizi = abs(open_p - close) / max(close, 0.01) < 0.005
        is_zt_today = False
        if idx >= 1:
            prev_close = float(df['close'].iloc[idx - 1])
            if prev_close > 0:
                is_zt_today = (close - prev_close) / prev_close >= limit_pct

        # ── 组合判断：三类信号 ──

        # 信号A: N字反包 — 缩量回调后放量突破回调高点（最经典）
        signal_a = (pb_info['is_shrinking'] and break_pb_high and
                    vol_expand and above_ma5 and days_since >= 1)

        # 信号B: 涨停双响炮 — 今天再次涨停（涨停复制成功！）
        signal_b = is_zt_today and not is_yizi and days_since >= 1 and pb_info['is_shrinking']

        # 信号C: 缩量回踩不破 — 缩量回踩均线后放量启动
        signal_c = (pb_info['is_shrinking'] and pb_info['is_above_zt_low'] and
                    vol_expand and vol_vs_prev >= VOL_EXPAND_RATIO and
                    close > zt_close * 0.98 and not break_pb_high)

        # 仅做涨停双响炮：禁用 N字反包 + 缩量回踩不破（实测回踩不破胜率0%，反包全年0买入）
        if self.only_double_cannon:
            signal_a = False
            signal_c = False

        if not (signal_a or signal_b or signal_c):
            return None

        # ── 过滤条件 ──
        # 过度放量（≥5倍量）= 出货嫌疑
        if vol_ratio >= 5.0:
            return None
        # 死亡换手（回调太深，跌幅>8%）
        if pb_info['zt_retrace_pct'] < -8:
            return None

        # 确定信号类型
        if signal_b:
            sig_type = '涨停双响炮'
        elif signal_a:
            sig_type = 'N字反包'
        else:
            sig_type = '缩量回踩不破'

        # 评分
        score = 50
        score += 20 if signal_b else (15 if signal_a else 10)  # 双响炮最强
        score += 10 if vol_ratio >= 2.0 else (5 if vol_ratio >= 1.5 else 0)
        score += 10 if pb_info['is_above_ma5'] else 0
        score += 5 if pb_info['is_above_ma10'] else 0
        score += 5 if pb_info['pullback_days'] <= 3 else 0  # 回调短=强势
        score = min(100, score)

        return {
            'sig_type': sig_type,
            'score': score,
            'break_pb_high': break_pb_high,
            'vol_ratio': round(vol_ratio, 1),
            'vol_vs_prev': round(vol_vs_prev, 1),
            'zt_days_ago': days_since,
            'zt_change_pct': zt_info['zt_change_pct'],
            'zt_retrace_pct': pb_info['zt_retrace_pct'],
            'pattern_type': pb_info['pattern_type'],
        }

    # ─── 卖出检查（龙哥式：每日判断"是否符合预期"） ───

    def _check_sell(self, code: str, hold: dict, check_date: date,
                    df_full: pd.DataFrame) -> Optional[dict]:
        """龙哥卖出规则 — 每日判断持仓是否「符合预期」。

        符合预期 = 继续拿；不及预期 = 次日开盘直接核。

        符合预期的标准（每天检查）：
          - 今日收阳（close > open）→ 多头仍在
          - 且量不腰斩（vol >= prev_vol * 0.5）→ 资金没撤
          - 或今日涨停 → 无条件继续持有

        不及预期（触发卖出）：
          1. 收阴 + 缩量（close < open AND vol < prev_vol * 0.7）→ 资金跑路
          2. 高开低走上影线（open > prev_close +1% but close < open）→ 诱多出货
          3. -5% 硬止损（安全底线，不变）
          4. 收盘跌破 MA5 → 短期趋势坏了
        """
        buy_date = hold['buy_date']
        buy_price = hold['buy_price']
        had_zt = hold.get('had_zt', False)

        mask = df_full['trade_date'].apply(
            lambda x: (x.date() if hasattr(x, 'date') else x) <= check_date
        )
        df = df_full[mask].copy()
        if df.empty or len(df) < 2:
            return None

        idx = len(df) - 1
        close = float(df['close'].iloc[idx])
        open_p = float(df['open'].iloc[idx])
        high = float(df['high'].iloc[idx])
        low = float(df['low'].iloc[idx])
        vol = float(df['volume'].iloc[idx])

        if idx >= 1:
            prev_close = float(df['close'].iloc[idx - 1])
            prev_vol = float(df['volume'].iloc[idx - 1])
        else:
            prev_close = close
            prev_vol = vol

        try:
            trading_days = self.cal.trading_days_between(buy_date, check_date)
        except Exception:
            trading_days = 1

        # ── 规则0: 今日涨停 → 符合预期，继续持有 ──
        limit_pct = self._limit_threshold(code)
        is_zt_today = prev_close > 0 and (close - prev_close) / prev_close >= limit_pct
        if is_zt_today:
            open_chg = (open_p - prev_close) / max(prev_close, 0.01)
            is_yizi = abs(open_chg) >= limit_pct * 0.95
            if not is_yizi:
                hold['had_zt'] = True
                hold['replica_success'] = True
                return None  # 涨停复制成功！坚定持有

        # ── 规则1: -5% 硬止损（安全底线永远不变） ──
        if buy_price > 0:
            loss_pct = (close - buy_price) / buy_price
            if loss_pct <= -0.05:
                return {
                    'sell_price': round(close, 2),
                    'sell_reason': '硬止损-5%',
                    'days_held': trading_days,
                }

        # ── 规则2: 收阴 + 缩量 → 不及预期，次日核按钮 ──
        is_red = close > open_p  # 收阳
        vol_collapse = prev_vol > 0 and vol < prev_vol * 0.7  # 量缩30%+

        if not is_red and vol_collapse:
            return {
                'sell_price': round(close, 2),
                'sell_reason': '收阴缩量·资金撤了',
                'days_held': trading_days,
            }

        # ── 规则3: 高开低走上影线 → 诱多出货 ──
        # 开盘高开 >1% but 收盘低于开盘（收阴），且上影线很长
        if prev_close > 0:
            open_chg = (open_p - prev_close) / prev_close
            upper_shadow = (high - max(open_p, close)) / max(close, 0.01)
            if open_chg > 0.01 and not is_red and upper_shadow > 0.03:
                return {
                    'sell_price': round(close, 2),
                    'sell_reason': f'高开{open_chg*100:.1f}%低走·诱多',
                    'days_held': trading_days,
                }

        # ── 规则4: 收盘跌破 MA5 → 短期趋势破坏 ──
        ma5 = float(df['ma5'].iloc[idx]) if 'ma5' in df.columns and not pd.isna(df['ma5'].iloc[idx]) else 0
        if ma5 > 0 and close < ma5:
            return {
                'sell_price': round(close, 2),
                'sell_reason': '跌破MA5·趋势坏了',
                'days_held': trading_days,
            }

        # ── 规则5: 曾涨停后断板 → 给一天机会反包，否则走 ──
        if had_zt:
            hold['awaiting_reversal'] = True
            hold['had_zt'] = False  # 重置标记，明天再检查
            if trading_days >= 2:  # 断板第二天还不反包
                return {
                    'sell_price': round(close, 2),
                    'sell_reason': '涨停后断板未反包',
                    'days_held': trading_days,
                }
            return None  # 给一天看反包

        # ── 符合预期：收阳 + 量不腰斩 → 坚定持有 ──
        # （没有被上述任何规则触发 = 今天表现正常，继续拿）
        return None

    # ─── 竞价确认 ───

    def _check_auction(self, code: str, buy_date: date,
                       df_full: pd.DataFrame) -> Optional[str]:
        """竞价确认，返回 None=通过, str=拒绝原因。"""
        mask = df_full['trade_date'].apply(
            lambda x: (x.date() if hasattr(x, 'date') else x) <= buy_date
        )
        df = df_full[mask].copy()
        if df.empty or len(df) < 2:
            return None

        idx = len(df) - 1
        open_p = float(df['open'].iloc[idx])
        prev_close = float(df['close'].iloc[idx - 1])
        vol_today = float(df['volume'].iloc[idx])
        vol_prev = float(df['volume'].iloc[idx - 1])

        open_chg = (open_p - prev_close) / prev_close * 100 if prev_close > 0 else 0
        if open_chg < -3:
            return f'竞价低开{open_chg:.1f}%'

        if vol_prev > 0 and vol_today < vol_prev * 0.5:
            return '竞价缩量>50%'

        return None

    def _check_open_limit_down(self, code: str, buy_date: date,
                               df_full: pd.DataFrame) -> bool:
        """检查买入日是否开盘跌停。"""
        mask = df_full['trade_date'].apply(
            lambda x: (x.date() if hasattr(x, 'date') else x) <= buy_date
        )
        df = df_full[mask].copy()
        if df.empty or len(df) < 2:
            return False
        idx = len(df) - 1
        open_p = float(df['open'].iloc[idx])
        prev_close = float(df['close'].iloc[idx - 1])
        limit_pct = self._limit_threshold(code)
        return open_p <= prev_close * (1 - limit_pct)

    # ═══════════════════════════════════════════════════════════════════════
    # 市场环境 & 板块共振
    # ═══════════════════════════════════════════════════════════════════════

    @staticmethod
    def _market_regime(idx_above_ma60: bool, ma60_rising: bool) -> str:
        """市场环境三分类。

        - bull:   上证 > MA60 且 MA60 向上 → 正常进攻
        - neutral: 上证在 MA60 ±3% 内 OR 信号矛盾 → 谨慎
        - bear:   上证 < MA60 且 MA60 向下 → 防御
        """
        if idx_above_ma60 and ma60_rising:
            return 'bull'
        elif not idx_above_ma60 and not ma60_rising:
            return 'bear'
        else:
            return 'neutral'

    def _calc_sector_heat(self, all_codes: List[str], stock_data: Dict[str, pd.DataFrame],
                          td: date, _get_idx_func) -> Dict[str, dict]:
        """计算当日各板块热度（轻量版：只统计涨停数，不做完整信号检测）。

        板块热度 = 板块内涨停数
        热板块 = 板块内 ≥3 只涨停 OR ≥2 只涨停 + 板块内有多只候选

        Returns: {sector_name: {'zt_count': N, 'is_hot': bool, 'heat_score': int}}
        """
        sector_zt = defaultdict(int)
        sector_total = defaultdict(int)  # 板块内总候选数

        for code in all_codes:
            if code not in stock_data:
                continue
            df = stock_data[code]
            idx = _get_idx_func(code, td)
            if idx < 1:
                continue

            sec = self._get_sector(code)
            if not sec:
                continue

            sector_total[sec] += 1

            # 轻量检测：只判断今日是否涨停
            close = float(df['close'].iloc[idx])
            prev_close = float(df['close'].iloc[idx - 1])
            limit_pct = self._limit_threshold(code)
            if prev_close > 0 and (close - prev_close) / prev_close >= limit_pct:
                # 排除一字板
                open_p = float(df['open'].iloc[idx])
                if abs(open_p - close) / max(close, 0.01) >= 0.005:
                    sector_zt[sec] += 1

        # 板块热度：涨停数 ≥ 3 → 热，或涨停数 ≥ 2 且板块候选多 → 温
        heat = {}
        for sec in set(list(sector_zt.keys()) + list(sector_total.keys())):
            zt_n = sector_zt.get(sec, 0)
            total_n = sector_total.get(sec, 0)
            is_hot = zt_n >= 3 or (zt_n >= 2 and total_n >= 10)
            heat[sec] = {
                'zt_count': zt_n,
                'total_count': total_n,
                'is_hot': is_hot,
                'heat_score': min(10, zt_n * 3),
            }
        return heat

    # ═══════════════════════════════════════════════════════════════════════
    # 主回测循环
    # ═══════════════════════════════════════════════════════════════════════

    def run(self, start_date: date = None, end_date: date = None) -> dict:
        """运行涨停复制战法回测。"""
        INITIAL_CAPITAL = 1_000_000.0
        MAX_POSITIONS = 10
        MAX_NEW_PER_DAY = 3
        PER_POSITION_PCT = 0.10

        if end_date is None:
            end_date = date.today() - timedelta(days=1)
        if start_date is None:
            start_date = end_date - timedelta(days=365)

        print(f'\n{"="*60}')
        print(f'  涨停复制战法 — 历史回测')
        print(f'  区间: {start_date} ~ {end_date}')
        print(f'  资金: {INITIAL_CAPITAL/1e4:.0f}万 | 最大持仓: {MAX_POSITIONS}只 | 单票: {PER_POSITION_PCT*100:.0f}%')
        print(f'{"="*60}\n')

        # ── Step 1: 获取候选池 ──
        universe = self.get_universe()
        if not universe:
            print('[ERROR] 候选池为空')
            return {}

        # ── Step 2: 预读所有候选股数据 ──
        print(f'预读 {len(universe)} 只候选股数据...')
        stock_data: Dict[str, pd.DataFrame] = {}
        stock_date_idx: Dict[str, Dict[date, int]] = {}
        t0 = _time.time()
        for i, code in enumerate(universe):
            if (i + 1) % 200 == 0:
                print(f'  读取 {i+1}/{len(universe)} (已加载 {len(stock_data)} 只, {_time.time() - t0:.0f}s)...')
            df = self._read_stock_full(code)
            if df is not None and len(df) >= MAVOL_PERIOD + 60:
                stock_data[code] = df
                dmap = {}
                for ri, td_val in enumerate(df['trade_date']):
                    if hasattr(td_val, 'date'):
                        dmap[td_val.date()] = ri
                    elif hasattr(td_val, 'strftime'):
                        dmap[td_val] = ri
                stock_date_idx[code] = dmap
        print(f'预读完成: {len(stock_data)} 只有效数据 ({_time.time() - t0:.0f}s)')

        # ── Step 3: 生成交易日列表 ──
        all_dates = []
        d = start_date
        while d <= end_date:
            if d.weekday() < 5:
                all_dates.append(d)
            d += timedelta(days=1)
        print(f'交易日: {len(all_dates)} 天 ({all_dates[0]} ~ {all_dates[-1]})')

        # ── Step 3b: 市场环境预计算（增强版） ──
        market_regime: Dict[date, str] = {}     # bull / neutral / bear
        market_bullish: Dict[date, bool] = {}
        try:
            sh_df = self.tdx.read_daily('999999', 'sh')
            if sh_df is not None and len(sh_df) >= 120:
                sh_df['ma60'] = sh_df['close'].rolling(60).mean()
                # MA60 方向（20日斜率）
                for td in all_dates:
                    mask = sh_df['trade_date'].apply(
                        lambda x: (x.date() if hasattr(x, 'date') else x) <= td
                    )
                    sub = sh_df[mask]
                    if len(sub) >= 60:
                        close = float(sub['close'].iloc[-1])
                        ma60 = float(sub['ma60'].iloc[-1])
                        idx_above = close > ma60 and not pd.isna(ma60)
                        market_bullish[td] = idx_above

                        # 判断 MA60 是否向上（对比20天前的MA60）
                        ma60_rising = True
                        if len(sub) >= 80:
                            ma60_20d_ago = float(sub['ma60'].iloc[-21])
                            if not pd.isna(ma60_20d_ago) and ma60_20d_ago > 0:
                                ma60_rising = ma60 > ma60_20d_ago
                        market_regime[td] = self._market_regime(idx_above, ma60_rising)
        except Exception:
            pass
        print(f'市场环境: 上证MA60数据 {len(market_regime)} 天')

        # ── Step 4: 逐日回测 ──
        cash = INITIAL_CAPITAL
        holdings: Dict[str, dict] = {}
        finished_trades: List[dict] = []
        skipped_trades: List[dict] = []
        daily_log: List[dict] = []
        total_signals = 0
        total_buys = 0
        portfolio_values = []

        def _get_idx(code: str, target_date: date) -> int:
            dmap = stock_date_idx.get(code)
            if dmap is None: return -1
            return dmap.get(target_date, -1)

        def _current_price(code: str, td: date) -> Optional[float]:
            idx = _get_idx(code, td)
            if idx >= 0:
                return float(stock_data[code]['close'].iloc[idx])
            return None

        t0 = _time.time()
        pending_sells: Dict[str, dict] = {}

        for di, td in enumerate(all_dates):
            if (di + 1) % 20 == 0:
                elapsed = _time.time() - t0
                eta = elapsed / (di + 1) * (len(all_dates) - di - 1) if di > 0 else 0
                pos_val = sum(
                    (h['shares'] * (_current_price(code, td) or h['buy_price']))
                    for code, h in holdings.items()
                )
                total_val = cash + pos_val
                print(f'  [{di+1}/{len(all_dates)}] {td}  '
                      f'持仓:{len(holdings)} 市值:{total_val/1e4:.1f}万 已平:{len(finished_trades)}  '
                      f'({elapsed:.0f}s, ETA {eta:.0f}s)', flush=True)

            # ── Step 4a: 执行昨日挂单的卖出 ──
            sold_today = []
            for code in list(pending_sells.keys()):
                sell_info = pending_sells[code]
                h = sell_info['hold']
                sell_idx = _get_idx(code, td)
                if sell_idx >= 0:
                    sell_price = float(stock_data[code]['open'].iloc[sell_idx])
                else:
                    sell_price = sell_info.get('fallback_price', h['buy_price'])

                buy_price = h['buy_price']
                shares = h['shares']
                sell_cost_rate = 0.0008
                proceeds = shares * sell_price * (1 - sell_cost_rate)
                cash += proceeds

                gross_ret = (sell_price - buy_price) / buy_price if buy_price > 0 else 0
                net_ret = gross_ret - TOTAL_COST

                trading_days = 0
                try:
                    trading_days = self.cal.trading_days_between(h['buy_date'], td)
                except Exception:
                    trading_days = sell_info.get('_days', 3)

                finished_trades.append({
                    'code': code,
                    'name': self._get_name(code),
                    'buy_date': h['buy_date'].strftime('%Y-%m-%d'),
                    'buy_price': round(buy_price, 2),
                    'sell_date': td.strftime('%Y-%m-%d'),
                    'sell_price': round(sell_price, 2),
                    'gross_ret': round(gross_ret * 100, 2),
                    'net_ret': round(net_ret * 100, 2),
                    'is_win': net_ret > 0,
                    'exit_reason': sell_info['sell_reason'],
                    'days_held': trading_days if trading_days > 0 else sell_info.get('_days', 1),
                    'had_zt': h.get('had_zt', False),
                    'replica_success': h.get('replica_success', False),
                    'sig_type': h.get('sig_type', ''),
                    'score': h.get('score', 0),
                    'vol_ratio': h.get('vol_ratio', 0),
                    'zt_days_ago': h.get('zt_days_ago', 0),
                    'zt_retrace_pct': h.get('zt_retrace_pct', 0),
                    'sector_name': h.get('sector_name', ''),
                    'sector_heat': h.get('sector_heat', 0),
                    'market_regime': h.get('market_regime', ''),
                })
                sold_today.append(code)
                del holdings[code]
                del pending_sells[code]

            # ── Step 4b: 检查持仓是否需要挂卖出单 ──
            for code in list(holdings.keys()):
                if code in pending_sells:
                    continue
                if code not in stock_data:
                    continue
                sell_sig = self._check_sell(code, holdings[code], td, stock_data[code])
                if sell_sig:
                    next_td_sell = self._next_trade_date(td, all_dates)
                    if next_td_sell is None:
                        sell_real_date = td
                    else:
                        sell_real_date = next_td_sell
                    pending_sells[code] = {
                        'hold': holdings[code],
                        'sell_date': sell_real_date,
                        'sell_reason': sell_sig['sell_reason'],
                        'fallback_price': sell_sig['sell_price'],
                        '_days': sell_sig['days_held'],
                    }

            # ── Step 4c: 检测涨停复制买入信号 ──
            buy_today = []
            regime = market_regime.get(td, 'neutral')
            is_bull = market_bullish.get(td, True)

            # — 市场环境动态参数 —
            if regime == 'bull':
                max_pos = 10
                max_new_day = 3
                min_score = 55
                require_sector = False
            elif regime == 'neutral':
                max_pos = 5
                max_new_day = 2
                min_score = 65
                require_sector = True   # 震荡市必须板块共振
            else:  # bear
                max_pos = 3
                max_new_day = 1
                min_score = 75
                require_sector = True   # 熊市必须板块共振 + 高门槛

            # — 计算当日板块热度 —
            sector_heat = self._calc_sector_heat(
                list(stock_data.keys()), stock_data, td, _get_idx
            )

            candidates_today = []

            for code, df in stock_data.items():
                if code in holdings or code in pending_sells:
                    continue
                if len(holdings) >= max_pos:
                    break

                idx = _get_idx(code, td)
                if idx < MAVOL_PERIOD + 30:
                    continue

                # 找最近涨停
                zt_info = self._find_recent_zt(df, idx, code)
                if zt_info is None:
                    continue

                # 分析回调状态
                pb_info = self._analyze_pullback(df, zt_info['zt_idx'], idx)
                if pb_info is None:
                    continue

                # 检测买入信号
                sig = self._check_buy_signal(df, idx, code, zt_info, pb_info)
                if sig is None:
                    continue

                total_signals += 1

                # — 板块共振过滤 —
                sec = self._get_sector(code)
                sec_info = sector_heat.get(sec, {})
                sector_bonus = 0
                if sec_info.get('is_hot'):
                    sector_bonus = sec_info.get('heat_score', 0)
                    sig['score'] = min(100, sig['score'] + sector_bonus)
                    sig['sector_name'] = sec
                    sig['sector_heat'] = sec_info['heat_score']
                elif require_sector:
                    # 震荡/熊市：板块不热直接跳过
                    continue

                # — 市场环境过滤 —
                if sig['score'] < min_score:
                    continue

                # — 熊市额外过滤：必须站上MA20（中期趋势确认） —
                if regime == 'bear':
                    ma20 = float(df['ma20'].iloc[idx]) if 'ma20' in df.columns and not pd.isna(df['ma20'].iloc[idx]) else 0
                    if ma20 <= 0 or float(df['close'].iloc[idx]) <= ma20:
                        continue

                # ═══════════════════════════════════════════════════════
                # 龙哥五条件手动筛选（缩小人机差距的核心）
                # ═══════════════════════════════════════════════════════

                close = float(df['close'].iloc[idx])
                vol = float(df['volume'].iloc[idx])

                # —— 条件⓵: 今日放量显著（龙哥"成交量近6个月最大量"的量化版）——
                # 实战中不是真的一定要=6个月最高，而是"量能显著放大"。
                # 用 今日量 >= 近20日均量 × 2.0（相对于自己近期明显放量）
                lookback_20 = min(20, idx)
                if lookback_20 >= 5:
                    vol_20_avg = sum(float(df['volume'].iloc[i]) for i in range(idx - lookback_20, idx)) / lookback_20
                    if vol_20_avg > 0:
                        vol_ratio_vs_20 = vol / vol_20_avg
                        if vol_ratio_vs_20 < 1.5:
                            continue  # 量不够显著 → 跳过
                        sig['score'] = min(100, sig['score'] + min(10, int(vol_ratio_vs_20)))

                # —— 条件⓶: 价格站上MA60（中期趋势向上，无大级别套牢盘）——
                # 龙哥"价格在新高附近"的本质：中期趋势确认，没被深套的筹码
                ma60 = float(df['ma60'].iloc[idx]) if 'ma60' in df.columns and not pd.isna(df['ma60'].iloc[idx]) else 0
                if ma60 > 0 and close < ma60:
                    continue  # 中期趋势向下，跳过
                if ma60 > 0:
                    above_ma60_pct = (close - ma60) / ma60 * 100
                    if above_ma60_pct > 10:
                        sig['score'] = min(100, sig['score'] + 5)  # 强势站上MA60

                # —— 条件⓷: 前日抗跌验证（龙哥选股中最高质量的信号）——
                # "昨天大盘普跌但个股独立翻红 = 资金逆势护盘"
                if idx >= 1:
                    prev_close = float(df['close'].iloc[idx - 1])
                    prev_chg = (prev_close - float(df['close'].iloc[idx - 2])) / max(float(df['close'].iloc[idx - 2]), 0.01) if idx >= 2 else 0
                    # 检测昨日市场是否普跌（用上证作为代理）
                    sh_idx = _get_idx('999999', td) if '999999' in stock_data else -1
                    market_down_yesterday = False
                    if sh_idx >= 2:
                        sh_prev_close = float(stock_data['999999']['close'].iloc[sh_idx - 1])
                        sh_prev2_close = float(stock_data['999999']['close'].iloc[sh_idx - 2])
                        market_down_yesterday = (sh_prev_close - sh_prev2_close) / max(sh_prev2_close, 0.01) < -0.005
                    if market_down_yesterday and prev_chg > 0.005:
                        # 大盘跌但个股涨 → 抗跌验证，加15分
                        sig['score'] = min(100, sig['score'] + 15)
                        sig['anti_fall_verified'] = True

                # 找下一交易日
                next_td = self._next_trade_date(td, all_dates)
                if next_td is None:
                    continue

                # 检查开盘跌停
                if self._check_open_limit_down(code, next_td, df):
                    skipped_trades.append({
                        'code': code, 'name': self._get_name(code),
                        'signal_date': td.strftime('%Y-%m-%d'),
                        'skip_reason': '开盘跌停',
                        'sig_type': sig['sig_type'],
                    })
                    continue

                # 竞价确认
                auction_reject = self._check_auction(code, next_td, df)
                if auction_reject:
                    skipped_trades.append({
                        'code': code, 'name': self._get_name(code),
                        'signal_date': td.strftime('%Y-%m-%d'),
                        'skip_reason': auction_reject,
                        'sig_type': sig['sig_type'],
                    })
                    continue

                candidates_today.append({
                    'code': code,
                    'score': sig['score'],
                    'sig_type': sig['sig_type'],
                    'vol_ratio': sig['vol_ratio'],
                    'zt_days_ago': sig['zt_days_ago'],
                    'zt_retrace_pct': sig['zt_retrace_pct'],
                    'close': float(df['close'].iloc[idx]),
                    'next_td': next_td,
                    'sector_name': sig.get('sector_name', ''),
                    'sector_heat': sig.get('sector_heat', 0),
                })

            # 按评分排序，限制每天最多 N 只新买入
            candidates_today.sort(key=lambda x: -x['score'])

            # — 龙哥规则：每日候选 ≤ 5，只做最强 —
            candidates_today = candidates_today[:5]

            # — 连亏3笔 → 强制空仓2天（龙哥铁律） —
            if len(finished_trades) >= 3:
                last3 = finished_trades[-3:]
                if sum(1 for t in last3 if not t['is_win']) >= 3:
                    # 检查最近一笔卖出是否在2天以内
                    last_sell_date_str = finished_trades[-1]['sell_date']
                    try:
                        last_sell_dt = datetime.strptime(last_sell_date_str, '%Y-%m-%d').date()
                        days_since = (td - last_sell_dt).days
                        if days_since <= 2:
                            candidates_today = []  # 空仓冷静期
                    except Exception:
                        pass

            available_slots = max(0, max_pos - len(holdings))
            max_new = min(max_new_day, available_slots)

            for c in candidates_today[:max_new]:
                code = c['code']
                next_td = c['next_td']

                next_idx = _get_idx(code, next_td)
                if next_idx < 0:
                    continue
                buy_open = float(stock_data[code]['open'].iloc[next_idx])

                position_capital = INITIAL_CAPITAL * PER_POSITION_PCT
                shares = int(position_capital / buy_open / 100) * 100
                if shares < 100:
                    continue
                buy_cost = shares * buy_open * (1 + 0.0003)
                if buy_cost > cash:
                    continue

                cash -= buy_cost

                holdings[code] = {
                    'code': code,
                    'buy_date': next_td,
                    'buy_price': buy_open,
                    'shares': shares,
                    'signal_date': td,
                    'sig_type': c['sig_type'],
                    'vol_ratio': c['vol_ratio'],
                    'zt_days_ago': c['zt_days_ago'],
                    'zt_retrace_pct': c['zt_retrace_pct'],
                    'had_zt': False,
                    'highest_close': buy_open,
                    'awaiting_reversal': False,
                    'replica_success': False,
                    'score': c['score'],
                    'sector_name': c.get('sector_name', ''),
                    'sector_heat': c.get('sector_heat', 0),
                    'market_regime': regime,
                }

                buy_today.append(code)
                total_buys += 1

            # ── 日统计 ──
            pos_val = sum(
                (h['shares'] * (_current_price(code, td) or h['buy_price']))
                for code, h in holdings.items()
            )
            total_val = cash + pos_val
            portfolio_values.append((td, total_val, cash, pos_val))

            daily_log.append({
                'date': td.strftime('%Y-%m-%d'),
                'signals': total_signals,
                'buys': len(buy_today),
                'sells': len(sold_today),
                'holdings': len(holdings),
                'portfolio_value': round(total_val, 2),
                'market_bull': is_bull,
                'market_regime': regime,
            })

        # ── 到期强制平仓 ──
        last_date = all_dates[-1]
        for code in list(holdings.keys()):
            h = holdings[code]
            sell_idx = _get_idx(code, last_date)
            if sell_idx >= 0:
                final_close = float(stock_data[code]['close'].iloc[sell_idx])
                buy_price = h['buy_price']
                shares = h['shares']
                proceeds = shares * final_close * (1 - 0.0008)
                cash += proceeds
                gross_ret = (final_close - buy_price) / buy_price if buy_price > 0 else 0
                net_ret = gross_ret - TOTAL_COST
                trading_days = 0
                try:
                    trading_days = self.cal.trading_days_between(h['buy_date'], last_date)
                except Exception:
                    trading_days = 5
                finished_trades.append({
                    'code': code,
                    'name': self._get_name(code),
                    'buy_date': h['buy_date'].strftime('%Y-%m-%d'),
                    'buy_price': round(buy_price, 2),
                    'sell_date': last_date.strftime('%Y-%m-%d'),
                    'sell_price': round(final_close, 2),
                    'gross_ret': round(gross_ret * 100, 2),
                    'net_ret': round(net_ret * 100, 2),
                    'is_win': net_ret > 0,
                    'exit_reason': '回测到期强平',
                    'days_held': trading_days if trading_days > 0 else 5,
                    'had_zt': h.get('had_zt', False),
                    'replica_success': h.get('replica_success', False),
                    'sig_type': h.get('sig_type', ''),
                    'score': h.get('score', 0),
                    'vol_ratio': h.get('vol_ratio', 0),
                    'zt_days_ago': h.get('zt_days_ago', 0),
                    'zt_retrace_pct': h.get('zt_retrace_pct', 0),
                    'sector_name': h.get('sector_name', ''),
                    'sector_heat': h.get('sector_heat', 0),
                    'market_regime': h.get('market_regime', ''),
                })
            del holdings[code]

        # ── 组合层统计 ──
        total_val = cash
        cumulative_return = (total_val / INITIAL_CAPITAL - 1) * 100 if INITIAL_CAPITAL > 0 else 0
        peak_value = INITIAL_CAPITAL
        max_drawdown = 0.0
        for _, tv, _, _ in portfolio_values:
            peak_value = max(peak_value, tv)
            dd = (peak_value - tv) / peak_value * 100 if peak_value > 0 else 0
            max_drawdown = max(max_drawdown, dd)

        elapsed = _time.time() - t0
        print(f'\n回测完成 ({elapsed:.0f}s): '
              f'{total_signals} 信号, {total_buys} 买入, '
              f'{len(finished_trades)} 笔交易')
        print(f'  最终市值: {total_val/1e4:.1f}万 | '
              f'累计收益: {cumulative_return:+.1f}% | '
              f'最大回撤: {max_drawdown:.1f}%')

        return {
            'trades': finished_trades,
            'skipped': skipped_trades,
            'daily_log': daily_log,
            'start_date': start_date.strftime('%Y-%m-%d'),
            'end_date': end_date.strftime('%Y-%m-%d'),
            'total_signals': total_signals,
            'total_buys': total_buys,
            'final_value': round(total_val, 2),
            'initial_capital': INITIAL_CAPITAL,
            'cumulative_return': round(cumulative_return, 2),
            'max_drawdown': round(max_drawdown, 2),
            'portfolio_values': portfolio_values,
        }

    def _next_trade_date(self, d: date, all_dates: List[date]) -> Optional[date]:
        for td in all_dates:
            if td > d:
                return td
        return None


# ═══════════════════════════════════════════════════════════════════════════
# Excel 导出
# ═══════════════════════════════════════════════════════════════════════════

def export_xlsx(results: dict, output_path: str):
    trades = results.get('trades', [])
    skipped = results.get('skipped', [])
    daily = results.get('daily_log', [])

    wb = Workbook()

    # ═══ Sheet 1: 交易明细 ═══
    ws1 = wb.active
    ws1.title = '交易明细'

    ws1.merge_cells('A1:R1')
    ws1['A1'] = f"涨停复制战法 — 交易明细 ({results['start_date']} ~ {results['end_date']})"
    ws1['A1'].font = TITLE_FONT
    ws1['A1'].alignment = CENTER

    headers = ['代码', '名称', '买入日', '买入价', '卖出日', '卖出价',
               '毛收益%', '净收益%', '结果', '持有天', '卖出原因',
               '曾涨停', '复制成功', '信号类型', '评分', '量比',
               '距涨停天', '回调深度%']
    for c, h in enumerate(headers, 1):
        ws1.cell(row=3, column=c, value=h)
    _style_header(ws1, 3, len(headers))

    for i, t in enumerate(trades):
        r = i + 4
        ws1.cell(row=r, column=1, value=t['code'])
        ws1.cell(row=r, column=2, value=t['name'])
        ws1.cell(row=r, column=3, value=t['buy_date'])
        ws1.cell(row=r, column=4, value=t['buy_price'])
        ws1.cell(row=r, column=5, value=t['sell_date'])
        ws1.cell(row=r, column=6, value=t['sell_price'])
        ws1.cell(row=r, column=7, value=t['gross_ret'])
        ws1.cell(row=r, column=8, value=t['net_ret'])
        ws1.cell(row=r, column=9, value='Win' if t['is_win'] else 'Loss')
        ws1.cell(row=r, column=10, value=t['days_held'])
        ws1.cell(row=r, column=11, value=t['exit_reason'])
        ws1.cell(row=r, column=12, value='是' if t.get('had_zt') else '否')
        ws1.cell(row=r, column=13, value='✅是' if t.get('replica_success') else '否')
        ws1.cell(row=r, column=14, value=t.get('sig_type', ''))
        ws1.cell(row=r, column=15, value=t.get('score', 0))
        ws1.cell(row=r, column=16, value=t.get('vol_ratio', 0))
        ws1.cell(row=r, column=17, value=t.get('zt_days_ago', 0))
        ws1.cell(row=r, column=18, value=t.get('zt_retrace_pct', 0))

        if t['is_win']:
            ws1.cell(row=r, column=8).fill = WIN_FILL
            ws1.cell(row=r, column=9).fill = WIN_FILL
        else:
            ws1.cell(row=r, column=8).fill = LOSS_FILL
            ws1.cell(row=r, column=9).fill = LOSS_FILL

        if t.get('replica_success'):
            ws1.cell(row=r, column=13).fill = WIN_FILL

    if trades:
        _style_data(ws1, 4, 3 + len(trades), len(headers))
    _auto_width(ws1)

    # ═══ Sheet 2: 统计汇总 ═══
    ws2 = wb.create_sheet('统计汇总')
    ws2.merge_cells('A1:D1')
    ws2['A1'] = '涨停复制战法 — 统计汇总'
    ws2['A1'].font = TITLE_FONT
    ws2['A1'].alignment = CENTER

    total = len(trades)
    wins = sum(1 for t in trades if t['is_win'])
    losses = total - wins
    win_rate = wins / max(total, 1) * 100
    net_rets = [t['net_ret'] for t in trades]
    avg_ret = np.mean(net_rets) if net_rets else 0
    avg_win = np.mean([r for r in net_rets if r > 0]) if wins > 0 else 0
    avg_loss = np.mean([r for r in net_rets if r <= 0]) if losses > 0 else 0
    max_win = max(net_rets) if net_rets else 0
    max_loss = min(net_rets) if net_rets else 0

    profit_factor = abs(sum(r for r in net_rets if r > 0) / min(sum(r for r in net_rets if r <= 0), -0.01)) if losses > 0 else 999
    avg_hold = np.mean([t['days_held'] for t in trades]) if trades else 0

    # 复制成功统计
    replica_success = sum(1 for t in trades if t.get('replica_success'))
    replica_wins = sum(1 for t in trades if t.get('replica_success') and t['is_win'])
    replica_win_rate = replica_wins / max(replica_success, 1) * 100

    initial_cap = results.get('initial_capital', 1000000.0)
    final_value = results.get('final_value', initial_cap)
    cum_ret = results.get('cumulative_return', 0.0)
    max_dd = results.get('max_drawdown', 0.0)

    portfolio_values = results.get('portfolio_values', [])
    if portfolio_values:
        peak_val = initial_cap
        max_dd_calc = 0.0
        for _, tv, _, _ in portfolio_values:
            peak_val = max(peak_val, tv)
            dd = (peak_val - tv) / peak_val * 100 if peak_val > 0 else 0
            max_dd_calc = max(max_dd_calc, dd)
        if max_dd_calc > 0:
            max_dd = max_dd_calc

    stats = [
        ('回测区间', f"{results['start_date']} ~ {results['end_date']}"),
        ('', ''),
        ('--- 资金模型 ---', ''),
        ('初始资金', f'{initial_cap/1e4:.0f}万'),
        ('最终市值', f'{final_value/1e4:.2f}万'),
        ('最大持仓', '10只（单票10%）'),
        ('每天最多新开', '3只'),
        ('', ''),
        ('--- 交易统计 ---', ''),
        ('信号总数', results.get('total_signals', 0)),
        ('实际买入', results.get('total_buys', 0)),
        ('跳过(竞价/跌停)', len(skipped)),
        ('总交易笔数', total),
        ('盈利笔数', wins),
        ('亏损笔数', losses),
        ('胜率', f'{win_rate:.1f}%'),
        ('', ''),
        ('平均净收益/笔', f'{avg_ret:+.2f}%'),
        ('平均盈利/笔', f'{avg_win:+.2f}%'),
        ('平均亏损/笔', f'{avg_loss:+.2f}%'),
        ('最大单笔盈利', f'{max_win:+.2f}%'),
        ('最大单笔亏损', f'{max_loss:+.2f}%'),
        ('盈亏比', f'{profit_factor:.2f}'),
        ('平均持有天数', f'{avg_hold:.1f}天'),
        ('', ''),
        ('--- 涨停复制专项 ---', ''),
        ('复制成功(持有期再涨停)', f'{replica_success}笔'),
        ('复制成功胜率', f'{replica_win_rate:.1f}%'),
        ('', ''),
        ('--- 组合表现 ---', ''),
        ('组合累计收益', f'{cum_ret:+.2f}%'),
        ('最大回撤', f'{max_dd:.2f}%'),
    ]

    for i, (label, val) in enumerate(stats):
        r = 3 + i
        ws2.cell(row=r, column=1, value=label).font = SUMMARY_LABEL_FONT
        ws2.cell(row=r, column=2, value=val).font = SUMMARY_VAL_FONT
        ws2.cell(row=r, column=1).border = THIN_BORDER
        ws2.cell(row=r, column=2).border = THIN_BORDER
        if '---' in str(label):
            ws2.cell(row=r, column=1).fill = PatternFill(start_color='F3F4F6', end_color='F3F4F6', fill_type='solid')
            ws2.cell(row=r, column=2).fill = PatternFill(start_color='F3F4F6', end_color='F3F4F6', fill_type='solid')

    # 大数字显示
    big_row = 3 + len(stats) + 2
    ws2.merge_cells(f'A{big_row}:B{big_row}')
    ws2.cell(row=big_row, column=1, value='胜率').font = SUMMARY_LABEL_FONT
    ws2.cell(row=big_row, column=3, value=f'{win_rate:.1f}%').font = GREEN_FONT if win_rate >= 50 else RED_FONT
    ws2.cell(row=big_row, column=1).border = THIN_BORDER
    ws2.cell(row=big_row, column=3).border = THIN_BORDER

    big_row2 = big_row + 1
    ws2.merge_cells(f'A{big_row2}:B{big_row2}')
    ws2.cell(row=big_row2, column=1, value='组合累计收益').font = SUMMARY_LABEL_FONT
    ws2.cell(row=big_row2, column=3, value=f'{cum_ret:+.2f}%').font = GREEN_FONT if cum_ret > 0 else RED_FONT
    ws2.cell(row=big_row2, column=1).border = THIN_BORDER
    ws2.cell(row=big_row2, column=3).border = THIN_BORDER

    big_row3 = big_row2 + 1
    ws2.merge_cells(f'A{big_row3}:B{big_row3}')
    ws2.cell(row=big_row3, column=1, value='最大回撤').font = SUMMARY_LABEL_FONT
    ws2.cell(row=big_row3, column=3, value=f'{max_dd:.2f}%').font = RED_FONT
    ws2.cell(row=big_row3, column=1).border = THIN_BORDER
    ws2.cell(row=big_row3, column=3).border = THIN_BORDER

    big_row4 = big_row3 + 1
    ws2.merge_cells(f'A{big_row4}:B{big_row4}')
    ws2.cell(row=big_row4, column=1, value='涨停复制成功率').font = SUMMARY_LABEL_FONT
    ws2.cell(row=big_row4, column=3, value=f'{replica_win_rate:.1f}%').font = GREEN_FONT if replica_win_rate >= 50 else RED_FONT
    ws2.cell(row=big_row4, column=1).border = THIN_BORDER
    ws2.cell(row=big_row4, column=3).border = THIN_BORDER

    _auto_width(ws2, min_width=14, max_width=30)
    ws2.column_dimensions['C'].width = 18

    # ═══ Sheet 3: 按信号类型统计 ═══
    ws3 = wb.create_sheet('信号类型统计')
    ws3.merge_cells('A1:F1')
    ws3['A1'] = '按涨停复制信号类型统计'
    ws3['A1'].font = TITLE_FONT
    ws3['A1'].alignment = CENTER

    sig_headers = ['信号类型', '笔数', '胜率%', '平均收益%', '总收益%', '复制成功率%']
    for c, h in enumerate(sig_headers, 1):
        ws3.cell(row=3, column=c, value=h)
    _style_header(ws3, 3, len(sig_headers))

    by_sig = defaultdict(list)
    for t in trades:
        sig = t.get('sig_type', '未知')
        by_sig[sig].append(t)

    row = 4
    for sig in ['N字反包', '涨停双响炮', '缩量回踩不破']:
        items = by_sig.get(sig, [])
        if not items:
            continue
        n = len(items)
        w = sum(1 for t in items if t['is_win'])
        wr = w / max(n, 1) * 100
        avg_r = np.mean([t['net_ret'] for t in items])
        total_r = sum(t['net_ret'] for t in items)
        rep_ok = sum(1 for t in items if t.get('replica_success'))
        rep_rate = rep_ok / max(n, 1) * 100

        ws3.cell(row=row, column=1, value=sig)
        ws3.cell(row=row, column=2, value=n)
        ws3.cell(row=row, column=3, value=f'{wr:.1f}%')
        ws3.cell(row=row, column=4, value=f'{avg_r:+.2f}%')
        ws3.cell(row=row, column=5, value=f'{total_r:+.2f}%')
        ws3.cell(row=row, column=6, value=f'{rep_rate:.1f}%')
        row += 1

    _style_data(ws3, 4, row - 1, len(sig_headers))
    _auto_width(ws3)

    # ═══ Sheet 4: 按卖出原因统计 ═══
    ws4 = wb.create_sheet('卖出原因统计')
    ws4.merge_cells('A1:F1')
    ws4['A1'] = '按卖出原因统计'
    ws4['A1'].font = TITLE_FONT
    ws4['A1'].alignment = CENTER

    reason_headers = ['卖出原因', '笔数', '胜率%', '平均收益%', '总收益%', '平均持有天']
    for c, h in enumerate(reason_headers, 1):
        ws4.cell(row=3, column=c, value=h)
    _style_header(ws4, 3, len(reason_headers))

    by_reason = defaultdict(list)
    for t in trades:
        reason = t['exit_reason']
        if '止损' in reason:
            reason = '止损-5%'
        elif '移动止盈' in reason:
            reason = '移动止盈'
        elif '断板' in reason:
            reason = '涨停后断板离场'
        elif '到期' in reason:
            reason = '持有到期'
        elif '强平' in reason:
            reason = '回测到期强平'
        by_reason[reason].append(t)

    row = 4
    for reason in sorted(by_reason.keys(), key=lambda x: -len(by_reason[x])):
        items = by_reason[reason]
        n = len(items)
        w = sum(1 for t in items if t['is_win'])
        wr = w / max(n, 1) * 100
        avg_r = np.mean([t['net_ret'] for t in items])
        total_r = sum(t['net_ret'] for t in items)
        avg_d = np.mean([t['days_held'] for t in items])

        ws4.cell(row=row, column=1, value=reason)
        ws4.cell(row=row, column=2, value=n)
        ws4.cell(row=row, column=3, value=f'{wr:.1f}%')
        ws4.cell(row=row, column=4, value=f'{avg_r:+.2f}%')
        ws4.cell(row=row, column=5, value=f'{total_r:+.2f}%')
        ws4.cell(row=row, column=6, value=f'{avg_d:.1f}')
        row += 1

    _style_data(ws4, 4, row - 1, len(reason_headers))
    _auto_width(ws4)

    # ═══ Sheet 5: 月度统计 ═══
    ws5 = wb.create_sheet('月度统计')
    ws5.merge_cells('A1:I1')
    ws5['A1'] = '按月统计（组合层级）'
    ws5['A1'].font = TITLE_FONT
    ws5['A1'].alignment = CENTER

    month_headers = ['月份', '笔数', '胜', '负', '胜率%', '平均收益%',
                     '组合月收益%', '月末市值(万)', '累计收益%']
    for c, h in enumerate(month_headers, 1):
        ws5.cell(row=3, column=c, value=h)
    _style_header(ws5, 3, len(month_headers))

    init_cap = results.get('initial_capital', 1000000.0)
    pv = results.get('portfolio_values', [])
    monthly_pv = {}
    for dt, tv, _, _ in pv:
        mk = dt.strftime('%Y-%m')
        if mk not in monthly_pv:
            monthly_pv[mk] = []
        monthly_pv[mk].append((dt, tv))

    by_month = defaultdict(list)
    for t in trades:
        month_key = t['sell_date'][:7]
        by_month[month_key].append(t)

    row = 4
    prev_month_end_val = init_cap

    for month_key in sorted(set(list(by_month.keys()) + list(monthly_pv.keys()))):
        items = by_month.get(month_key, [])
        pvs = monthly_pv.get(month_key, [])
        n = len(items)
        w = sum(1 for t in items if t['is_win']) if n > 0 else 0
        l = n - w
        wr = w / max(n, 1) * 100
        avg_r = np.mean([t['net_ret'] for t in items]) if n > 0 else 0

        if pvs:
            month_end_val = pvs[-1][1]
        else:
            month_end_val = prev_month_end_val
        month_ret = (month_end_val / prev_month_end_val - 1) * 100 if prev_month_end_val > 0 else 0
        cum_ret = (month_end_val / init_cap - 1) * 100

        ws5.cell(row=row, column=1, value=month_key)
        ws5.cell(row=row, column=2, value=n if n > 0 else 0)
        ws5.cell(row=row, column=3, value=w)
        ws5.cell(row=row, column=4, value=l)
        ws5.cell(row=row, column=5, value=f'{wr:.1f}%' if n > 0 else '-')
        ws5.cell(row=row, column=6, value=f'{avg_r:+.2f}%' if n > 0 else '-')
        ws5.cell(row=row, column=7, value=f'{month_ret:+.2f}%')
        ws5.cell(row=row, column=8, value=f'{month_end_val/1e4:.1f}')
        ws5.cell(row=row, column=9, value=f'{cum_ret:.1f}%')

        if month_ret > 0:
            ws5.cell(row=row, column=7).fill = WIN_FILL
        else:
            ws5.cell(row=row, column=7).fill = LOSS_FILL
        if cum_ret > 0:
            ws5.cell(row=row, column=9).fill = WIN_FILL
        else:
            ws5.cell(row=row, column=9).fill = LOSS_FILL

        prev_month_end_val = month_end_val
        row += 1

    _style_data(ws5, 4, row - 1, len(month_headers))
    _auto_width(ws5)

    # ═══ Sheet 6: 日度日志 ═══
    ws6 = wb.create_sheet('日度日志')
    ws6.merge_cells('A1:E1')
    ws6['A1'] = '日度持仓变化'
    ws6['A1'].font = TITLE_FONT
    ws6['A1'].alignment = CENTER

    daily_headers = ['日期', '累计信号', '买入', '卖出', '持仓数']
    for c, h in enumerate(daily_headers, 1):
        ws6.cell(row=3, column=c, value=h)
    _style_header(ws6, 3, len(daily_headers))

    for i, d in enumerate(daily):
        r = i + 4
        ws6.cell(row=r, column=1, value=d['date'])
        ws6.cell(row=r, column=2, value=d['signals'])
        ws6.cell(row=r, column=3, value=d['buys'])
        ws6.cell(row=r, column=4, value=d['sells'])
        ws6.cell(row=r, column=5, value=d['holdings'])

    if daily:
        _style_data(ws6, 4, 3 + len(daily), len(daily_headers))
    _auto_width(ws6)

    # ═══ Sheet 7: 跳过记录 ═══
    if skipped:
        ws7 = wb.create_sheet('跳过记录')
        ws7.merge_cells('A1:F1')
        ws7['A1'] = '竞价确认/开盘跌停 跳过的买入信号'
        ws7['A1'].font = TITLE_FONT
        ws7['A1'].alignment = CENTER

        skip_headers = ['代码', '名称', '信号日', '跳过原因', '信号类型']
        for c, h in enumerate(skip_headers, 1):
            ws7.cell(row=3, column=c, value=h)
        _style_header(ws7, 3, len(skip_headers))

        for i, s in enumerate(skipped):
            r = i + 4
            ws7.cell(row=r, column=1, value=s['code'])
            ws7.cell(row=r, column=2, value=s['name'])
            ws7.cell(row=r, column=3, value=s['signal_date'])
            ws7.cell(row=r, column=4, value=s['skip_reason'])
            ws7.cell(row=r, column=5, value=s.get('sig_type', ''))
        _style_data(ws7, 4, 3 + len(skipped), len(skip_headers))
        _auto_width(ws7)

    wb.save(output_path)
    print(f'\n[OK] Excel saved: {output_path}')


# ═══════════════════════════════════════════════════════════════════════════
# Main
# ═══════════════════════════════════════════════════════════════════════════

if __name__ == '__main__':
    import argparse
    p = argparse.ArgumentParser(description='涨停复制战法回测')
    p.add_argument('--start', type=str, default=None, help='开始日期 YYYY-MM-DD (默认一年前)')
    p.add_argument('--end', type=str, default=None, help='结束日期 YYYY-MM-DD (默认昨天)')
    p.add_argument('--output', type=str, default=None, help='输出 xlsx 路径')
    p.add_argument('--only-ssxp', action='store_true',
                   help='仅做涨停双响炮（禁用 N字反包 + 缩量回踩不破）')
    args = p.parse_args()

    start_date = datetime.strptime(args.start, '%Y-%m-%d').date() if args.start else None
    end_date = datetime.strptime(args.end, '%Y-%m-%d').date() if args.end else None
    output_path = args.output or os.path.join(OUTPUT_DIR, 'zt_replica_backtest_result.xlsx')

    bt = ZTReplicaBacktest(only_double_cannon=args.only_ssxp)
    results = bt.run(start_date=start_date, end_date=end_date)

    if results:
        trades = results['trades']
        wins = sum(1 for t in trades if t['is_win'])
        rets = [t['net_ret'] for t in trades]
        print(f'\n[对比] {"双响炮only" if args.only_ssxp else "全形态"} '
              f'信号{results.get("total_signals",0)} 买入{results.get("total_buys",0)} '
              f'交易{len(trades)} 胜{wins} 胜率{wins/max(len(trades),1)*100:.1f}% '
              f'笔均{np.mean(rets):+.2f}% 累计{results.get("cumulative_return",0):+.2f}% '
              f'回撤{results.get("max_drawdown",0):.2f}%')
        export_xlsx(results, output_path)
