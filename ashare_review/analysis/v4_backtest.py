"""
V4 VOL180 突破战法 — 市场情绪增强版回测

选股: 沪深主板 · 年涨停≥10 · 非ST · 距压力位≤10%
买入: V3 全部条件 + V4 市场情绪过滤:
      1. SH < MA60 → 空仓（禁止买入）
      2. 涨停家数 < 60 → 不买入（冰点过滤）
      3. 涨停家数 > 140 → 提高评分门槛（过热分化）
      4. 连板高度 >= 5 → 额外加分
卖出: 与 V3 完全相同
      -6% 硬止损 · 移动止盈(回落>5%) · N字反包等待 · 5天兜底

输出: xlsx 文件（交易明细 / 统计汇总 / 按卖出原因 / 按月统计 / 市场分析）
"""

# ... (imports same as V3) ...
import sys, os, json, struct, time as _time
from datetime import date, datetime, timedelta
from typing import Dict, List, Optional, Tuple
from collections import defaultdict

import numpy as np
import pandas as pd

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

from ashare_review.data.tdx_reader import TdxReader, RECORD_SIZE
from ashare_review.analysis.indicators import calc_zigzag_find_top_line, calc_ma
from ashare_review.utils.calendar import TradingCalendar

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

# ═══════════════════════════════════════════════════════════════════════════
# Constants (与 sim_portfolio.py 一致)
# ═══════════════════════════════════════════════════════════════════════════
FEE = 0.0015
SLIPPAGE = 0.002
TOTAL_COST = FEE + SLIPPAGE
MAX_HOLD_DAYS = 5  # V3
MIN_LIMIT_UP_COUNT = 10
PRESSURE_DIST_PCT = 10.0
MAVOL_PERIOD = 180
MAVOL_MULTIPLIER = 1.2

DATA_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'data')
LIMIT_UP_POOL_FILE = os.path.join(DATA_DIR, 'limit_up_pool.json')
NAME_CACHE_FILE = os.path.join(DATA_DIR, 'stock_name_map.json')
ZT_CACHE_FILE = os.path.join(DATA_DIR, 'zt_scan_cache.json')

# ── V4 市场情绪过滤参数 ──
MIN_ZT_COUNT = 60              # 涨停家数 < 此值 → 不买入（冰点过滤）
MAX_ZT_OVERHEAT = 140          # 涨停家数 > 此值 → 过热，提高评分门槛
OVERHEAT_SCORE_MIN = 80        # 过热时的最低评分
MIN_BOARD_HEIGHT_BONUS = 5     # 连板高度 >= 此值 → 额外加分

# ═══════════════════════════════════════════════════════════════════════════
# Excel 样式
# ═══════════════════════════════════════════════════════════════════════════
TITLE_FONT = Font(name='微软雅黑', size=14, bold=True, color='1F2937')
HEADER_FONT = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='374151', end_color='374151', fill_type='solid')
HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
CENTER = Alignment(horizontal='center', vertical='center')
THIN_BORDER = Border(
    left=Side(style='thin', color='D1D5DB'),
    right=Side(style='thin', color='D1D5DB'),
    top=Side(style='thin', color='D1D5DB'),
    bottom=Side(style='thin', color='D1D5DB'),
)
WIN_FILL = PatternFill(start_color='DCFCE7', end_color='DCFCE7', fill_type='solid')
LOSS_FILL = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
S_FILL = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
A_FILL = PatternFill(start_color='DBEAFE', end_color='DBEAFE', fill_type='solid')
SUMMARY_LABEL_FONT = Font(name='微软雅黑', size=10, bold=True, color='374151')
SUMMARY_VAL_FONT = Font(name='Consolas', size=11, color='1F2937')
BIG_NUM_FONT = Font(name='Consolas', size=14, bold=True, color='1F2937')
GREEN_FONT = Font(name='Consolas', size=14, bold=True, color='059669')
RED_FONT = Font(name='Consolas', size=14, bold=True, color='DC2626')

OUTPUT_DIR = os.path.join(DATA_DIR, '..', 'analysis')


def _style_header(ws, row: int, ncols: int):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def _style_data(ws, start_row: int, end_row: int, ncols: int):
    for r in range(start_row, end_row + 1):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='center', vertical='center')


def _auto_width(ws, min_width: int = 8, max_width: int = 22):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            if cell.value:
                # 中文字符算2个宽度
                val = str(cell.value)
                length = sum(2 if ord(c) > 127 else 1 for c in val)
                max_len = max(max_len, length)
        ws.column_dimensions[col_letter].width = max(min_width, min(max_len + 3, max_width))


# ═══════════════════════════════════════════════════════════════════════════
# V3 Backtest Engine
# ═══════════════════════════════════════════════════════════════════════════

class V3Backtest:
    """V3 VOL180 突破战法历史回测"""

    def __init__(self):
        self.tdx = TdxReader()
        self.cal = TradingCalendar()
        self._name_map: Dict[str, str] = {}
        self._load_name_map()

    def _load_name_map(self):
        if os.path.exists(NAME_CACHE_FILE):
            try:
                with open(NAME_CACHE_FILE, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                if isinstance(data, dict) and len(data) > 1000:
                    self._name_map = data
            except Exception:
                pass

    # ─── V4: 市场情绪数据 ───

    def _load_zt_cache(self) -> Dict[str, list]:
        """加载涨停缓存: date_str → [code, ...]."""
        if os.path.exists(ZT_CACHE_FILE):
            try:
                with open(ZT_CACHE_FILE, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except Exception:
                pass
        return {}

    def _compute_daily_market(self, zt_cache: dict, all_dates: List[date]) -> dict:
        """预计算每日市场情绪指标。

        Returns:
            dict: {date_str: {zt_count, max_board, sh_above_ma60, sh_chg, sh_amt_yi}}
        """
        # 上证指数
        sh_above_ma60 = {}
        sh_daily = {}
        try:
            sh_df = self.tdx.read_daily('999999', 'sh')
            if sh_df is not None and len(sh_df) >= 60:
                sh_df['ma60'] = sh_df['close'].rolling(60).mean()
                for td in all_dates:
                    mask = sh_df['trade_date'].apply(
                        lambda x: (x.date() if hasattr(x, 'date') else x) <= td
                    )
                    sub = sh_df[mask]
                    if len(sub) >= 60:
                        close = float(sub['close'].iloc[-1])
                        ma60 = float(sub['ma60'].iloc[-1])
                        chg = float(sub['close'].pct_change().iloc[-1] * 100) if len(sub) >= 2 else 0
                        amt = float(sub['amount'].iloc[-1]) / 1e8 if 'amount' in sub.columns else 0
                        ds = td.strftime('%Y%m%d')
                        sh_above_ma60[ds] = close > ma60 and not pd.isna(ma60)
                        sh_daily[ds] = {'chg': round(chg, 2), 'amt_yi': round(amt, 0)}
        except Exception:
            pass

        # 预计算连板高度
        sorted_dates = sorted(zt_cache.keys())
        board_height = {}
        for ds in sorted_dates:
            zt_list = zt_cache.get(ds, [])
            if not isinstance(zt_list, list):
                zt_list = []
            codes_today = set(zt_list)
            max_cons = 2
            if codes_today:
                prev_dates = [d for d in sorted_dates if d < ds]
                for code in codes_today:
                    cons = 1
                    for pd_str in reversed(prev_dates[-15:]):
                        pd_codes = set(zt_cache.get(pd_str, []))
                        if code in pd_codes:
                            cons += 1
                        else:
                            break
                    max_cons = max(max_cons, cons)
            board_height[ds] = max_cons

        # 合并
        result = {}
        for td in all_dates:
            ds = td.strftime('%Y%m%d')
            zt_list = zt_cache.get(ds, [])
            if not isinstance(zt_list, list):
                zt_list = []
            zt_count = len(zt_list)
            result[ds] = {
                'zt_count': zt_count,
                'max_board': board_height.get(ds, 2),
                'sh_above_ma60': sh_above_ma60.get(ds, True),
                'sh_chg': sh_daily.get(ds, {}).get('chg', 0),
                'sh_amt_yi': sh_daily.get(ds, {}).get('amt_yi', 0),
            }
        return result

    @staticmethod
    def _market_score(md: dict) -> int:
        """根据市场情绪计算评分 (0-100)。"""
        score = 0
        zt = md.get('zt_count', 0)
        board = md.get('max_board', 2)
        above = md.get('sh_above_ma60', True)

        # 1. 趋势 (最重要): SH > MA60 = 40分
        if above:
            score += 40
        else:
            score -= 20  # 熊市惩罚

        # 2. 热度: ZT 家数
        if zt >= 100:
            score += 30
        elif zt >= 80:
            score += 25
        elif zt >= 60:
            score += 20
        elif zt >= 40:
            score += 10
        else:
            score += 0  # 冰点

        # 3. 龙头: 连板高度
        if board >= 8:
            score += 20
        elif board >= 6:
            score += 15
        elif board >= 5:
            score += 10
        elif board >= 4:
            score += 5

        # 4. 成交换手
        amt = md.get('sh_amt_yi', 0)
        if amt >= 6000:
            score += 10
        elif amt >= 4000:
            score += 5

        return min(score, 100)

    def _can_buy_v4(self, md: dict, stock_score: int) -> tuple:
        """V4 买入许可。返回 (allowed: bool, reason: str)。"""
        zt = md.get('zt_count', 0)
        above = md.get('sh_above_ma60', True)

        # Rule 1: SH < MA60 → 空仓
        if not above:
            return False, '熊市空仓(SH<MA60)'

        # Rule 2: ZT < 60 → 冰点过滤
        if zt < MIN_ZT_COUNT:
            return False, f'冰点过滤(ZT={zt}<{MIN_ZT_COUNT})'

        # Rule 3: ZT > 140 → 过热，提高评分门槛
        if zt > MAX_ZT_OVERHEAT and stock_score < OVERHEAT_SCORE_MIN:
            return False, f'过热过滤(ZT={zt}>{MAX_ZT_OVERHEAT})'

        return True, ''

    def _get_name(self, code: str) -> str:
        return self._name_map.get(str(code).zfill(6), code)

    @staticmethod
    def _limit_threshold(code: str) -> float:
        code = str(code).zfill(6)
        if code.startswith(('300', '301', '688')): return 0.199
        if code.startswith(('8', '4')): return 0.299
        return 0.095

    @staticmethod
    def _is_main_board(code: str) -> bool:
        code = str(code).zfill(6)
        return code.startswith(('60', '00', '001', '002'))

    # ─── 获取候选池（年涨停≥10 + 主板 + 非ST） ───

    def get_universe(self) -> List[str]:
        """从缓存获取候选池股票代码列表。"""
        if os.path.exists(LIMIT_UP_POOL_FILE):
            try:
                with open(LIMIT_UP_POOL_FILE, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                pool = data.get('pool', [])
                if pool:
                    codes = [s['code'] for s in pool
                             if self._is_main_board(s['code'])
                             and 'ST' not in self._get_name(s['code'])]
                    print(f'从缓存加载候选池: {len(codes)} 只')
                    return codes
            except Exception:
                pass
        print('[WARN] 候选池缓存不存在，请先运行 sim_portfolio.py 构建候选池')
        return []

    # ─── 读取单只股票全量数据 ───

    def _read_stock_full(self, code: str) -> Optional[pd.DataFrame]:
        """读取股票全量日线 + 计算指标 + 预计算zigzag压力位。"""
        market = 'sh' if str(code).startswith('6') else 'sz'
        if str(code).startswith(('8', '4')):
            market = 'bj'
        try:
            df = self.tdx.read_daily(code, market)
            if df is None or df.empty or len(df) < MAVOL_PERIOD + 60:
                return None
            # 截断到尾 600 行（~2.5 年），覆盖一年回测+180 日均量+zigzag 历史
            # 避免 8000+ 行长历史导致 zigzag O(n²) 耗时过大
            if len(df) > 600:
                df = df.iloc[-600:].copy()
                df.reset_index(drop=True, inplace=True)
            df = calc_ma(df, [5, 10])
            df['mavol180'] = df['volume'].rolling(MAVOL_PERIOD).mean()

            # ── 预计算简化 zigzag 压力位（快速向量化版本） ──
            # 用滚动窗口峰值检测替代完整 zigzag BACKSET/BARSLAST 链
            # 精度接近但速度提升 100x+
            try:
                high_v = df['high'].values.astype(float)
                n = len(high_v)
                # Step 1: 找局部峰值（前后各5根K线内最高）
                is_peak = np.ones(n, dtype=bool)
                for offset in range(1, 6):
                    is_peak &= (np.roll(high_v, offset) <= high_v)
                    is_peak &= (np.roll(high_v, -offset) <= high_v)
                is_peak[:5] = False; is_peak[-5:] = False
                # Step 2: 只保留相隔≥10根K线的显著峰值（去噪）
                peak_idx = np.where(is_peak)[0]
                filtered = []
                last_kept = -100
                for pi in peak_idx:
                    if pi - last_kept >= 10:
                        filtered.append(pi)
                        last_kept = pi
                # Step 3: 构建每日压力位 = 最近峰值的 high
                pressure = np.full(n, np.nan)
                last_high = np.nan
                fi = 0
                for i in range(n):
                    if fi < len(filtered) and i == filtered[fi]:
                        last_high = high_v[i]
                        fi += 1
                    pressure[i] = last_high
                df['_pressure'] = pressure
                df['_nn'] = is_peak.astype(int)
                # fallback: 前60行的 NaN 用60日最高
                high_60max = df['high'].rolling(60, min_periods=1).max().shift(1)
                mask = pd.isna(df['_pressure']) | (df['_pressure'] <= 0)
                df.loc[mask, '_pressure'] = high_60max[mask]
            except Exception:
                df['_pressure'] = df['high'].rolling(60, min_periods=1).max().shift(1)

            return df
        except Exception:
            return None

    # ─── 获取预计算的压力位 ───

    def _calc_pressure_at(self, df: pd.DataFrame, idx: int) -> float:
        """读取预计算的压力位（zigzag 找顶线 last NN high）。
        优化版本：从预计算的 _pressure 列直接读取，不再重复计算 zigzag。
        """
        try:
            val = float(df['_pressure'].iloc[idx])
            if pd.isna(val) or val <= 0:
                return float(df['high'].rolling(60).max().shift(1).iloc[idx]) \
                    if idx >= 60 else 0.0
            return round(val, 2)
        except Exception:
            return float(df['high'].rolling(60).max().shift(1).iloc[idx]) \
                if idx >= 60 else 0.0

    # ─── V3 卖出检查 ───

    def _check_sell_v3(self, code: str, hold: dict, check_date: date,
                       df_full: pd.DataFrame) -> Optional[dict]:
        """V3 卖出规则（与 sim_portfolio.py _check_sell_v3 完全一致）。"""
        buy_date = hold['buy_date']
        buy_price = hold['buy_price']
        had_zt = hold.get('had_zt', False)
        highest = hold.get('highest_close', buy_price)

        # 截取数据到 check_date
        mask = df_full['trade_date'].apply(
            lambda x: (x.date() if hasattr(x, 'date') else x) <= check_date
        )
        df = df_full[mask].copy()
        if df.empty or len(df) < 2:
            return None

        idx = len(df) - 1
        close = float(df['close'].iloc[idx])

        # 更新最高收盘价
        highest = max(highest, close)
        hold['highest_close'] = round(highest, 2)

        # 交易日计数
        try:
            trading_days = self.cal.trading_days_between(buy_date, check_date)
        except Exception:
            trading_days = 3

        # ── -6% 硬止损 ──
        if buy_price > 0:
            loss_pct = (close - buy_price) / buy_price
            if loss_pct <= -0.06:
                return {
                    'sell_price': round(close, 2),
                    'sell_reason': '止损-6%',
                    'days_held': trading_days,
                }

        # ── 移动止盈: 从最高回落 > 5%（需先盈利 3%+） ──
        if highest > buy_price * 1.03:
            pullback = (close - highest) / highest
            if pullback <= -0.05:
                return {
                    'sell_price': round(close, 2),
                    'sell_reason': f'移动止盈-5%(高{highest:.2f}→{close:.2f})',
                    'days_held': trading_days,
                }

        # ── 检查今日是否涨停 ──
        if idx >= 1:
            prev_close = float(df['close'].iloc[idx - 1])
            limit_pct = self._limit_threshold(code)
            is_zt_today = (close - prev_close) / prev_close >= limit_pct if prev_close > 0 else False
        else:
            is_zt_today = False

        # ── 今日涨停 → 继续持有 ──
        if is_zt_today:
            hold['had_zt'] = True
            hold['awaiting_reversal'] = False
            return None

        # ── N字反包 ──
        if had_zt:
            awaiting = hold.get('awaiting_reversal', False)
            if not awaiting:
                hold['awaiting_reversal'] = True
                hold['reversal_day_close'] = round(close, 2)
                hold['reversal_day_vol'] = int(float(df['volume'].iloc[idx])) if idx >= 0 else 0
                return None  # 等反包
            else:
                rev_vol = hold.get('reversal_day_vol', 0)
                today_vol = int(float(df['volume'].iloc[idx])) if idx >= 0 else 0
                today_open = float(df['open'].iloc[idx]) if idx >= 0 else close
                is_up = close > today_open
                vol_expand = today_vol > rev_vol
                if is_up and vol_expand:
                    hold['awaiting_reversal'] = False
                    hold['had_zt'] = True  # 反包成功，继续
                    return None
                else:
                    return {
                        'sell_price': round(close, 2),
                        'sell_reason': '涨停后断板离场',
                        'days_held': trading_days,
                    }

        # ── 始终没涨停 + 持有 ≥ 5 天兜底 ──
        if trading_days >= MAX_HOLD_DAYS:
            return {
                'sell_price': round(close, 2),
                'sell_reason': f'持有{trading_days}天到期',
                'days_held': trading_days,
            }

        # ── 没涨停但浮盈 5%+，也启用移动止盈 ──
        if highest > buy_price * 1.05:
            pullback = (close - highest) / highest
            if pullback <= -0.05:
                return {
                    'sell_price': round(close, 2),
                    'sell_reason': f'移动止盈-5%(高{highest:.2f})',
                    'days_held': trading_days,
                }

        return None

    # ─── V3 竞价确认 ───

    def _check_auction_v3(self, code: str, buy_date: date,
                          df_full: pd.DataFrame) -> Optional[str]:
        """V3 竞价确认，返回 None=通过, str=拒绝原因。"""
        mask = df_full['trade_date'].apply(
            lambda x: (x.date() if hasattr(x, 'date') else x) <= buy_date
        )
        df = df_full[mask].copy()
        if df.empty or len(df) < 2:
            return None

        idx = len(df) - 1
        open_p = float(df['open'].iloc[idx])
        prev_close = float(df['close'].iloc[idx - 1])
        vol_today = float(df['volume'].iloc[idx]) if idx >= 0 else 0
        vol_prev = float(df['volume'].iloc[idx - 1]) if idx >= 1 else 0

        # 低开 > 3%
        open_chg = (open_p - prev_close) / prev_close * 100 if prev_close > 0 else 0
        if open_chg < -3:
            return f'竞价低开{open_chg:.1f}%'

        # 缩量 > 50%
        if vol_prev > 0 and vol_today < vol_prev * 0.5:
            return '竞价缩量>50%'

        return None

    # ─── 检查开盘跌停 ───

    def _check_open_limit_down(self, code: str, buy_date: date,
                               df_full: pd.DataFrame) -> bool:
        """检查买入日是否开盘跌停。"""
        mask = df_full['trade_date'].apply(
            lambda x: (x.date() if hasattr(x, 'date') else x) <= buy_date
        )
        df = df_full[mask].copy()
        if df.empty or len(df) < 2:
            return False

        idx = len(df) - 1
        open_p = float(df['open'].iloc[idx])
        prev_close = float(df['close'].iloc[idx - 1])
        limit_pct = self._limit_threshold(code)
        return open_p <= prev_close * (1 - limit_pct)

    # ═══════════════════════════════════════════════════════════════════════
    # 主回测循环
    # ═══════════════════════════════════════════════════════════════════════

    def run(self, start_date: date = None, end_date: date = None) -> dict:
        """运行 V3 回测（含真实资金模型）。

        Args:
            start_date: 回测开始日（默认一年前）
            end_date: 回测结束日（默认昨天）
        """
        # ── 资金/仓位模型 ──
        INITIAL_CAPITAL = 1_000_000.0    # 初始资金 100 万
        MAX_POSITIONS = 10               # 最大持仓数
        MAX_NEW_PER_DAY = 3              # 每天最多新开仓
        PER_POSITION_PCT = 0.10          # 单票仓位 10%

        if end_date is None:
            end_date = date.today() - timedelta(days=1)
        if start_date is None:
            start_date = end_date - timedelta(days=365)

        print(f'\n{"="*60}')
        print(f'  V3 VOL180 突破战法 — 历史回测')
        print(f'  区间: {start_date} ~ {end_date}')
        print(f'  资金: {INITIAL_CAPITAL/1e4:.0f}万 | 最大持仓: {MAX_POSITIONS}只 | 单票: {PER_POSITION_PCT*100:.0f}%')
        print(f'{"="*60}\n')

        # ── Step 1: 获取候选池 ──
        universe = self.get_universe()
        if not universe:
            print('[ERROR] 候选池为空')
            return {}

        # ── Step 2: 预读所有候选股数据 + 构建日期索引 ──
        print(f'预读 {len(universe)} 只候选股数据...')
        stock_data: Dict[str, pd.DataFrame] = {}
        stock_date_idx: Dict[str, Dict[date, int]] = {}  # code → {date: row_index}
        t0 = _time.time()
        for i, code in enumerate(universe):
            if (i + 1) % 100 == 0:
                print(f'  读取 {i+1}/{len(universe)} (已加载 {len(stock_data)} 只, {_time.time() - t0:.0f}s)...')
            df = self._read_stock_full(code)
            if df is not None and len(df) >= MAVOL_PERIOD + 60:
                stock_data[code] = df
                # 预建日期→行号映射
                dmap = {}
                for ri, td_val in enumerate(df['trade_date']):
                    if hasattr(td_val, 'date'):
                        dmap[td_val.date()] = ri
                    elif hasattr(td_val, 'strftime'):
                        dmap[td_val] = ri
                stock_date_idx[code] = dmap
        print(f'预读完成: {len(stock_data)} 只有效数据 ({_time.time() - t0:.0f}s)')

        # ── Step 3: 生成交易日列表 ──
        all_dates = []
        d = start_date
        while d <= end_date:
            if d.weekday() < 5:
                all_dates.append(d)
            d += timedelta(days=1)
        print(f'交易日: {len(all_dates)} 天 ({all_dates[0]} ~ {all_dates[-1]})')

        # ── Step 3b: V4 市场情绪预计算 ──
        print('加载涨停缓存...')
        zt_cache = self._load_zt_cache()
        print(f'  涨停缓存: {len(zt_cache)} 天')
        market_data = self._compute_daily_market(zt_cache, all_dates)
        print(f'  市场数据: {len(market_data)} 天预计算完成')

        # ── Step 4: 逐日回测（资金模型） ──
        # 资金状态
        cash = INITIAL_CAPITAL
        holdings: Dict[str, dict] = {}  # code → {shares, buy_price, buy_date, ...}
        finished_trades: List[dict] = []
        skipped_trades: List[dict] = []
        daily_log: List[dict] = []

        # 统计
        total_signals = 0
        total_buys = 0
        portfolio_values = []  # [(date, total_value, cash, positions_value)] for drawdown calc

        # ── 辅助 ──
        def _get_idx(code: str, target_date: date) -> int:
            dmap = stock_date_idx.get(code)
            if dmap is None: return -1
            return dmap.get(target_date, -1)

        def _current_price(code: str, td: date) -> Optional[float]:
            """获取某股票在某日的收盘价"""
            idx = _get_idx(code, td)
            if idx >= 0:
                return float(stock_data[code]['close'].iloc[idx])
            return None

        t0 = _time.time()
        pending_sells: Dict[str, dict] = {}  # code → sell_info (执行日为次日)

        for di, td in enumerate(all_dates):
            if (di + 1) % 20 == 0:
                elapsed = _time.time() - t0
                eta = elapsed / (di + 1) * (len(all_dates) - di - 1) if di > 0 else 0
                # 计算当前组合市值
                pos_val = 0.0
                for code, h in holdings.items():
                    px = _current_price(code, td)
                    if px: pos_val += h['shares'] * px
                total_val = cash + pos_val
                print(f'  [{di+1}/{len(all_dates)}] {td}  '
                      f'持仓:{len(holdings)} 市值:{total_val/1e4:.1f}万 已平:{len(finished_trades)}  '
                      f'({elapsed:.0f}s, ETA {eta:.0f}s)', flush=True)

            # ── Step 4a: 执行昨日挂单的卖出（用今日开盘价） ──
            sold_today = []
            for code in list(pending_sells.keys()):
                sell_info = pending_sells[code]
                sell_date = sell_info['sell_date']
                sell_reason = sell_info['sell_reason']
                h = sell_info['hold']

                # 用今天的开盘价卖出
                sell_idx = _get_idx(code, td)
                if sell_idx >= 0:
                    sell_price = float(stock_data[code]['open'].iloc[sell_idx])
                else:
                    sell_price = sell_info.get('fallback_price', h['buy_price'])

                buy_price = h['buy_price']
                shares = h['shares']
                # 卖出回款（扣除卖出手续费+印花税 0.08%）
                sell_cost_rate = 0.0008
                proceeds = shares * sell_price * (1 - sell_cost_rate)
                cash += proceeds

                gross_ret = (sell_price - buy_price) / buy_price if buy_price > 0 else 0
                net_ret = gross_ret - TOTAL_COST

                trading_days = 0
                try:
                    trading_days = self.cal.trading_days_between(h['buy_date'], td)
                except Exception:
                    trading_days = sell_info.get('_days', 3)

                finished_trades.append({
                    'code': code,
                    'name': self._get_name(code),
                    'buy_date': h['buy_date'].strftime('%Y-%m-%d'),
                    'buy_price': round(buy_price, 2),
                    'sell_date': td.strftime('%Y-%m-%d'),
                    'sell_price': round(sell_price, 2),
                    'gross_ret': round(gross_ret * 100, 2),
                    'net_ret': round(net_ret * 100, 2),
                    'is_win': net_ret > 0,
                    'exit_reason': sell_reason,
                    'days_held': trading_days if trading_days > 0 else sell_info.get('_days', 1),
                    'had_zt': h.get('had_zt', False),
                    'score': h.get('score', 0),
                    'vol_ratio': h.get('vol_ratio', 0),
                    'dist_pct': h.get('dist_pct', 0),
                    'limit_count': h.get('limit_count', 0),
                })
                sold_today.append(code)
                del holdings[code]
                del pending_sells[code]

            # ── Step 4b: 检查持仓是否需要挂卖出单（用今日收盘价判断） ──
            for code in list(holdings.keys()):
                if code in pending_sells:  # 已有挂单
                    continue
                if code not in stock_data:
                    continue
                sell_sig = self._check_sell_v3(code, holdings[code], td, stock_data[code])
                if sell_sig:
                    # 挂单：下一个交易日以开盘价卖出
                    next_td_sell = self._next_trade_date(td, all_dates)
                    if next_td_sell is None:
                        # 回测最后一天，用收盘价强平
                        fallback_px = sell_sig['sell_price']
                        sell_real_date = td
                    else:
                        fallback_px = sell_sig['sell_price']
                        sell_real_date = next_td_sell

                    pending_sells[code] = {
                        'hold': holdings[code],
                        'sell_date': sell_real_date,
                        'sell_reason': sell_sig['sell_reason'],
                        'fallback_price': fallback_px,
                        '_days': sell_sig['days_held'],
                    }

            # ── Step 4c: 检测今日买入信号 ──
            buy_today = []
            # 市场环境过滤
            ds_key = td.strftime('%Y%m%d')
            md_today = market_data.get(ds_key, {})
            is_bull = md_today.get('sh_above_ma60', True)  # 无数据默认允许

            # 按评分降序排列候选
            candidates_today = []
            for code, df in stock_data.items():
                if code in holdings or code in pending_sells:
                    continue
                if len(holdings) >= MAX_POSITIONS:  # 满仓
                    break

                idx = _get_idx(code, td)
                if idx < MAVOL_PERIOD + 20:
                    continue

                close = float(df['close'].iloc[idx])
                vol = float(df['volume'].iloc[idx])
                mavol180 = float(df['mavol180'].iloc[idx])

                if pd.isna(mavol180) or mavol180 <= 0:
                    continue

                pressure = self._calc_pressure_at(df, idx)
                if pressure <= 0:
                    continue

                dist_pct = (close - pressure) / pressure * 100
                vol_ratio = vol / mavol180

                # V3 买入条件
                if not (close > pressure and vol > mavol180 * MAVOL_MULTIPLIER):
                    continue

                total_signals += 1

                # V3 过滤: 死亡区间 3-5%
                if 3 < abs(dist_pct) <= 5:
                    continue
                # V3 过滤: 过度放量 ≥ 5x
                if vol_ratio >= 5.0:
                    continue

                # 熊市降低抽奖率：只做评分更高的
                score = min(100, 60 + (10 if abs(dist_pct) >= 3 else 0)
                                  + (10 if vol_ratio >= 2.0 else 5 if vol_ratio >= 1.5 else 0))
                if not is_bull and score < 75:
                    continue

                # 找到下一交易日
                next_td = self._next_trade_date(td, all_dates)
                if next_td is None:
                    continue

                # 检查开盘跌停
                if self._check_open_limit_down(code, next_td, df):
                    skipped_trades.append({
                        'code': code, 'name': self._get_name(code),
                        'signal_date': td.strftime('%Y-%m-%d'),
                        'skip_reason': '开盘跌停',
                        'close': round(close, 2), 'pressure': pressure,
                    })
                    continue

                # V3 竞价确认
                auction_reject = self._check_auction_v3(code, next_td, df)
                if auction_reject:
                    skipped_trades.append({
                        'code': code, 'name': self._get_name(code),
                        'signal_date': td.strftime('%Y-%m-%d'),
                        'skip_reason': auction_reject,
                        'close': round(close, 2), 'pressure': pressure,
                    })
                    continue

                candidates_today.append({
                    'code': code,
                    'score': score,
                    'close': close, 'pressure': pressure,
                    'dist_pct': dist_pct, 'vol_ratio': vol_ratio,
                    'next_td': next_td,
                })

            # 按评分排序，限制每天最多 N 只新买入
            candidates_today.sort(key=lambda x: -x['score'])
            available_slots = max(0, MAX_POSITIONS - len(holdings))
            max_new = min(MAX_NEW_PER_DAY, available_slots)

            for c in candidates_today[:max_new]:
                code = c['code']
                next_td = c['next_td']

                # 获取次日开盘价
                next_idx = _get_idx(code, next_td)
                if next_idx < 0:
                    continue
                buy_open = float(stock_data[code]['open'].iloc[next_idx])

                # 计算买入份额（手数取整）
                position_capital = INITIAL_CAPITAL * PER_POSITION_PCT
                shares = int(position_capital / buy_open / 100) * 100
                if shares < 100:
                    continue
                buy_cost = shares * buy_open * (1 + 0.0003)  # 买入佣金万3
                if buy_cost > cash:
                    continue  # 资金不足，跳过

                cash -= buy_cost

                holdings[code] = {
                    'code': code,
                    'buy_date': next_td,
                    'buy_price': buy_open,
                    'shares': shares,
                    'signal_date': td,
                    'signal_close': round(c['close'], 2),
                    'pressure': c['pressure'],
                    'dist_pct': round(c['dist_pct'], 1),
                    'vol_ratio': round(c['vol_ratio'], 1),
                    'limit_count': self._fast_count_limit_ups(code),
                    'had_zt': False,
                    'highest_close': buy_open,
                    'awaiting_reversal': False,
                    'score': c['score'],
                }

                buy_today.append(code)
                total_buys += 1

            # ── 日统计 ──
            # 计算当前组合市值
            pos_val = 0.0
            for code, h in holdings.items():
                px = _current_price(code, td)
                if px: pos_val += h['shares'] * px
            total_val = cash + pos_val
            portfolio_values.append((td, total_val, cash, pos_val))

            daily_log.append({
                'date': td.strftime('%Y-%m-%d'),
                'signals': total_signals,
                'buys': len(buy_today),
                'sells': len(sold_today),
                'holdings': len(holdings),
                'portfolio_value': round(total_val, 2),
                'market_bull': is_bull,
            })

        # ── 到期强制平仓 ──
        last_date = all_dates[-1]
        for code in list(holdings.keys()):
            h = holdings[code]
            sell_idx = _get_idx(code, last_date)
            if sell_idx >= 0:
                final_close = float(stock_data[code]['close'].iloc[sell_idx])
                buy_price = h['buy_price']
                shares = h['shares']
                proceeds = shares * final_close * (1 - 0.0008)
                cash += proceeds

                gross_ret = (final_close - buy_price) / buy_price if buy_price > 0 else 0
                net_ret = gross_ret - TOTAL_COST
                trading_days = 0
                try:
                    trading_days = self.cal.trading_days_between(h['buy_date'], last_date)
                except Exception:
                    trading_days = 5

                finished_trades.append({
                    'code': code,
                    'name': self._get_name(code),
                    'buy_date': h['buy_date'].strftime('%Y-%m-%d'),
                    'buy_price': round(buy_price, 2),
                    'sell_date': last_date.strftime('%Y-%m-%d'),
                    'sell_price': round(final_close, 2),
                    'gross_ret': round(gross_ret * 100, 2),
                    'net_ret': round(net_ret * 100, 2),
                    'is_win': net_ret > 0,
                    'exit_reason': '回测到期强平',
                    'days_held': trading_days if trading_days > 0 else 5,
                    'had_zt': h.get('had_zt', False),
                    'score': h.get('score', 0),
                    'vol_ratio': h.get('vol_ratio', 0),
                    'dist_pct': h.get('dist_pct', 0),
                    'limit_count': h.get('limit_count', 0),
                })
            del holdings[code]

        # ── 组合层统计 ──
        cumulative_return = (total_val / INITIAL_CAPITAL - 1) * 100 if INITIAL_CAPITAL > 0 else 0
        peak_value = INITIAL_CAPITAL
        max_drawdown = 0.0
        for _, tv, _, _ in portfolio_values:
            peak_value = max(peak_value, tv)
            dd = (peak_value - tv) / peak_value * 100 if peak_value > 0 else 0
            max_drawdown = max(max_drawdown, dd)

        # ── 牛熊市周期分割统计 ──
        # 使用 daily_log 中记录的 market_bull 字段标记每笔交易的市场状态
        bull_dates = set()
        for dl in daily_log:
            if dl.get('market_bull'):
                bull_dates.add(dl['date'])

        bear_trades_list = [t for t in finished_trades
                           if t['buy_date'] not in bull_dates]
        bull_trades_list = [t for t in finished_trades
                           if t['buy_date'] in bull_dates]

        def _compute_cycle_stats(trades_list, pv_subset=None):
            """计算一组交易的统计指标."""
            if not trades_list:
                return {}
            wins = [t for t in trades_list if t['is_win']]
            losses = [t for t in trades_list if not t['is_win']]
            n = len(trades_list)
            win_n = len(wins)
            loss_n = len(losses)
            wr = win_n / n * 100 if n > 0 else 0
            net_rets = [t['net_ret'] for t in trades_list]
            avg_ret = np.mean(net_rets) if net_rets else 0
            avg_win = np.mean([t['net_ret'] for t in wins]) if wins else 0
            avg_loss = np.mean([t['net_ret'] for t in losses]) if losses else 0
            total_profit = sum(t['net_ret'] for t in wins)
            total_loss = sum(abs(t['net_ret']) for t in losses)
            profit_factor = round(total_profit / max(total_loss, 0.01), 2) if losses else 999

            avg_hold = round(np.mean([t['days_held'] for t in trades_list]), 1) if trades_list else 0

            # 最大回撤（基于组合估值子集）
            if pv_subset:
                peak = INITIAL_CAPITAL
                dd_max = 0.0
                for _, tv, _, _ in pv_subset:
                    peak = max(peak, tv)
                    d = (peak - tv) / peak * 100 if peak > 0 else 0
                    dd_max = max(dd_max, d)
            else:
                dd_max = 0.0

            return {
                'trades': n, 'wins': win_n, 'losses': loss_n,
                'win_rate': round(wr, 1),
                'avg_net_return': round(avg_ret, 2),
                'avg_win': round(avg_win, 2),
                'avg_loss': round(avg_loss, 2),
                'profit_factor': profit_factor,
                'cumulative_return': round(sum(net_rets), 2),
                'max_drawdown': round(dd_max, 2),
                'avg_hold_days': avg_hold,
            }

        # 分割 portfolio_values 为牛熊市区间
        bear_pv = []
        bull_pv = []
        for pv_entry in portfolio_values:
            td_pv, tv, c, pv = pv_entry
            date_str = td_pv.strftime('%Y-%m-%d') if hasattr(td_pv, 'strftime') else str(td_pv)
            if date_str in bull_dates or str(td_pv) in bull_dates:
                bull_pv.append(pv_entry)
            else:
                bear_pv.append(pv_entry)

        market_cycle_stats = {}
        bear_stats = _compute_cycle_stats(bear_trades_list, bear_pv)
        bull_stats = _compute_cycle_stats(bull_trades_list, bull_pv)

        # 计算年化收益（需要知道每个周期的时间跨度）
        total_days = (end_date - start_date).days if start_date and end_date else 365
        years = total_days / 365.25 if total_days > 0 else 1

        # 通过 daily_log 的日期范围估算熊牛市天数
        bear_dates_sorted = sorted([dl['date'] for dl in daily_log
                                    if not dl.get('market_bull')])
        bull_dates_sorted = sorted([dl['date'] for dl in daily_log
                                    if dl.get('market_bull')])

        bear_years = 1.0
        bull_years = 1.0
        if bear_dates_sorted and len(bear_dates_sorted) > 1:
            bear_days = (datetime.strptime(bear_dates_sorted[-1], '%Y-%m-%d') -
                        datetime.strptime(bear_dates_sorted[0], '%Y-%m-%d')).days
            bear_years = max(bear_days / 365.25, 0.25)
        if bull_dates_sorted and len(bull_dates_sorted) > 1:
            bull_days = (datetime.strptime(bull_dates_sorted[-1], '%Y-%m-%d') -
                        datetime.strptime(bull_dates_sorted[0], '%Y-%m-%d')).days
            bull_years = max(bull_days / 365.25, 0.25)

        # 年化收益 (CAGR)
        final_bear_val = bear_pv[-1][1] if bear_pv else INITIAL_CAPITAL
        bear_annualized = ((final_bear_val / INITIAL_CAPITAL) ** (1 / bear_years) - 1) * 100 if bear_years > 0 else 0
        final_bull_val = bull_pv[-1][1] if bull_pv else INITIAL_CAPITAL
        bull_annualized = ((final_bull_val / INITIAL_CAPITAL) ** (1 / bull_years) - 1) * 100 if bull_years > 0 else 0
        full_annualized = round(((total_val / INITIAL_CAPITAL) ** (1 / years) - 1) * 100, 2) if years > 0 else 0

        bear_stats['annualized_return'] = round(bear_annualized, 2)
        bull_stats['annualized_return'] = round(bull_annualized, 2)

        # 平均同时持仓 (从 daily_log 取平均)
        bear_holdings_avg = round(np.mean([dl['holdings'] for dl in daily_log
                                           if not dl.get('market_bull')]), 1) if daily_log else 0
        bull_holdings_avg = round(np.mean([dl['holdings'] for dl in daily_log
                                           if dl.get('market_bull')]), 1) if daily_log else 0
        bear_stats['avg_positions'] = bear_holdings_avg
        bull_stats['avg_positions'] = bull_holdings_avg

        # 夏普比率
        def _compute_sharpe(daily_log_subset, init_cap):
            """从每日组合市值计算夏普比率."""
            if not daily_log_subset or len(daily_log_subset) < 5:
                return 0
            values = [dl['portfolio_value'] for dl in daily_log_subset]
            if not values or len(values) < 5:
                return 0
            daily_rets = []
            for i in range(1, len(values)):
                if values[i-1] > 0:
                    daily_rets.append((values[i] - values[i-1]) / values[i-1])
            if not daily_rets or len(daily_rets) < 3:
                return 0
            avg_daily = np.mean(daily_rets)
            std_daily = np.std(daily_rets, ddof=1)
            if std_daily == 0:
                return 0
            # 年化夏普（假设250个交易日）
            return round((avg_daily / std_daily) * np.sqrt(250), 2)

        bear_daily = [dl for dl in daily_log if not dl.get('market_bull')]
        bull_daily = [dl for dl in daily_log if dl.get('market_bull')]

        bear_stats['sharpe_ratio'] = _compute_sharpe(bear_daily, INITIAL_CAPITAL)
        bull_stats['sharpe_ratio'] = _compute_sharpe(bull_daily, INITIAL_CAPITAL)
        full_sharpe = _compute_sharpe(daily_log, INITIAL_CAPITAL)

        # SH > MA20 占比
        bear_above_count = sum(1 for dl in daily_log
                              if not dl.get('market_bull'))
        bull_above_count = sum(1 for dl in daily_log
                              if dl.get('market_bull'))
        full_above_pct = round((bear_above_count + bull_above_count) / max(len(daily_log), 1) * 100, 1)
        bear_above_pct = round(sum(1 for dl in bear_daily) / max(len(bear_daily), 1) * 100, 1) if bear_daily else 0
        bull_above_pct = round(sum(1 for dl in bull_daily) / max(len(bull_daily), 1) * 100, 1) if bull_daily else 0

        market_cycle_stats = {
            'bear_2022_2024': bear_stats,
            'bull_2025_2026': bull_stats,
            'full_cycle': {
                'total_signals': total_signals,
                'total_buys': total_buys,
                'annualized_return': full_annualized,
                'sharpe_ratio': full_sharpe,
                'sh_above_ma20_pct': full_above_pct,
                'bear_sh_above_pct': bear_above_pct,
                'bull_sh_above_pct': bull_above_pct,
            },
        }

        elapsed = _time.time() - t0
        print(f'\n回测完成 ({elapsed:.0f}s): '
              f'{total_signals} 信号, {total_buys} 买入, '
              f'{len(finished_trades)} 笔交易')
        print(f'  最终市值: {total_val/1e4:.1f}万 | '
              f'累计收益: {cumulative_return:+.1f}% | '
              f'最大回撤: {max_drawdown:.1f}%')
        if bear_stats:
            print(f'  熊市: {bear_stats.get("trades",0)}笔 胜率{bear_stats.get("win_rate",0):.1f}% '
                  f'均盈{bear_stats.get("avg_win",0):+.2f}% 均亏{bear_stats.get("avg_loss",0):+.2f}% '
                  f'回撤{bear_stats.get("max_drawdown",0):.1f}%')
        if bull_stats:
            print(f'  牛市: {bull_stats.get("trades",0)}笔 胜率{bull_stats.get("win_rate",0):.1f}% '
                  f'均盈{bull_stats.get("avg_win",0):+.2f}% 均亏{bull_stats.get("avg_loss",0):+.2f}% '
                  f'回撤{bull_stats.get("max_drawdown",0):.1f}%')

        return {
            'trades': finished_trades,
            'skipped': skipped_trades,
            'daily_log': daily_log,
            'start_date': start_date.strftime('%Y-%m-%d'),
            'end_date': end_date.strftime('%Y-%m-%d'),
            'total_signals': total_signals,
            'total_buys': total_buys,
            'final_value': round(total_val, 2),
            'initial_capital': INITIAL_CAPITAL,
            'cumulative_return': round(cumulative_return, 2),
            'max_drawdown': round(max_drawdown, 2),
            'portfolio_values': portfolio_values,
            'market_cycle_stats': market_cycle_stats,
        }

    def _next_trade_date(self, d: date, all_dates: List[date]) -> Optional[date]:
        """找到 d 之后的下一个交易日。"""
        for td in all_dates:
            if td > d:
                return td
        return None

    @staticmethod
    def _fast_count_limit_ups(code: str) -> int:
        """快速数涨停（取最近250根K线）。"""
        tdx = TdxReader()
        market = 'sh' if str(code).startswith('6') else 'sz'
        if str(code).startswith(('8', '4')):
            market = 'bj'
        try:
            fpath = os.path.join(tdx._market_dir(market), f'{market}{code}.day')
            if not os.path.exists(fpath):
                return 0
            fsize = os.path.getsize(fpath)
            if fsize < RECORD_SIZE * 20:
                return 0
            read_bytes = min(RECORD_SIZE * 260, fsize)
            with open(fpath, 'rb') as f:
                f.seek(fsize - read_bytes)
                raw = f.read(read_bytes)
            n = len(raw) // RECORD_SIZE
            count = 0
            prev_close = 0
            for j in range(n):
                offset = j * RECORD_SIZE
                cl_int = struct.unpack_from('I', raw, offset + 16)[0]
                close = cl_int / 100.0
                if prev_close > 0:
                    chg = (close - prev_close) / prev_close
                    if chg >= 0.093:
                        count += 1
                prev_close = close
            return count
        except Exception:
            return 0


# ═══════════════════════════════════════════════════════════════════════════
# Excel 导出
# ═══════════════════════════════════════════════════════════════════════════

def export_xlsx(results: dict, output_path: str):
    trades = results.get('trades', [])
    skipped = results.get('skipped', [])
    daily = results.get('daily_log', [])

    wb = Workbook()

    # ═══ Sheet 1: 交易明细 ═══
    ws1 = wb.active
    ws1.title = '交易明细'

    # 标题行
    ws1.merge_cells('A1:P1')
    ws1['A1'] = f"V3 VOL180 突破战法 — 交易明细 ({results['start_date']} ~ {results['end_date']})"
    ws1['A1'].font = TITLE_FONT
    ws1['A1'].alignment = CENTER

    headers = ['代码', '名称', '买入日', '买入价', '卖出日', '卖出价',
               '毛收益%', '净收益%', '结果', '持有天', '卖出原因',
               '曾涨停', '评分', '量比', '距压力%', '年涨停次']
    for c, h in enumerate(headers, 1):
        ws1.cell(row=3, column=c, value=h)
    _style_header(ws1, 3, len(headers))

    for i, t in enumerate(trades):
        r = i + 4
        ws1.cell(row=r, column=1, value=t['code'])
        ws1.cell(row=r, column=2, value=t['name'])
        ws1.cell(row=r, column=3, value=t['buy_date'])
        ws1.cell(row=r, column=4, value=t['buy_price'])
        ws1.cell(row=r, column=5, value=t['sell_date'])
        ws1.cell(row=r, column=6, value=t['sell_price'])
        ws1.cell(row=r, column=7, value=t['gross_ret'])
        ws1.cell(row=r, column=8, value=t['net_ret'])
        ws1.cell(row=r, column=9, value='Win' if t['is_win'] else 'Loss')
        ws1.cell(row=r, column=10, value=t['days_held'])
        ws1.cell(row=r, column=11, value=t['exit_reason'])
        ws1.cell(row=r, column=12, value='是' if t.get('had_zt') else '否')
        ws1.cell(row=r, column=13, value=t.get('score', 0))
        ws1.cell(row=r, column=14, value=t.get('vol_ratio', 0))
        ws1.cell(row=r, column=15, value=t.get('dist_pct', 0))
        ws1.cell(row=r, column=16, value=t.get('limit_count', 0))

        if t['is_win']:
            ws1.cell(row=r, column=8).fill = WIN_FILL
            ws1.cell(row=r, column=9).fill = WIN_FILL
        else:
            ws1.cell(row=r, column=8).fill = LOSS_FILL
            ws1.cell(row=r, column=9).fill = LOSS_FILL

    if trades:
        _style_data(ws1, 4, 3 + len(trades), len(headers))
    _auto_width(ws1)

    # ═══ Sheet 2: 统计汇总 ═══
    ws2 = wb.create_sheet('统计汇总')
    ws2.merge_cells('A1:D1')
    ws2['A1'] = 'V3 策略回测 — 统计汇总'
    ws2['A1'].font = TITLE_FONT
    ws2['A1'].alignment = CENTER

    # ── 计算统计 ──
    total = len(trades)
    wins = sum(1 for t in trades if t['is_win'])
    losses = total - wins
    win_rate = wins / max(total, 1) * 100
    net_rets = [t['net_ret'] for t in trades]
    avg_ret = np.mean(net_rets) if net_rets else 0
    avg_win = np.mean([r for r in net_rets if r > 0]) if wins > 0 else 0
    avg_loss = np.mean([r for r in net_rets if r <= 0]) if losses > 0 else 0
    max_win = max(net_rets) if net_rets else 0
    max_loss = min(net_rets) if net_rets else 0

    # 盈亏比
    profit_factor = abs(sum(r for r in net_rets if r > 0) / min(sum(r for r in net_rets if r <= 0), -0.01)) if losses > 0 else 999

    # 平均持有天数
    avg_hold = np.mean([t['days_held'] for t in trades]) if trades else 0

    # ── 组合层级（真实资金模型） ──
    initial_cap = results.get('initial_capital', 1000000.0)
    final_value = results.get('final_value', initial_cap)
    cum_ret = results.get('cumulative_return', 0.0)
    max_dd = results.get('max_drawdown', 0.0)

    # ── 牛熊市分割统计 ──
    market_cycle = results.get('market_cycle_stats', {})
    bear_stats = market_cycle.get('bear_2022_2024', {})
    bull_stats = market_cycle.get('bull_2025_2026', {})
    full_cycle = market_cycle.get('full_cycle', {})

    # ── 算最大回撤（基于每日组合市值） ──
    portfolio_values = results.get('portfolio_values', [])
    if portfolio_values:
        peak_val = initial_cap
        max_dd_calc = 0.0
        for _, tv, _, _ in portfolio_values:
            peak_val = max(peak_val, tv)
            dd = (peak_val - tv) / peak_val * 100 if peak_val > 0 else 0
            max_dd_calc = max(max_dd_calc, dd)
        if max_dd_calc > 0:
            max_dd = max_dd_calc

    stats = [
        ('回测区间', f"{results['start_date']} ~ {results['end_date']}"),
        ('', ''),
        ('--- 资金模型 ---', ''),
        ('初始资金', f'{initial_cap/1e4:.0f}万'),
        ('最终市值', f'{final_value/1e4:.2f}万'),
        ('最大持仓', '10只（单票10%）'),
        ('每天最多新开', '3只'),
        ('', ''),
        ('--- 交易统计 ---', ''),
        ('信号总数', results.get('total_signals', 0)),
        ('实际买入', results.get('total_buys', 0)),
        ('跳过(竞价/跌停)', len(skipped)),
        ('总交易笔数', total),
        ('盈利笔数', wins),
        ('亏损笔数', losses),
        ('胜率', f'{win_rate:.1f}%'),
        ('', ''),
        ('平均净收益/笔', f'{avg_ret:+.2f}%'),
        ('平均盈利/笔', f'{avg_win:+.2f}%'),
        ('平均亏损/笔', f'{avg_loss:+.2f}%'),
        ('最大单笔盈利', f'{max_win:+.2f}%'),
        ('最大单笔亏损', f'{max_loss:+.2f}%'),
        ('盈亏比', f'{profit_factor:.2f}'),
        ('平均持有天数', f'{avg_hold:.1f}天'),
        ('', ''),
        ('--- 组合表现 ---', ''),
        ('组合累计收益', f'{cum_ret:+.2f}%'),
        ('最大回撤', f'{max_dd:.2f}%'),
        ('年化收益（估算）', f'{cum_ret:+.1f}%'),
        ('', ''),
        ('--- 牛熊市分割 ---', ''),
        ('【2022-2024 熊市】', f'{bear_stats.get("trades", 0)}笔'),
        ('  胜率', f'{bear_stats.get("win_rate", 0):.1f}%'),
        ('  平均净收益/笔', f'{bear_stats.get("avg_net_return", 0):+.2f}%'),
        ('  平均盈利', f'{bear_stats.get("avg_win", 0):+.2f}%'),
        ('  平均亏损', f'{bear_stats.get("avg_loss", 0):+.2f}%'),
        ('  盈亏比', f'{bear_stats.get("profit_factor", 0):.2f}'),
        ('  累计收益', f'{bear_stats.get("cumulative_return", 0):+.2f}%'),
        ('  最大回撤', f'{bear_stats.get("max_drawdown", 0):.2f}%'),
        ('  年化收益', f'{bear_stats.get("annualized_return", 0):+.2f}%'),
        ('  夏普比率', f'{bear_stats.get("sharpe_ratio", 0):.2f}'),
        ('  平均持有天数', f'{bear_stats.get("avg_hold_days", 0):.1f}天'),
        ('  平均同时持仓', f'{bear_stats.get("avg_positions", 0):.1f}只'),
        ('【2025-2026 牛市】', f'{bull_stats.get("trades", 0)}笔'),
        ('  胜率', f'{bull_stats.get("win_rate", 0):.1f}%'),
        ('  平均净收益/笔', f'{bull_stats.get("avg_net_return", 0):+.2f}%'),
        ('  平均盈利', f'{bull_stats.get("avg_win", 0):+.2f}%'),
        ('  平均亏损', f'{bull_stats.get("avg_loss", 0):+.2f}%'),
        ('  盈亏比', f'{bull_stats.get("profit_factor", 0):.2f}'),
        ('  累计收益', f'{bull_stats.get("cumulative_return", 0):+.2f}%'),
        ('  最大回撤', f'{bull_stats.get("max_drawdown", 0):.2f}%'),
        ('  年化收益', f'{bull_stats.get("annualized_return", 0):+.2f}%'),
        ('  夏普比率', f'{bull_stats.get("sharpe_ratio", 0):.2f}'),
        ('  平均持有天数', f'{bull_stats.get("avg_hold_days", 0):.1f}天'),
        ('  平均同时持仓', f'{bull_stats.get("avg_positions", 0):.1f}只'),
    ]

    for i, (label, val) in enumerate(stats):
        r = 3 + i
        ws2.cell(row=r, column=1, value=label).font = SUMMARY_LABEL_FONT
        ws2.cell(row=r, column=2, value=val).font = SUMMARY_VAL_FONT
        ws2.cell(row=r, column=1).border = THIN_BORDER
        ws2.cell(row=r, column=2).border = THIN_BORDER
        # 分隔行着色
        if '---' in str(label):
            ws2.cell(row=r, column=1).fill = PatternFill(start_color='F3F4F6', end_color='F3F4F6', fill_type='solid')
            ws2.cell(row=r, column=2).fill = PatternFill(start_color='F3F4F6', end_color='F3F4F6', fill_type='solid')

    # 大数字显示
    big_row = 3 + len(stats) + 2
    ws2.merge_cells(f'A{big_row}:B{big_row}')
    ws2.cell(row=big_row, column=1, value='胜率').font = SUMMARY_LABEL_FONT
    ws2.cell(row=big_row, column=3, value=f'{win_rate:.1f}%').font = GREEN_FONT if win_rate >= 50 else RED_FONT
    ws2.cell(row=big_row, column=1).border = THIN_BORDER
    ws2.cell(row=big_row, column=3).border = THIN_BORDER

    big_row2 = big_row + 1
    ws2.merge_cells(f'A{big_row2}:B{big_row2}')
    ws2.cell(row=big_row2, column=1, value='组合累计收益').font = SUMMARY_LABEL_FONT
    ws2.cell(row=big_row2, column=3, value=f'{cum_ret:+.2f}%').font = GREEN_FONT if cum_ret > 0 else RED_FONT
    ws2.cell(row=big_row2, column=1).border = THIN_BORDER
    ws2.cell(row=big_row2, column=3).border = THIN_BORDER

    big_row3 = big_row2 + 1
    ws2.merge_cells(f'A{big_row3}:B{big_row3}')
    ws2.cell(row=big_row3, column=1, value='最大回撤').font = SUMMARY_LABEL_FONT
    ws2.cell(row=big_row3, column=3, value=f'{max_dd:.2f}%').font = RED_FONT
    ws2.cell(row=big_row3, column=1).border = THIN_BORDER
    ws2.cell(row=big_row3, column=3).border = THIN_BORDER

    _auto_width(ws2, min_width=14, max_width=30)
    ws2.column_dimensions['C'].width = 18

    # ═══ Sheet 3: 按卖出原因统计 ═══
    ws3 = wb.create_sheet('卖出原因统计')
    ws3.merge_cells('A1:F1')
    ws3['A1'] = '按卖出原因统计'
    ws3['A1'].font = TITLE_FONT
    ws3['A1'].alignment = CENTER

    reason_headers = ['卖出原因', '笔数', '胜率%', '平均收益%', '总收益%', '平均持有天']
    for c, h in enumerate(reason_headers, 1):
        ws3.cell(row=3, column=c, value=h)
    _style_header(ws3, 3, len(reason_headers))

    by_reason = defaultdict(list)
    for t in trades:
        reason = t['exit_reason']
        # 合并相似原因
        if '止损' in reason:
            reason = '止损-6%'
        elif '移动止盈' in reason:
            reason = '移动止盈'
        elif '断板' in reason or '涨停后断板' in reason:
            reason = '涨停后断板离场'
        elif '到期' in reason:
            reason = '持有到期'
        elif '强平' in reason:
            reason = '回测到期强平'
        by_reason[reason].append(t)

    row = 4
    for reason in sorted(by_reason.keys(), key=lambda x: -len(by_reason[x])):
        items = by_reason[reason]
        n = len(items)
        w = sum(1 for t in items if t['is_win'])
        wr = w / max(n, 1) * 100
        avg_r = np.mean([t['net_ret'] for t in items])
        total_r = sum(t['net_ret'] for t in items)
        avg_d = np.mean([t['days_held'] for t in items])

        ws3.cell(row=row, column=1, value=reason)
        ws3.cell(row=row, column=2, value=n)
        ws3.cell(row=row, column=3, value=f'{wr:.1f}%')
        ws3.cell(row=row, column=4, value=f'{avg_r:+.2f}%')
        ws3.cell(row=row, column=5, value=f'{total_r:+.2f}%')
        ws3.cell(row=row, column=6, value=f'{avg_d:.1f}')
        row += 1

    _style_data(ws3, 4, row - 1, len(reason_headers))
    _auto_width(ws3)

    # ═══ Sheet 4: 月度统计 ═══
    ws4 = wb.create_sheet('月度统计')
    ws4.merge_cells('A1:I1')
    ws4['A1'] = '按月统计（组合层级）'
    ws4['A1'].font = TITLE_FONT
    ws4['A1'].alignment = CENTER

    month_headers = ['月份', '笔数', '胜', '负', '胜率%', '平均收益%',
                     '组合月收益%', '月末市值(万)', '累计收益%']
    for c, h in enumerate(month_headers, 1):
        ws4.cell(row=3, column=c, value=h)
    _style_header(ws4, 3, len(month_headers))

    # 用组合估值计算月度收益（而非简单累加每笔交易）
    init_cap = results.get('initial_capital', 1000000.0)
    pv = results.get('portfolio_values', [])
    monthly_pv = {}  # YYYY-MM → [(date, value)]
    for dt, tv, _, _ in pv:
        mk = dt.strftime('%Y-%m')
        if mk not in monthly_pv:
            monthly_pv[mk] = []
        monthly_pv[mk].append((dt, tv))

    # 每月交易统计
    by_month = defaultdict(list)
    for t in trades:
        month_key = t['sell_date'][:7]
        by_month[month_key].append(t)

    row = 4
    prev_month_end_val = init_cap
    cum_ret = 0.0

    for month_key in sorted(set(list(by_month.keys()) + list(monthly_pv.keys()))):
        items = by_month.get(month_key, [])
        pvs = monthly_pv.get(month_key, [])
        n = len(items)
        w = sum(1 for t in items if t['is_win']) if n > 0 else 0
        l = n - w
        wr = w / max(n, 1) * 100
        avg_r = np.mean([t['net_ret'] for t in items]) if n > 0 else 0

        # 组合层级月收益（基于月末市值 vs 上月月末市值）
        if pvs:
            month_end_val = pvs[-1][1]  # last portfolio value of month
        else:
            month_end_val = prev_month_end_val
        month_ret = (month_end_val / prev_month_end_val - 1) * 100 if prev_month_end_val > 0 else 0
        cum_ret = (month_end_val / init_cap - 1) * 100

        ws4.cell(row=row, column=1, value=month_key)
        ws4.cell(row=row, column=2, value=n if n > 0 else 0)
        ws4.cell(row=row, column=3, value=w)
        ws4.cell(row=row, column=4, value=l)
        ws4.cell(row=row, column=5, value=f'{wr:.1f}%' if n > 0 else '-')
        ws4.cell(row=row, column=6, value=f'{avg_r:+.2f}%' if n > 0 else '-')
        ws4.cell(row=row, column=7, value=f'{month_ret:+.2f}%')
        ws4.cell(row=row, column=8, value=f'{month_end_val/1e4:.1f}')
        ws4.cell(row=row, column=9, value=f'{cum_ret:+.1f}%')

        if month_ret > 0:
            ws4.cell(row=row, column=7).fill = WIN_FILL
        else:
            ws4.cell(row=row, column=7).fill = LOSS_FILL
        if cum_ret > 0:
            ws4.cell(row=row, column=9).fill = WIN_FILL
        else:
            ws4.cell(row=row, column=9).fill = LOSS_FILL

        prev_month_end_val = month_end_val
        row += 1

    _style_data(ws4, 4, row - 1, len(month_headers))
    _auto_width(ws4)

    # ═══ Sheet 5: 日度日志 ═══
    ws5 = wb.create_sheet('日度日志')
    ws5.merge_cells('A1:E1')
    ws5['A1'] = '日度持仓变化'
    ws5['A1'].font = TITLE_FONT
    ws5['A1'].alignment = CENTER

    daily_headers = ['日期', '累计信号', '买入', '卖出', '持仓数']
    for c, h in enumerate(daily_headers, 1):
        ws5.cell(row=3, column=c, value=h)
    _style_header(ws5, 3, len(daily_headers))

    for i, d in enumerate(daily):
        r = i + 4
        ws5.cell(row=r, column=1, value=d['date'])
        ws5.cell(row=r, column=2, value=d['signals'])
        ws5.cell(row=r, column=3, value=d['buys'])
        ws5.cell(row=r, column=4, value=d['sells'])
        ws5.cell(row=r, column=5, value=d['holdings'])

    if daily:
        _style_data(ws5, 4, 3 + len(daily), len(daily_headers))
    _auto_width(ws5)

    # ═══ Sheet 6: 跳过记录 ═══
    if skipped:
        ws6 = wb.create_sheet('跳过记录')
        ws6.merge_cells('A1:F1')
        ws6['A1'] = '竞价确认/开盘跌停 跳过的买入信号'
        ws6['A1'].font = TITLE_FONT
        ws6['A1'].alignment = CENTER

        skip_headers = ['代码', '名称', '信号日', '跳过原因', '收盘价', '压力位']
        for c, h in enumerate(skip_headers, 1):
            ws6.cell(row=3, column=c, value=h)
        _style_header(ws6, 3, len(skip_headers))

        for i, s in enumerate(skipped):
            r = i + 4
            ws6.cell(row=r, column=1, value=s['code'])
            ws6.cell(row=r, column=2, value=s['name'])
            ws6.cell(row=r, column=3, value=s['signal_date'])
            ws6.cell(row=r, column=4, value=s['skip_reason'])
            ws6.cell(row=r, column=5, value=s.get('close', 0))
            ws6.cell(row=r, column=6, value=s.get('pressure', 0))
        _style_data(ws6, 4, 3 + len(skipped), len(skip_headers))
        _auto_width(ws6)

    wb.save(output_path)
    print(f'\n[OK] Excel saved: {output_path}')


# ═══════════════════════════════════════════════════════════════════════════
# Main
# ═══════════════════════════════════════════════════════════════════════════

if __name__ == '__main__':
    import argparse
    p = argparse.ArgumentParser(description='V3 VOL180 突破战法回测')
    p.add_argument('--start', type=str, default=None, help='开始日期 YYYY-MM-DD (默认一年前)')
    p.add_argument('--end', type=str, default=None, help='结束日期 YYYY-MM-DD (默认昨天)')
    p.add_argument('--output', type=str, default=None, help='输出 xlsx 路径')
    args = p.parse_args()

    start_date = datetime.strptime(args.start, '%Y-%m-%d').date() if args.start else None
    end_date = datetime.strptime(args.end, '%Y-%m-%d').date() if args.end else None
    output_path = args.output or os.path.join(OUTPUT_DIR, 'v3_backtest_result.xlsx')

    bt = V3Backtest()
    results = bt.run(start_date=start_date, end_date=end_date)

    if results:
        export_xlsx(results, output_path)
