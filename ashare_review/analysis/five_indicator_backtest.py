"""冰点抄底 + 启动突破 T+1 回测 — 年化版，导出 xlsx

买卖规则（严格T+1）:
  - 买入: T日收盘价
  - 卖出: T+1日收盘价
  - 仅统计 评分>=80 的高质量标的

数据源:
  - akshare 涨停池API 支持最近约20个交易日
  - 更早日期自动回退到 TDX 本地文件扫描（有磁盘缓存，首次慢、后续秒开）

用法:
    python -m ashare_review.analysis.five_indicator_backtest --days 250
"""
import sys, os, struct, argparse, json, time
from datetime import date, datetime, timedelta
from typing import Dict, List, Optional, Set
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

from ashare_review.data.tdx_reader import TdxReader, RECORD_SIZE
from ashare_review.data.akshare_fetcher import AkshareFetcher
from ashare_review.screening.five_indicator import (
    StartBreakoutScreener, StartBreakoutScreenerV2, IceBottomScreener,
)

# ─── Excel 样式 ───────────────────────────────────────────────────────────

HEADER_FONT = Font(name='Microsoft YaHei', bold=True, color='FFFFFF', size=11)
HEADER_FILL = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
WIN_FILL    = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
LOSS_FILL   = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
TITLE_FONT  = Font(name='Microsoft YaHei', bold=True, size=14, color='2F5496')
SUB_FONT    = Font(name='Microsoft YaHei', bold=True, size=11)
NUM_FONT    = Font(name='Consolas', size=10)
THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9'),
)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
GRAY_FILL = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
BLUE_FILL = PatternFill(start_color='E8F0FE', end_color='E8F0FE', fill_type='solid')
WARN_FILL = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')


class Backtest:
    def __init__(self):
        self.tdx = TdxReader()
        self.ak = AkshareFetcher()
        # TDX 涨停扫描磁盘缓存路径
        self._cache_path = os.path.join(
            os.path.dirname(os.path.dirname(__file__)), 'data', 'zt_scan_cache.json')
        self._zt_map: Dict[str, Set[str]] = {}  # date_str -> {codes}

    # ─── 交易日 ─────────────────────────────────────────────────────

    def _get_dates(self, n: int) -> List[date]:
        d = date.today() - timedelta(days=1)
        dates = []
        while len(dates) < n + 1:
            if d.weekday() < 5: dates.append(d)
            d -= timedelta(days=1)
        return list(reversed(dates))

    # ─── 候选池 ─────────────────────────────────────────────────────

    @staticmethod
    def _limit(code: str) -> float:
        code = str(code).zfill(6)
        if code.startswith(('300', '301', '688')): return 0.199
        if code.startswith(('8', '4')): return 0.299
        return 0.095

    def _ensure_zt_scan(self, dates: List[date]):
        """确保 TDX 涨停扫描缓存覆盖所有目标日期（首次慢，后续秒开）"""
        needed = {d.strftime('%Y%m%d') for d in dates}

        # 加载磁盘缓存
        if os.path.exists(self._cache_path):
            try:
                with open(self._cache_path) as f:
                    cached = json.load(f)
                # 只保留需要的日期
                self._zt_map = {k: set(v) for k, v in cached.items() if k in needed}
                print(f'Loaded ZT cache: {len(self._zt_map)} dates')
            except Exception:
                pass

        missing = needed - set(self._zt_map.keys())
        if not missing:
            return

        print(f'Scanning TDX for {len(missing)} dates (this is slow, cached for future)...')
        stocks = self.tdx.list_stocks()
        total = len(stocks)
        # 初始化
        for d in missing:
            self._zt_map[d] = set()
        t0 = time.time()

        for si, (code, market) in enumerate(stocks):
            if si == 0: t0 = time.time()
            if (si + 1) % 1500 == 0:
                elapsed = time.time() - t0
                eta = elapsed / (si + 1) * (total - si - 1)
                print(f'  {si+1}/{total} ({elapsed:.0f}s, ETA {eta:.0f}s)...', flush=True)

            threshold = self._limit(code)
            fpath = os.path.join(self.tdx._market_dir(market), f'{market}{code}.day')
            if not os.path.exists(fpath): continue
            try:
                fsize = os.path.getsize(fpath)
                if fsize < RECORD_SIZE * 30: continue
                read_bytes = min(RECORD_SIZE * 500, fsize)
                with open(fpath, 'rb') as f:
                    f.seek(fsize - read_bytes)
                    raw = f.read(read_bytes)
                n_records = len(raw) // RECORD_SIZE
                if n_records < 2: continue

                prev_close = None
                for j, (dt, _, _, _, cl, _, _, _) in enumerate(
                    struct.iter_unpack('IIIIIfII', raw)
                ):
                    ds = str(dt)
                    if ds in missing:
                        cp = cl / 100.0
                        if prev_close and prev_close > 0:
                            if (cp - prev_close) / prev_close >= threshold:
                                self._zt_map[ds].add(code)
                    prev_close = cl / 100.0
            except Exception:
                continue

        # 保存缓存
        os.makedirs(os.path.dirname(self._cache_path), exist_ok=True)
        with open(self._cache_path, 'w') as f:
            json.dump({k: list(v) for k, v in self._zt_map.items()}, f)
        print(f'  Cache saved. {sum(1 for v in self._zt_map.values() if v)} dates have ZT data.')

    def _get_candidates(self, ds: str) -> List[dict]:
        """先试 akshare，不行用 TDX 缓存"""
        try:
            pool = self.ak.get_limit_up_pool(trade_date=ds)
            if pool:
                return [{'code': lu.code, 'name': lu.name,
                         'board_type': lu.board_type, 'consecutive': lu.consecutive}
                        for lu in pool if lu.board_type != '一字板']
        except Exception:
            pass
        codes = self._zt_map.get(ds, set())
        return [{'code': c, 'name': '', 'board_type': '', 'consecutive': 1} for c in codes]

    # ─── T+1 收益 ──────────────────────────────────────────────────

    def _t1(self, code: str, entry_date: str) -> Optional[dict]:
        market = 'sh' if str(code).startswith('6') else (
            'bj' if str(code).startswith(('4', '8')) else 'sz')
        try:
            df = self.tdx.read_daily(str(code).zfill(6), market)
            if df.empty or len(df) < 2: return None
            entry_dt = datetime.strptime(entry_date, '%Y%m%d').date()
            tds = []
            for i in range(len(df)):
                td = df['trade_date'].iloc[i]
                if isinstance(td, (datetime,)): tds.append(td.date())
                elif isinstance(td, date): tds.append(td)
                else: tds.append(None)
            idx = None
            for i, d in enumerate(tds):
                if d == entry_dt: idx = i; break
            if idx is None or idx + 1 >= len(df): return None
            ep = float(df['close'].iloc[idx])
            xp = float(df['close'].iloc[idx + 1])
            return {'entry_price': round(ep, 2), 'exit_price': round(xp, 2),
                    'exit_date': str(tds[idx + 1]),
                    'ret_pct': round((xp - ep) / ep * 100, 2)}
        except Exception:
            return None

    # ─── 主循环 ────────────────────────────────────────────────────

    def run(self, lookback: int = 250):
        trade_dates = self._get_dates(lookback)
        valid = trade_dates[:-1]
        print(f'Backtest: {trade_dates[0]} ~ {trade_dates[-1]}, {len(valid)} days')

        # 确保 TDX 扫描缓存
        self._ensure_zt_scan(trade_dates)

        all_trades = []
        daily_log = []
        t0 = time.time()

        for i, td in enumerate(valid):
            ds = td.strftime('%Y%m%d')
            if (i + 1) % 20 == 0 or i == 0:
                elapsed = time.time() - t0
                eta = elapsed / (i + 1) * (len(valid) - i - 1) if i > 0 else 0
                print(f'[{i+1}/{len(valid)}] {ds} ({elapsed:.0f}s, ETA {eta:.0f}s)...', flush=True)

            candidates = self._get_candidates(ds)
            if not candidates:
                daily_log.append({'date': ds, 'trades': 0, 'wins': 0, 'losses': 0, 'ret_sum': 0})
                continue

            day_trades = []
            for name, ScreenerCls in [('启动+突破(V1基线)', StartBreakoutScreener),
                                      ('启动+突破V2(证据驱动)', StartBreakoutScreenerV2)]:
                # 注: IceBottomScreener 内部调用 get_market_breadth() 扫描全市场，
                # 回测逐日跑时太慢，需预计算 breadth 后再加入
                screener = ScreenerCls(self.tdx, self.ak)
                class _LU:
                    def __init__(self, c):
                        self.code = c['code']; self.name = c.get('name', '')
                        self.board_type = c.get('board_type', ''); self.consecutive = c.get('consecutive', 1)
                screener.ak.get_limit_up_pool = lambda trade_date=None: [_LU(c) for c in candidates]

                try:
                    results = screener.screen(trade_date=ds)
                except Exception:
                    results = []
                for r in results:
                    if r.score < 80: continue
                    t1 = self._t1(r.code, ds)
                    if t1 is None: continue
                    t = {'date': ds, 'strategy': name, 'code': r.code, 'name': r.name,
                         'score': r.score, 'entry_price': t1['entry_price'],
                         'exit_price': t1['exit_price'], 'ret_pct': t1['ret_pct'],
                         'exit_date': t1['exit_date'], 'is_win': t1['ret_pct'] > 0}
                    all_trades.append(t); day_trades.append(t)

            wins = sum(1 for t in day_trades if t['is_win'])
            daily_log.append({'date': ds, 'trades': len(day_trades),
                              'wins': wins, 'losses': len(day_trades) - wins,
                              'ret_sum': round(sum(t['ret_pct'] for t in day_trades), 2)})

        df = pd.DataFrame(all_trades)
        ddf = pd.DataFrame(daily_log)
        print(f'Done: {len(df)} trades, {sum(1 for d in daily_log if d["trades"]==0)} zero-trade days')
        return df, ddf, len(valid)

    def _summarize(self, df: pd.DataFrame) -> dict:
        out = {}
        for name in df['strategy'].unique():
            s = df[df['strategy'] == name]
            w = s[s['is_win']]; l = s[~s['is_win']]
            n = len(s); wr = len(w) / n * 100 if n else 0
            s2 = s.copy(); s2['week'] = pd.to_datetime(s2['date']).dt.isocalendar().week
            wk = s2.groupby('week').agg(n=('is_win', 'count'), w=('is_win', 'sum'),
                                        r=('ret_pct', 'sum'))
            out[name] = {
                'trades': n, 'wins': len(w), 'losses': len(l),
                'win_rate': round(wr, 1), 'avg_ret': round(s['ret_pct'].mean(), 2),
                'avg_win': round(w['ret_pct'].mean(), 2) if len(w) else 0,
                'avg_loss': round(l['ret_pct'].mean(), 2) if len(l) else 0,
                'cum_ret': round(s['ret_pct'].sum(), 2),
                'pl_ratio': round(w['ret_pct'].sum() / max(abs(l['ret_pct'].sum()), 0.01), 2),
                'weekly_wr': round(wk['w'].sum() / max(wk['n'].sum(), 1) * 100, 1),
                'weekly_trades': round(wk['n'].mean(), 1),
            }
        return out


# ═══════════════════════════════════════════════════════════════════════════════
# Excel 导出
# ═══════════════════════════════════════════════════════════════════════════════

def export_xlsx(df, ddf, summary, lookback, total_days, path):
    wb = Workbook()

    # Sheet 1: 汇总
    ws = wb.active; ws.title = '汇总'
    ws.merge_cells('A1:H1')
    ws['A1'] = f'冰点抄底+启动突破 高分(>=80) T+1 回测 ({lookback}天)'
    ws['A1'].font = TITLE_FONT; ws['A1'].alignment = CENTER
    ws.merge_cells('A3:H3')
    ws['A3'] = f'回测天数: {total_days} | 总交易: {len(df)}笔 | 仅评分>=80'
    ws['A3'].font = SUB_FONT
    hdrs = ['战法', '交易数', '胜', '负', '胜率%', '均收益%', '均盈%', '均亏%',
            '累计收益%', '盈亏比', '周均胜率%', '周均笔数']
    for c, h in enumerate(hdrs, 1):
        ws.cell(row=5, column=c, value=h)
    _style_row(ws, 5, len(hdrs), HEADER_FONT, HEADER_FILL)
    row = 6
    for name in ['启动+突破', '冰点抄底']:
        s = summary.get(name)
        if not s: continue
        for c, v in enumerate([name, s['trades'], s['wins'], s['losses'],
                               s['win_rate'], s['avg_ret'], s['avg_win'], s['avg_loss'],
                               s['cum_ret'], s['pl_ratio'], s['weekly_wr'], s['weekly_trades']], 1):
            ws.cell(row=row, column=c, value=v)
        wr = ws.cell(row=row, column=5)
        if s['win_rate'] >= 52: wr.fill = WIN_FILL
        elif s['win_rate'] < 48: wr.fill = LOSS_FILL
        row += 1
    _style_data(ws, 6, row - 1, len(hdrs))
    _auto_width(ws)

    # Sheet 2: 逐日记录（每天都有，含0交易日）
    ws2 = wb.create_sheet('逐日记录')
    ws2.merge_cells('A1:J1')
    ws2['A1'] = '逐日交易记录（每一天都有，含无交易日期）'
    ws2['A1'].font = TITLE_FONT; ws2['A1'].alignment = CENTER
    hdrs2 = ['日期', '战法', '代码', '名称', '评分', '入场价', '出场价', '收益%', '出场日', '结果']
    for c, h in enumerate(hdrs2, 1):
        ws2.cell(row=3, column=c, value=h)
    _style_row(ws2, 3, len(hdrs2), HEADER_FONT, HEADER_FILL)

    ddf_s = ddf.sort_values('date') if not ddf.empty else ddf
    cum = 0.0; r = 4
    for _, di in ddf_s.iterrows():
        ds = di['date']; day_trades = df[df['date'] == ds] if not df.empty else pd.DataFrame()
        if len(day_trades) == 0:
            ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
            c = ws2.cell(row=r, column=1, value=f'{ds}  |  无交易')
            c.font = Font(name='Microsoft YaHei', italic=True, color='999999', size=10)
            c.alignment = Alignment(horizontal='left', vertical='center')
            for cc in range(1, 11): ws2.cell(row=r, column=cc).border = THIN_BORDER; ws2.cell(row=r, column=cc).fill = GRAY_FILL
            r += 1
        else:
            dw = day_trades['is_win'].sum(); dt = len(day_trades); dr = day_trades['ret_pct'].sum(); cum += dr
            ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
            txt = f'{ds}  |  {dt}笔  |  胜{int(dw)}负{int(dt-dw)}  |  胜率{dw/dt*100:.0f}%  |  日收益{dr:+.2f}%  |  累计{cum:+.1f}%'
            c = ws2.cell(row=r, column=1, value=txt)
            c.font = Font(name='Microsoft YaHei', bold=True, size=10, color='2F5496')
            c.alignment = Alignment(horizontal='left', vertical='center')
            bg = BLUE_FILL if dw >= dt * 0.5 else WARN_FILL
            for cc in range(1, 11): ws2.cell(row=r, column=cc).border = THIN_BORDER; ws2.cell(row=r, column=cc).fill = bg
            r += 1
            for _, t in day_trades.iterrows():
                ws2.cell(row=r, column=1, value=t['date']); ws2.cell(row=r, column=2, value=t['strategy'])
                ws2.cell(row=r, column=3, value=t['code']); ws2.cell(row=r, column=4, value=t['name'])
                ws2.cell(row=r, column=5, value=t['score']); ws2.cell(row=r, column=6, value=t['entry_price'])
                ws2.cell(row=r, column=7, value=t['exit_price']); ws2.cell(row=r, column=8, value=t['ret_pct'])
                ws2.cell(row=r, column=9, value=t['exit_date'])
                rc = ws2.cell(row=r, column=10, value='Win' if t['is_win'] else 'Loss')
                if t['is_win']: rc.fill = WIN_FILL; ws2.cell(row=r, column=8).fill = WIN_FILL
                else: rc.fill = LOSS_FILL; ws2.cell(row=r, column=8).fill = LOSS_FILL
                r += 1
    _style_data(ws2, 4, r - 1, len(hdrs2))
    _auto_width(ws2); ws2.freeze_panes = 'A4'

    # Sheet 3: 每周汇总
    ws3 = wb.create_sheet('每周汇总')
    ws3['A1'] = '每周汇总'; ws3['A1'].font = TITLE_FONT; ws3['A1'].alignment = CENTER
    if not df.empty:
        df3 = df.copy(); df3['yw'] = pd.to_datetime(df3['date']).dt.strftime('%Y-W%V')
        wk = df3.groupby('yw').agg(n=('is_win', 'count'), w=('is_win', 'sum'),
                                   r=('ret_pct', 'sum')).reset_index()
        wk = wk.sort_values('yw'); wk['l'] = wk['n'] - wk['w']; wk['wr'] = (wk['w'] / wk['n'] * 100).round(1)
        for c, h in enumerate(['周', '交易数', '胜', '负', '胜率%', '周收益%'], 1):
            ws3.cell(row=3, column=c, value=h)
        _style_row(ws3, 3, 6, HEADER_FONT, HEADER_FILL)
        cum3 = 0
        for i, (_, wrow) in enumerate(wk.iterrows()):
            r3 = i + 4; cum3 += wrow['r']
            ws3.cell(row=r3, column=1, value=wrow['yw']); ws3.cell(row=r3, column=2, value=wrow['n'])
            ws3.cell(row=r3, column=3, value=wrow['w']); ws3.cell(row=r3, column=4, value=wrow['l'])
            ws3.cell(row=r3, column=5, value=wrow['wr']); ws3.cell(row=r3, column=6, value=round(cum3, 2))
            if wrow['wr'] >= 55: ws3.cell(row=r3, column=5).fill = WIN_FILL
            elif wrow['wr'] < 45: ws3.cell(row=r3, column=5).fill = LOSS_FILL
        _style_data(ws3, 4, len(wk) + 3, 6)
    _auto_width(ws3)

    wb.save(path)
    print(f'Saved: {path}')


def _style_row(ws, row, ncols, font, fill):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = font; cell.fill = fill; cell.alignment = CENTER; cell.border = THIN_BORDER

def _style_data(ws, sr, er, nc):
    for r in range(sr, er + 1):
        for c in range(1, nc + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = THIN_BORDER; cell.alignment = CENTER; cell.font = NUM_FONT

def _auto_width(ws):
    for col in ws.columns:
        mx = max((len(str(c.value or '')) for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(mx + 4, 30)


# ═══════════════════════════════════════════════════════════════════════════════

if __name__ == '__main__':
    p = argparse.ArgumentParser(description='冰点+启动 T+1回测')
    p.add_argument('--days', type=int, default=250); p.add_argument('--out', type=str, default=None)
    a = p.parse_args()

    bt = Backtest()
    df, ddf, total_days = bt.run(lookback=a.days)

    summary = bt._summarize(df) if not df.empty else {}
    print(f'\n{"="*60}')
    if not df.empty:
        print(f'  Trades: {len(df)} | WR: {df["is_win"].mean()*100:.1f}% | '
              f'Avg: {df["ret_pct"].mean():+.2f}% | Cum: {df["ret_pct"].sum():+.1f}%')
    for n in ['启动+突破', '冰点抄底']:
        s = summary.get(n)
        if s: print(f'  {n}: {s["trades"]:4d} trades | WR={s["win_rate"]:.1f}% | '
                     f'Avg={s["avg_ret"]:+.2f}% | Cum={s["cum_ret"]:+.1f}% | P/L={s["pl_ratio"]:.2f}')
    print(f'{"="*60}')

    out = a.out or os.path.join(os.path.dirname(os.path.dirname(__file__)),
                                'data', f'ice_breakout_backtest_{a.days}d.xlsx')
    os.makedirs(os.path.dirname(out), exist_ok=True)
    try:
        export_xlsx(df, ddf, summary, a.days, total_days, out)
    except PermissionError:
        alt = out.replace('.xlsx', '_2.xlsx')
        print(f'File locked, saving to: {alt}')
        export_xlsx(df, ddf, summary, a.days, total_days, alt)
