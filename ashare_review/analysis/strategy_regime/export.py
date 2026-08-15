"""共享统计 + xlsx 导出 — 供三个战法独立回测脚本和合并编排器复用"""
import os
import json
from collections import defaultdict
from typing import Dict, List

import numpy as np
import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SNAME_CN = {'V3启动突破': '启动突破V3', '1进2': '1进2接力', '冰点抄底': '冰点抄底'}

TITLE = Font(name='微软雅黑', size=14, bold=True, color='1F2937')
HDR = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
HDR_FILL = PatternFill('solid', start_color='374151', end_color='374151')
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
BORDER = Border(*[Side(style='thin', color='D1D5DB')] * 4)
WIN_FILL = PatternFill('solid', start_color='DCFCE7', end_color='DCFCE7')
LOSS_FILL = PatternFill('solid', start_color='FEE2E2', end_color='FEE2E2')
HIGHLIGHT = PatternFill('solid', start_color='FEF3C7', end_color='FEF3C7')

REGIMES = ['强势趋势', '题材轮动', '冰点超跌', '震荡观望', '弱市回调', '退潮下跌']


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HDR; cell.fill = HDR_FILL; cell.alignment = CENTER; cell.border = BORDER


def strat_stats(trades: List[Dict]) -> Dict:
    """单策略统计（逐笔等权，已扣成本）"""
    valid = [t for t in trades if not t.get('skipped_gap')]
    wins = [t for t in valid if t.get('is_win')]
    losses = [t for t in valid if not t.get('is_win')]
    rets = [t.get('net_ret', 0) for t in valid]
    gross = [t.get('gross_ret', 0) for t in valid]
    return {
        'n': len(valid), 'wins': len(wins), 'losses': len(losses),
        'skipped': len(trades) - len(valid),
        'win_rate': round(len(wins) / max(len(valid), 1) * 100, 1),
        'avg_ret': round(float(np.mean(rets)) if rets else 0, 2),
        'avg_win': round(float(np.mean([t['net_ret'] for t in wins])) if wins else 0, 2),
        'avg_loss': round(float(np.mean([t['net_ret'] for t in losses])) if losses else 0, 2),
        'pf': round(sum(t['net_ret'] for t in wins) / max(abs(sum(t['net_ret'] for t in losses)), 0.01), 2),
        'cum_ret': round(sum(rets), 2),  # 等权累计（非复利）
        'total_gross': round(sum(gross), 2),
    }


def _regime_map(state_df: pd.DataFrame) -> Dict[str, str]:
    """date(str YYYY-MM-DD) → regime"""
    return {str(r['date']): r['regime'] for _, r in state_df.iterrows()}


def tag_trades(trades: List[Dict], regime_map: Dict[str, str],
               default_regime: str = None) -> List[Dict]:
    """给交易打上行情标签（按 signal_date / buy_date）"""
    for t in trades:
        d = t.get('signal_date') or t.get('buy_date') or ''
        t['regime'] = regime_map.get(d, default_regime or '未知')
    return trades


def regime_matrix(state_df: pd.DataFrame, strategy_trades: Dict[str, List[Dict]]) -> pd.DataFrame:
    """战法 × 行情 矩阵"""
    rows = []
    for regime in REGIMES:
        n_days = int((state_df['regime'] == regime).sum())
        row = {'regime': regime, 'days': n_days,
               'pct_days': round(n_days / len(state_df) * 100, 1)}
        for sname, trades in strategy_trades.items():
            sub = [t for t in trades if t.get('regime') == regime]
            if sub:
                st = strat_stats(sub)
                row[f'{sname}_n'] = st['n']
                row[f'{sname}_wr'] = st['win_rate']
                row[f'{sname}_avg'] = st['avg_ret']
                row[f'{sname}_pf'] = st['pf']
                row[f'{sname}_cum'] = st['cum_ret']
            else:
                row[f'{sname}_n'] = 0
                row[f'{sname}_wr'] = np.nan
                row[f'{sname}_avg'] = np.nan
                row[f'{sname}_pf'] = np.nan
                row[f'{sname}_cum'] = np.nan
        best = None; best_score = -999
        for sname in strategy_trades:
            n = row.get(f'{sname}_n', 0)
            avg = row.get(f'{sname}_avg', np.nan)
            if n >= 3 and not pd.isna(avg) and avg > best_score:
                best_score = avg; best = sname
        row['best_strategy'] = best or '-'
        rows.append(row)
    return pd.DataFrame(rows)


def method_notes(sname: str) -> List[tuple]:
    """返回该战法的详细方法说明（含代码引用）。"""
    if sname == 'V3启动突破':
        return [
            ('一、策略概述',
             '启动突破 V3 = VOL180 量价突破 + 找顶线压力位突破。选沪深主板强势股（年涨停≥10），'
             '在缩量整理后放量突破压力位时次日竞价确认进场，持有到 -6%止损 / 移动止盈 / N字反包 / 5天到期。'
             '对应你系统里的 V3_STRATEGY.md 与 Vol180SimPortfolio(mode="v3")。'),
            ('二、候选池',
             '默认用因果候选池: 逐日判定"前期强势"(近250交易日涨停≥10，只看当日之前历史，无未来函数)，'
             '回测期间 ever-eligible 约 1034 只（修复静态池幸存者偏差）。'
             '代码: strategy_regime/causal_universe.py CausalUniverse；v3_backtest.py get_universe():146 为静态池对照。'),
            ('三、数据预处理',
             '每只股票截断尾部 600 行（约2.5年），避免完整历史 O(n²)；'
             '计算 MAVOL180 = MA(volume,180)；预计算快速向量化 zigzag 找顶线 _pressure 压力位'
             '（前后5根局部峰值检测 + 相隔≥10根去噪 + 前段 60 日高回退，比完整 BACKSET 链快约 100 倍，精度接近）。'
             '代码: _read_stock_full():166-223。'),
            ('四、买入信号',
             'T日收盘: close > 找顶线压力位 且 volume > MAVOL180 × 1.2（v3_backtest.py:635）。'
             '过滤: ①距压力位 3~5% 死亡区间跳过 ②量比 ≥5x 过度放量跳过 ③熊市(上证<MA60)要求评分≥75。'
             '评分 = 60 + 距压力位≥3%加10 + 量比≥2加10/≥1.5加5 + 年涨停≥20加10/≥15加5。'),
            ('五、竞价确认（T+1）',
             '_check_auction_v3():353 — 低开>3% 放弃 / 当日量<前日50% 放弃；_check_open_limit_down():382 — 开盘跌停放弃。'
             '全部通过 → T+1 开盘价买入。'),
            ('六、卖出（_check_sell_v3:243）',
             '① 收盘价 ≤ 买入价×0.94 → -6% 硬止损（次日开盘执行，跳空会放大亏损）；'
             '② 移动止盈: 持仓最高收盘回落 >5% 且已浮盈 ≥3% → 锁利；'
             '③ N字反包: 涨停后断板不立即卖，等一天——次日放量收阳(反包成功)继续持有，否则离场；'
             '④ 始终未涨停 + 持有≥5天 → 到期卖出。'),
            ('七、资金模型',
             '100万本金 · 最多10持仓 · 单票10%仓位 · 每日最多3新仓 · 卖出次日开盘执行（v3_backtest.py run():410）。'
             '本回测: 组合累计收益 +119.8%、最大回撤 63.7%。'),
            ('八、代码位置',
             '主引擎: ashare_review/analysis/v3_backtest.py（V3Backtest 类，run() 主循环 :402）\n'
             '独立入口: ashare_review/analysis/strategy_regime/backtest_v3.py\n'
             '线上对照: ashare_review/tools/sim_portfolio.py Vol180SimPortfolio（V3 状态机）'),
            ('九、近似与局限',
             '找顶线用快速向量化 zigzag（非完整 BACKSET 链）；涨停阈值按主板9.5%/创业板19.9%；'
             '候选池已改为逐日因果判定（无幸存者偏差）；仍未纳入板块强度/龙头识别；'
             '2025-2026 是突破策略顺风年，+75% 含顺风成分；79.95% 最大回撤反映模型常年满仓的激进风格，'
             '实盘需配合你 S/A/B/C 仓位分级降杠杆。'),
        ]
    if sname == '1进2':
        return [
            ('一、策略概述',
             '1进2 接力 = 首板（今日涨停昨日未涨停）→ 次日竞价高开确认后买入做"第二板"，'
             '涨停则晋级持有、断板即走。复刻复盘页《今日一进二精选》(report/daily.py _select_top_picks:910) 的选股逻辑，'
             '对应你系统里的 1进2 战法（1-into-2-strategy.md）。'),
            ('二、数据源',
             'TDX 本地日线。首板检测: 扫描全部 .day 文件构建 日期→涨停 索引'
             '(one_two_backtest.py TdxLimitUpIndex)，再判断"今日涨停且昨日未涨停"(_was_lup_yesterday)。'
             '股性=近250交易日涨停次数(_limit_up_count)。'),
            ('三、选股（_select_picks:89，复刻 _select_top_picks）',
             '硬性: 仅首板 + 非一字板 + 沪深主板(60/00/001/002) + 非ST + 股价3-15元。\n'
             '评分: 主板+3 / 低价+10 / 首板+8 / 股性(年涨停≥15加8、≥10加5、≥5加2) / '
             '封板质量(上影线<20%加10≈封成比) / 早盘封板(高开1~5%加8≈涨停时间) / 成交额>5亿加5。'
             '评分≥8 → 每日取 Top 8。'),
            ('四、买入（_simulate:171）',
             'T+1 开盘价买入，竞价确认: 高开 3%~7% 才参与——<3% 视为竞价未确认放弃(卖出纪律"弱转强需高开")，'
             '>7% 视为买不到跳过。全年 1918 个候选被此过滤掉 1561 个(81%)。'),
            ('五、卖出',
             '连板跟踪: 买入后若当日涨停 → 晋级继续持有(2板成功)；再涨停 → 3板；'
             '未涨停: ①盘中冲高 ≥+7% 不封 → 冲高止盈(战法"拉到7-8点必须卖") ②首日未晋级 → 收盘卖 '
             '③晋级后断板 → 断板卖 ④ -5% 无条件止损 ⑤ 最多持有3天。'),
            ('六、关键参数',
             '股价3-15元 · 竞价确认高开3%~7% · 止盈+7% · 止损-5% · 最多3天 · 每日Top8 · 成本0.35%。'),
            ('七、代码位置',
             '引擎: ashare_review/analysis/strategy_regime/one_two_backtest.py（OneTwoBacktest，run():270）\n'
             '独立入口: strategy_regime/backtest_one_two.py\n'
             '选股源对照: ashare_review/report/daily.py _select_top_picks:910'),
            ('八、近似与局限',
             'akshare 的封单额/涨停时间/流通市值无历史 → 用 上影线≈封成比、高开幅度≈涨停时间 近似，'
             '流通市值过滤省略；竞价量无法用日线还原(实际还需竞价量≥昨日爆量50%)；'
             '买入所有 Top8(择优 1-3 只会更好)；2025-2026 接力整体中性(+0.0%)，只在冰点超跌行情下显著为正。'),
        ]
    if sname == '冰点抄底':
        return [
            ('一、策略概述',
             '冰点抄底 = 全市场情绪冰点 + 缠论下跌衰竭 + 反转确认后，买前期强势股的深度超跌修复。'
             '核心是"买在下跌衰竭"(缠论一买/二买)，不接下跌中继的飞刀。对应你系统里的冰点战法'
             '(pattern-volume-framework.md 量价框架 + 缠论 integration)。'),
            ('二、市场级别: 冰点检测',
             '每个交易日统计 上涨家数/涨停家数(market_state.py bulk_breadth 一次性扫描全市场)。'
             '普通冰点: 上涨≤1200 或 涨停≤30；极冰点: 上涨≤800 或 涨停≤20。'
             '2025-2026 全年仅 13 天情绪触冰(5.4%)——冰点本身稀有。'),
            ('三、市场级别: 缠论反转确认（find_reversal_days:71）',
             '冰点段(连续冰点日合并)后 5 日内找反转确认: 上证涨幅≥1.5%大阳 或 (涨停≥80 且 上涨≥3000)；'
             '加二次确认: 次日上证不跌≥0.5%(过滤单日脉冲/诱多)。每段冰点只确认一次 → 全年 13 个反转确认日。'
             '缠论加持: 行情分类里的 sh_trend_now(上证当下笔方向) 与 detect_beichi(背驰) 标记下跌衰竭。'),
            ('四、个股选股（_oversold_candidates:171）',
             '候选 = 因果候选池(前期强势: 逐日近250日涨停≥10，无幸存者偏差，约1034只主板过滤后)。硬性: '
             '①距120日高点超跌≥30% ②股价≤40元(高价超跌股易阴跌) ③缩量企稳: 反转前一周均量<60日最大量×0.6'
             '(战法核心"卖盘释放，筹码沉淀") ④底部特征: 站上MA5 或 近5日缠论底分型。'
             '按超跌幅度取 Top5。'),
            ('五、买入',
             '反转确认日收盘价买入（缠论二买，右侧确认；比冰点当天抄底更安全）。'),
            ('六、卖出（_simulate:213，缠论: 卖点永远在下跌中产生）',
             '① 反弹 ≥+8% 止盈；② 反弹到前20日高点(压力位)止盈；③ 移动止盈: 最高收盘回落>5%；'
             '④ 跌破买入价-6% 止损；⑤ 最长持有10天到期。'),
            ('七、关键参数',
             '冰点阈值1200/30 · 反转大阳≥1.5% · 超跌≥30% · 缩量<0.6 · 股价≤40 · 止盈+8% · 止损-6% · 最多10天 · 每反转日Top5。'),
            ('八、代码位置',
             '引擎: ashare_review/analysis/strategy_regime/ice_backtest.py（IceBottomBacktest，run():273）\n'
             '缠论: strategy_regime/chan.py（分型:33 笔:53 走势:144 背驰:182 当下笔:213）\n'
             '独立入口: strategy_regime/backtest_ice.py'),
            ('九、近似与局限',
             '主力资金指标(calc_main_capital)未纳入(需额外数据)；缩量企稳为战法近似；'
             '全年样本仅 50 笔、13 个反转日，统计显著性和结论需谨慎；'
             '2025 上半年为上涨市，3 个反转日找不到超跌≥30%的强势股(没货可抄)——冰点抄底本质是择时型战法，等待是常态。'),
        ]
    return [('方法说明', sname)]


# ═══════════════════════════════════════════════════════════════════════
# 单战法独立 xlsx
# ═══════════════════════════════════════════════════════════════════════
def export_strategy_xlsx(sname: str, trades: List[Dict], state_df: pd.DataFrame,
                         output: str, period: str = '2025-08 ~ 2026-08',
                         meta_note: str = None, notes_extra: List[tuple] = None):
    """为一个战法生成独立的完整回测 xlsx（结论/逐笔/月度/行情分布/方法说明）"""
    wb = Workbook()

    # ── Sheet1 结论总览 ──
    ws = wb.active
    ws.title = '结论总览'
    ws['A1'] = f'{SNAME_CN[sname]} 回测（{period}，TDX 本地数据）'
    ws['A1'].font = TITLE
    if meta_note:
        ws['A2'] = meta_note
        ws['A2'].font = Font(name='微软雅黑', size=10, color='B45309', bold=True)
        ws['A2'].alignment = Alignment(wrap_text=True)
    r = 4
    ws.cell(row=r, column=1, value='一、整体表现（已扣成本0.35%）').font = Font(bold=True)
    r += 1
    st = strat_stats(trades)
    headers = ['交易数', '竞价过滤/跳过', '胜率%', '笔均收益%', '平均盈利%', '平均亏损%', '盈亏比', '等权累计%']
    for i, h in enumerate(headers, 1):
        ws.cell(row=r, column=i, value=h)
    style_header(ws, r, len(headers))
    r += 1
    vals = [st['n'], st['skipped'], st['win_rate'], st['avg_ret'], st['avg_win'],
            st['avg_loss'], st['pf'], st['cum_ret']]
    for i, v in enumerate(vals, 1):
        c = ws.cell(row=r, column=i, value=v); c.border = BORDER
    ws.cell(row=r, column=4).fill = WIN_FILL if st['avg_ret'] > 0 else LOSS_FILL
    r += 2

    # 卖出原因分布
    from collections import Counter
    ws.cell(row=r, column=1, value='二、卖出原因分布').font = Font(bold=True)
    r += 1
    for i, h in enumerate(['卖出原因', '笔数', '占比%', '平均收益%'], 1):
        ws.cell(row=r, column=i, value=h)
    style_header(ws, r, 4)
    r += 1
    reasons = Counter(t.get('exit_reason', '') for t in trades if not t.get('skipped_gap'))
    for reason, cnt in reasons.most_common():
        sub = [t for t in trades if t.get('exit_reason') == reason and not t.get('skipped_gap')]
        avg = round(sum(t.get('net_ret', 0) for t in sub) / len(sub), 2) if sub else 0
        for i, v in enumerate([reason, cnt, round(cnt / max(st['n'], 1) * 100, 1), avg], 1):
            ws.cell(row=r, column=i, value=v).border = BORDER
        r += 1
    r += 1

    # 行情分布
    ws.cell(row=r, column=1, value='三、按行情类型分布（该战法在每种行情下的表现）').font = Font(bold=True)
    r += 1
    for i, h in enumerate(['行情类型', '天数', '交易数', '胜率%', '笔均收益%', '盈亏比'], 1):
        ws.cell(row=r, column=i, value=h)
    style_header(ws, r, 6)
    r += 1
    for regime in REGIMES:
        n_days = int((state_df['regime'] == regime).sum())
        sub = [t for t in trades if t.get('regime') == regime]
        if sub:
            sst = strat_stats(sub)
            vals = [regime, n_days, sst['n'], sst['win_rate'], sst['avg_ret'], sst['pf']]
        else:
            vals = [regime, n_days, 0, '-', '-', '-']
        for i, v in enumerate(vals, 1):
            ws.cell(row=r, column=i, value=v).border = BORDER
        if sub:
            ws.cell(row=r, column=5).fill = WIN_FILL if sst['avg_ret'] > 0 else LOSS_FILL
        r += 1

    # 最新行情状态
    r += 1
    ws.cell(row=r, column=1, value='四、最新行情状态').font = Font(bold=True)
    r += 1
    last = state_df.iloc[-1]
    keys = ['date', 'sh_close', 'gz_close', 'sh_chg', 'gz_chg', 'up_count', 'limit_up',
            'emotion', 'sh_trend_now', 'regime', 'recommend']
    for i, k in enumerate(keys, 1):
        ws.cell(row=r, column=i, value=k)
    style_header(ws, r, len(keys))
    r += 1
    for i, k in enumerate(keys, 1):
        ws.cell(row=r, column=i, value=last.get(k)).border = BORDER

    # ── Sheet2 逐笔交易 ──
    ws = wb.create_sheet('逐笔交易')
    tcols = ['signal_date', 'buy_date', 'sell_date', 'code', 'name', 'buy_price',
             'sell_price', 'net_ret', 'gross_ret', 'exit_reason', 'days_held',
             'regime', 'signal_regime', 'position_w']
    hdr_list = [c for c in tcols if c in (trades[0].keys() if trades else {})]
    for i, c in enumerate(hdr_list, 1):
        ws.cell(row=1, column=i, value=c)
    style_header(ws, 1, len(hdr_list))
    for r, t in enumerate(sorted(trades, key=lambda x: x.get('signal_date', '')), 2):
        for i, c in enumerate(hdr_list, 1):
            cell = ws.cell(row=r, column=i, value=t.get(c))
            cell.border = BORDER
        net = t.get('net_ret', 0)
        if net > 0:
            ws.cell(row=r, column=hdr_list.index('net_ret') + 1).fill = WIN_FILL
        elif net < 0:
            ws.cell(row=r, column=hdr_list.index('net_ret') + 1).fill = LOSS_FILL
    ws.freeze_panes = 'A2'

    # ── Sheet3 月度汇总 ──
    ws = wb.create_sheet('月度汇总')
    headers = ['月份', '交易数', '胜率%', '笔均收益%', '平均盈利%', '平均亏损%', '盈亏比', '等权累计%']
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    style_header(ws, 1, len(headers))
    by_month = defaultdict(list)
    for t in trades:
        if t.get('skipped_gap'):
            continue
        d = t.get('signal_date') or t.get('buy_date') or ''
        by_month[d[:7]].append(t)
    r = 2
    for m in sorted(by_month.keys()):
        ms_ = strat_stats(by_month[m])
        vals = [m, ms_['n'], ms_['win_rate'], ms_['avg_ret'], ms_['avg_win'],
                ms_['avg_loss'], ms_['pf'], ms_['cum_ret']]
        for i, v in enumerate(vals, 1):
            ws.cell(row=r, column=i, value=v).border = BORDER
        r += 1

    # ── Sheet4 方法说明 ──
    ws = wb.create_sheet('方法说明')
    notes = [('回测区间', period)]
    notes += method_notes(sname)
    notes.append(('数据近似', 'akshare 封单额/涨停时间/流通市值无历史 → TDX 可算量近似；收益已扣0.35%成本'))
    if notes_extra:
        notes += notes_extra
    for i, (k, v) in enumerate(notes, 1):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws.cell(row=i, column=2, value=v)
        ws.cell(row=i, column=2).alignment = Alignment(wrap_text=True, vertical='top')
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 95

    os.makedirs(os.path.dirname(output), exist_ok=True)
    wb.save(output)
    print(f'[xlsx] {SNAME_CN[sname]} → {output}')
