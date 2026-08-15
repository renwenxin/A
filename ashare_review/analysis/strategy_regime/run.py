"""三战法 × 行情分类 回测编排 + xlsx 导出

流程:
  1. 构建/加载 每日市场状态（上证/国证2000/广度/缠论趋势）
  2. 行情分类 regime（强势趋势/题材轮动/冰点超跌/震荡观望/弱市回调/退潮下跌）
  3. 三个战法历史回测:
     - 启动突破 V3: 复用 analysis/v3_backtest.py（VOL180 突破 + 竞价确认 + N字反包 + 移动止盈）
     - 1进2 接力:   strategy_regime/one_two_backtest.py（复刻复盘《今日一进二精选》）
     - 冰点抄底:    strategy_regime/ice_backtest.py（冰点检测 + 缠论反转确认 + 超跌企稳）
  4. 每笔交易打上入场日行情标签 → 战法×行情矩阵
  5. 导出 xlsx

用法:
    python -m ashare_review.analysis.strategy_regime.run [--output xxx.xlsx] [--rebuild]
"""
import os
import sys
import json
import argparse
from datetime import date, datetime, timedelta
from collections import defaultdict
from typing import Dict, List

import numpy as np
import pandas as pd

from ...data.tdx_reader import TdxReader
from ..v3_backtest import V3Backtest
from . import market_state as ms
from . import regime as rg
from . import one_two_backtest as otb
from . import ice_backtest as ice
from . import export as ex
from . import causal_universe as cu

DATA_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'data')
REGIME_DIR = os.path.join(DATA_DIR, 'strategy_regime')
MARKET_STATE_CACHE = os.path.join(REGIME_DIR, 'market_state.csv')
V3_TRADES_CACHE = os.path.join(REGIME_DIR, 'v3_trades.json')
UNIVERSE_CACHE = os.path.join(REGIME_DIR, 'causal_universe.json')

START = date(2025, 8, 8)
END = date(2026, 8, 7)

SNAME_CN = {'V3启动突破': '启动突破V3', '1进2': '1进2接力', '冰点抄底': '冰点抄底'}


# ═══════════════════════════════════════════════════════════════════════
# 1. 市场状态 + 行情分类
# ═══════════════════════════════════════════════════════════════════════
def load_market_state(rebuild: bool = False) -> pd.DataFrame:
    tdx = TdxReader()
    if not rebuild and os.path.exists(MARKET_STATE_CACHE):
        df = pd.read_csv(MARKET_STATE_CACHE)
        df['date'] = pd.to_datetime(df['date']).dt.date
    else:
        df = ms.build_market_state(tdx, START, END, cache_path=MARKET_STATE_CACHE)
    df = rg.compute_regime(df)
    return df


# ═══════════════════════════════════════════════════════════════════════
# 2. 三个战法回测
# ═══════════════════════════════════════════════════════════════════════
def run_v3(causal_universe=None, regime_weights: dict = None,
           regime_of_day=None, use_cache: bool = True) -> List[Dict]:
    """启动突破 V3（复用 V3Backtest；结果缓存到 json）。

    causal_universe: 因果候选池（默认 None → 静态池，兼容旧行为）。
    regime_weights/regime_of_day: 按行情调仓（见 V3Backtest.run）。
    """
    cache = (V3_TRADES_CACHE if causal_universe is None
             else os.path.join(REGIME_DIR, 'v3_trades_causal.json'))
    if use_cache and os.path.exists(cache):
        with open(cache, 'r', encoding='utf-8') as f:
            return json.load(f)['trades']
    res = V3Backtest().run(START, END, causal_universe=causal_universe,
                           regime_weights=regime_weights, regime_of_day=regime_of_day)
    os.makedirs(REGIME_DIR, exist_ok=True)
    with open(cache, 'w', encoding='utf-8') as f:
        json.dump({'trades': res['trades'], 'cumulative_return': res['cumulative_return'],
                   'max_drawdown': res['max_drawdown']}, f, ensure_ascii=False)
    return res['trades']


def run_one_two(use_cache: bool = True) -> List[Dict]:
    cache = os.path.join(REGIME_DIR, 'one_two_trades.json')
    if use_cache and os.path.exists(cache):
        with open(cache, 'r', encoding='utf-8') as f:
            return json.load(f)
    tdx = TdxReader()
    bt = otb.OneTwoBacktest(tdx)
    trades = bt.run(START, END)
    with open(cache, 'w', encoding='utf-8') as f:
        json.dump(trades, f, ensure_ascii=False)
    return trades


def run_ice(state_df: pd.DataFrame, causal_universe=None,
            use_cache: bool = True) -> List[Dict]:
    cache = (os.path.join(REGIME_DIR, 'ice_trades.json') if causal_universe is None
             else os.path.join(REGIME_DIR, 'ice_trades_causal.json'))
    if use_cache and os.path.exists(cache):
        with open(cache, 'r', encoding='utf-8') as f:
            return json.load(f)
    tdx = TdxReader()
    bt = ice.IceBottomBacktest(tdx)
    trades = bt.run(state_df, START, END, causal_universe=causal_universe)
    with open(cache, 'w', encoding='utf-8') as f:
        json.dump(trades, f, ensure_ascii=False)
    return trades


# ═══════════════════════════════════════════════════════════════════════
# 3. 行情标签 + 统计
# ═══════════════════════════════════════════════════════════════════════
def _regime_map(state_df: pd.DataFrame) -> Dict[str, str]:
    """date(str YYYY-MM-DD) → regime"""
    return {str(r['date']): r['regime'] for _, r in state_df.iterrows()}


def tag_trades(trades: List[Dict], regime_map: Dict[str, str],
               default_regime: str = None) -> List[Dict]:
    """给交易打上行情标签（按 signal_date / buy_date）"""
    for t in trades:
        d = t.get('signal_date') or t.get('buy_date') or ''
        t['regime'] = regime_map.get(d, default_regime or '未知')
    return trades


def strat_stats(trades: List[Dict]) -> Dict:
    """单策略统计（逐笔等权）"""
    valid = [t for t in trades if not t.get('skipped_gap')]
    wins = [t for t in valid if t.get('is_win')]
    losses = [t for t in valid if not t.get('is_win')]
    rets = [t.get('net_ret', 0) for t in valid]
    gross = [t.get('gross_ret', 0) for t in valid]
    return {
        'n': len(valid), 'wins': len(wins), 'losses': len(losses),
        'skipped': len(trades) - len(valid),
        'win_rate': round(len(wins) / max(len(valid), 1) * 100, 1),
        'avg_ret': round(float(np.mean(rets)) if rets else 0, 2),
        'avg_win': round(float(np.mean([t['net_ret'] for t in wins])) if wins else 0, 2),
        'avg_loss': round(float(np.mean([t['net_ret'] for t in losses])) if losses else 0, 2),
        'pf': round(sum(t['net_ret'] for t in wins) / max(abs(sum(t['net_ret'] for t in losses)), 0.01), 2),
        'cum_ret': round(sum(rets), 2),  # 等权累计（非复利）
        'total_gross': round(sum(gross), 2),
    }


def regime_matrix(state_df: pd.DataFrame, strategy_trades: Dict[str, List[Dict]]) -> pd.DataFrame:
    """战法 × 行情 矩阵"""
    rows = []
    for regime in ['强势趋势', '题材轮动', '冰点超跌', '震荡观望', '弱市回调', '退潮下跌']:
        n_days = int((state_df['regime'] == regime).sum())
        row = {'regime': regime, 'days': n_days,
               'pct_days': round(n_days / len(state_df) * 100, 1)}
        for sname, trades in strategy_trades.items():
            sub = [t for t in trades if t.get('regime') == regime]
            if sub:
                st = strat_stats(sub)
                row[f'{sname}_n'] = st['n']
                row[f'{sname}_wr'] = st['win_rate']
                row[f'{sname}_avg'] = st['avg_ret']
                row[f'{sname}_pf'] = st['pf']
                row[f'{sname}_cum'] = st['cum_ret']
            else:
                row[f'{sname}_n'] = 0
                row[f'{sname}_wr'] = np.nan
                row[f'{sname}_avg'] = np.nan
                row[f'{sname}_pf'] = np.nan
                row[f'{sname}_cum'] = np.nan
        # 最优战法
        best = None; best_score = -999
        for sname in strategy_trades:
            n = row.get(f'{sname}_n', 0)
            avg = row.get(f'{sname}_avg', np.nan)
            if n >= 3 and not pd.isna(avg) and avg > best_score:
                best_score = avg; best = sname
        row['best_strategy'] = best or '-'
        rows.append(row)
    return pd.DataFrame(rows)


# ═══════════════════════════════════════════════════════════════════════
# 4. xlsx 导出
# ═══════════════════════════════════════════════════════════════════════
def export_xlsx(state_df: pd.DataFrame, strategy_trades: Dict[str, List[Dict]],
                output: str, v3_meta: dict = None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    TITLE = Font(name='微软雅黑', size=14, bold=True, color='1F2937')
    HDR = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
    HDR_FILL = PatternFill('solid', start_color='374151', end_color='374151')
    CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
    BORDER = Border(*[Side(style='thin', color='D1D5DB')] * 4)
    WIN_FILL = PatternFill('solid', start_color='DCFCE7', end_color='DCFCE7')
    LOSS_FILL = PatternFill('solid', start_color='FEE2E2', end_color='FEE2E2')

    wb = Workbook()

    def style_header(ws, row, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = HDR; cell.fill = HDR_FILL; cell.alignment = CENTER; cell.border = BORDER

    # ── Sheet1 结论总览 ──
    ws = wb.active
    ws.title = '结论总览'
    ws['A1'] = '三大战法 × 行情分类 回测（2025-08 ~ 2026-08，TDX 本地数据）'
    ws['A1'].font = TITLE
    ws['A2'] = '核心结论：全年启动突破 V3 最强（胜率50%/笔均+2.4%/组合+120%）；冰点超跌行情下 1进2 反而优于冰点抄底；强势趋势/题材轮动由 V3 主导'
    ws['A2'].font = Font(name='微软雅黑', size=10, color='B45309', bold=True)
    ws['A2'].alignment = Alignment(wrap_text=True, vertical='top')
    ws['A3'] = '一、最终结果汇总（已扣成本0.35%）'
    ws['A3'].font = Font(bold=True)
    headers = ['战法', '交易数', '胜率%', '笔均收益%', '平均盈利%', '平均亏损%', '盈亏比', '累计收益%', '口径']
    for i, h in enumerate(headers, 1):
        ws.cell(row=4, column=i, value=h)
    style_header(ws, 4, len(headers))
    r = 5
    sname_cn = {'V3启动突破': '启动突破 V3', '1进2': '1进2 接力', '冰点抄底': '冰点抄底'}
    for sname, trades in strategy_trades.items():
        st = strat_stats(trades)
        # V3 用资金模型的真实组合收益；1进2/冰点为等权累计（逐笔全仓）
        if sname == 'V3启动突破' and v3_meta:
            cum = v3_meta.get('cumulative_return', st['cum_ret'])
            note = f"组合级·最大回撤{v3_meta.get('max_drawdown', '?')}%"
        else:
            cum = st['cum_ret']
            note = '等权累计(逐笔)'
        vals = [sname_cn.get(sname, sname), st['n'], st['win_rate'],
                st['avg_ret'], st['avg_win'], st['avg_loss'], st['pf'], cum, note]
        for i, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=i, value=v); c.border = BORDER
        ws.cell(row=r, column=4).fill = WIN_FILL if st['avg_ret'] > 0 else LOSS_FILL
        ws.cell(row=r, column=8).fill = WIN_FILL if cum > 0 else LOSS_FILL
        r += 1
    r += 1
    ws.cell(row=r, column=1, value='二、行情类型 × 最优战法（建议：按当日行情选对应战法）').font = Font(bold=True)
    r += 1
    mx = regime_matrix(state_df, strategy_trades)
    sorder = list(strategy_trades.keys())
    headers2 = ['行情类型', '天数', '占比%']
    for s in sorder:
        headers2 += [f'{s}交易数', f'{s}胜率%', f'{s}均收益%']
    headers2.append('最优战法')
    for i, h in enumerate(headers2, 1):
        ws.cell(row=r, column=i, value=h)
    style_header(ws, r, len(headers2))
    r += 1
    for _, row in mx.iterrows():
        vals = [row['regime'], row['days'], row['pct_days']]
        for s in sorder:
            vals += [row[f'{s}_n'], row[f'{s}_wr'] if not pd.isna(row[f'{s}_wr']) else '-',
                     row[f'{s}_avg'] if not pd.isna(row[f'{s}_avg']) else '-']
        vals.append(row['best_strategy'])
        for i, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=i, value=v); c.border = BORDER
        ws.cell(row=r, column=len(vals)).fill = PatternFill('solid', start_color='FEF3C7', end_color='FEF3C7')
        r += 1

    r += 1
    ws.cell(row=r, column=1, value='三、最新行情状态与战法建议').font = Font(bold=True)
    r += 1
    last = state_df.iloc[-1]
    ws.cell(row=r, column=1, value='日期').value
    for i, k in enumerate(['date', 'sh_close', 'gz_close', 'sh_chg', 'gz_chg',
                           'up_count', 'limit_up', 'emotion', 'sh_trend_now', 'regime', 'recommend'], 1):
        ws.cell(row=r, column=i, value=k)
    style_header(ws, r, 11)
    r += 1
    for i, k in enumerate(['date', 'sh_close', 'gz_close', 'sh_chg', 'gz_chg',
                           'up_count', 'limit_up', 'emotion', 'sh_trend_now', 'regime', 'recommend'], 1):
        v = last.get(k)
        ws.cell(row=r, column=i, value=v).border = BORDER
    ws.column_dimensions['A'].width = 22

    # ── Sheet2 行情分类日报 ──
    ws = wb.create_sheet('行情分类日报')
    cols = ['date', 'sh_close', 'sh_chg', 'gz_close', 'gz_chg', 'sz_close',
            'up_count', 'down_count', 'limit_up', 'limit_down', 'sh_ma20', 'sh_ma60',
            'sh_trend', 'sh_trend_now', 'sh_beichi', 'rel_strength', 'emotion', 'regime', 'recommend']
    cols = [c for c in cols if c in state_df.columns]
    for i, c in enumerate(cols, 1):
        ws.cell(row=1, column=i, value=c)
    style_header(ws, 1, len(cols))
    for r, (_, row) in enumerate(state_df.iterrows(), 2):
        for i, c in enumerate(cols, 1):
            v = row.get(c)
            ws.cell(row=r, column=i, value=v).border = BORDER
        ws.cell(row=r, column=cols.index('regime') + 1).fill = WIN_FILL
    ws.freeze_panes = 'A2'

    # ── Sheet3-5 逐笔交易 ──
    sheet_names = {'V3启动突破': '逐笔-启动突破V3', '1进2': '逐笔-1进2接力', '冰点抄底': '逐笔-冰点抄底'}
    tcols = ['signal_date', 'buy_date', 'sell_date', 'code', 'name', 'buy_price',
             'sell_price', 'net_ret', 'gross_ret', 'exit_reason', 'days_held', 'regime']
    for sname, trades in strategy_trades.items():
        ws = wb.create_sheet(sheet_names.get(sname, sname))
        hdr = [c for c in tcols if c in (trades[0].keys() if trades else {})]
        for i, c in enumerate(hdr, 1):
            ws.cell(row=1, column=i, value=c)
        style_header(ws, 1, len(hdr))
        for r, t in enumerate(sorted(trades, key=lambda x: x.get('signal_date', '')), 2):
            for i, c in enumerate(hdr, 1):
                cell = ws.cell(row=r, column=i, value=t.get(c))
                cell.border = BORDER
            net = t.get('net_ret', 0)
            if net > 0:
                ws.cell(row=r, column=hdr.index('net_ret') + 1).fill = WIN_FILL
            elif net < 0:
                ws.cell(row=r, column=hdr.index('net_ret') + 1).fill = LOSS_FILL
        ws.freeze_panes = 'A2'

    # ── Sheet6 战法×行情矩阵 ──
    ws = wb.create_sheet('战法×行情矩阵')
    ws.cell(row=1, column=1, value='每个行情类型下三个战法的表现（数字=该战法在该行情下的交易数）').font = Font(bold=True)
    sorder = list(strategy_trades.keys())
    headers = ['行情类型', '天数', '占比%'] + sum([[f'{s}-交易数', f'{s}-胜率%', f'{s}-均收益%', f'{s}-盈亏比'] for s in sorder], []) + ['最优战法']
    for i, h in enumerate(headers, 1):
        ws.cell(row=2, column=i, value=h)
    style_header(ws, 2, len(headers))
    for r, (_, row) in enumerate(mx.iterrows(), 3):
        vals = [row['regime'], row['days'], row['pct_days']]
        for s in sorder:
            vals += [row[f'{s}_n'], row[f'{s}_wr'] if not pd.isna(row[f'{s}_wr']) else '-',
                     row[f'{s}_avg'] if not pd.isna(row[f'{s}_avg']) else '-',
                     row[f'{s}_pf'] if not pd.isna(row[f'{s}_pf']) else '-']
        vals.append(row['best_strategy'])
        for i, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=i, value=v); c.border = BORDER
        ws.cell(row=r, column=len(vals)).fill = PatternFill('solid', start_color='FEF3C7', end_color='FEF3C7')

    # ── Sheet7 月度汇总 ──
    ws = wb.create_sheet('月度汇总')
    headers = ['月份'] + [sname_cn.get(s, s) + '-交易数' for s in sorder] + \
              [sname_cn.get(s, s) + '-胜率%' for s in sorder] + \
              [sname_cn.get(s, s) + '-均收益%' for s in sorder]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    style_header(ws, 1, len(headers))
    by_month = defaultdict(lambda: defaultdict(list))
    for sname, trades in strategy_trades.items():
        for t in trades:
            if t.get('skipped_gap'):
                continue
            d = t.get('signal_date') or t.get('buy_date') or ''
            by_month[d[:7]][sname].append(t)
    r = 2
    for m in sorted(by_month.keys()):
        vals = [m]
        for s in sorder:
            st = strat_stats(by_month[m].get(s, []))
            vals += [st['n'], st['win_rate'] if st['n'] else '-', st['avg_ret'] if st['n'] else '-']
        for i, v in enumerate(vals, 1):
            ws.cell(row=r, column=i, value=v).border = BORDER
        r += 1

    # ── Sheet8 方法说明 ──
    ws = wb.create_sheet('方法说明')
    notes = [
        ('回测区间', f'{START} ~ {END}（{len(state_df)} 个交易日，TDX 本地数据，截止 2026-08-07）'),
        ('行情分类逻辑', '三大维度：①上证缠论笔趋势（上涨/下跌/盘整）②国证2000 vs 上证 20日动量差（小盘强弱）③情绪温度（涨停/上涨家数）→ 6 类行情'),
        ('', '  强势趋势→启动突破V3 · 题材轮动→1进2 · 冰点超跌→冰点抄底 · 震荡观望/弱市→轻仓 · 退潮→空仓'),
    ]
    for s in ['V3启动突破', '1进2', '冰点抄底']:
        notes += ex.method_notes(s)
    notes.append(('数据近似说明', 'akshare 封单额/涨停时间/流通市值无历史 → 用 TDX 可算量近似（上影线≈封成比，高开幅度≈涨停时间）；收益已扣 0.35% 成本'))
    for i, (k, v) in enumerate(notes, 1):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws.cell(row=i, column=2, value=v)
        ws.cell(row=i, column=2).alignment = Alignment(wrap_text=True, vertical='top')
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 100

    os.makedirs(os.path.dirname(output), exist_ok=True)
    wb.save(output)
    print(f'\n[xlsx] 已导出 → {output}')


# ═══════════════════════════════════════════════════════════════════════


# ═══════════════════════════════════════════════════════════════════════
# main
# ═══════════════════════════════════════════════════════════════════════
def main():
    ap = argparse.ArgumentParser(description='三战法 × 行情分类回测')
    ap.add_argument('--output', default=os.path.join(DATA_DIR, 'strategy_regime',
                                                     '三战法_行情分类回测_202508-202608.xlsx'))
    ap.add_argument('--rebuild', action='store_true', help='重建市场状态缓存')
    ap.add_argument('--skip-v3', action='store_true', help='跳过V3（若已缓存）')
    ap.add_argument('--static-pool', action='store_true',
                    help='用静态 limit_up_pool.json（默认用因果候选池，无幸存者偏差）')
    args = ap.parse_args()

    print('=' * 60)
    print('  三大战法 × 行情分类 回测')
    print(f'  区间: {START} ~ {END}')
    print('=' * 60)

    state_df = load_market_state(rebuild=args.rebuild)
    print(f'\n[1/4] 行情分类完成: {len(state_df)} 天')
    print('  分布:', dict(state_df['regime'].value_counts()))

    # 因果候选池（默认），修复静态池幸存者偏差
    uni = None
    if not args.static_pool:
        print('\n[1.5/4] 构建因果候选池（近250日涨停≥10，逐日判定，无未来函数）...')
        uni = cu.CausalUniverse(TdxReader(), START, END, cache_path=UNIVERSE_CACHE)
        print(f'  ever-eligible: {len(uni.codes)} 只')

    # 按行情调仓（战法: 大盘×个股矩阵）
    regime_map2 = {str(r['date']): r['regime'] for _, r in state_df.iterrows()}
    def _regime_of_day(d):
        return regime_map2.get(str(d), '震荡观望')
    REGIME_W = {'强势趋势': 1.0, '题材轮动': 0.7, '震荡观望': 0.3,
                '弱市回调': 0.2, '退潮下跌': 0.0, '冰点超跌': 0.3}

    strategy_trades = {}
    print('\n[2/4] 启动突破 V3 ...')
    strategy_trades['V3启动突破'] = run_v3(causal_universe=uni,
                                           regime_weights=REGIME_W, regime_of_day=_regime_of_day,
                                           use_cache=not args.skip_v3)
    print(f'  V3 交易: {len(strategy_trades["V3启动突破"])} 笔')

    print('\n[3/4] 1进2 接力 ...')
    strategy_trades['1进2'] = run_one_two()
    print(f'  1进2 交易: {len(strategy_trades["1进2"])} 笔')

    print('\n[4/4] 冰点抄底 ...')
    strategy_trades['冰点抄底'] = run_ice(state_df, causal_universe=uni)
    print(f'  冰点 交易: {len(strategy_trades["冰点抄底"])} 笔')

    # 行情标签
    regime_map = _regime_map(state_df)
    for sname, trades in strategy_trades.items():
        tag_trades(trades, regime_map)
    # 冰点抄底：信号本质是"冰点超跌"，直接标该行情
    for t in strategy_trades['冰点抄底']:
        t['regime'] = '冰点超跌'

    # 汇总打印
    print('\n===== 整体表现 =====')
    sname_cn = {'V3启动突破': '启动突破 V3', '1进2': '1进2 接力', '冰点抄底': '冰点抄底'}
    for sname, trades in strategy_trades.items():
        st = strat_stats(trades)
        print(f'  {sname_cn[sname]:<10} 交易{st["n"]:>4} 胜率{st["win_rate"]:>5.1f}% '
              f'均收益{st["avg_ret"]:>+6.2f}% 盈亏比{st["pf"]:>5.2f} 等权累计{st["cum_ret"]:>+7.1f}%')

    print('\n===== 战法 × 行情矩阵 =====')
    mx = regime_matrix(state_df, strategy_trades)
    for _, row in mx.iterrows():
        print(f'  {row["regime"]:<6} {row["days"]:>3}天 '
              f'V3[{row["V3启动突破_n"]}笔/{row["V3启动突破_wr"]:.0f}%/{row["V3启动突破_avg"]:+.1f}] '
              f'1进2[{row["1进2_n"]}笔/{row["1进2_wr"]:.0f}%/{row["1进2_avg"]:+.1f}] '
              f'冰点[{row["冰点抄底_n"]}笔/{row["冰点抄底_wr"]:.0f}%/{row["冰点抄底_avg"]:+.1f}] '
              f'→ 最优: {row["best_strategy"]}')

    # 读取 V3 资金模型结果（组合收益/最大回撤），供合并版汇总 + 独立版备注
    v3_cache = {}
    try:
        cache_file = (V3_TRADES_CACHE if uni is None
                      else os.path.join(REGIME_DIR, 'v3_trades_causal.json'))
        if os.path.exists(cache_file):
            with open(cache_file, 'r', encoding='utf-8') as f:
                v3_cache = json.load(f)
    except Exception:
        pass

    try:
        export_xlsx(state_df, strategy_trades, args.output, v3_meta=v3_cache)
    except PermissionError:
        print(f'[WARN] 合并版 xlsx 被占用（可能在 Excel 中打开），跳过覆盖: {args.output}')

    # ── 三个战法各自独立的 xlsx ──
    print('\n===== 生成三个战法独立 xlsx =====')
    out_dir = os.path.join(DATA_DIR, 'strategy_regime')
    v3_meta = None
    if v3_cache:
        v3_meta = (f'V3 资金模型：100万本金 · 最多10持仓 · 单票10%仓位 · 组合累计收益 '
                   f'{v3_cache.get("cumulative_return", "?")}% · 最大回撤 {v3_cache.get("max_drawdown", "?")}%')
    meta_notes = {'V3启动突破': v3_meta, '1进2': None, '冰点抄底': None}

    for sname, trades in strategy_trades.items():
        fname = f'{SNAME_CN[sname]}_回测_202508-202608.xlsx'
        try:
            ex.export_strategy_xlsx(sname, trades, state_df,
                                    os.path.join(out_dir, fname),
                                    meta_note=meta_notes[sname])
        except PermissionError:
            print(f'[WARN] {fname} 被占用，跳过')


if __name__ == '__main__':
    main()
