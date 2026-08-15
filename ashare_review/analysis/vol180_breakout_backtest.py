"""MAVOL180成交量突破 + 压力位突破 回测

选股逻辑（与网站 启动+突破 筛选器一致）:
  1. 沪深主板（60/00/001/002开头）
  2. 近一年涨停 > 10次
  3. 非ST、非*ST
  4. 距压力位（60日高点）下方10%以内

买入逻辑:
  1. 价格突破压力位（收盘价 > 前一日60日高点）
  2. 成交量突破 MAVOL180（当日量 > 180日均量 × 1.2）
  两个条件同时满足 → 次日开盘买入

卖出逻辑:
  1. 连板就继续持有（涨停板不断，就一直拿）
  2. 断板就卖出（不再涨停的当天收盘卖出）
  3. 最多持有三天（如果始终没有涨停，第三天收盘卖出）
"""
import sys, os, struct, json, argparse, time, pickle
from datetime import date, datetime, timedelta
from typing import Dict, List, Optional, Tuple
from collections import defaultdict
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

from ashare_review.data.tdx_reader import TdxReader, RECORD_SIZE
from ashare_review.analysis.indicators import calc_ma, calc_zigzag_find_top_line

# ─── 常量 ──────────────────────────────────────────────────────────────
FEE = 0.0015
SLIPPAGE_BUY = 0.001
SLIPPAGE_SELL = 0.001
TOTAL_COST = FEE + SLIPPAGE_BUY + SLIPPAGE_SELL
MAX_HOLD_BASELINE = 3

DATA_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'data')
GAINERS_CACHE = os.path.join(DATA_DIR, 'gainers_7pct.json')
STOCK_CACHE_DIR = os.path.join(DATA_DIR, 'stock_cache_v2')
os.makedirs(DATA_DIR, exist_ok=True)
os.makedirs(STOCK_CACHE_DIR, exist_ok=True)


class Vol180BreakoutBacktest:
    """MAVOL180 成交量突破 + 压力位突破 回测系统 (优化版)"""

    def __init__(self):
        self.tdx = TdxReader()
        self._stock_cache: Dict[str, pd.DataFrame] = {}  # code -> df with all indicators

    @staticmethod
    def _limit_threshold(code: str) -> float:
        code = str(code).zfill(6)
        if code.startswith(('300', '301', '688')):
            return 0.199
        if code.startswith(('8', '4')):
            return 0.299
        return 0.095

    @staticmethod
    def _is_main_board(code: str) -> bool:
        """判断是否沪深主板（上海主板+深圳主板）"""
        return code.startswith(('600', '601', '603', '605', '000', '001', '002'))

    @staticmethod
    def _is_stock_st(name: str) -> bool:
        """判断是否ST/*ST股票"""
        return bool(name) and ('ST' in str(name) or '*ST' in str(name))

    def _build_limit_up_count(self, gainers: Dict[str, List[dict]]) -> Dict[str, int]:
        """从 gainers 缓存统计每只股票过去一年的涨停次数(>=9.5%)"""
        counts: Dict[str, int] = defaultdict(int)
        for ds, lst in gainers.items():
            for info in lst:
                if info.get('change_pct', 0) >= 9.5:
                    counts[info['code']] += 1
        return dict(counts)

    @staticmethod
    def _trade_dates(n: int, end_date: date = None) -> List[date]:
        """生成回溯 N 个交易日的日期列表"""
        d = end_date or (date.today() - timedelta(days=1))
        dates = []
        while len(dates) < n + 30:
            if d.weekday() < 5:
                dates.append(d)
            d -= timedelta(days=1)
        return list(reversed(dates))

    def _get_stock_name(self, code: str) -> str:
        """从TDX数据读取股票名称"""
        market = 'sh' if code.startswith('6') else 'sz'
        if code.startswith(('8', '4')):
            market = 'bj'
        try:
            df = self.tdx.read_daily(code, market)
            if df is not None and not df.empty and 'name' in df.columns:
                return str(df['name'].iloc[0])
        except Exception:
            pass
        return ''
        d = end_date or (date.today() - timedelta(days=1))
        dates = []
        while len(dates) < n + 30:
            if d.weekday() < 5:
                dates.append(d)
            d -= timedelta(days=1)
        return list(reversed(dates))

    # ═══════════════════════════════════════════════════════════════════
    # Phase 1: 预加载 + 缓存
    # ═══════════════════════════════════════════════════════════════════

    def _load_gainers_cache(self, all_dates: List[date]) -> Dict[str, List[dict]]:
        """加载或构建 7%+涨幅股缓存"""
        all_ds = {d.strftime('%Y%m%d') for d in all_dates}

        if os.path.exists(GAINERS_CACHE):
            with open(GAINERS_CACHE, 'r', encoding='utf-8') as f:
                cache = json.load(f)
            if all_ds.issubset(set(cache.keys())):
                return {k: v for k, v in cache.items() if k in all_ds}

        print('Building gainers cache...')
        result: Dict[str, list] = {d: [] for d in all_ds}
        stocks = self.tdx.list_stocks()
        total = len(stocks)
        t0 = time.time()

        for si, (code, market) in enumerate(stocks):
            if (si + 1) % 1500 == 0:
                e = time.time() - t0
                print(f'  Scan {si+1}/{total} ({e:.0f}s)...', flush=True)
            if code.startswith(('8', '4')):
                continue
            fpath = os.path.join(self.tdx._market_dir(market), f'{market}{code}.day')
            if not os.path.exists(fpath):
                continue
            try:
                fsize = os.path.getsize(fpath)
                if fsize < RECORD_SIZE * 70:
                    continue
                # Read enough records to cover the full date range
                n_dates_needed = len(all_ds) + 60  # extra buffer
                read_bytes_needed = RECORD_SIZE * max(600, n_dates_needed)
                read_bytes = min(read_bytes_needed, fsize)
                with open(fpath, 'rb') as f:
                    f.seek(fsize - read_bytes)
                    raw = f.read(read_bytes)
                n_records = len(raw) // RECORD_SIZE
                if n_records < 70:
                    continue
                prev_close = None
                for dt, op, hi, lo, cl_int, amt, vol, _ in struct.iter_unpack('IIIIIfII', raw):
                    ds = str(dt)
                    close = cl_int / 100.0
                    chg = (close - prev_close) / prev_close if prev_close and prev_close > 0 else 0
                    if ds in all_ds and chg >= 0.07:
                        result[ds].append({
                            'code': code, 'close': round(close, 2),
                            'change_pct': round(chg * 100, 2),
                            'open': op / 100.0, 'high': hi / 100.0,
                            'low': lo / 100.0, 'volume': vol,
                            'amount': amt if isinstance(amt, float) else float(amt),
                            'prev_close': round(prev_close, 2) if prev_close else 0,
                            'is_zt': chg >= self._limit_threshold(code),
                        })
                    prev_close = close
            except Exception:
                continue

        with open(GAINERS_CACHE, 'w', encoding='utf-8') as f:
            json.dump(result, f)
        n_dates = sum(1 for v in result.values() if v)
        print(f'  Cache saved: {n_dates} dates with 7%+ gainers')
        return result

    def _preload_stocks(self, gainers: Dict[str, List[dict]], eligible_only: bool = False) -> int:
        """收集所有出现过7%+涨幅的股票代码，预加载完整数据+指标，缓存到磁盘。

        如果 eligible_only=True，只加载 self._eligible_codes 中的股票。
        """
        # 收集所有唯一代码
        all_codes = set()
        for day_list in gainers.values():
            for info in day_list:
                all_codes.add(info['code'])

        eligible = getattr(self, '_eligible_codes', None)
        if eligible_only and eligible is not None:
            codes_to_load = all_codes & eligible  # 交集：出现过7%+且在合格池中
            print(f'\nPhase 1: {len(all_codes)} 只候选 → 过滤后 {len(codes_to_load)} 只（主板+涨停>10）')
        else:
            codes_to_load = all_codes
            print(f'\nPhase 1: 预加载 {len(codes_to_load)} 只股票的数据和指标...')

        loaded = 0
        skipped = 0
        t0 = time.time()

        for i, code in enumerate(sorted(codes_to_load)):
            cache_path = os.path.join(STOCK_CACHE_DIR, f'{code}.pkl')
            if os.path.exists(cache_path):
                # 已有缓存，直接加载
                try:
                    with open(cache_path, 'rb') as f:
                        self._stock_cache[code] = pickle.load(f)
                    loaded += 1
                    continue
                except Exception:
                    pass

            # 读取并计算指标
            market = 'sh' if code.startswith('6') else 'sz'
            if code.startswith(('8', '4')):
                market = 'bj'
            try:
                df = self.tdx.read_daily(code, market)
                if df.empty or len(df) < 180:
                    skipped += 1
                    continue

                # ── ST 检查 ──
                # 从TDX数据中获取股票名称
                if 'name' in df.columns:
                    name = str(df['name'].iloc[0]) if not df.empty else ''
                else:
                    name = ''
                if self._is_stock_st(name):
                    skipped += 1
                    continue

                # 快速指标计算
                df = calc_ma(df, [5, 10, 20])

                # MAVOL180 = MA(volume, 180) * 1.2（对应副图指标 MAVOL180 红线）
                df['mavol180'] = df['volume'].rolling(180).mean() * 1.2

                # 截断到最近 400 行再做 zigzag（大幅提速，400行足够覆盖回测周期）
                df_zig = df.iloc[-400:].copy() if len(df) > 400 else df.copy()
                df_zig = calc_zigzag_find_top_line(df_zig)
                df['resistance'] = np.nan
                df.loc[df_zig.index, 'resistance'] = df_zig['find_top_line'].values

                # 将 trade_date 列统一为 YYYYMMDD 格式的索引
                if 'trade_date' in df.columns:
                    df.index = [str(d)[:10].replace('-', '') if hasattr(d, 'strftime')
                               else str(d)[:10].replace('-', '') for d in df['trade_date']]
                    df.index.name = 'date'

                # 只保留需要的列，减少内存
                keep_cols = ['open', 'high', 'low', 'close', 'volume',
                            'ma5', 'ma10', 'ma20', 'mavol180', 'resistance']
                df = df[[c for c in keep_cols if c in df.columns]]

                self._stock_cache[code] = df

                # 保存到磁盘缓存
                try:
                    with open(cache_path, 'wb') as f:
                        pickle.dump(df, f, protocol=pickle.HIGHEST_PROTOCOL)
                except Exception:
                    pass

                loaded += 1
                if (loaded) % 200 == 0:
                    e = time.time() - t0
                    rem = len(codes_to_load) - loaded
                    print(f'  Loaded {loaded}/{len(codes_to_load)} ({e:.0f}s, ETA {e/loaded*rem:.0f}s)...',
                          flush=True)
            except Exception:
                skipped += 1

        elapsed = time.time() - t0
        print(f'  Phase 1 完成: {loaded} 只已加载, {skipped} 只跳过 ({elapsed:.0f}s)')
        return loaded

    # ═══════════════════════════════════════════════════════════════════
    # Phase 2: 逐日检查 + 模拟交易
    # ═══════════════════════════════════════════════════════════════════

    def _get_stock_row(self, code: str, date_str: str) -> Optional[dict]:
        """从缓存中获取某只股票某一天的数据"""
        df = self._stock_cache.get(code)
        if df is None:
            return None
        try:
            if date_str in df.index:
                row = df.loc[date_str]
                return {
                    'open': float(row['open']), 'high': float(row['high']),
                    'low': float(row['low']), 'close': float(row['close']),
                    'volume': float(row['volume']), 'ma5': float(row.get('ma5', 0)),
                    'ma10': float(row.get('ma10', 0)),
                    'mavol180': float(row.get('mavol180', 0)),
                    'resistance': float(row.get('resistance', 0)),
                }
            # fallback: 搜索最接近的日期
            for idx in df.index:
                if str(idx)[:10] == date_str[:10]:
                    row = df.loc[idx]
                    return {
                        'open': float(row['open']), 'high': float(row['high']),
                        'low': float(row['low']), 'close': float(row['close']),
                        'volume': float(row['volume']), 'ma5': float(row.get('ma5', 0)),
                        'ma10': float(row.get('ma10', 0)),
                        'mavol180': float(row.get('mavol180', 0)),
                        'resistance': float(row.get('resistance', 0)),
                    }
        except Exception:
            pass
        return None

    def _check_signal(self, info: dict, trade_date: str) -> Optional[dict]:
        """检查买入条件（纯内存操作）"""
        code = info['code']
        row = self._get_stock_row(code, trade_date)
        if row is None:
            return None

        close = row['close']
        vol = row['volume']
        resistance = row['resistance']
        mavol180 = row['mavol180']

        # ── 条件1: 收盘价突破 zigzag 找顶线（最近的压力位） ──
        if np.isnan(resistance) or resistance <= 0:
            return None
        if close <= resistance:
            return None

        resist_label = '找顶线'

        # ── 条件2: 成交量 > MAVOL180 ──
        if np.isnan(mavol180) or mavol180 <= 0:
            return None
        if vol <= mavol180:
            return None

        vol_ratio = vol / mavol180
        breakthrough_pct = (close - resistance) / resistance * 100

        # ── 辅助: 非一字板 ──
        if row['open'] >= close * 1.095:
            return None

        # ── 硬性要求: 前一日收盘价必须在压力位下方（真突破） ──
        score = 20.0  # 基础分降低，留出区分空间
        reasons = []

        # ── 硬性要求: 前一日收盘价必须在压力位下方（真突破） ──
        prev_date = self._prev_trade_date(trade_date)
        prev_below_resistance = False
        prev_dist = 0
        if prev_date:
            prev_row = self._get_stock_row(code, prev_date)
            if prev_row:
                prev_res = prev_row['resistance']
                if prev_res > 0 and prev_row['close'] <= prev_res:
                    prev_below_resistance = True
                    prev_dist = (prev_res - prev_row['close']) / prev_res * 100
                    if 0 < prev_dist <= 10:
                        # 距压力位越近突破越有价值
                        if prev_dist <= 3:
                            score += 25  # 紧贴压力位，蓄势充分
                            reasons.append(f'紧贴{resist_label}{prev_dist:.1f}%→突破{breakthrough_pct:+.1f}%')
                        elif prev_dist <= 5:
                            score += 20
                            reasons.append(f'距{resist_label}{prev_dist:.1f}%→突破{breakthrough_pct:+.1f}%')
                        else:
                            score += 15
                            reasons.append(f'距{resist_label}{prev_dist:.1f}%→突破{breakthrough_pct:+.1f}%')
                    elif prev_dist <= 0:
                        return None
                    else:
                        return None
        if not prev_below_resistance:
            return None

        # 量能突破 MAVOL180（大分级）
        if vol_ratio >= 3.0:
            score += 20
            reasons.append(f'爆量{vol_ratio:.1f}倍MAVOL180')
        elif vol_ratio >= 2.0:
            score += 15
            reasons.append(f'显著放量{vol_ratio:.1f}倍MAVOL180')
        elif vol_ratio >= 1.5:
            score += 10
            reasons.append(f'放量{vol_ratio:.1f}倍MAVOL180')
        else:
            score += 5
            reasons.append(f'突破MAVOL180({vol_ratio:.1f}倍)')

        # 均线状态
        if row['ma5'] > 0 and row['ma10'] > 0:
            if row['ma5'] > row['ma10']:
                score += 8
                reasons.append('MA5>MA10多头')
            # 额外：站上MA5
            if close > row['ma5']:
                score += 5
                reasons.append('站上MA5')

        # 当日涨幅
        chg = info.get('change_pct', 0)
        if chg >= 9.5:
            score += 12
            reasons.append(f'涨停突破{chg:.1f}%')
        elif chg >= 7:
            score += 6
            reasons.append(f'大阳突破{chg:.1f}%')
        else:
            score += 2
            reasons.append(f'涨幅{chg:.1f}%')

        return {
            'code': code, 'score': round(score),  # 不设上限，满分约110+
            'raw_score': round(score),
            'close': close, 'volume': vol,
            'resistance': round(resistance, 2), 'resist_label': resist_label,
            'mavol180': round(mavol180, 0),
            'vol_ratio': round(vol_ratio, 2),
            'breakthrough_pct': round(breakthrough_pct, 2),
            'change_pct': chg, 'reasons': '; '.join(reasons),
            'ma5': round(row['ma5'], 2), 'ma10': round(row['ma10'], 2),
            'prev_dist_pct': round(prev_dist, 1),  # 前一日距压力位距离
        }

    def _prev_trade_date(self, date_str: str) -> Optional[str]:
        """获取前一交易日（简化版，周一到周五）"""
        try:
            d = datetime.strptime(date_str, '%Y%m%d')
            d -= timedelta(days=1)
            while d.weekday() >= 5:
                d -= timedelta(days=1)
            return d.strftime('%Y%m%d')
        except Exception:
            return None

    def _next_trade_date(self, date_str: str) -> Optional[str]:
        """获取下一交易日"""
        try:
            d = datetime.strptime(date_str, '%Y%m%d')
            d += timedelta(days=1)
            while d.weekday() >= 5:
                d += timedelta(days=1)
            return d.strftime('%Y%m%d')
        except Exception:
            return None

    # ═══════════════════════════════════════════════════════════════════
    # 卖出模拟
    # ═══════════════════════════════════════════════════════════════════

    def _simulate_sell(self, code: str, buy_date: str, buy_price: float,
                       signal: dict, all_dates_str: List[str]) -> Optional[dict]:
        """连板持有 / 断板卖出 / 跌破-6%止损 / 最多3天"""
        try:
            bd_idx = all_dates_str.index(buy_date)
        except ValueError:
            return None

        limit_threshold = self._limit_threshold(code)
        STOP_LOSS_PCT = -0.06  # 相对买入价跌幅≥6% → 止损
        had_limit_up = False

        for day_offset in range(1, 30):
            check_idx = bd_idx + day_offset
            if check_idx >= len(all_dates_str):
                sell_date = all_dates_str[-1]
                row = self._get_stock_row(code, sell_date)
                if row is None:
                    return None
                sell_price = row['close']
                days_held = check_idx - bd_idx
                gross = (sell_price - buy_price) / buy_price
                return {
                    'buy_date': buy_date, 'buy_price': round(buy_price, 2),
                    'sell_date': sell_date, 'sell_price': round(sell_price, 2),
                    'days_held': days_held,
                    'gross_ret': round(gross * 100, 2),
                    'net_ret': round((gross - TOTAL_COST) * 100, 2),
                    'is_win': gross > TOTAL_COST,
                    'exit_reason': '数据不足强平', 'had_limit_up': had_limit_up,
                }

            check_date = all_dates_str[check_idx]
            row = self._get_stock_row(code, check_date)
            if row is None:
                continue

            close = row['close']
            # 前一交易日收盘价
            prev_date = all_dates_str[check_idx - 1]
            prev_row = self._get_stock_row(code, prev_date)
            if prev_row is None:
                continue
            prev_close = prev_row['close']

            chg = (close - prev_close) / prev_close if prev_close > 0 else 0
            is_zt = chg >= limit_threshold
            days_held = day_offset

            # ── 硬止损: 收盘价相对买入价跌幅 ≥ 6% → 立即卖出 ──
            loss_from_buy = (close - buy_price) / buy_price
            if loss_from_buy <= STOP_LOSS_PCT:
                gross = loss_from_buy
                return {
                    'buy_date': buy_date, 'buy_price': round(buy_price, 2),
                    'sell_date': check_date, 'sell_price': round(close, 2),
                    'days_held': days_held,
                    'gross_ret': round(gross * 100, 2),
                    'net_ret': round((gross - TOTAL_COST) * 100, 2),
                    'is_win': False,
                    'exit_reason': '止损-6%', 'had_limit_up': had_limit_up,
                    'last_chg': round(chg * 100, 2),
                }

            if is_zt:
                had_limit_up = True
                continue  # 连板 → 继续持有

            # ── 不是涨停 → 判断卖出 ──
            if had_limit_up:
                # 断板卖出
                gross = (close - buy_price) / buy_price
                return {
                    'buy_date': buy_date, 'buy_price': round(buy_price, 2),
                    'sell_date': check_date, 'sell_price': round(close, 2),
                    'days_held': days_held,
                    'gross_ret': round(gross * 100, 2),
                    'net_ret': round((gross - TOTAL_COST) * 100, 2),
                    'is_win': gross > TOTAL_COST,
                    'exit_reason': '断板卖出', 'had_limit_up': had_limit_up,
                    'last_chg': round(chg * 100, 2),
                }
            elif days_held >= MAX_HOLD_BASELINE:
                # 到期卖出
                gross = (close - buy_price) / buy_price
                return {
                    'buy_date': buy_date, 'buy_price': round(buy_price, 2),
                    'sell_date': check_date, 'sell_price': round(close, 2),
                    'days_held': days_held,
                    'gross_ret': round(gross * 100, 2),
                    'net_ret': round((gross - TOTAL_COST) * 100, 2),
                    'is_win': gross > TOTAL_COST,
                    'exit_reason': f'持有{days_held}天到期', 'had_limit_up': had_limit_up,
                    'last_chg': round(chg * 100, 2),
                }

        return None

    # ═══════════════════════════════════════════════════════════════════
    # 主循环
    # ═══════════════════════════════════════════════════════════════════

    def run(self, lookback: int = 250, end_date: date = None):
        total_dates = lookback + 30
        all_dates_raw = self._trade_dates(total_dates, end_date=end_date)
        signal_dates = all_dates_raw[:-20]
        all_dates_str = [d.strftime('%Y%m%d') for d in all_dates_raw]

        print(f'回测周期: {signal_dates[0]} ~ {signal_dates[-1]}, {len(signal_dates)} 个信号日')
        print(f'选股: 沪深主板 | 年涨停>10次 | 非ST | 距压力位<10%')
        print(f'买入: 价格突破压力位 AND 成交量>MAVOL180')
        print(f'卖出: 连板持有 / 断板卖出 / 最多{MAX_HOLD_BASELINE}天到期')

        # Phase 1: 加载 gainers → 计算涨停次数 → 构建合格池 → 预加载股票
        t0 = time.time()
        gainers = self._load_gainers_cache(all_dates_raw)

        # ── 计算每只股票近一年涨停次数（从 gainers 缓存） ──
        print('计算涨停次数...')
        limit_up_counts = self._build_limit_up_count(gainers)
        print(f'  有涨停记录的股票: {len(limit_up_counts)} 只')

        # ── 构建合格股票池（主板 + 涨停>10 + 后续检查非ST） ──
        eligible_codes = set()
        for code, cnt in limit_up_counts.items():
            if self._is_main_board(code) and cnt > 10:
                eligible_codes.add(code)
        print(f'  合格股票池（主板+涨停>10）: {len(eligible_codes)} 只')

        # 只预加载合格池中的股票
        self._eligible_codes = eligible_codes
        n_loaded = self._preload_stocks(gainers, eligible_only=True)

        # Phase 2: 逐日扫描
        print(f'\nPhase 2: 逐日扫描交易信号...')
        all_trades = []
        daily_log = []
        signal_count = 0
        t0 = time.time()

        for i, td in enumerate(signal_dates):
            ds = td.strftime('%Y%m%d')
            if (i + 1) % 50 == 0 or i == 0:
                e = time.time() - t0
                eta = e / (i + 1) * (len(signal_dates) - i - 1) if i > 0 else 0
                n_t = len(all_trades)
                print(f'  [{i+1}/{len(signal_dates)}] {ds} '
                      f'信号累计{signal_count} | 交易{n_t} | {e:.0f}s ETA{eta:.0f}s',
                      flush=True)

            candidates = gainers.get(ds, [])
            day_trades = []

            for info in candidates:
                code = info['code']
                # ── 预过滤: 必须在合格池中（主板+涨停>10） ──
                if code not in eligible_codes:
                    continue
                signal = self._check_signal(info, ds)
                if signal is None:
                    continue
                signal_count += 1

                # 买入日 = 信号次日
                buy_date = self._next_trade_date(ds)
                if buy_date is None or buy_date not in all_dates_str:
                    continue

                # 获取买入日开盘价
                buy_row = self._get_stock_row(signal['code'], buy_date)
                if buy_row is None:
                    continue
                buy_open = buy_row['open']

                # 开盘涨停 → 买不到
                if buy_open >= signal['close'] * 1.095:
                    continue
                # 高开超5% → 追高风险
                if buy_open > signal['close'] * 1.05:
                    continue

                # 模拟卖出
                trade = self._simulate_sell(signal['code'], buy_date, buy_open, signal, all_dates_str)
                if trade is None:
                    continue

                trade['signal_date'] = ds
                trade['code'] = signal['code']
                trade['score'] = signal['score']
                trade['signal_close'] = signal['close']
                trade['signal_vol'] = signal['volume']
                trade['resistance'] = signal['resistance']
                trade['resist_label'] = signal.get('resist_label', '')
                trade['mavol180'] = signal['mavol180']
                trade['vol_ratio'] = signal['vol_ratio']
                trade['breakthrough_pct'] = signal['breakthrough_pct']
                trade['signal_chg'] = signal['change_pct']
                trade['reasons'] = signal['reasons']
                trade['ma5'] = signal.get('ma5', 0)
                trade['ma10'] = signal.get('ma10', 0)

                all_trades.append(trade)
                day_trades.append(trade)

            wins = sum(1 for t in day_trades if t['is_win'])
            daily_log.append({
                'date': ds,
                'signals': len(day_trades),
                'wins': wins,
                'losses': len(day_trades) - wins,
                'ret_sum': round(sum(t['net_ret'] for t in day_trades), 2),
            })

        elapsed = time.time() - t0
        df = pd.DataFrame(all_trades)
        ddf = pd.DataFrame(daily_log)

        if not df.empty:
            print(f'\nPhase 2 完成 ({elapsed:.0f}s)')
            print(f'  触发信号: {signal_count}')
            print(f'  完成交易: {len(df)}')
            print(f'  胜率: {df["is_win"].mean()*100:.1f}%')
            print(f'  均收益: {df["net_ret"].mean():+.2f}%')
        else:
            print(f'\nPhase 2 完成 ({elapsed:.0f}s) — 无符合条件的交易')

        return df, ddf, signal_count

    # ═══════════════════════════════════════════════════════════════════
    # 统计
    # ═══════════════════════════════════════════════════════════════════

    def _summarize(self, df: pd.DataFrame) -> dict:
        if df.empty:
            return {'trades': 0}

        wins = df[df['is_win']]
        losses = df[~df['is_win']]
        n = len(df)
        wr = len(wins) / n * 100 if n > 0 else 0

        # 累计收益：每笔1万元独立，总收益求和（非复利，因为每笔独立本金）
        cum_sum = df['net_ret'].sum()  # 总收益百分比之和
        # 逐笔复利参考（资金曲线用）
        equity = 1.0
        equity_peak = 1.0
        max_dd = 0.0
        for _, t in df.iterrows():
            equity *= (1 + t['net_ret'] / 100)
            equity_peak = max(equity_peak, equity)
            max_dd = max(max_dd, (equity_peak - equity) / equity_peak * 100)
        cum_ret = cum_sum  # 使用总收益求和作为主要累计指标

        exit_layers = {}
        for reason in df['exit_reason'].unique():
            sub = df[df['exit_reason'] == reason]
            sw = sub[sub['is_win']]
            exit_layers[reason] = {
                'trades': len(sub),
                'win_rate': round(len(sw) / len(sub) * 100, 1) if len(sub) > 0 else 0,
                'avg_ret': round(sub['net_ret'].mean(), 2),
            }

        with_zt = df[df['had_limit_up'] == True] if 'had_limit_up' in df.columns else pd.DataFrame()
        without_zt = df[df['had_limit_up'] == False] if 'had_limit_up' in df.columns else pd.DataFrame()

        # ── 按评分分层统计 ──
        score_layers = {}
        score_bins = [(90, 200, '90+'), (80, 90, '80-90'), (70, 80, '70-80'), (60, 70, '60-70'), (0, 60, '<60')]
        for lo, hi, label in score_bins:
            sub = df[(df['score'] >= lo) & (df['score'] < hi)]
            if len(sub) == 0:
                continue
            sw = sub[sub['is_win']]
            score_layers[label] = {
                'trades': len(sub),
                'win_rate': round(len(sw) / len(sub) * 100, 1) if len(sub) > 0 else 0,
                'avg_ret': round(sub['net_ret'].mean(), 2),
                'cum_ret': round(sub['net_ret'].sum(), 2),
            }

        df['month'] = pd.to_datetime(df['signal_date']).dt.strftime('%Y-%m')
        month_layers = {}
        for m in sorted(df['month'].unique()):
            sub = df[df['month'] == m]
            sw = sub[sub['is_win']]
            month_layers[m] = {
                'trades': len(sub),
                'win_rate': round(len(sw) / len(sub) * 100, 1) if len(sub) > 0 else 0,
                'avg_ret': round(sub['net_ret'].mean(), 2),
                'cum_ret': round(sub['net_ret'].sum(), 2),
            }

        return {
            'trades': n, 'wins': len(wins), 'losses': len(losses),
            'win_rate': round(wr, 1),
            'avg_ret': round(df['net_ret'].mean(), 2),
            'avg_win': round(wins['net_ret'].mean(), 2) if len(wins) > 0 else 0,
            'avg_loss': round(losses['net_ret'].mean(), 2) if len(losses) > 0 else 0,
            'median_ret': round(df['net_ret'].median(), 2),
            'cum_ret': round(cum_ret, 2), 'max_dd': round(max_dd, 2),
            'pl_ratio': round(abs(wins['net_ret'].mean() / losses['net_ret'].mean()), 2)
                        if len(losses) > 0 and len(wins) > 0 and losses['net_ret'].mean() != 0 else 0,
            'max_win': round(df['net_ret'].max(), 2),
            'max_loss': round(df['net_ret'].min(), 2),
            'avg_days': round(df['days_held'].mean(), 1) if 'days_held' in df.columns else 0,
            'exit_layers': exit_layers, 'month_layers': month_layers,
            'score_layers': score_layers,
            'with_zt_trades': len(with_zt),
            'with_zt_wr': round(with_zt['is_win'].mean() * 100, 1) if len(with_zt) > 0 else 0,
            'with_zt_avg': round(with_zt['net_ret'].mean(), 2) if len(with_zt) > 0 else 0,
            'without_zt_trades': len(without_zt),
            'without_zt_wr': round(without_zt['is_win'].mean() * 100, 1) if len(without_zt) > 0 else 0,
            'without_zt_avg': round(without_zt['net_ret'].mean(), 2) if len(without_zt) > 0 else 0,
            'return_dist': {
                'gt_20': int((df['net_ret'] > 20).sum()),
                '10_to_20': int(((df['net_ret'] > 10) & (df['net_ret'] <= 20)).sum()),
                '5_to_10': int(((df['net_ret'] > 5) & (df['net_ret'] <= 10)).sum()),
                '0_to_5': int(((df['net_ret'] > 0) & (df['net_ret'] <= 5)).sum()),
                'neg5_to_0': int(((df['net_ret'] > -5) & (df['net_ret'] <= 0)).sum()),
                'neg10_to_neg5': int(((df['net_ret'] > -10) & (df['net_ret'] <= -5)).sum()),
                'lt_neg10': int((df['net_ret'] <= -10).sum()),
            },
        }


# ═══════════════════════════════════════════════════════════════════════
# Excel 导出
# ═══════════════════════════════════════════════════════════════════════

HEADER_FONT = Font(name='Microsoft YaHei', bold=True, color='FFFFFF', size=11)
HEADER_FILL = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
WIN_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
LOSS_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
TITLE_FONT = Font(name='Microsoft YaHei', bold=True, size=14, color='2F5496')
SUB_FONT = Font(name='Microsoft YaHei', bold=True, size=11)
NUM_FONT = Font(name='Consolas', size=10)
THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9'),
)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
ZT_FILL = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')


def _style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT; cell.fill = HEADER_FILL
        cell.alignment = CENTER; cell.border = THIN_BORDER


def _style_data(ws, sr, er, nc):
    for r in range(sr, er + 1):
        for c in range(1, nc + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = THIN_BORDER; cell.alignment = CENTER; cell.font = NUM_FONT


def _auto_width(ws):
    for col in ws.columns:
        mx = max((len(str(c.value or '')) for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(mx + 4, 35)


def export_xlsx(df, ddf, summary, signal_count, path):
    wb = Workbook()

    # ════════════════════════════════════════════════════════════════
    # Sheet 1: 汇总
    # ════════════════════════════════════════════════════════════════
    ws = wb.active; ws.title = '汇总'
    ws['A1'] = '成交量突破MAVOL180 + 压力位突破 回测报告'
    ws['A1'].font = TITLE_FONT; ws['A1'].alignment = CENTER
    ws.merge_cells('A1:H1')

    ws['A3'] = (f'买入: 价格突破找顶线 AND 成交量>MAVOL180 | '
                f'卖出: 连板持有 / 断板卖出 / 最多3天 | '
                f'手续费: {TOTAL_COST*100:.2f}% | 每笔1万元')
    ws['A3'].font = Font(name='Microsoft YaHei', italic=True, color='666666', size=9)
    ws.merge_cells('A3:H3')

    row = 5
    ws.cell(row=row, column=1, value='核心指标').font = SUB_FONT
    row += 1
    for c, h in enumerate(['指标', '数值', '指标', '数值'], 1):
        ws.cell(row=row, column=c, value=h)
    _style_header(ws, row, 4); row += 1

    metrics = [
        ('触发信号数', signal_count, '完成交易', summary.get('trades', 0)),
        ('胜率', f'{summary.get("win_rate", 0):.1f}%', '平均收益率', f'{summary.get("avg_ret", 0):+.2f}%'),
        ('中位数收益', f'{summary.get("median_ret", 0):+.2f}%', '盈亏比', f'{summary.get("pl_ratio", 0):.2f}'),
        ('累计收益率', f'{summary.get("cum_ret", 0):+.2f}%', '最大回撤', f'{summary.get("max_dd", 0):.2f}%'),
        ('平均盈利', f'{summary.get("avg_win", 0):+.2f}%', '平均亏损', f'{summary.get("avg_loss", 0):+.2f}%'),
        ('单笔最大盈利', f'{summary.get("max_win", 0):+.2f}%', '单笔最大亏损', f'{summary.get("max_loss", 0):+.2f}%'),
        ('平均持有天数', f'{summary.get("avg_days", 0):.1f}天', '盈利笔数', summary.get('wins', 0)),
    ]
    for i, (k1, v1, k2, v2) in enumerate(metrics):
        ws.cell(row=row + i, column=1, value=k1); ws.cell(row=row + i, column=2, value=v1)
        ws.cell(row=row + i, column=3, value=k2); ws.cell(row=row + i, column=4, value=v2)
    _style_data(ws, row, row + len(metrics) - 1, 4)
    row += len(metrics) + 1

    # 连板效应
    ws.cell(row=row, column=1, value='连板效应分析').font = SUB_FONT; row += 1
    for c, h in enumerate(['类型', '笔数', '胜率', '均收益'], 1):
        ws.cell(row=row, column=c, value=h)
    _style_header(ws, row, 4); row += 1
    zt_data = [
        ('经历过连板', summary.get('with_zt_trades', 0),
         f'{summary.get("with_zt_wr", 0):.1f}%', f'{summary.get("with_zt_avg", 0):+.2f}%'),
        ('未经历连板', summary.get('without_zt_trades', 0),
         f'{summary.get("without_zt_wr", 0):.1f}%', f'{summary.get("without_zt_avg", 0):+.2f}%'),
    ]
    for i, (label, cnt, wr, avg) in enumerate(zt_data):
        ws.cell(row=row + i, column=1, value=label)
        ws.cell(row=row + i, column=2, value=cnt)
        ws.cell(row=row + i, column=3, value=wr)
        ws.cell(row=row + i, column=4, value=avg)
        if i == 0: ws.cell(row=row + i, column=1).fill = ZT_FILL
    _style_data(ws, row, row + 1, 4); row += 3

    # 收益分布
    ws.cell(row=row, column=1, value='收益分布').font = SUB_FONT; row += 1
    for c, h in enumerate(['区间', '笔数', '占比'], 1):
        ws.cell(row=row, column=c, value=h)
    _style_header(ws, row, 3); row += 1
    rd = summary.get('return_dist', {})
    total_n = max(summary.get('trades', 1), 1)
    dist_data = [
        ('> 20%', rd.get('gt_20', 0)), ('10% ~ 20%', rd.get('10_to_20', 0)),
        ('5% ~ 10%', rd.get('5_to_10', 0)), ('0 ~ 5%', rd.get('0_to_5', 0)),
        ('-5% ~ 0', rd.get('neg5_to_0', 0)), ('-10% ~ -5%', rd.get('neg10_to_neg5', 0)),
        ('< -10%', rd.get('lt_neg10', 0)),
    ]
    for label, cnt in dist_data:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=cnt)
        ws.cell(row=row, column=3, value=f'{cnt/total_n*100:.1f}%')
        if cnt > 0:
            if label.startswith('>') or '~' in label and not label.startswith('-'):
                ws.cell(row=row, column=1).fill = WIN_FILL
            elif label.startswith('<') or label.startswith('-'):
                ws.cell(row=row, column=1).fill = LOSS_FILL
        row += 1
    _style_data(ws, row - len(dist_data), row - 1, 3); row += 1

    # 退出原因
    ws.cell(row=row, column=1, value='按退出原因').font = SUB_FONT; row += 1
    for c, h in enumerate(['退出原因', '笔数', '胜率', '均收益'], 1):
        ws.cell(row=row, column=c, value=h)
    _style_header(ws, row, 4); row += 1
    for reason, stats in summary.get('exit_layers', {}).items():
        ws.cell(row=row, column=1, value=reason)
        ws.cell(row=row, column=2, value=stats['trades'])
        ws.cell(row=row, column=3, value=f'{stats["win_rate"]:.1f}%')
        ws.cell(row=row, column=4, value=f'{stats["avg_ret"]:+.2f}%')
        row += 1
    n_exit = len(summary.get('exit_layers', {}))
    if n_exit > 0:
        _style_data(ws, row - n_exit, row - 1, 4)
    row += 1

    # ── 按评分分层 ──
    ws.cell(row=row, column=1, value='📊 按评分分层').font = SUB_FONT; row += 1
    for c, h in enumerate(['分数段', '笔数', '胜率', '均收益', '累计收益'], 1):
        ws.cell(row=row, column=c, value=h)
    _style_header(ws, row, 5); row += 1
    for label, stats in summary.get('score_layers', {}).items():
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=stats['trades'])
        ws.cell(row=row, column=3, value=f'{stats["win_rate"]:.1f}%')
        ws.cell(row=row, column=4, value=f'{stats["avg_ret"]:+.2f}%')
        ws.cell(row=row, column=5, value=f'{stats["cum_ret"]:+.2f}%')
        if stats['avg_ret'] > 0:
            ws.cell(row=row, column=4).fill = WIN_FILL
        elif stats['avg_ret'] < 0:
            ws.cell(row=row, column=4).fill = LOSS_FILL
        if label == '90+':
            ws.cell(row=row, column=1).fill = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')
        row += 1
    n_score = len(summary.get('score_layers', {}))
    if n_score > 0:
        _style_data(ws, row - n_score, row - 1, 5)
    row += 1

    # 按月
    ws.cell(row=row, column=1, value='按月表现').font = SUB_FONT; row += 1
    for c, h in enumerate(['月份', '笔数', '胜率', '均收益', '月累计收益'], 1):
        ws.cell(row=row, column=c, value=h)
    _style_header(ws, row, 5); row += 1
    for m, stats in summary.get('month_layers', {}).items():
        ws.cell(row=row, column=1, value=m)
        ws.cell(row=row, column=2, value=stats['trades'])
        ws.cell(row=row, column=3, value=f'{stats["win_rate"]:.1f}%')
        ws.cell(row=row, column=4, value=f'{stats["avg_ret"]:+.2f}%')
        ws.cell(row=row, column=5, value=f'{stats["cum_ret"]:+.2f}%')
        if stats['cum_ret'] > 0: ws.cell(row=row, column=5).fill = WIN_FILL
        elif stats['cum_ret'] < 0: ws.cell(row=row, column=5).fill = LOSS_FILL
        row += 1
    n_month = len(summary.get('month_layers', {}))
    if n_month > 0:
        _style_data(ws, row - n_month, row - 1, 5)

    _auto_width(ws)

    # ════════════════════════════════════════════════════════════════
    # Sheet 2: 交易明细
    # ════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet('交易明细')
    ws2['A1'] = '逐笔交易明细'; ws2['A1'].font = TITLE_FONT; ws2['A1'].alignment = CENTER
    ws2.merge_cells('A1:U1')

    trade_headers = [
        '序号', '信号日', '买入日', '代码', '评分', '信号涨幅%',
        '信号价', '突破幅度%', '量比MAVOL180', '压力位', 'MAVOL180',
        '买入价', '卖出日', '卖出价', '持有天数', '净收益%',
        '结果', '退出原因', '是否连板', '卖出日涨跌%', '信号理由',
    ]
    ncols = len(trade_headers)
    for c, h in enumerate(trade_headers, 1):
        ws2.cell(row=3, column=c, value=h)
    _style_header(ws2, 3, ncols)

    r = 4
    for i, (_, t) in enumerate(df.iterrows()):
        ws2.cell(row=r, column=1, value=i + 1)
        ws2.cell(row=r, column=2, value=t.get('signal_date', ''))
        ws2.cell(row=r, column=3, value=t.get('buy_date', ''))
        ws2.cell(row=r, column=4, value=t.get('code', ''))
        ws2.cell(row=r, column=5, value=t.get('score', 0))
        ws2.cell(row=r, column=6, value=f'{t.get("signal_chg", 0):.1f}%')
        ws2.cell(row=r, column=7, value=t.get('signal_close', 0))
        ws2.cell(row=r, column=8, value=f'{t.get("breakthrough_pct", 0):+.1f}%')
        ws2.cell(row=r, column=9, value=f'{t.get("vol_ratio", 0):.1f}x')
        ws2.cell(row=r, column=10, value=t.get('resistance', 0))
        ws2.cell(row=r, column=11, value=t.get('mavol180', 0))
        ws2.cell(row=r, column=12, value=t.get('buy_price', 0))
        ws2.cell(row=r, column=13, value=t.get('sell_date', ''))
        ws2.cell(row=r, column=14, value=t.get('sell_price', 0))
        ws2.cell(row=r, column=15, value=t.get('days_held', 0))
        ws2.cell(row=r, column=16, value=t.get('net_ret', 0))
        ws2.cell(row=r, column=17, value='Win' if t.get('is_win') else 'Loss')
        ws2.cell(row=r, column=18, value=t.get('exit_reason', ''))
        ws2.cell(row=r, column=19, value='是' if t.get('had_limit_up') else '否')
        ws2.cell(row=r, column=20, value=f'{t.get("last_chg", 0):.1f}%' if t.get('last_chg') is not None else '')
        ws2.cell(row=r, column=21, value=t.get('reasons', ''))

        if t.get('is_win'):
            ws2.cell(row=r, column=16).fill = WIN_FILL
            ws2.cell(row=r, column=17).fill = WIN_FILL
        else:
            ws2.cell(row=r, column=16).fill = LOSS_FILL
            ws2.cell(row=r, column=17).fill = LOSS_FILL
        if t.get('had_limit_up'):
            ws2.cell(row=r, column=19).fill = ZT_FILL
        r += 1

    _style_data(ws2, 4, r - 1, ncols)
    _auto_width(ws2); ws2.freeze_panes = 'A4'

    # ════════════════════════════════════════════════════════════════
    # Sheet 3: 每日统计
    # ════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet('每日统计')
    ws3['A1'] = '每日交易统计'; ws3['A1'].font = TITLE_FONT; ws3['A1'].alignment = CENTER
    ws3.merge_cells('A1:G1')

    daily_headers = ['日期', '信号数', '胜', '负', '胜率%', '日收益%', '累计收益%']
    for c, h in enumerate(daily_headers, 1):
        ws3.cell(row=3, column=c, value=h)
    _style_header(ws3, 3, len(daily_headers))

    r = 4; cum = 0.0
    for _, di in ddf.iterrows():
        ws3.cell(row=r, column=1, value=di['date'])
        ws3.cell(row=r, column=2, value=di['signals'])
        ws3.cell(row=r, column=3, value=di['wins'])
        ws3.cell(row=r, column=4, value=di['losses'])
        wr = di['wins'] / max(di['signals'], 1) * 100
        ws3.cell(row=r, column=5, value=round(wr, 1))
        ws3.cell(row=r, column=6, value=di['ret_sum'])
        cum += di['ret_sum']
        ws3.cell(row=r, column=7, value=round(cum, 2))
        if di['signals'] > 0:
            if wr >= 60: ws3.cell(row=r, column=5).fill = WIN_FILL
            elif wr < 40: ws3.cell(row=r, column=5).fill = LOSS_FILL
            if di['ret_sum'] > 0: ws3.cell(row=r, column=6).fill = WIN_FILL
            elif di['ret_sum'] < 0: ws3.cell(row=r, column=6).fill = LOSS_FILL
        r += 1

    _style_data(ws3, 4, r - 1, len(daily_headers))
    _auto_width(ws3); ws3.freeze_panes = 'A4'

    wb.save(path)
    print(f'\n已保存: {path}')


# ═══════════════════════════════════════════════════════════════════════
# CLI
# ═══════════════════════════════════════════════════════════════════════

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='MAVOL180 成交量突破 + 压力位突破 回测')
    parser.add_argument('--days', type=int, default=250, help='回测信号日数量')
    parser.add_argument('--end-date', type=str, default=None, help='回测截止日 YYYY-MM-DD')
    args = parser.parse_args()

    end_dt = datetime.strptime(args.end_date, '%Y-%m-%d').date() if args.end_date else None
    bt = Vol180BreakoutBacktest()
    df, ddf, signal_count = bt.run(lookback=args.days, end_date=end_dt)

    if df.empty:
        print('\n未找到符合条件的交易信号。')
        print('可能原因: 候选池仅7%+涨幅股 / 数据不足180天 / 市场环境不配合')
        import sys; sys.exit(0)

    summary = bt._summarize(df)

    print(f'\n{"="*65}')
    print(f'  回测结果摘要')
    print(f'{"="*65}')
    print(f'  触发信号: {signal_count}  |  完成交易: {summary["trades"]}  |  胜率: {summary["win_rate"]:.1f}%')
    print(f'  均收益: {summary["avg_ret"]:+.2f}%  |  中位数: {summary["median_ret"]:+.2f}%')
    print(f'  累计: {summary["cum_ret"]:+.2f}%  |  最大回撤: {summary["max_dd"]:.2f}%  |  盈亏比: {summary["pl_ratio"]:.2f}')
    print(f'  连板交易: {summary["with_zt_trades"]}笔(胜率{summary["with_zt_wr"]:.1f}% 均{summary["with_zt_avg"]:+.2f}%)')
    print(f'  非连板:   {summary["without_zt_trades"]}笔(胜率{summary["without_zt_wr"]:.1f}% 均{summary["without_zt_avg"]:+.2f}%)')

    for reason, stats in summary.get('exit_layers', {}).items():
        print(f'  {reason}: {stats["trades"]}笔 胜率{stats["win_rate"]:.1f}% 均{stats["avg_ret"]:+.2f}%')

    print(f'\n  按评分分层:')
    print(f'  {"分数段":<10s} {"笔数":>6s} {"胜率":>8s} {"均收益":>9s} {"累计收益":>10s}')
    print(f'  {"-"*45}')
    for label, stats in summary.get('score_layers', {}).items():
        print(f'  {label:<10s} {stats["trades"]:>6d} {stats["win_rate"]:>7.1f}% '
              f'{stats["avg_ret"]:>+8.2f}% {stats["cum_ret"]:>+9.1f}%')

    rd = summary.get('return_dist', {})
    print(f'  收益分布: >20%:{rd.get("gt_20",0)} | 10~20%:{rd.get("10_to_20",0)} | '
          f'5~10%:{rd.get("5_to_10",0)} | 0~5%:{rd.get("0_to_5",0)} | '
          f'-5~0:{rd.get("neg5_to_0",0)} | -10~-5:{rd.get("neg10_to_neg5",0)} | <-10%:{rd.get("lt_neg10",0)}')

    out_path = os.path.join(DATA_DIR, f'vol180_breakout_backtest_{args.days}d.xlsx')
    export_xlsx(df, ddf, summary, signal_count, out_path)
