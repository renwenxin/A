"""启动突破战法 完整回测 — 按复盘.txt v2 规范

逻辑:
  第一层: 板块环境过滤(4条件) → 只在强势板块内选股
  第二层: 个股技术信号(8条件) → T日突破信号股
  第三层: T+1~T+5回踩确认(5条件) → 次日开盘买入
  第四层: 持有3天卖出(主) / 动态止盈止损(辅)

统计: 胜率/盈亏比/累计收益/最大回撤/分层统计/逐日逐周/图表
"""
import sys, os, struct, json, argparse, time
from datetime import date, datetime, timedelta
from typing import Dict, List, Optional, Tuple
from collections import defaultdict
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

from ashare_review.data.tdx_reader import TdxReader, RECORD_SIZE
from ashare_review.analysis.indicators import enrich_all, calc_ma

# ─── 常量 ──────────────────────────────────────────────────────────────
FEE = 0.0015          # 手续费 0.15% 来回
SLIPPAGE_BUY = 0.001  # 买入滑点 +0.1%
SLIPPAGE_SELL = 0.001 # 卖出滑点 -0.1%
TOTAL_COST = FEE + SLIPPAGE_BUY + SLIPPAGE_SELL  # 0.35%
HOLD_DAYS = 3
MAX_WAIT = 5          # 信号后最多等5天
PER_STOCK_CAPITAL = 10000  # 每只股固定1万元
ZHONGJUN_AMOUNT = 2_000_000_000  # 中军成交额阈值: 20亿(元)
DYNAMIC_HOLD_MAX = 5  # 动态卖出最长持有天数


class StartBreakoutBacktest:
    def __init__(self, skip_sector_filter: bool = False):
        self.tdx = TdxReader()
        self.skip_sector_filter = skip_sector_filter
        self._cache_dir = os.path.join(
            os.path.dirname(os.path.dirname(__file__)), 'data')
        os.makedirs(self._cache_dir, exist_ok=True)
        self.industry_map: Dict[str, str] = {}
        self._gainers_cache: Dict = None  # 懒加载  # code -> industry_name

    # ─── 交易日 ────────────────────────────────────────────────────

    @staticmethod
    def _trade_dates(n: int, end_date: date = None) -> List[date]:
        d = end_date or (date.today() - timedelta(days=1))
        dates = []
        while len(dates) < n + MAX_WAIT + DYNAMIC_HOLD_MAX + 10:
            if d.weekday() < 5: dates.append(d)
            d -= timedelta(days=1)
        return list(reversed(dates))

    # ─── 行业映射 (从涨停池历史数据构建) ──────────────────────────

    def _build_industry_map(self, trade_dates: List[date]) -> Dict[str, str]:
        """从历史涨停池数据构建 code→行业名称 映射。

        使用 ak.stock_zt_pool_em() 采样过去交易日的涨停池，
        从 '所属行业' 字段提取每只股票的行业分类。
        采样每5天一次，约50次API调用，缓存永久有效。
        """
        cache_path = os.path.join(self._cache_dir, 'industry_map.json')
        if os.path.exists(cache_path):
            try:
                with open(cache_path, 'r', encoding='utf-8') as f:
                    imap = json.load(f)
                if len(imap) > 500:
                    print(f'Loaded industry map: {len(imap)} stocks')
                    return imap
                print(f'Industry map too small ({len(imap)}), rebuilding...')
            except Exception:
                pass

        # 清除代理环境变量（部分 eastmoney 推流主机被代理拦截）
        for k in ('HTTP_PROXY', 'HTTPS_PROXY', 'http_proxy', 'https_proxy', 'ALL_PROXY'):
            os.environ.pop(k, None)

        # 采样: 取最近60个交易日（API仅支持近期数据）
        sample_dates = trade_dates[-60:] if len(trade_dates) > 60 else trade_dates
        print(f'Building industry map from {len(sample_dates)} sample dates '
              f'({sample_dates[0]} ~ {sample_dates[-1]})...')

        imap: Dict[str, str] = {}
        try:
            import akshare as ak
        except ImportError:
            print('[WARN] akshare not available, sector filter disabled')
            return {}

        for i, d in enumerate(sample_dates):
            ds = d.strftime('%Y%m%d')
            try:
                df = ak.stock_zt_pool_em(date=ds)
                if df is not None and not df.empty:
                    code_col = next((c for c in df.columns if c == '代码'), None)
                    ind_col = next((c for c in df.columns if c == '所属行业'), None)
                    if code_col and ind_col:
                        for _, row in df.iterrows():
                            code = str(row[code_col]).strip().zfill(6)
                            industry = str(row[ind_col]).strip()
                            if len(code) == 6 and code.isdigit() and industry and industry != 'nan':
                                if code not in imap:
                                    imap[code] = industry
                time.sleep(0.4)  # 限流
            except Exception:
                continue
            if (i + 1) % 10 == 0 or i == 0:
                print(f'  Sample {i+1}/{len(sample_dates)}: {len(imap)} stocks mapped...',
                      flush=True)

        print(f'  Industry map built: {len(imap)} stocks')

        # 缓存
        try:
            with open(cache_path, 'w', encoding='utf-8') as f:
                json.dump(imap, f, ensure_ascii=False)
        except Exception:
            pass

        return imap

    # ─── 涨停阈值 ──────────────────────────────────────────────────

    @staticmethod
    def _limit_threshold(code: str) -> float:
        code = str(code).zfill(6)
        if code.startswith(('300', '301', '688')): return 0.199
        if code.startswith(('8', '4')): return 0.299
        return 0.095

    # ─── 预扫描: 7%+涨幅股 + 行业统计 ─────────────────────────────

    def _pre_scan_gainers(self, dates: List[date], industry_map: Dict[str, str]
                          ) -> Tuple[Dict[str, List[dict]], Dict[str, dict]]:
        """扫描全市场, 返回:
        1. date→7%涨幅股列表 (同旧版)
        2. date→{industry: {sum_gain, count, zt_count, zhongjun}} 行业统计
        """
        cache_path = os.path.join(self._cache_dir, 'gainers_7pct.json')
        sector_cache_path = os.path.join(self._cache_dir, 'sector_daily_stats.json')
        all_ds = {d.strftime('%Y%m%d') for d in dates}

        # 尝试加载 gainers 缓存
        cached_gainers = None
        if os.path.exists(cache_path):
            with open(cache_path) as f:
                cached_gainers = json.load(f)
            if not all_ds.issubset(set(cached_gainers.keys())):
                cached_gainers = None

        # 尝试加载 sector_stats 缓存
        cached_sector = None
        if os.path.exists(sector_cache_path):
            with open(sector_cache_path, 'r', encoding='utf-8') as f:
                cached_sector = json.load(f)
            if not all_ds.issubset(set(cached_sector.keys())):
                cached_sector = None

        if cached_gainers is not None and cached_sector is not None:
            print(f'Loaded gainers cache: {len(cached_gainers)} dates')
            print(f'Loaded sector stats cache: {len(cached_sector)} dates')
            return (
                {k: v for k, v in cached_gainers.items() if k in all_ds},
                {k: v for k, v in cached_sector.items() if k in all_ds},
            )

        if cached_gainers:
            print(f'Cache has gainers ({len(cached_gainers)} dates) but sector stats missing, rebuilding both...')

        result: Dict[str, list] = {d: [] for d in all_ds}
        # sector_stats: {ds: {industry: {sum_gain, count, zt_count, best_zj_code, best_zj_amt, best_zj_gain}}}
        sector_stats: Dict[str, dict] = {d: {} for d in all_ds}

        stocks = self.tdx.list_stocks()
        total = len(stocks)
        t0 = time.time()
        has_imap = len(industry_map) > 0

        for si, (code, market) in enumerate(stocks):
            if (si + 1) % 1500 == 0:
                e = time.time() - t0
                print(f'  Scan {si+1}/{total} ({e:.0f}s, ETA {e/(si+1)*(total-si-1):.0f}s)...', flush=True)

            # 跳过北交所、ST类
            if code.startswith(('8', '4')): continue

            industry = industry_map.get(code, '') if has_imap else ''

            fpath = os.path.join(self.tdx._market_dir(market), f'{market}{code}.day')
            if not os.path.exists(fpath): continue
            try:
                fsize = os.path.getsize(fpath)
                if fsize < RECORD_SIZE * 70: continue
                read_bytes = min(RECORD_SIZE * 600, fsize)
                with open(fpath, 'rb') as f:
                    f.seek(fsize - read_bytes)
                    raw = f.read(read_bytes)
                n_records = len(raw) // RECORD_SIZE
                if n_records < 70: continue

                prev_close = None
                for j, (dt, op, hi, lo, cl_int, amt, vol, _) in enumerate(
                    struct.iter_unpack('IIIIIfII', raw)
                ):
                    ds = str(dt)
                    close = cl_int / 100.0
                    chg = (close - prev_close) / prev_close if prev_close and prev_close > 0 else 0
                    is_zt = chg >= self._limit_threshold(code)
                    is_7pct = chg >= 0.07
                    amount = amt if isinstance(amt, float) else float(amt)

                    if ds in all_ds:
                        # --- 行业统计累加 ---
                        if industry:
                            sec = sector_stats[ds]
                            if industry not in sec:
                                sec[industry] = {
                                    'sum_gain': 0.0, 'count': 0, 'zt_count': 0,
                                    'best_zj_code': '', 'best_zj_amt': 0.0, 'best_zj_gain': 0.0,
                                }
                            s = sec[industry]
                            s['sum_gain'] += chg * 100  # 百分比
                            s['count'] += 1
                            if is_zt:
                                s['zt_count'] += 1
                            # 中军候选: 涨幅≥5% 且 成交额≥20亿
                            if chg >= 0.05 and amount >= ZHONGJUN_AMOUNT:
                                if amount > s['best_zj_amt']:
                                    s['best_zj_amt'] = amount
                                    s['best_zj_code'] = code
                                    s['best_zj_gain'] = chg * 100

                        # --- 7%+涨幅股记录 ---
                        if is_zt or is_7pct:
                            result[ds].append({
                                'code': code,
                                'close': round(close, 2),
                                'change_pct': round(chg * 100, 2),
                                'open': op / 100.0,
                                'high': hi / 100.0,
                                'low': lo / 100.0,
                                'volume': vol,
                                'amount': amount,
                                'prev_close': round(prev_close, 2) if prev_close else 0,
                                'is_zt': is_zt,
                                'industry': industry,
                            })

                    prev_close = close
            except Exception:
                continue

        # 后处理: 计算行业平均涨幅
        for ds in all_ds:
            for ind, s in sector_stats[ds].items():
                s['avg_gain'] = round(s['sum_gain'] / s['count'], 2) if s['count'] > 0 else 0.0
                s['has_zhongjun'] = s['best_zj_amt'] >= ZHONGJUN_AMOUNT

        # 缓存 gainers
        os.makedirs(os.path.dirname(cache_path), exist_ok=True)
        with open(cache_path, 'w') as f:
            json.dump(result, f)
        # 缓存 sector_stats
        with open(sector_cache_path, 'w', encoding='utf-8') as f:
            json.dump(sector_stats, f, ensure_ascii=False)

        n_dates = sum(1 for v in result.values() if v)
        print(f'  Cache saved. {n_dates} dates with 7%+ gainers, '
              f'{sum(1 for v in sector_stats.values() if v)} dates with sector stats.')
        return result, sector_stats

    # ─── 读取单只股票日线（带日期过滤 + 指标） ──────────────────────

    def _read_stock(self, code: str, up_to_date: str = None) -> Optional[pd.DataFrame]:
        market = 'sh' if str(code).startswith('6') else 'sz'
        try:
            df = self.tdx.read_daily(str(code).zfill(6), market)
            if df.empty or len(df) < 60: return None
            df = enrich_all(df)
            df = calc_ma(df, [5, 10, 20])
            if up_to_date:
                try:
                    target = datetime.strptime(up_to_date, '%Y%m%d').date()
                except ValueError:
                    target = None
                if target:
                    df = df[df['trade_date'].apply(
                        lambda x: (x.date() if hasattr(x, 'date') else x) <= target
                    )]
            return df if len(df) >= 40 else None
        except Exception:
            return None

    # ─── 板块环境过滤 (4条件) ──────────────────────────────────────

    def _sector_pass(self, industry: str, date_str: str, sector_stats: Dict[str, dict],
                     prev_dates: List[str]) -> bool:
        """检查板块是否满足4个环境条件。

        条件:
        1. 板块当日平均涨幅 ≥ 2%
        2. 板块内涨停家数 ≥ 3只
        3. 板块内有中军（成交额≥20亿 且 涨幅≥5%）
        4. 近3日板块涨幅≥0的天数 ≥ 2天（持续走强）
        """
        if self.skip_sector_filter or not industry:
            return True

        today = sector_stats.get(date_str, {}).get(industry)
        if today is None:
            return False

        # 条件1: 板块涨幅 ≥ 2%
        if today.get('avg_gain', 0) < 2.0:
            return False

        # 条件2: 涨停 ≥ 3只
        if today.get('zt_count', 0) < 3:
            return False

        # 条件3: 有中军
        if not today.get('has_zhongjun', False):
            return False

        # 条件4: 近3日 ≥2天涨幅为正
        pos_days = 0
        for ds in prev_dates[-3:]:
            if ds not in sector_stats:
                continue
            sec = sector_stats[ds].get(industry)
            if sec and sec.get('avg_gain', 0) > 0:
                pos_days += 1
        if pos_days < 2:
            return False

        return True

    # ─── 市场状态采集（用于 Regime Research，不影响评分） ────────

    def _get_market_state(self, trade_date: str = None) -> dict:
        """获取指定交易日的市场环境数据。

        使用 gainers_7pct 缓存（7%+涨幅股数作为市场热度代理变量）
        和上证指数 MA60 方向，不新增实时 API 调用。
        """
        state = {'sh_ma60_up': 0, 'up_ratio': 0,
                 'limit_up_num': 0, 'limit_down_num': 0}

        if not trade_date:
            return state

        try:
            # 市场热度：7%+涨幅股数（内存缓存）
            if self._gainers_cache is None:
                cache_path = os.path.join(self._cache_dir, 'gainers_7pct.json')
                if os.path.exists(cache_path):
                    with open(cache_path, 'r', encoding='utf-8') as f:
                        self._gainers_cache = json.load(f)
                else:
                    self._gainers_cache = {}
            key = trade_date.replace('-', '') if '-' in trade_date else trade_date
            day_gainers = self._gainers_cache.get(key, [])
            state['limit_up_num'] = len(day_gainers) if day_gainers else 0

            # 上证指数 MA60 方向 + 涨跌
            df_sh = self.tdx.read_daily('999999', 'sh', up_to_date=trade_date)
            if df_sh is not None and len(df_sh) >= 60:
                close = float(df_sh['close'].iloc[-1])
                ma60 = float(df_sh['close'].rolling(60).mean().iloc[-1])
                state['sh_ma60_up'] = 1 if close > ma60 else 0
                if len(df_sh) >= 2:
                    prev = float(df_sh['close'].iloc[-2])
                    state['up_ratio'] = 1.0 if close > prev else 0.0
        except Exception:
            pass
        return state

    def _get_cached_limit_ups(self, trade_date: str = None) -> list:
        """从缓存获取指定日期的涨停股列表。"""
        cache_path = os.path.join(self._cache_dir, 'zt_scan_cache.json')
        if not os.path.exists(cache_path):
            return []
        try:
            with open(cache_path, 'r', encoding='utf-8') as f:
                cache = json.load(f)
            key = trade_date.replace('-', '') if trade_date else ''
            if key in cache:
                return cache[key]
            # 尝试不同日期格式
            for k in cache:
                if k.startswith(key) or key.startswith(k):
                    return cache[k]
        except Exception:
            pass
        return []

    # ─── T日: 信号检测 (8条件 + 板块过滤 + 增强评分) ─────────────

    def _check_signal(self, info: dict, trade_date: str,
                      sector_stats: Dict[str, dict],
                      prev_dates: List[str]) -> Optional[dict]:
        """对7%+涨幅股检查8条件+板块过滤。返回评分dict或None"""
        code = info['code']
        industry = info.get('industry', '')

        # ── 板块环境过滤 ──
        if not self._sector_pass(industry, trade_date, sector_stats, prev_dates):
            return None

        df = self._read_stock(code, up_to_date=trade_date)
        if df is None or df.empty: return None
        idx = len(df) - 1

        close   = float(df['close'].iloc[idx])
        open_p  = float(df['open'].iloc[idx])
        high    = float(df['high'].iloc[idx])
        low     = float(df['low'].iloc[idx])
        vol     = float(df['volume'].iloc[idx])
        amount  = float(df.get('amount', pd.Series([0])).iloc[idx]) if 'amount' in df.columns else vol * close

        score = 0
        conditions = []

        # 条件1: 当日涨幅 >= 7% (已由预扫描保证)
        chg = info['change_pct']
        conditions.append(f'涨幅{chg:.1f}%')

        # 条件2: 放量突破 — vol >= 前5日均量1.5倍（前5日不含当天）
        vol_ma5 = float(df['volume'].rolling(5).mean().shift(1).iloc[idx])
        if pd.isna(vol_ma5) or vol_ma5 <= 0:
            return None
        vol_ratio = vol / vol_ma5
        vol_ok = vol_ratio >= 1.5
        if not vol_ok:
            return None  # 核心条件不满足
        # 换手率评分: 温和放量(1.5~3x)最优, 过度放量(>5x)有害
        if 1.5 <= vol_ratio <= 3.0:
            score += 15
            conditions.append(f'温和放量{vol_ratio:.1f}倍')
        elif 3.0 < vol_ratio <= 5.0:
            score += 10
            conditions.append(f'放量{vol_ratio:.1f}倍')
        else:
            # >5x 过度放量，不加分
            conditions.append(f'过度放量{vol_ratio:.1f}倍')

        # 条件3: 收盘价突破60日平台
        # 优先: 突破前60日高点（大平台突破）
        high_60 = float(df['high'].iloc[max(0, idx-60):idx].max())
        breakout_60 = close >= high_60
        # 备选: 突破30日盘整平台（横盘≥20天，波动<15%，今日突破30日高点）
        high_30 = float(df['high'].iloc[max(0, idx-30):idx].max())
        high_30_range = float(df['high'].iloc[max(0, idx-30):idx].max()) - float(df['high'].iloc[max(0, idx-30):idx].min())
        high_30_min = float(df['high'].iloc[max(0, idx-30):idx].min())
        range_pct = high_30_range / high_30_min if high_30_min > 0 else 1
        is_range = range_pct < 0.15  # 30日盘整波动<15%
        breakout_30_platform = close >= high_30 and is_range

        if breakout_60:
            score += 20
            conditions.append('突破60日高点')
        elif breakout_30_platform:
            score += 15
            conditions.append('突破30日盘整平台')
        else:
            return None

        # 条件4: 实体阳线 + 上影<=实体1/2
        body = abs(close - open_p)
        upper_shadow = high - max(close, open_p)
        yang = close > open_p
        shadow_ok = upper_shadow <= body / 2 if body > 0 else upper_shadow <= 0.005 * close
        if yang and shadow_ok:
            score += 5
            conditions.append('阳线实体')
        # 不强制

        # 条件5: 5日 > 10日 > 20日均线 + 趋势向上
        ma5  = float(df['ma5'].iloc[idx])
        ma10 = float(df['ma10'].iloc[idx])
        ma20 = float(df['ma20'].iloc[idx])
        ma_bull = ma5 > ma10 > ma20
        if not ma_bull:
            return None
        score += 10
        conditions.append('均线多头')

        # 条件5b: 趋势过滤 — MA20向上 且 MA60向上（5日窗口比较）
        if 'ma60' in df.columns and idx >= 5:
            ma20_5d = float(df['ma20'].iloc[idx-5])
            ma60_cur = float(df['ma60'].iloc[idx])
            ma60_5d = float(df['ma60'].iloc[idx-5])
            trend_up = ma20 > ma20_5d and ma60_cur > ma60_5d
            if trend_up:
                score += 10
                conditions.append('趋势向上')
            else:
                return None  # 强制：不做下降趋势里的突破

        # 条件6: 流通市值 20-80亿（收紧范围，避免大盘股和妖股）
        avg_price = close
        est_cap = avg_price * vol_ma5 * 250 / 1e8
        cap_ok = 20 <= est_cap <= 80
        if cap_ok:
            score += 5
            conditions.append(f'市值~{est_cap:.0f}亿')
        # 不强制

        # 条件7: 非ST/非北交所 (已在前置过滤)
        score += 5
        conditions.append('非ST/北交所')

        # 条件8: 近10日有涨停或8%+大阳线
        pct_chg = df['close'].pct_change()
        recent_big = (pct_chg.iloc[max(0, idx-10):idx+1] >= 0.08).any()
        if recent_big:
            score += 10
            conditions.append('近10日有大阳线')
        # 不强制

        # ── 至此为止是 V1 基础分 ──
        # V2 评分移至回踩确认阶段：_calc_v2_score_at_pullback()
        v1_score = min(score, 100)

        # ── 留存数值特征（供回踩阶段V2评分使用） ──
        pullback_features = {}
        # V2因子: 距250日高点
        if 'ma250' in df.columns:
            high_250_all = float(df['high'].iloc[max(0, idx-250):idx].max())
            dist_250 = (close - high_250_all) / high_250_all if high_250_all > 0 else 0
            pullback_features['dist_250d'] = dist_250
        else:
            pullback_features['dist_250d'] = 0

        # V2因子: 前60日涨幅
        if idx >= 60:
            close_60d = float(df['close'].iloc[idx-60])
            chg_60d = (close - close_60d) / close_60d * 100
            pullback_features['chg_60d'] = chg_60d
        else:
            pullback_features['chg_60d'] = 0

        pullback_features['vol_ratio'] = vol_ratio
        pullback_features['est_cap'] = est_cap

        # V1层级
        if v1_score >= 90: v1_tier = 'S'
        elif v1_score >= 80: v1_tier = 'A'
        elif v1_score >= 70: v1_tier = 'B'
        elif v1_score >= 60: v1_tier = 'C'
        else: v1_tier = 'D'

        return {
            'code': code,
            'v1_score': v1_score,          # V1基础分（不含回踩质量）
            'v1_tier': v1_tier,
            'pullback_features': pullback_features,  # 数值特征（供V2回踩评分用）
            'close': close, 'volume': vol,
            'ma5': ma5, 'ma10': ma10, 'ma20': ma20,
            'high_break': max(high_60, high_30),
            'industry': industry,
            'conditions': '; '.join(conditions),
            'market_state': self._get_market_state(trade_date=trade_date),
        }

    # ─── T+1~T+5: 回踩确认 ──────────────────────────────────────────

    def _find_pullback(self, signal: dict, signal_date: str,
                       all_dates: List[str]) -> Optional[dict]:
        """在信号日后5天内找回踩确认。返回 dict(pullback) 或 None

        pullback = {
            'buy_date', 'buy_price', 'pullback_days',
            'check_date',  # 回踩确认日（T+N）
            'check_idx_in_all_dates',
            'pb_low', 'pb_close', 'pb_open', 'pb_high', 'pb_vol',  # 回踩日数据
            'trigger_type',  # 'hammer'/'doji'/'engulf'
        }
        """
        code = signal['code']
        signal_close = signal['close']

        try:
            sd_idx = all_dates.index(signal_date)
        except ValueError:
            return None

        for offset in range(1, MAX_WAIT + 1):
            check_idx = sd_idx + offset
            if check_idx >= len(all_dates): break
            check_date = all_dates[check_idx]

            df = self._read_stock(code, up_to_date=check_date)
            if df is None or df.empty: continue
            idx = len(df) - 1

            close  = float(df['close'].iloc[idx])
            low    = float(df['low'].iloc[idx])
            open_p = float(df['open'].iloc[idx])
            high   = float(df['high'].iloc[idx])
            vol    = float(df['volume'].iloc[idx])
            ma5    = float(df['ma5'].iloc[idx])
            ma10   = float(df['ma10'].iloc[idx])

            # 条件1: 最低价触及MA5或MA10，收盘价站回
            touched_ma = (low <= ma5 * 1.01) or (low <= ma10 * 1.01)
            above_ma = close > ma5 or close > ma10
            pullback = touched_ma and above_ma
            if not pullback: continue

            # 条件2: 缩量 — vol <= 信号日vol*60% AND vol <= 前5日均量*80%（前5日不含当天）
            vol_ma5_now_series = df['volume'].rolling(5).mean().shift(1)
            vol_ma5_now = float(vol_ma5_now_series.iloc[idx])
            vol_shrink = (vol <= signal['volume'] * 0.6) and (not pd.isna(vol_ma5_now) and vol <= vol_ma5_now * 0.8)
            if not vol_shrink: continue

            # 条件3: K线止跌 — 下影线>=实体0.5倍 或 十字星 或 阳包阴
            body = abs(close - open_p)
            lower_shadow = min(close, open_p) - low
            is_hammer = lower_shadow >= body * 0.5 if body > 0 else lower_shadow > 0
            is_doji = body <= close * 0.005
            engulf = False
            if idx >= 1:
                prev_c = float(df['close'].iloc[idx-1])
                prev_o = float(df['open'].iloc[idx-1])
                prev_body = prev_c - prev_o
                today_body = close - open_p
                engulf = (prev_body < 0) and (today_body > 0) and (close > prev_o) and (open_p < prev_c)
            stop_signal = is_hammer or is_doji or engulf
            if not stop_signal: continue

            # 确认次日开盘买入
            if check_idx + 1 >= len(all_dates): break
            buy_date = all_dates[check_idx + 1]

            # 读买入日开盘
            df2 = self._read_stock(code, up_to_date=buy_date)
            if df2 is None or df2.empty: continue
            buy_open = float(df2['open'].iloc[-1])

            # 开盘涨停(买不进)跳过
            if buy_open >= signal_close * 1.095: continue

            # 买入价 > 信号日收盘+2% 跳过
            if buy_open > signal_close * 1.02: continue

            # 价格跳空高开超信号日3%跳过
            if buy_open > signal_close * 1.03: continue

            return {
                'buy_date': buy_date, 'buy_price': buy_open,
                'pullback_days': offset,
                'check_date': check_date,
                'check_idx_in_all_dates': check_idx,
                'pb_open': open_p, 'pb_high': high, 'pb_low': low,
                'pb_close': close, 'pb_vol': vol,
                'trigger_type': 'hammer' if is_hammer else ('doji' if is_doji else 'engulf'),
            }

        return None

    # ─── V2评分: 在回踩确认后计算（信号质量 + 回踩质量） ──────────

    def _calc_v2_score_at_pullback(self, signal: dict, pb: dict) -> dict:
        """回踩确认后计算V2综合评分。

        V2 = 信号强度(0-40) + 位置优势(0-30) + 回踩质量(0-30)，满分100。

        因子验证来源（backtest分析）:
        - 距年高: 区分度+127% (最大)
        - 60日动量: 区分度+28%
        - 回踩质量: 回踩幅度/缩量程度/K线形态类型/MA支撑强度
        """
        pf = signal.get('pullback_features', {})
        v2_score = 0
        v2_factors = {}

        # ── 信号强度 (0-40) — 基于V1基础分换算 ──
        v1_score = signal.get('v1_score', 0)
        signal_strength = min(v1_score / 100 * 40, 40)
        v2_score += signal_strength
        v2_factors['信号强度'] = round(signal_strength, 1)

        # ── 位置优势 (0-30) — 距年高 + 60日动量 ──
        dist_250 = pf.get('dist_250d', 0)
        if dist_250 >= 0:
            v2_score += 15
            v2_factors['距年高'] = 15
        elif dist_250 > -0.03:
            v2_score += 10
            v2_factors['距年高'] = 10
        elif dist_250 > -0.08:
            v2_score += 5
            v2_factors['距年高'] = 5
        else:
            v2_factors['距年高'] = 0

        chg_60d = pf.get('chg_60d', 0)
        if chg_60d > 50:
            v2_score += 10
            v2_factors['60日涨'] = 10
        elif chg_60d > 30:
            v2_score += 8
            v2_factors['60日涨'] = 8
        elif chg_60d > 10:
            v2_score += 3
            v2_factors['60日涨'] = 3
        else:
            v2_factors['60日涨'] = 0

        # 新高强度加分
        if chg_60d > 50 and dist_250 >= 0:
            v2_score += 5  # 共振加分
            v2_factors['位置共振'] = 5
        else:
            v2_factors['位置共振'] = 0

        # ── 回踩质量 (0-30) ──
        pullback_quality = 0
        signal_close = signal['close']
        pb_close = pb['pb_close']
        pb_low = pb['pb_low']
        pb_vol = pb['pb_vol']
        signal_vol = signal['volume']

        # 1. 回踩幅度 (0-10) — 温和回踩最优
        pullback_pct = (signal_close - pb_low) / signal_close * 100
        if 1.0 <= pullback_pct <= 3.0:
            pullback_quality += 10
            v2_factors['回踩幅度'] = 10
        elif 3.0 < pullback_pct <= 5.0:
            pullback_quality += 7
            v2_factors['回踩幅度'] = 7
        elif 0.5 <= pullback_pct < 1.0:
            pullback_quality += 5  # 回踩太浅，洗盘不充分
            v2_factors['回踩幅度'] = 5
        elif 5.0 < pullback_pct <= 8.0:
            pullback_quality += 4  # 回踩偏深，有破位风险
            v2_factors['回踩幅度'] = 4
        else:
            pullback_quality += 0  # 没怎么回踩或回踩过深
            v2_factors['回踩幅度'] = 0

        # 2. 缩量程度 (0-8) — 缩量=浮筹清洗干净
        vol_shrink_ratio = pb_vol / signal_vol if signal_vol > 0 else 1.0
        if vol_shrink_ratio <= 0.40:
            pullback_quality += 8
            v2_factors['缩量'] = 8
        elif vol_shrink_ratio <= 0.50:
            pullback_quality += 6
            v2_factors['缩量'] = 6
        elif vol_shrink_ratio <= 0.60:
            pullback_quality += 4
            v2_factors['缩量'] = 4
        elif vol_shrink_ratio <= 0.80:
            pullback_quality += 2
            v2_factors['缩量'] = 2
        else:
            v2_factors['缩量'] = 0

        # 3. K线止跌形态类型 (0-7)
        tt = pb.get('trigger_type', '')
        if tt == 'engulf':
            pullback_quality += 7
            v2_factors['止跌K线'] = 7
        elif tt == 'hammer':
            pullback_quality += 5
            v2_factors['止跌K线'] = 5
        elif tt == 'doji':
            pullback_quality += 3
            v2_factors['止跌K线'] = 3
        else:
            v2_factors['止跌K线'] = 0

        # 4. 回踩天数 (0-5) — T+2/T+3最佳（不早不晚）
        pb_days = pb.get('pullback_days', 1)
        if 2 <= pb_days <= 3:
            pullback_quality += 5
            v2_factors['回踩天数'] = 5
        elif pb_days == 1 or pb_days == 4:
            pullback_quality += 3
            v2_factors['回踩天数'] = 3
        else:
            pullback_quality += 1
            v2_factors['回踩天数'] = 1

        v2_score += pullback_quality

        # ── V2层级 ──
        v2_score = min(round(v2_score), 100)
        if v2_score >= 90: tier = 'S'
        elif v2_score >= 80: tier = 'A'
        elif v2_score >= 70: tier = 'B'
        elif v2_score >= 60: tier = 'C'
        else: tier = 'D'

        return {
            'score': v2_score,
            'tier': tier,
            'v2_factors': v2_factors,
            'pullback_pct': round(pullback_pct, 2),
            'vol_shrink_ratio': round(vol_shrink_ratio, 2),
        }

    # ─── 持有+卖出: 固定持有N天 ──────────────────────────────────

    def _simulate_hold(self, code: str, buy_date: str, buy_price: float,
                       all_dates: List[str], hold_days: int = 3) -> Optional[dict]:
        """持有N天 → 第N天收盘卖出。返回交易结果或None"""
        try:
            bd_idx = all_dates.index(buy_date)
        except ValueError:
            return None

        sell_idx = bd_idx + hold_days
        if sell_idx >= len(all_dates):
            sell_idx = len(all_dates) - 1

        sell_date = all_dates[sell_idx]
        df = self._read_stock(code, up_to_date=sell_date)
        if df is None or df.empty: return None

        sell_close = float(df['close'].iloc[-1])

        gross = (sell_close - buy_price) / buy_price
        net = gross - TOTAL_COST

        return {
            'buy_date': buy_date,
            'buy_price': round(buy_price, 2),
            'sell_date': sell_date,
            'sell_price': round(sell_close, 2),
            'gross_ret': round(gross * 100, 2),
            'net_ret': round(net * 100, 2),
            'is_win': net > 0,
            'exit_mode': f'fixed_{hold_days}d',
            'exit_reason': f'hold_{hold_days}d',
        }

    # ─── 持有+卖出: 动态止盈止损 ──────────────────────────────────

    def _simulate_hold_dynamic(self, code: str, buy_date: str, buy_price: float,
                               all_dates: List[str]) -> Optional[dict]:
        """动态止盈止损:
        - 止损: 日内最低价触及买入价-5% → 次日开盘卖出
        - 止盈: 浮盈≥+10% → 卖一半，剩余持有至第3日收盘
        - 强制: 第5日收盘未触发 → 强制卖出
        """
        try:
            bd_idx = all_dates.index(buy_date)
        except ValueError:
            return None

        stop_loss_price = buy_price * 0.95   # -5%
        take_profit_price = buy_price * 1.10  # +10%

        # 从买入次日开始逐日检查
        for day_offset in range(1, DYNAMIC_HOLD_MAX + 1):
            check_idx = bd_idx + day_offset
            if check_idx >= len(all_dates):
                # 无后续数据, 以最后一天收盘强制卖出
                sell_idx = len(all_dates) - 1
                sell_date = all_dates[sell_idx]
                df = self._read_stock(code, up_to_date=sell_date)
                if df is None or df.empty: return None
                sell_price = float(df['close'].iloc[-1])
                gross = (sell_price - buy_price) / buy_price
                return {
                    'buy_date': buy_date, 'buy_price': round(buy_price, 2),
                    'sell_date': sell_date, 'sell_price': round(sell_price, 2),
                    'gross_ret': round(gross * 100, 2),
                    'net_ret': round((gross - TOTAL_COST) * 100, 2),
                    'is_win': gross > TOTAL_COST,
                    'exit_mode': 'dynamic', 'exit_reason': 'force_end',
                }

            check_date = all_dates[check_idx]
            df = self._read_stock(code, up_to_date=check_date)
            if df is None or df.empty: continue
            idx = len(df) - 1

            day_high = float(df['high'].iloc[idx])
            day_low = float(df['low'].iloc[idx])
            day_close = float(df['close'].iloc[idx])
            day_open = float(df['open'].iloc[idx])

            # 检查止损: 日内最低价触及止损价
            if day_low <= stop_loss_price:
                # 次日开盘卖出
                exit_idx = check_idx + 1
                if exit_idx >= len(all_dates):
                    sell_price = day_close
                    sell_date = check_date
                else:
                    sell_date = all_dates[exit_idx]
                    df2 = self._read_stock(code, up_to_date=sell_date)
                    if df2 is None or df2.empty:
                        sell_price = day_close
                        sell_date = check_date
                    else:
                        sell_price = float(df2['open'].iloc[-1])
                gross = (sell_price - buy_price) / buy_price
                return {
                    'buy_date': buy_date, 'buy_price': round(buy_price, 2),
                    'sell_date': sell_date, 'sell_price': round(sell_price, 2),
                    'gross_ret': round(gross * 100, 2),
                    'net_ret': round((gross - TOTAL_COST) * 100, 2),
                    'is_win': False,
                    'exit_mode': 'dynamic', 'exit_reason': 'stop_loss',
                }

            # 检查止盈: 日内最高价触及止盈价
            if day_high >= take_profit_price:
                # 一半仓位以止盈价卖出, 另一半持有至第3日
                half_profit = (take_profit_price - buy_price) / buy_price
                # 剩余仓位持有至买入后第3天
                remain_idx = bd_idx + HOLD_DAYS
                if remain_idx >= len(all_dates):
                    remain_idx = len(all_dates) - 1
                remain_date = all_dates[remain_idx]
                df3 = self._read_stock(code, up_to_date=remain_date)
                if df3 is None or df3.empty:
                    remain_price = take_profit_price
                else:
                    remain_price = float(df3['close'].iloc[-1])
                remain_profit = (remain_price - buy_price) / buy_price
                # 综合收益: 一半止盈 + 一半持有到期
                gross = 0.5 * half_profit + 0.5 * remain_profit
                return {
                    'buy_date': buy_date, 'buy_price': round(buy_price, 2),
                    'sell_date': remain_date, 'sell_price': round(remain_price, 2),
                    'gross_ret': round(gross * 100, 2),
                    'net_ret': round((gross - TOTAL_COST) * 100, 2),
                    'is_win': gross > TOTAL_COST,
                    'exit_mode': 'dynamic', 'exit_reason': 'take_profit_half',
                }

            # 第5天强制卖出
            if day_offset == DYNAMIC_HOLD_MAX:
                sell_date = check_date
                sell_price = day_close
                gross = (sell_price - buy_price) / buy_price
                return {
                    'buy_date': buy_date, 'buy_price': round(buy_price, 2),
                    'sell_date': sell_date, 'sell_price': round(sell_price, 2),
                    'gross_ret': round(gross * 100, 2),
                    'net_ret': round((gross - TOTAL_COST) * 100, 2),
                    'is_win': gross > TOTAL_COST,
                    'exit_mode': 'dynamic', 'exit_reason': 'force_day5',
                }

        # 不应到达此处
        return None

    # ─── 持有+卖出: 跌破MA卖出 ────────────────────────────────────

    def _simulate_ma_exit(self, code: str, buy_date: str, buy_price: float,
                          all_dates: List[str], ma_period: int = 5,
                          max_hold: int = 20) -> Optional[dict]:
        """持有直到收盘价跌破MA(N) → 当日收盘卖出。最长持有max_hold天。"""
        try:
            bd_idx = all_dates.index(buy_date)
        except ValueError:
            return None

        for day_offset in range(1, max_hold + 1):
            check_idx = bd_idx + day_offset
            if check_idx >= len(all_dates):
                # 末尾强平
                sell_idx = len(all_dates) - 1
                sell_date = all_dates[sell_idx]
                df = self._read_stock(code, up_to_date=sell_date)
                if df is None or df.empty: return None
                sell_price = float(df['close'].iloc[-1])
                gross = (sell_price - buy_price) / buy_price
                return {'buy_date': buy_date, 'buy_price': round(buy_price, 2),
                        'sell_date': sell_date, 'sell_price': round(sell_price, 2),
                        'gross_ret': round(gross * 100, 2),
                        'net_ret': round((gross - TOTAL_COST) * 100, 2),
                        'is_win': gross > TOTAL_COST,
                        'exit_mode': f'ma{ma_period}', 'exit_reason': 'force_end'}

            check_date = all_dates[check_idx]
            df = self._read_stock(code, up_to_date=check_date)
            if df is None or df.empty: continue
            idx = len(df) - 1
            close = float(df['close'].iloc[idx])

            # 计算MA
            ma_col = f'ma{ma_period}'
            if ma_col not in df.columns:
                df = calc_ma(df, [ma_period])
            ma_val = float(df[ma_col].iloc[idx])

            # 跌破MA → 收盘卖出
            if close < ma_val or day_offset == max_hold:
                sell_date = check_date
                sell_price = close
                gross = (sell_price - buy_price) / buy_price
                return {'buy_date': buy_date, 'buy_price': round(buy_price, 2),
                        'sell_date': sell_date, 'sell_price': round(sell_price, 2),
                        'gross_ret': round(gross * 100, 2),
                        'net_ret': round((gross - TOTAL_COST) * 100, 2),
                        'is_win': gross > TOTAL_COST,
                        'exit_mode': f'ma{ma_period}',
                        'exit_reason': f'below_ma{ma_period}' if close < ma_val else 'max_hold'}

        return None

    # ─── 持有+卖出: 止盈+MA跟踪 ──────────────────────────────────

    def _simulate_tp_ma_exit(self, code: str, buy_date: str, buy_price: float,
                             all_dates: List[str], tp_pct: float = 0.10,
                             ma_period: int = 5, max_hold: int = 20) -> Optional[dict]:
        """+10%止盈 或 跌破MA5卖出，哪个先触发。最长持有max_hold天。"""
        try:
            bd_idx = all_dates.index(buy_date)
        except ValueError:
            return None

        tp_price = buy_price * (1 + tp_pct)

        for day_offset in range(1, max_hold + 1):
            check_idx = bd_idx + day_offset
            if check_idx >= len(all_dates):
                sell_idx = len(all_dates) - 1
                sell_date = all_dates[sell_idx]
                df = self._read_stock(code, up_to_date=sell_date)
                if df is None or df.empty: return None
                sell_price = float(df['close'].iloc[-1])
                gross = (sell_price - buy_price) / buy_price
                return {'buy_date': buy_date, 'buy_price': round(buy_price, 2),
                        'sell_date': sell_date, 'sell_price': round(sell_price, 2),
                        'gross_ret': round(gross * 100, 2),
                        'net_ret': round((gross - TOTAL_COST) * 100, 2),
                        'is_win': gross > TOTAL_COST,
                        'exit_mode': f'tp{tp_pct*100:.0f}_ma{ma_period}', 'exit_reason': 'force_end'}

            check_date = all_dates[check_idx]
            df = self._read_stock(code, up_to_date=check_date)
            if df is None or df.empty: continue
            idx = len(df) - 1
            close = float(df['close'].iloc[idx])
            high = float(df['high'].iloc[idx])

            # 计算MA
            ma_col = f'ma{ma_period}'
            if ma_col not in df.columns:
                df = calc_ma(df, [ma_period])
            ma_val = float(df[ma_col].iloc[idx])

            # 检查止盈
            if high >= tp_price:
                sell_date = check_date
                sell_price = tp_price  # 以止盈价卖出
                gross = (sell_price - buy_price) / buy_price
                return {'buy_date': buy_date, 'buy_price': round(buy_price, 2),
                        'sell_date': sell_date, 'sell_price': round(sell_price, 2),
                        'gross_ret': round(gross * 100, 2),
                        'net_ret': round((gross - TOTAL_COST) * 100, 2),
                        'is_win': True,
                        'exit_mode': f'tp{tp_pct*100:.0f}_ma{ma_period}', 'exit_reason': 'take_profit'}

            # 检查MA跌破
            if close < ma_val:
                sell_date = check_date
                sell_price = close
                gross = (sell_price - buy_price) / buy_price
                return {'buy_date': buy_date, 'buy_price': round(buy_price, 2),
                        'sell_date': sell_date, 'sell_price': round(sell_price, 2),
                        'gross_ret': round(gross * 100, 2),
                        'net_ret': round((gross - TOTAL_COST) * 100, 2),
                        'is_win': gross > TOTAL_COST,
                        'exit_mode': f'tp{tp_pct*100:.0f}_ma{ma_period}',
                        'exit_reason': f'below_ma{ma_period}'}

            # 最大持有
            if day_offset == max_hold:
                sell_date = check_date
                sell_price = close
                gross = (sell_price - buy_price) / buy_price
                return {'buy_date': buy_date, 'buy_price': round(buy_price, 2),
                        'sell_date': sell_date, 'sell_price': round(sell_price, 2),
                        'gross_ret': round(gross * 100, 2),
                        'net_ret': round((gross - TOTAL_COST) * 100, 2),
                        'is_win': gross > TOTAL_COST,
                        'exit_mode': f'tp{tp_pct*100:.0f}_ma{ma_period}', 'exit_reason': 'max_hold'}

        return None

    # ─── 主循环 ────────────────────────────────────────────────────

    def run(self, lookback: int = 250, exit_mode: str = 'fixed',
            hold_days: int = 3, end_date: date = None):
        """主回测循环

        Args:
            lookback: 回测天数
            exit_mode: 'fixed'/'dynamic'/'ma5'/'ma10'/'tp_ma5'/'all'
            hold_days: 固定持有天数（fixed模式用，默认3天）
            end_date: 回测截止日（None=昨天，用于样本外验证）
        """
        buf = MAX_WAIT
        # 信号日数量 = lookback（约250~300个交易日）
        # 尾部保留 buf 天作为回踩确认 + 退出缓冲
        total = lookback + buf + 10  # +10 额外安全余量
        all_dates_raw = self._trade_dates(total, end_date=end_date)
        signal_dates = all_dates_raw[:-(buf + 5)]
        all_dates_str = [d.strftime('%Y%m%d') for d in all_dates_raw]

        print(f'Backtest: {signal_dates[0]} ~ {signal_dates[-1]}, {len(signal_dates)} days')

        # 步骤0: 构建行业映射
        self.industry_map = self._build_industry_map(all_dates_raw)

        # 步骤1: 预扫描
        gainers, sector_stats = self._pre_scan_gainers(all_dates_raw, self.industry_map)

        all_trades_fixed = []
        all_trades_dynamic = []
        daily_log = []
        signal_stats = {'total': 0, 'pulled_back': 0, 'sector_filtered': 0,
                        'all_checked': 0,  # 所有经过板块检查的信号
                        'pullback_days': defaultdict(int)}
        t0 = time.time()

        for i, td in enumerate(signal_dates):
            ds = td.strftime('%Y%m%d')
            if (i + 1) % 25 == 0 or i == 0:
                e = time.time() - t0
                eta = e / (i + 1) * (len(signal_dates) - i - 1) if i > 0 else 0
                print(f'[{i+1}/{len(signal_dates)}] {ds} ({e:.0f}s, ETA {eta:.0f}s)...', flush=True)

            candidates = gainers.get(ds, [])
            day_trades_fixed = []
            day_trades_dynamic = []

            # 前N日日期列表 (用于板块持续逻辑)
            ds_idx = all_dates_str.index(ds) if ds in all_dates_str else i
            prev_dates = all_dates_str[max(0, ds_idx-3):ds_idx+1]

            for info in candidates:
                signal = self._check_signal(info, ds, sector_stats, prev_dates)
                if signal is None:
                    # 确定是不是被板块过滤掉的
                    if info.get('industry') and not self.skip_sector_filter:
                        # 检查是否因为板块不通过
                        if not self._sector_pass(info['industry'], ds, sector_stats, prev_dates):
                            signal_stats['sector_filtered'] += 1
                    continue
                signal_stats['total'] += 1

                # 找回踩
                pb = self._find_pullback(signal, ds, all_dates_str)
                if pb is None: continue
                buy_date = pb['buy_date']
                buy_price = pb['buy_price']
                pullback_days = pb['pullback_days']
                signal_stats['pulled_back'] += 1
                signal_stats['pullback_days'][str(pullback_days)] += 1

                # ── V2评分: 回踩确认后计算（含回踩质量因子） ──
                v2_result = self._calc_v2_score_at_pullback(signal, pb)

                # V2资格过滤: 低于60分的跳过（与StartBreakoutScreenerV2一致）
                if v2_result['score'] < 60:
                    continue

                # 市场状态（用于 Regime Research，不影响评分）
                market_state = signal.get('market_state', {})

                base_info = {
                    'signal_date': ds,
                    'code': signal['code'],
                    'score': v2_result['score'],           # V2综合评分
                    'v1_score': signal.get('v1_score', 0), # V1基础分
                    'tier': v2_result['tier'],
                    'v1_tier': signal.get('v1_tier', ''),
                    'v2_factors': v2_result['v2_factors'],
                    'features': signal.get('pullback_features', {}),
                    'signal_close': signal['close'],
                    'industry': signal.get('industry', ''),
                    'pullback_days': pullback_days,
                    'pullback_pct': v2_result['pullback_pct'],
                    'vol_shrink_ratio': v2_result['vol_shrink_ratio'],
                    'limit_up_num': market_state.get('limit_up_num'),
                    'limit_down_num': market_state.get('limit_down_num'),
                    'sh_ma60_up': market_state.get('sh_ma60_up'),
                    'up_ratio': market_state.get('up_ratio'),
                }

                # 根据exit_mode选择卖出方式
                if exit_mode in ('fixed', 'all'):
                    trade = self._simulate_hold(signal['code'], buy_date, buy_price,
                                                all_dates_str, hold_days=hold_days)
                    if trade is not None:
                        trade.update(base_info)
                        all_trades_fixed.append(trade)
                        day_trades_fixed.append(trade)

                if exit_mode in ('dynamic', 'all'):
                    trade = self._simulate_hold_dynamic(signal['code'], buy_date, buy_price, all_dates_str)
                    if trade is not None:
                        trade.update(base_info)
                        all_trades_dynamic.append(trade)
                        day_trades_dynamic.append(trade)

                if exit_mode in ('ma5', 'all'):
                    trade = self._simulate_ma_exit(signal['code'], buy_date, buy_price,
                                                   all_dates_str, ma_period=5)
                    if trade is not None:
                        trade.update(base_info)
                        if not hasattr(self, '_trades_ma5'): self._trades_ma5 = []
                        self._trades_ma5.append(trade)

                if exit_mode in ('ma10', 'all'):
                    trade = self._simulate_ma_exit(signal['code'], buy_date, buy_price,
                                                   all_dates_str, ma_period=10)
                    if trade is not None:
                        trade.update(base_info)
                        if not hasattr(self, '_trades_ma10'): self._trades_ma10 = []
                        self._trades_ma10.append(trade)

                if exit_mode in ('tp_ma5', 'all'):
                    trade = self._simulate_tp_ma_exit(signal['code'], buy_date, buy_price,
                                                      all_dates_str, tp_pct=0.10, ma_period=5)
                    if trade is not None:
                        trade.update(base_info)
                        if not hasattr(self, '_trades_tp_ma5'): self._trades_tp_ma5 = []
                        self._trades_tp_ma5.append(trade)

            # 日统计(用第一个有交易的模式)
            day_trades = day_trades_fixed or day_trades_dynamic or \
                         getattr(self, '_trades_ma5', [])[-1:] or \
                         getattr(self, '_trades_ma10', [])[-1:] or \
                         getattr(self, '_trades_tp_ma5', [])[-1:]
            wins = sum(1 for t in day_trades if t['is_win'])
            daily_log.append({
                'date': ds, 'trades': len(day_trades),
                'wins': wins, 'losses': len(day_trades) - wins,
                'ret_sum': round(sum(t['net_ret'] for t in day_trades), 2),
            })

        df_fixed = pd.DataFrame(all_trades_fixed)
        df_dynamic = pd.DataFrame(all_trades_dynamic)
        df_ma5 = pd.DataFrame(getattr(self, '_trades_ma5', []))
        df_ma10 = pd.DataFrame(getattr(self, '_trades_ma10', []))
        df_tp_ma5 = pd.DataFrame(getattr(self, '_trades_tp_ma5', []))
        ddf = pd.DataFrame(daily_log)
        back_pct = signal_stats['pulled_back'] / max(signal_stats['total'], 1) * 100
        sector_pct = signal_stats['sector_filtered'] / max(signal_stats['total'] + signal_stats['sector_filtered'], 1) * 100
        print(f'\nDone: {signal_stats["total"]} signals, '
              f'{signal_stats["sector_filtered"]} sector-filtered ({sector_pct:.1f}%), '
              f'{signal_stats["pulled_back"]} pullbacks ({back_pct:.1f}%), '
              f'{len(df_fixed)} fixed, {len(df_ma5)} ma5, {len(df_ma10)} ma10, {len(df_tp_ma5)} tp_ma5')
        return df_fixed, df_dynamic, df_ma5, df_ma10, df_tp_ma5, ddf, len(signal_dates), signal_stats

    # ─── 前向模拟: 今日模拟买入/卖出/持仓 ──────────────────────────

    @staticmethod
    def _trading_days_between(start_date: str, end_date: str,
                               all_dates: List[str]) -> int:
        """计算两个交易日期之间的交易日数（不含起始日，含结束日）。"""
        try:
            si = all_dates.index(start_date)
            ei = all_dates.index(end_date)
            return max(0, ei - si)
        except ValueError:
            return 999

    def _check_position_sell(self, pos: dict, today_str: str,
                              all_dates: List[str], exit_mode: str,
                              hold_days: int) -> Optional[dict]:
        """检查持仓是否需要今日卖出。

        Args:
            pos: 持仓信息 {code, buy_date, buy_price, ...}
            today_str: 今日日期 YYYYMMDD
            all_dates: 全部交易日列表
            exit_mode: 'fixed'/'dynamic'/'ma5'/'ma10'
            hold_days: 固定持有天数

        Returns:
            None 如果不需要卖出，否则 {sell_price, sell_date, sell_reason}
        """
        code = pos['code']
        buy_date = pos['buy_date']
        buy_price = pos['buy_price']
        days_held = self._trading_days_between(buy_date, today_str, all_dates)

        df = self._read_stock(code, up_to_date=today_str)
        if df is None or df.empty:
            return None
        idx = len(df) - 1
        close = float(df['close'].iloc[idx])
        high = float(df['high'].iloc[idx])
        low = float(df['low'].iloc[idx])

        if exit_mode == 'fixed':
            if days_held >= hold_days:
                return {'sell_price': round(close, 2), 'sell_date': today_str,
                        'sell_reason': f'hold_{hold_days}d'}

        elif exit_mode == 'dynamic':
            # 止损: 日内最低触及 -5%
            if low <= buy_price * 0.95:
                return {'sell_price': round(buy_price * 0.95, 2),
                        'sell_date': today_str, 'sell_reason': 'stop_loss'}
            # 止盈: 日内最高触及 +10%
            if high >= buy_price * 1.10:
                return {'sell_price': round(buy_price * 1.10, 2),
                        'sell_date': today_str, 'sell_reason': 'take_profit'}
            # 强制第5天卖出
            if days_held >= DYNAMIC_HOLD_MAX:
                return {'sell_price': round(close, 2), 'sell_date': today_str,
                        'sell_reason': 'force_day5'}

        elif exit_mode in ('ma5', 'ma10'):
            ma_period = 5 if exit_mode == 'ma5' else 10
            if f'ma{ma_period}' not in df.columns:
                df = calc_ma(df, [ma_period])
            ma_val = float(df[f'ma{ma_period}'].iloc[idx])
            if close < ma_val:
                return {'sell_price': round(close, 2), 'sell_date': today_str,
                        'sell_reason': f'below_ma{ma_period}'}
            # MA模式最大持有20天
            if days_held >= 20:
                return {'sell_price': round(close, 2), 'sell_date': today_str,
                        'sell_reason': 'max_hold_20d'}

        return None

    def run_simulation(self, lookback: int = 250, exit_mode: str = 'fixed',
                       hold_days: int = 3, end_date: date = None,
                       position_size: float = 10000.0):
        """前向模拟 — 按回测逻辑逐日推进，追踪模拟持仓。

        与回测使用完全相同的信号检测、回踩确认、V2评分、卖出规则，
        但按时间顺序逐日处理，维持组合状态。

        Args:
            lookback: 回看天数
            exit_mode: 'fixed'/'dynamic'/'ma5'/'ma10'
            hold_days: 固定持有天数（fixed模式）
            end_date: 模拟截止日（默认昨天）
            position_size: 每只股票固定金额（元）

        Returns:
            (today_buys_df, today_sells_df, holdings_df, history_df, summary)
        """
        buf = MAX_WAIT + DYNAMIC_HOLD_MAX + 10
        total = lookback + buf + 10
        all_dates_raw = self._trade_dates(total, end_date=end_date)
        signal_dates = all_dates_raw[:-(buf + 5)]
        all_dates_str = [d.strftime('%Y%m%d') for d in all_dates_raw]

        print(f'\n{"="*70}')
        print(f'  启动突破战法 — 前向模拟')
        print(f'  回看: {lookback}天 | 退出: {exit_mode}')
        if exit_mode == 'fixed':
            print(f'  固定持有: {hold_days}天 | 每笔: {position_size:.0f}元')
        print(f'  区间: {signal_dates[0]} ~ {signal_dates[-1]}')
        print(f'{"="*70}')

        # ── 预扫描 ──
        self.industry_map = self._build_industry_map(all_dates_raw)
        gainers, sector_stats = self._pre_scan_gainers(all_dates_raw, self.industry_map)

        # ── 状态 ──
        portfolio: Dict[str, dict] = {}       # code → 当前持仓
        pending_buys: List[dict] = []          # 回踩确认，等待买入日
        watch_signals: Dict[str, list] = {}    # code → [{signal_date, signal}]
        history: List[dict] = []               # 已完成交易
        name_cache: Dict[str, str] = {}        # code → name

        sim_buy_today: List[dict] = []         # 最终日 模拟买入
        sim_sell_today: List[dict] = []        # 最终日 模拟卖出

        total_signals = 0
        total_pullbacks = 0
        t0 = time.time()

        for i, td in enumerate(signal_dates):
            ds = td.strftime('%Y%m%d')
            is_last = (i == len(signal_dates) - 1)

            if (i + 1) % 50 == 0 or i == 0:
                e = time.time() - t0
                eta = e / (i + 1) * (len(signal_dates) - i - 1) if i > 0 else 0
                print(f'  [{i+1}/{len(signal_dates)}] {ds}  '
                      f'持仓:{len(portfolio)}  挂单:{len(pending_buys)}  '
                      f'观察:{len(watch_signals)}  '
                      f'({e:.0f}s, ETA {eta:.0f}s)', flush=True)

            # ── Step 1: 执行今日到期的买入（buy_date == today） ──
            buys_executed = []
            for pb in pending_buys:
                if pb['buy_date'] == ds:
                    code = pb['code']
                    # 获取名称
                    name = name_cache.get(code, '')
                    if not name:
                        # Try to get from TDX
                        try:
                            df_name = self._read_stock(code, up_to_date=ds)
                            if df_name is not None and not df_name.empty:
                                pass  # name stays as code
                        except Exception:
                            pass
                        name = code

                    portfolio[code] = {
                        'code': code,
                        'name': name,
                        'buy_date': ds,
                        'buy_price': pb['buy_price'],
                        'signal_date': pb.get('signal_date', ''),
                        'score': pb.get('score', 0),
                        'tier': pb.get('tier', ''),
                        'v1_score': pb.get('v1_score', 0),
                        'v1_tier': pb.get('v1_tier', ''),
                        'industry': pb.get('industry', ''),
                        'pullback_days': pb.get('pullback_days', 0),
                        'pullback_pct': pb.get('pullback_pct', 0),
                        'vol_shrink_ratio': pb.get('vol_shrink_ratio', 0),
                        'signal_close': pb.get('signal_close', 0),
                        'v2_factors': pb.get('v2_factors', {}),
                        'position_size': position_size,
                        'shares': position_size / pb['buy_price'],
                        'exit_mode': exit_mode,
                        'hold_days': hold_days,
                    }
                    buys_executed.append(pb)

            for pb in buys_executed:
                pending_buys.remove(pb)
                sim_buy_today.append(pb)

            # ── Step 2: 检测今日新信号 ──
            candidates = gainers.get(ds, [])
            ds_idx_in_all = all_dates_str.index(ds) if ds in all_dates_str else i
            prev_dates = all_dates_str[max(0, ds_idx_in_all - 3):ds_idx_in_all + 1]

            for info in candidates:
                code = info['code']
                # 已在持仓/观察/挂单中 → 跳过
                if code in portfolio or code in watch_signals:
                    continue
                if any(pb['code'] == code for pb in pending_buys):
                    continue

                signal = self._check_signal(info, ds, sector_stats, prev_dates)
                if signal is None:
                    continue

                total_signals += 1

                # 缓存名称
                if code not in name_cache:
                    name_cache[code] = info.get('name', '') or code

                # 加入观察列表
                if code not in watch_signals:
                    watch_signals[code] = []
                watch_signals[code].append({
                    'signal_date': ds,
                    'signal': signal,
                })

            # ── Step 3: 检查观察列表中信号的回踩 ──
            for code in list(watch_signals.keys()):
                if code in portfolio:
                    del watch_signals[code]
                    continue

                for ws in list(watch_signals.get(code, [])):
                    signal_date = ws['signal_date']
                    signal = ws['signal']

                    try:
                        sd_idx = all_dates_str.index(signal_date)
                    except ValueError:
                        watch_signals[code].remove(ws)
                        continue

                    days_since = all_dates_str.index(ds) - sd_idx
                    if days_since > MAX_WAIT:
                        watch_signals[code].remove(ws)
                        continue
                    if days_since < 1:
                        continue  # T+0, 至少T+1才能回踩

                    # 找 回踩确认
                    pb = self._find_pullback(signal, signal_date, all_dates_str)
                    if pb is None:
                        continue

                    # V2评分
                    v2_result = self._calc_v2_score_at_pullback(signal, pb)
                    if v2_result['score'] < 60:
                        watch_signals[code].remove(ws)
                        continue

                    total_pullbacks += 1

                    # 排入买入计划
                    pending_buys.append({
                        'code': code,
                        'name': name_cache.get(code, code),
                        'buy_date': pb['buy_date'],
                        'buy_price': pb['buy_price'],
                        'signal_date': signal_date,
                        'score': v2_result['score'],
                        'tier': v2_result['tier'],
                        'v1_score': signal.get('v1_score', 0),
                        'v1_tier': signal.get('v1_tier', ''),
                        'industry': signal.get('industry', ''),
                        'pullback_days': pb['pullback_days'],
                        'pullback_pct': v2_result['pullback_pct'],
                        'vol_shrink_ratio': v2_result['vol_shrink_ratio'],
                        'signal_close': signal['close'],
                        'v2_factors': v2_result.get('v2_factors', {}),
                    })
                    watch_signals[code].remove(ws)

            # 清理空的观察条目
            for code in list(watch_signals.keys()):
                if not watch_signals.get(code):
                    del watch_signals[code]

            # ── Step 4: 检查持仓卖出条件 ──
            sells_executed = []
            for code in list(portfolio.keys()):
                pos = portfolio[code]
                sell_signal = self._check_position_sell(
                    pos, ds, all_dates_str, exit_mode, hold_days)

                if sell_signal:
                    sell_price = sell_signal['sell_price']
                    sell_date = sell_signal['sell_date']
                    sell_reason = sell_signal['sell_reason']

                    gross_ret = (sell_price - pos['buy_price']) / pos['buy_price']
                    net_ret = gross_ret - TOTAL_COST

                    trade_record = {
                        **pos,
                        'sell_date': sell_date,
                        'sell_price': sell_price,
                        'gross_ret': round(gross_ret * 100, 2),
                        'net_ret': round(net_ret * 100, 2),
                        'is_win': net_ret > 0,
                        'exit_mode': exit_mode,
                        'exit_reason': sell_reason,
                        'days_held': self._trading_days_between(
                            pos['buy_date'], sell_date, all_dates_str),
                    }
                    history.append(trade_record)
                    sells_executed.append(trade_record)
                    del portfolio[code]

            sim_sell_today.extend(sells_executed)

            # 非最后一天清空当日记录（只保留最终日的buy/sell）
            if not is_last:
                sim_buy_today.clear()
                sim_sell_today.clear()

        # ── 后处理: 获取持仓的当前价格和MA ──
        holdings_list = []
        last_date = all_dates_str[-1] if all_dates_str else signal_dates[-1].strftime('%Y%m%d')

        for code, pos in portfolio.items():
            df = self._read_stock(code, up_to_date=last_date)
            current_price = pos['buy_price']
            ma5_v = ma10_v = ma20_v = 0.0
            if df is not None and not df.empty:
                current_price = float(df['close'].iloc[-1])
                try:
                    if 'ma5' in df.columns: ma5_v = float(df['ma5'].iloc[-1])
                    if 'ma10' in df.columns: ma10_v = float(df['ma10'].iloc[-1])
                    if 'ma20' in df.columns: ma20_v = float(df['ma20'].iloc[-1])
                except Exception:
                    pass

            unrealized_pnl = (current_price - pos['buy_price']) / pos['buy_price'] * 100
            days_held = self._trading_days_between(pos['buy_date'], last_date, all_dates_str)

            holdings_list.append({
                **pos,
                'current_price': round(current_price, 2),
                'unrealized_pnl_pct': round(unrealized_pnl, 2),
                'unrealized_pnl_yuan': round(pos.get('position_size', position_size)
                                             * unrealized_pnl / 100, 2),
                'days_held': days_held,
                'ma5': round(ma5_v, 2),
                'ma10': round(ma10_v, 2),
                'ma20': round(ma20_v, 2),
            })

        # ── 转 DataFrame ──
        # 将待执行买入也加入今日买入（它们将在今天/下个交易日执行）
        all_buys_today = sim_buy_today + pending_buys
        df_buys = pd.DataFrame(all_buys_today) if all_buys_today else pd.DataFrame()
        df_sells = pd.DataFrame(sim_sell_today) if sim_sell_today else pd.DataFrame()
        df_holdings = pd.DataFrame(holdings_list) if holdings_list else pd.DataFrame()
        df_history = pd.DataFrame(history) if history else pd.DataFrame()

        # ── 汇总统计 ──
        total_completed = len(history)
        wins = sum(1 for h in history if h.get('is_win'))
        cum_equity = 1.0
        for h in history:
            cum_equity *= (1 + h.get('net_ret', 0) / 100)
        cum_ret = (cum_equity - 1) * 100

        sim_summary = {
            'total_days': len(signal_dates),
            'total_signals': total_signals,
            'total_pullbacks': total_pullbacks,
            'total_trades': total_completed,
            'wins': wins,
            'losses': total_completed - wins,
            'win_rate': round(wins / max(total_completed, 1) * 100, 1),
            'avg_ret': round(sum(h.get('net_ret', 0) for h in history)
                             / max(total_completed, 1), 2) if history else 0,
            'avg_win': round(sum(h.get('net_ret', 0) for h in history if h.get('is_win'))
                             / max(wins, 1), 2) if wins > 0 else 0,
            'avg_loss': round(sum(h.get('net_ret', 0) for h in history
                                   if not h.get('is_win'))
                              / max(total_completed - wins, 1), 2)
            if total_completed - wins > 0 else 0,
            'cum_ret': round(cum_ret, 2),
            'sim_buy_today': len(sim_buy_today),
            'sim_sell_today': len(sim_sell_today),
            'holdings': len(holdings_list),
            'pending_buys': len(pending_buys),
            'last_date': last_date,
        }

        # ── 控制台输出 ──
        self._print_simulation_summary(sim_summary, df_buys, df_sells,
                                        df_holdings, df_history, exit_mode)

        return df_buys, df_sells, df_holdings, df_history, sim_summary

    @staticmethod
    def _print_simulation_summary(sim_summary: dict, df_buys: 'pd.DataFrame',
                                   df_sells: 'pd.DataFrame',
                                   df_holdings: 'pd.DataFrame',
                                   df_history: 'pd.DataFrame',
                                   exit_mode: str = ''):
        """打印模拟结果的三板块总结。"""
        last_date = sim_summary.get('last_date', '')

        # ═══ 今日模拟买入 ═══
        print(f'\n{"="*80}')
        print(f'  📈 今日模拟买入  ({last_date})  —  共 {sim_summary["sim_buy_today"]} 只')
        print(f'{"="*80}')
        if df_buys.empty:
            print(f'  (无买入信号)')
        else:
            print(f'  {"代码":<8s} {"名称":<8s} {"评分":>5s} {"等级":>4s} '
                  f'{"买入价":>8s} {"信号日":>10s} {"回踩":>4s} {"回踩%":>7s} '
                  f'{"行业":<10s}')
            print(f'  {"-"*72}')
            for _, row in df_buys.iterrows():
                print(f'  {row.get("code",""):<8s} {row.get("name",""):<8s} '
                      f'{row.get("score",0):>5.0f} {row.get("tier",""):>4s} '
                      f'{row.get("buy_price",0):>8.2f} {row.get("signal_date",""):>10s} '
                      f'{row.get("pullback_days",0):>4.0f} '
                      f'{row.get("pullback_pct",0):>6.1f}% '
                      f'{row.get("industry",""):<10s}')

        # ═══ 今日模拟卖出 ═══
        print(f'\n{"="*80}')
        print(f'  📉 今日模拟卖出  ({last_date})  —  共 {sim_summary["sim_sell_today"]} 只')
        print(f'{"="*80}')
        if df_sells.empty:
            print(f'  (无卖出信号)')
        else:
            print(f'  {"代码":<8s} {"名称":<8s} {"买入价":>8s} {"卖出价":>8s} '
                  f'{"收益%":>7s} {"持有天":>6s} {"卖出原因":<14s} {"评分":>5s} {"等级":>4s}')
            print(f'  {"-"*78}')
            for _, row in df_sells.iterrows():
                print(f'  {row.get("code",""):<8s} {row.get("name",""):<8s} '
                      f'{row.get("buy_price",0):>8.2f} {row.get("sell_price",0):>8.2f} '
                      f'{row.get("net_ret",0):>+6.2f}% {row.get("days_held",0):>6.0f} '
                      f'{row.get("exit_reason",""):<14s} '
                      f'{row.get("score",0):>5.0f} {row.get("tier",""):>4s}')

        # ═══ 模拟持仓 ═══
        print(f'\n{"="*80}')
        print(f'  💼 模拟持仓  ({last_date})  —  共 {sim_summary["holdings"]} 只')
        print(f'{"="*80}')
        if df_holdings.empty:
            print(f'  (当前无持仓)')
        else:
            print(f'  {"代码":<8s} {"名称":<8s} {"买入价":>8s} {"现价":>8s} '
                  f'{"浮盈%":>7s} {"持有天":>6s} {"MA5":>8s} {"MA10":>8s} '
                  f'{"评分":>5s} {"等级":>4s} {"行业":<10s}')
            print(f'  {"-"*88}')
            for _, row in df_holdings.iterrows():
                print(f'  {row.get("code",""):<8s} {row.get("name",""):<8s} '
                      f'{row.get("buy_price",0):>8.2f} {row.get("current_price",0):>8.2f} '
                      f'{row.get("unrealized_pnl_pct",0):>+6.2f}% '
                      f'{row.get("days_held",0):>6.0f} '
                      f'{row.get("ma5",0):>8.2f} {row.get("ma10",0):>8.2f} '
                      f'{row.get("score",0):>5.0f} {row.get("tier",""):>4s} '
                      f'{row.get("industry",""):<10s}')

        # ═══ 总 结 ═══
        holding_value = 0.0
        for _, row in df_holdings.iterrows():
            holding_value += row.get('position_size', 10000)

        total_pnl = sum(h.get('net_ret', 0) for _, h in df_history.iterrows()) if not df_history.empty else 0
        if not df_holdings.empty:
            for _, row in df_holdings.iterrows():
                total_pnl += row.get('unrealized_pnl_yuan', 0)

        print(f'\n{"="*80}')
        print(f'  📊 模拟账户总结')
        print(f'{"="*80}')
        print(f'  回看区间: {sim_summary["total_days"]}天')
        print(f'  信号总数: {sim_summary["total_signals"]}  |  '
              f'回踩确认: {sim_summary["total_pullbacks"]}')
        print(f'  已完成交易: {sim_summary["total_trades"]}笔  |  '
              f'胜 {sim_summary["wins"]} 负 {sim_summary["losses"]}  |  '
              f'胜率: {sim_summary.get("win_rate",0):.1f}%')
        if sim_summary["total_trades"] > 0:
            print(f'  均收益: {sim_summary.get("avg_ret",0):+.2f}%  |  '
                  f'均盈利: {sim_summary.get("avg_win",0):+.2f}%  |  '
                  f'均亏损: {sim_summary.get("avg_loss",0):+.2f}%  |  '
                  f'复利累计: {sim_summary.get("cum_ret",0):+.1f}%')
        print(f'  当前持仓: {sim_summary["holdings"]}只 (市值约 {holding_value:.0f}元)  |  '
              f'待执行买入: {sim_summary["pending_buys"]}只')
        print(f'{"="*80}\n')

    # ─── 模拟结果 Excel 导出 ────────────────────────────────────────

    def export_simulation_xlsx(self, df_buys: 'pd.DataFrame',
                                df_sells: 'pd.DataFrame',
                                df_holdings: 'pd.DataFrame',
                                df_history: 'pd.DataFrame',
                                sim_summary: dict, path: str):
        """导出模拟结果到 Excel，四个工作表对应三个面板+历史。"""
        if df_buys.empty and df_sells.empty and df_holdings.empty and df_history.empty:
            print('[WARN] No simulation data to export')
            return

        wb = Workbook()
        last_date = sim_summary.get('last_date', '')

        # ── Sheet 1: 今日模拟买入 ──
        ws1 = wb.active
        ws1.title = '今日模拟买入'
        ws1['A1'] = f'今日模拟买入 ({last_date})'
        ws1['A1'].font = TITLE_FONT; ws1['A1'].alignment = CENTER
        ws1.merge_cells('A1:K1')

        buy_hdrs = ['代码', '名称', 'V2评分', '等级', 'V1评分', '买入价',
                    '信号日', '买入日', '回踩天数', '回踩幅度%', '缩量比', '行业']
        for c, h in enumerate(buy_hdrs, 1):
            ws1.cell(row=3, column=c, value=h)
        _style_row(ws1, 3, len(buy_hdrs))

        if not df_buys.empty:
            for i, (_, row) in enumerate(df_buys.iterrows()):
                r = i + 4
                ws1.cell(row=r, column=1, value=row.get('code', ''))
                ws1.cell(row=r, column=2, value=row.get('name', ''))
                ws1.cell(row=r, column=3, value=row.get('score', 0))
                tier = row.get('tier', '')
                ws1.cell(row=r, column=4, value=tier)
                ws1.cell(row=r, column=5, value=row.get('v1_score', 0))
                ws1.cell(row=r, column=6, value=row.get('buy_price', 0))
                ws1.cell(row=r, column=7, value=row.get('signal_date', ''))
                ws1.cell(row=r, column=8, value=row.get('buy_date', ''))
                ws1.cell(row=r, column=9, value=row.get('pullback_days', 0))
                ws1.cell(row=r, column=10, value=row.get('pullback_pct', 0))
                ws1.cell(row=r, column=11, value=row.get('vol_shrink_ratio', 0))
                ws1.cell(row=r, column=12, value=row.get('industry', ''))
                if tier == 'S': ws1.cell(row=r, column=4).fill = S_FILL
                elif tier == 'A': ws1.cell(row=r, column=4).fill = A_FILL
            _style_data(ws1, 4, 3 + len(df_buys), len(buy_hdrs))
        _auto_width(ws1)

        # ── Sheet 2: 今日模拟卖出 ──
        ws2 = wb.create_sheet('今日模拟卖出')
        ws2['A1'] = f'今日模拟卖出 ({last_date})'
        ws2['A1'].font = TITLE_FONT; ws2['A1'].alignment = CENTER
        ws2.merge_cells('A1:L1')

        sell_hdrs = ['代码', '名称', '买入日', '买入价', '卖出日', '卖出价',
                     '净收益%', '结果', '持有天', '卖出原因', '评分', '等级']
        for c, h in enumerate(sell_hdrs, 1):
            ws2.cell(row=3, column=c, value=h)
        _style_row(ws2, 3, len(sell_hdrs))

        if not df_sells.empty:
            for i, (_, row) in enumerate(df_sells.iterrows()):
                r = i + 4
                ws2.cell(row=r, column=1, value=row.get('code', ''))
                ws2.cell(row=r, column=2, value=row.get('name', ''))
                ws2.cell(row=r, column=3, value=row.get('buy_date', ''))
                ws2.cell(row=r, column=4, value=row.get('buy_price', 0))
                ws2.cell(row=r, column=5, value=row.get('sell_date', ''))
                ws2.cell(row=r, column=6, value=row.get('sell_price', 0))
                ws2.cell(row=r, column=7, value=row.get('net_ret', 0))
                is_win = row.get('is_win', False)
                ws2.cell(row=r, column=8, value='Win' if is_win else 'Loss')
                ws2.cell(row=r, column=9, value=row.get('days_held', 0))
                ws2.cell(row=r, column=10, value=row.get('exit_reason', ''))
                ws2.cell(row=r, column=11, value=row.get('score', 0))
                ws2.cell(row=r, column=12, value=row.get('tier', ''))
                if is_win:
                    ws2.cell(row=r, column=7).fill = WIN_FILL
                    ws2.cell(row=r, column=8).fill = WIN_FILL
                else:
                    ws2.cell(row=r, column=7).fill = LOSS_FILL
                    ws2.cell(row=r, column=8).fill = LOSS_FILL
            _style_data(ws2, 4, 3 + len(df_sells), len(sell_hdrs))
        _auto_width(ws2)

        # ── Sheet 3: 模拟持仓 ──
        ws3 = wb.create_sheet('模拟持仓')
        ws3['A1'] = f'模拟持仓 ({last_date})'
        ws3['A1'].font = TITLE_FONT; ws3['A1'].alignment = CENTER
        ws3.merge_cells('A1:M1')

        hold_hdrs = ['代码', '名称', '买入日', '买入价', '现价', '浮盈%',
                     '持有天', 'MA5', 'MA10', 'MA20', '评分', '等级', '行业']
        for c, h in enumerate(hold_hdrs, 1):
            ws3.cell(row=3, column=c, value=h)
        _style_row(ws3, 3, len(hold_hdrs))

        if not df_holdings.empty:
            for i, (_, row) in enumerate(df_holdings.iterrows()):
                r = i + 4
                ws3.cell(row=r, column=1, value=row.get('code', ''))
                ws3.cell(row=r, column=2, value=row.get('name', ''))
                ws3.cell(row=r, column=3, value=row.get('buy_date', ''))
                ws3.cell(row=r, column=4, value=row.get('buy_price', 0))
                ws3.cell(row=r, column=5, value=row.get('current_price', 0))
                pnl = row.get('unrealized_pnl_pct', 0)
                ws3.cell(row=r, column=6, value=pnl)
                ws3.cell(row=r, column=7, value=row.get('days_held', 0))
                ws3.cell(row=r, column=8, value=row.get('ma5', 0))
                ws3.cell(row=r, column=9, value=row.get('ma10', 0))
                ws3.cell(row=r, column=10, value=row.get('ma20', 0))
                ws3.cell(row=r, column=11, value=row.get('score', 0))
                tier = row.get('tier', '')
                ws3.cell(row=r, column=12, value=tier)
                ws3.cell(row=r, column=13, value=row.get('industry', ''))
                if pnl > 0:
                    ws3.cell(row=r, column=6).fill = WIN_FILL
                elif pnl < 0:
                    ws3.cell(row=r, column=6).fill = LOSS_FILL
                if tier == 'S': ws3.cell(row=r, column=12).fill = S_FILL
                elif tier == 'A': ws3.cell(row=r, column=12).fill = A_FILL
            _style_data(ws3, 4, 3 + len(df_holdings), len(hold_hdrs))
        _auto_width(ws3)

        # ── Sheet 4: 交易历史 ──
        ws4 = wb.create_sheet('交易历史')
        ws4['A1'] = '模拟交易历史'
        ws4['A1'].font = TITLE_FONT; ws4['A1'].alignment = CENTER
        ws4.merge_cells('A1:N1')

        hist_hdrs = ['代码', '名称', '信号日', '买入日', '买入价', '卖出日', '卖出价',
                     '净收益%', '结果', '持有天', '卖出原因', '评分', '等级', '行业']
        for c, h in enumerate(hist_hdrs, 1):
            ws4.cell(row=3, column=c, value=h)
        _style_row(ws4, 3, len(hist_hdrs))

        if not df_history.empty:
            for i, (_, row) in enumerate(df_history.iterrows()):
                r = i + 4
                ws4.cell(row=r, column=1, value=row.get('code', ''))
                ws4.cell(row=r, column=2, value=row.get('name', ''))
                ws4.cell(row=r, column=3, value=row.get('signal_date', ''))
                ws4.cell(row=r, column=4, value=row.get('buy_date', ''))
                ws4.cell(row=r, column=5, value=row.get('buy_price', 0))
                ws4.cell(row=r, column=6, value=row.get('sell_date', ''))
                ws4.cell(row=r, column=7, value=row.get('sell_price', 0))
                ws4.cell(row=r, column=8, value=row.get('net_ret', 0))
                is_win = row.get('is_win', False)
                ws4.cell(row=r, column=9, value='Win' if is_win else 'Loss')
                ws4.cell(row=r, column=10, value=row.get('days_held', 0))
                ws4.cell(row=r, column=11, value=row.get('exit_reason', ''))
                ws4.cell(row=r, column=12, value=row.get('score', 0))
                tier = row.get('tier', '')
                ws4.cell(row=r, column=13, value=tier)
                ws4.cell(row=r, column=14, value=row.get('industry', ''))
                if is_win:
                    ws4.cell(row=r, column=8).fill = WIN_FILL
                    ws4.cell(row=r, column=9).fill = WIN_FILL
                else:
                    ws4.cell(row=r, column=8).fill = LOSS_FILL
                    ws4.cell(row=r, column=9).fill = LOSS_FILL
                if tier == 'S': ws4.cell(row=r, column=13).fill = S_FILL
                elif tier == 'A': ws4.cell(row=r, column=13).fill = A_FILL
            _style_data(ws4, 4, 3 + len(df_history), len(hist_hdrs))
        _auto_width(ws4)

        # ── Sheet 5: 汇总 ──
        ws5 = wb.create_sheet('模拟汇总')
        ws5['A1'] = '模拟账户汇总'
        ws5['A1'].font = TITLE_FONT; ws5['A1'].alignment = CENTER
        ws5.merge_cells('A1:C1')

        rows_data = [
            ('模拟截止日', last_date),
            ('回看天数', sim_summary.get('total_days', 0)),
            ('信号总数', sim_summary.get('total_signals', 0)),
            ('回踩确认', sim_summary.get('total_pullbacks', 0)),
            ('已完成交易', sim_summary.get('total_trades', 0)),
            ('盈利笔数', sim_summary.get('wins', 0)),
            ('亏损笔数', sim_summary.get('losses', 0)),
            ('胜率', f'{sim_summary.get("win_rate", 0):.1f}%'),
            ('均收益', f'{sim_summary.get("avg_ret", 0):+.2f}%'),
            ('均盈利', f'{sim_summary.get("avg_win", 0):+.2f}%'),
            ('均亏损', f'{sim_summary.get("avg_loss", 0):+.2f}%'),
            ('复利累计', f'{sim_summary.get("cum_ret", 0):+.1f}%'),
            ('当前持仓', sim_summary.get('holdings', 0)),
            ('待执行买入', sim_summary.get('pending_buys', 0)),
        ]
        for i, (k, v) in enumerate(rows_data):
            ws5.cell(row=3 + i, column=1, value=k).font = SUB_FONT
            ws5.cell(row=3 + i, column=2, value=v).font = NUM_FONT
            ws5.cell(row=3 + i, column=1).border = THIN_BORDER
            ws5.cell(row=3 + i, column=2).border = THIN_BORDER
        _auto_width(ws5)

        wb.save(path)
        print(f'Simulation Excel saved: {path}')

    # ─── 统计 ──────────────────────────────────────────────────────

    def _summarize(self, df: pd.DataFrame) -> dict:
        if df.empty: return {}
        wins = df[df['is_win']]; losses = df[~df['is_win']]
        n = len(df); wr = len(wins) / n * 100

        # 累计收益(复利) + 最大回撤
        equity = 1.0
        equity_peak = 1.0
        max_dd = 0.0
        cum_list = []
        for _, t in df.iterrows():
            r = t['net_ret'] / 100
            equity *= (1 + r)
            equity_peak = max(equity_peak, equity)
            dd = (equity_peak - equity) / equity_peak * 100
            max_dd = max(max_dd, dd)
            cum_list.append((equity - 1) * 100)

        cum_ret = (equity - 1) * 100

        # 分层 (按 tier)
        layers = {}
        tier_order = [('S', 90, 101), ('A', 80, 90), ('B', 70, 80), ('C', 60, 70), ('D', 0, 60)]
        for tier, lo, hi in tier_order:
            sub = df[(df['score'] >= lo) & (df['score'] < hi)]
            if len(sub) == 0: continue
            sw = sub[sub['is_win']]
            layers[tier] = {
                'trades': len(sub), 'win_rate': round(len(sw) / len(sub) * 100, 1),
                'avg_ret': round(sub['net_ret'].mean(), 2),
                'cum_ret': round((sub['net_ret'] / 100 + 1).prod() * 100 - 100, 2),
            }

        # 按回踩天数分层
        pullback_layers = {}
        if 'pullback_days' in df.columns:
            for pd_days in range(1, MAX_WAIT + 1):
                sub = df[df['pullback_days'] == pd_days]
                if len(sub) == 0: continue
                sw = sub[sub['is_win']]
                pullback_layers[f'T+{pd_days}'] = {
                    'trades': len(sub), 'win_rate': round(len(sw) / len(sub) * 100, 1),
                    'avg_ret': round(sub['net_ret'].mean(), 2),
                }

        return {
            'trades': n, 'wins': len(wins), 'losses': len(losses),
            'win_rate': round(wr, 1),
            'avg_ret': round(df['net_ret'].mean(), 2),
            'avg_win': round(wins['net_ret'].mean(), 2) if len(wins) else 0,
            'avg_loss': round(losses['net_ret'].mean(), 2) if len(losses) else 0,
            'cum_ret': round(cum_ret, 2),
            'max_dd': round(max_dd, 2),
            'pl_ratio': round(abs(wins['net_ret'].mean() / losses['net_ret'].mean()), 2)
                        if len(losses) and len(wins) and losses['net_ret'].mean() != 0 else 0,
            'max_win': round(df['net_ret'].max(), 2),
            'max_loss': round(df['net_ret'].min(), 2),
            'layers': layers,
            'pullback_layers': pullback_layers,
            'cum_list': cum_list,
            'return_dist': {
                'gt_20': int((df['net_ret'] > 20).sum()),
                '10_to_20': int(((df['net_ret'] > 10) & (df['net_ret'] <= 20)).sum()),
                '5_to_10': int(((df['net_ret'] > 5) & (df['net_ret'] <= 10)).sum()),
                '0_to_5': int(((df['net_ret'] > 0) & (df['net_ret'] <= 5)).sum()),
                'neg5_to_0': int(((df['net_ret'] > -5) & (df['net_ret'] <= 0)).sum()),
                'neg10_to_neg5': int(((df['net_ret'] > -10) & (df['net_ret'] <= -5)).sum()),
                'lt_neg10': int((df['net_ret'] <= -10).sum()),
            },
        }


# ═══════════════════════════════════════════════════════════════════════════════
# Excel 样式
# ═══════════════════════════════════════════════════════════════════════════════

HEADER_FONT = Font(name='Microsoft YaHei', bold=True, color='FFFFFF', size=11)
HEADER_FILL = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
WIN_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
LOSS_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
TITLE_FONT = Font(name='Microsoft YaHei', bold=True, size=14, color='2F5496')
SUB_FONT = Font(name='Microsoft YaHei', bold=True, size=11)
NUM_FONT = Font(name='Consolas', size=10)
THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9'),
)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
GRAY_FILL = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
BLUE_FILL = PatternFill(start_color='E8F0FE', end_color='E8F0FE', fill_type='solid')
WARN_FILL = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')
S_FILL = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')   # 金色 S级
A_FILL = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')   # 绿色 A级


def _style_row(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT; cell.fill = HEADER_FILL; cell.alignment = CENTER; cell.border = THIN_BORDER

def _style_data(ws, sr, er, nc):
    for r in range(sr, er + 1):
        for c in range(1, nc + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = THIN_BORDER; cell.alignment = CENTER; cell.font = NUM_FONT

def _auto_width(ws):
    for col in ws.columns:
        mx = max((len(str(c.value or '')) for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(mx + 4, 30)


def export_xlsx(df, df_dynamic, ddf, summary, summary_dynamic, ss, path):
    wb = Workbook()

    # ==================================================================
    # Sheet 1: 汇总 (固定持有 + 动态 对比)
    # ==================================================================
    ws = wb.active; ws.title = '汇总'
    ws['A1'] = '启动突破战法 回测报告'; ws['A1'].font = TITLE_FONT; ws['A1'].alignment = CENTER
    ws.merge_cells('A1:J1')

    # 信号统计行
    ws.merge_cells('A3:J3')
    ws['A3'] = (f'信号: {ss["total"]} | 板块过滤: {ss.get("sector_filtered", 0)} '
                f'({ss.get("sector_filtered",0)/max(ss["total"]+ss.get("sector_filtered",0),1)*100:.1f}%) | '
                f'回踩确认: {ss["pulled_back"]} '
                f'({ss["pulled_back"]/max(ss["total"],1)*100:.1f}%) | '
                f'交易: {summary.get("trades",0)}笔 | 胜率: {summary.get("win_rate",0):.1f}%')
    ws['A3'].font = SUB_FONT

    # 双列对比: 固定持有 vs 动态
    for col_offset, (label, s) in enumerate([('固定持有3天', summary), ('动态止盈止损', summary_dynamic or {})]):
        col_start = 1 + col_offset * 5
        c = col_start
        ws.cell(row=5, column=c, value=label).font = SUB_FONT
        hdrs = ['指标', '数值']
        for i, h in enumerate(hdrs):
            ws.cell(row=6, column=c+i, value=h)
        _style_row(ws, 6, 2)

        if not s:
            ws.cell(row=7, column=c, value='无数据')
            continue

        rows_data = [
            ('总交易次数', s.get('trades', 0)),
            ('盈利笔数', s.get('wins', 0)),
            ('亏损笔数', s.get('losses', 0)),
            ('胜率', f'{s.get("win_rate",0):.1f}%'),
            ('平均收益率', f'{s.get("avg_ret",0):+.2f}%'),
            ('平均盈利', f'{s.get("avg_win",0):+.2f}%'),
            ('平均亏损', f'{s.get("avg_loss",0):+.2f}%'),
            ('累计收益率', f'{s.get("cum_ret",0):+.2f}%'),
            ('盈亏比', f'{s.get("pl_ratio",0):.2f}'),
            ('最大回撤', f'{s.get("max_dd",0):.2f}%'),
            ('单笔最大盈利', f'{s.get("max_win",0):+.2f}%'),
            ('单笔最大亏损', f'{s.get("max_loss",0):+.2f}%'),
            ('回踩成功率',
             f'{ss["pulled_back"]/max(ss["total"],1)*100:.1f}%'),
        ]
        for i, (k, v) in enumerate(rows_data):
            ws.cell(row=7+i, column=c, value=k)
            ws.cell(row=7+i, column=c+1, value=v)
        _style_data(ws, 7, 7 + len(rows_data) - 1, 2)

    # ==================================================================
    # 分层统计 (按评分等级)
    # ==================================================================
    r2 = 7 + 14 + 2
    ws.cell(row=r2, column=1, value='分层统计（按评分等级）').font = SUB_FONT
    hdrs2 = ['等级', '交易数', '胜率%', '均收益%', '累计收益%']
    for c, h in enumerate(hdrs2, 1): ws.cell(row=r2+1, column=c, value=h)
    _style_row(ws, r2+1, 5)
    for i, (tier, l) in enumerate(s.get('layers', {}).items()):
        ws.cell(row=r2+2+i, column=1, value=f'{tier}级')
        ws.cell(row=r2+2+i, column=2, value=l['trades'])
        ws.cell(row=r2+2+i, column=3, value=l['win_rate'])
        ws.cell(row=r2+2+i, column=4, value=l['avg_ret'])
        ws.cell(row=r2+2+i, column=5, value=l.get('cum_ret', 0))
        # 颜色
        if tier == 'S': ws.cell(row=r2+2+i, column=1).fill = S_FILL
        elif tier == 'A': ws.cell(row=r2+2+i, column=1).fill = A_FILL
    _style_data(ws, r2+2, r2+1+len(s.get('layers', {})), 5)

    # 按回踩天数分层
    r3 = r2 + 2 + len(s.get('layers', {})) + 2
    pb = s.get('pullback_layers', {})
    if pb:
        ws.cell(row=r3, column=1, value='按回踩天数分布').font = SUB_FONT
        hdrs3 = ['回踩天数', '交易数', '胜率%', '均收益%']
        for c, h in enumerate(hdrs3, 1): ws.cell(row=r3+1, column=c, value=h)
        _style_row(ws, r3+1, 4)
        for i, (name, l) in enumerate(pb.items()):
            ws.cell(row=r3+2+i, column=1, value=name)
            ws.cell(row=r3+2+i, column=2, value=l['trades'])
            ws.cell(row=r3+2+i, column=3, value=l['win_rate'])
            ws.cell(row=r3+2+i, column=4, value=l['avg_ret'])
        _style_data(ws, r3+2, r3+1+len(pb), 4)
    _auto_width(ws)

    # ==================================================================
    # Sheet 2: 逐日记录 (固定持有)
    # ==================================================================
    _write_daily_sheet(wb, df, ddf, '逐日记录(固定)')

    # ==================================================================
    # Sheet 3: 逐日记录 (动态) — 如果有
    # ==================================================================
    if df_dynamic is not None and not df_dynamic.empty:
        _write_daily_sheet(wb, df_dynamic, ddf, '逐日记录(动态)')

    # ==================================================================
    # Sheet 4: 每周汇总
    # ==================================================================
    ws4 = wb.create_sheet('每周汇总')
    ws4['A1'] = '每周汇总'; ws4['A1'].font = TITLE_FONT; ws4['A1'].alignment = CENTER
    if not df.empty:
        df4 = df.copy()
        df4['yw'] = pd.to_datetime(df4['signal_date']).dt.strftime('%Y-W%V')
        wk = df4.groupby('yw').agg(n=('is_win', 'count'), w=('is_win', 'sum'),
                                   r=('net_ret', 'sum')).reset_index().sort_values('yw')
        wk['l'] = wk['n'] - wk['w']; wk['wr'] = (wk['w'] / wk['n'] * 100).round(1)
        for c, h in enumerate(['周', '交易数', '胜', '负', '胜率%', '周收益%'], 1):
            ws4.cell(row=3, column=c, value=h)
        _style_row(ws4, 3, 6)
        cum4 = 0
        for i, (_, wr2) in enumerate(wk.iterrows()):
            r4 = i + 4; cum4 += wr2['r']
            ws4.cell(row=r4, column=1, value=wr2['yw']); ws4.cell(row=r4, column=2, value=wr2['n'])
            ws4.cell(row=r4, column=3, value=wr2['w']); ws4.cell(row=r4, column=4, value=wr2['l'])
            ws4.cell(row=r4, column=5, value=wr2['wr']); ws4.cell(row=r4, column=6, value=round(cum4, 2))
            if wr2['wr'] >= 55: ws4.cell(row=r4, column=5).fill = WIN_FILL
            elif wr2['wr'] < 45: ws4.cell(row=r4, column=5).fill = LOSS_FILL
        _style_data(ws4, 4, len(wk) + 3, 6)
    _auto_width(ws4)

    # ==================================================================
    # Sheet 5: 信号统计
    # ==================================================================
    ws5 = wb.create_sheet('信号统计')
    ws5['A1'] = '每日信号与回踩统计'; ws5['A1'].font = TITLE_FONT; ws5['A1'].alignment = CENTER
    hdrs5 = ['日期', '信号总数', '板块过滤', '回踩确认', '回踩率%', '实际交易', '胜', '负', '胜率%', '日收益%']
    for c, h in enumerate(hdrs5, 1): ws5.cell(row=3, column=c, value=h)
    _style_row(ws5, 3, len(hdrs5))

    for i, (_, di) in enumerate(ddf.iterrows()):
        r5 = i + 4
        ws5.cell(row=r5, column=1, value=di['date'])
        ws5.cell(row=r5, column=2, value=di['trades'])  # 实际交易的信号
        ws5.cell(row=r5, column=3, value='')  # 板块过滤数(每日无记录)
        ws5.cell(row=r5, column=4, value='')  # 回踩确认数(每日无记录)
        ws5.cell(row=r5, column=5, value='')
        ws5.cell(row=r5, column=6, value=di['trades'])
        ws5.cell(row=r5, column=7, value=di['wins'])
        ws5.cell(row=r5, column=8, value=di['losses'])
        wr = di['wins'] / max(di['trades'], 1) * 100
        ws5.cell(row=r5, column=9, value=round(wr, 1))
        ws5.cell(row=r5, column=10, value=di['ret_sum'])
        if di['trades'] > 0 and wr >= 60:
            ws5.cell(row=r5, column=9).fill = WIN_FILL
        elif di['trades'] > 0 and wr < 40:
            ws5.cell(row=r5, column=9).fill = LOSS_FILL
    _style_data(ws5, 4, len(ddf) + 3, len(hdrs5))
    _auto_width(ws5); ws5.freeze_panes = 'A4'

    wb.save(path)
    print(f'Saved: {path}')


def _write_daily_sheet(wb, df, ddf, title):
    """写入逐日记录工作表"""
    ws = wb.create_sheet(title)
    ws['A1'] = title; ws['A1'].font = TITLE_FONT; ws['A1'].alignment = CENTER
    ws.merge_cells('A1:P1')
    hdrs = ['信号日', '买入日', '代码', 'V2评分', '等级', 'V1评分', '行业', '回踩天',
            '回踩幅度%', '缩量比', '信号价', '买入价', '卖出日', '卖出价', '净收益%', '结果',
            '涨停数', '跌停数', 'MA60方向', '上涨比']
    ncols = len(hdrs)
    for c, h in enumerate(hdrs, 1): ws.cell(row=3, column=c, value=h)
    _style_row(ws, 3, ncols)

    ddf_s = ddf.sort_values('date') if not ddf.empty else ddf
    cum = 0.0; r = 4
    for _, di in ddf_s.iterrows():
        ds = di['date']; dt = df[df['signal_date'] == ds] if not df.empty else pd.DataFrame()
        if len(dt) == 0:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
            c = ws.cell(row=r, column=1, value=f'{ds}  |  无交易')
            c.font = Font(name='Microsoft YaHei', italic=True, color='999999', size=10)
            c.alignment = Alignment(horizontal='left', vertical='center')
            for cc in range(1, ncols+1):
                ws.cell(row=r, column=cc).border = THIN_BORDER
                ws.cell(row=r, column=cc).fill = GRAY_FILL
            r += 1
        else:
            dw = dt['is_win'].sum(); dtt = len(dt); dr = dt['net_ret'].sum(); cum += dr
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=12)
            txt = (f'{ds}  |  {dtt}笔  |  胜{int(dw)}负{int(dtt-dw)}  |  '
                   f'胜率{dw/dtt*100:.0f}%  |  日收益{dr:+.2f}%  |  累计{cum:+.1f}%')
            c = ws.cell(row=r, column=1, value=txt)
            c.font = Font(name='Microsoft YaHei', bold=True, size=10, color='2F5496')
            c.alignment = Alignment(horizontal='left', vertical='center')
            bg = BLUE_FILL if dw >= dtt * 0.5 else WARN_FILL
            for cc in range(1, ncols+1):
                ws.cell(row=r, column=cc).border = THIN_BORDER
                ws.cell(row=r, column=cc).fill = bg
            r += 1
            for _, t in dt.iterrows():
                ws.cell(row=r, column=1, value=t['signal_date'])
                ws.cell(row=r, column=2, value=t['buy_date'])
                ws.cell(row=r, column=3, value=t['code'])
                ws.cell(row=r, column=4, value=t['score'])
                ws.cell(row=r, column=5, value=t.get('tier', ''))
                ws.cell(row=r, column=6, value=t.get('v1_score', ''))
                ws.cell(row=r, column=7, value=t.get('industry', ''))
                ws.cell(row=r, column=8, value=t.get('pullback_days', ''))
                ws.cell(row=r, column=9, value=t.get('pullback_pct', ''))
                ws.cell(row=r, column=10, value=t.get('vol_shrink_ratio', ''))
                ws.cell(row=r, column=11, value=t['signal_close'])
                ws.cell(row=r, column=12, value=t['buy_price'])
                ws.cell(row=r, column=13, value=t.get('sell_date', t.get('sell_date', '')))
                ws.cell(row=r, column=14, value=t.get('sell_price', t.get('sell_price', '')))
                ws.cell(row=r, column=15, value=t['net_ret'])
                rc = ws.cell(row=r, column=16, value='Win' if t['is_win'] else 'Loss')
                if t['is_win']:
                    rc.fill = WIN_FILL; ws.cell(row=r, column=15).fill = WIN_FILL
                else:
                    rc.fill = LOSS_FILL; ws.cell(row=r, column=15).fill = LOSS_FILL
                # 市场状态列 (Q~T)
                for col_off, key in [(17, 'limit_up_num'), (18, 'limit_down_num'), (19, 'sh_ma60_up'), (20, 'up_ratio')]:
                    v = t.get(key, '')
                    if pd.notna(v) if hasattr(pd, 'notna') else v is not None:
                        ws.cell(row=r, column=col_off, value=v)
                # tier颜色
                tier = t.get('tier', '')
                if tier == 'S': ws.cell(row=r, column=5).fill = S_FILL
                elif tier == 'A': ws.cell(row=r, column=5).fill = A_FILL
                r += 1
    _style_data(ws, 4, r - 1, ncols)
    _auto_width(ws); ws.freeze_panes = 'A4'


# ═══════════════════════════════════════════════════════════════════════════════
# matplotlib 图表
# ═══════════════════════════════════════════════════════════════════════════════

def plot_charts(df, output_dir, label=''):
    """生成权益曲线图和收益分布直方图"""
    try:
        import matplotlib
        matplotlib.use('Agg')
        import matplotlib.pyplot as plt
        import matplotlib.ticker as ticker

        # 中文字体
        plt.rcParams['font.sans-serif'] = ['Microsoft YaHei', 'SimHei', 'DejaVu Sans']
        plt.rcParams['axes.unicode_minus'] = False
    except ImportError:
        print('[WARN] matplotlib not available, skipping charts')
        return

    if df.empty:
        print('[WARN] No trades to plot')
        return

    tag = f'_{label}' if label else ''
    fig, axes = plt.subplots(2, 2, figsize=(16, 10))
    fig.suptitle(f'启动突破战法 回测分析{tag}', fontsize=16, fontweight='bold')

    # ── 左上: 权益曲线（复利计算，与 _summarize 一致） ──
    ax1 = axes[0, 0]
    equity = 1.0; eq = []
    for _, t in df.iterrows():
        r = t['net_ret'] / 100
        equity *= (1 + r)
        eq.append((equity - 1) * 100)
    cum_ret = eq[-1] if eq else 0
    eq_arr = np.array(eq)
    ax1.plot(range(len(eq)), eq_arr, color='#2F5496', linewidth=1.2, label='复利累计收益%')
    ax1.fill_between(range(len(eq)), 0, eq_arr, where=eq_arr >= 0,
                     color='#C6EFCE', alpha=0.5)
    ax1.fill_between(range(len(eq)), 0, eq_arr, where=eq_arr < 0,
                     color='#FFC7CE', alpha=0.5)
    ax1.axhline(y=0, color='#999999', linewidth=0.5, linestyle='--')
    ax1.set_xlabel('交易序号'); ax1.set_ylabel('累计收益 (%)')
    ax1.set_title(f'复利权益曲线 (累计: {cum_ret:+.1f}%)')
    ax1.legend(loc='upper left')
    ax1.grid(True, alpha=0.3)

    # ── 右上: 收益分布直方图 ──
    ax2 = axes[0, 1]
    rets = df['net_ret'].dropna()
    bins = max(20, min(60, len(rets) // 10))
    n_bins, _, patches = ax2.hist(rets, bins=bins, edgecolor='white', alpha=0.8, color='#2F5496')
    # 正收益绿色, 负收益红色
    for patch, left_edge in zip(patches, np.linspace(rets.min(), rets.max(), bins+1)[:-1]):
        if left_edge >= 0:
            patch.set_facecolor('#C6EFCE')
        else:
            patch.set_facecolor('#FFC7CE')
    ax2.axvline(x=0, color='#999999', linewidth=0.5, linestyle='--')
    ax2.axvline(x=rets.mean(), color='#2F5496', linewidth=1.5, linestyle='-',
                label=f'均值: {rets.mean():+.2f}%')
    ax2.set_xlabel('单笔收益率 (%)')
    ax2.set_ylabel('频次')
    ax2.set_title(f'收益分布 (胜率: {df["is_win"].mean()*100:.1f}%)')
    ax2.legend(loc='upper right')
    ax2.grid(True, alpha=0.3, axis='y')

    # ── 左下: 按评分等级胜率 ──
    ax3 = axes[1, 0]
    if 'tier' in df.columns:
        tiers_order = ['S', 'A', 'B', 'C', 'D']
        tier_data = []
        for t in tiers_order:
            sub = df[df['tier'] == t]
            if len(sub) > 0:
                tier_data.append({'tier': t, 'count': len(sub),
                                  'wr': sub['is_win'].mean() * 100,
                                  'avg_ret': sub['net_ret'].mean()})
        if tier_data:
            tdf = pd.DataFrame(tier_data)
            colors_tier = {'S': '#FFD700', 'A': '#90EE90', 'B': '#87CEEB', 'C': '#DDA0DD', 'D': '#FFB6C1'}
            bar_colors = [colors_tier.get(t, '#2F5496') for t in tdf['tier']]
            bars = ax3.bar(range(len(tdf)), tdf['avg_ret'], color=bar_colors, edgecolor='white')
            ax3.set_xticks(range(len(tdf))); ax3.set_xticklabels(tdf['tier'])
            ax3.set_xlabel('评分等级'); ax3.set_ylabel('平均收益 (%)')
            ax3.set_title('按评分等级 平均收益')
            ax3.axhline(y=0, color='#999999', linewidth=0.5, linestyle='--')
            ax3.grid(True, alpha=0.3, axis='y')
            for bi, (b, v) in enumerate(zip(bars, tdf['avg_ret'])):
                cnt = int(tdf.iloc[bi]['count'])
                ax3.text(b.get_x() + b.get_width()/2., v + (0.3 if v >= 0 else -1.0),
                        f'{v:+.1f}%\n({cnt}笔)',
                        ha='center', va='bottom' if v >= 0 else 'top', fontsize=8)

    # ── 右下: 滚动10笔胜率 ──
    ax4 = axes[1, 1]
    window = min(20, max(5, len(df) // 5))
    rolling_wr = df['is_win'].rolling(window=window).mean() * 100
    ax4.plot(range(len(rolling_wr)), rolling_wr, color='#2F5496', linewidth=1.0)
    ax4.axhline(y=50, color='#999999', linewidth=0.5, linestyle='--', label='50%基准')
    ax4.fill_between(range(len(rolling_wr)), 50, rolling_wr,
                     where=rolling_wr >= 50, color='#C6EFCE', alpha=0.3)
    ax4.fill_between(range(len(rolling_wr)), 50, rolling_wr,
                     where=rolling_wr < 50, color='#FFC7CE', alpha=0.3)
    ax4.set_xlabel('交易序号'); ax4.set_ylabel(f'滚动{window}笔胜率 (%)')
    ax4.set_title(f'滚动{window}笔胜率')
    ax4.legend(loc='upper right')
    ax4.grid(True, alpha=0.3)
    ax4.set_ylim(0, 100)

    plt.tight_layout()
    chart_path = os.path.join(output_dir, f'start_breakout_charts{tag}.png')
    plt.savefig(chart_path, dpi=150, bbox_inches='tight')
    plt.close()
    print(f'Charts saved: {chart_path}')


# ═══════════════════════════════════════════════════════════════════════════════
# CLI
# ═══════════════════════════════════════════════════════════════════════════════

if __name__ == '__main__':
    p = argparse.ArgumentParser(description='启动突破战法回测 v3')
    p.add_argument('--days', type=int, default=250, help='回测天数')
    p.add_argument('--hold', type=int, default=5, help='固定持有天数 (默认5, 仅fixed模式)')
    p.add_argument('--exit-mode', choices=['fixed', 'dynamic', 'ma5', 'ma10', 'tp_ma5', 'all'],
                   default='all', help='卖出方式 (默认all=五种全部对比)')
    p.add_argument('--skip-sector-filter', action='store_true',
                   help='跳过板块环境过滤 (对比测试用)')
    p.add_argument('--end-date', type=str, default=None,
                   help='回测截止日 YYYY-MM-DD (默认昨天, 样本外用2025-06-30)')
    p.add_argument('--sim', action='store_true',
                   help='前向模拟模式: 输出今日模拟买入/卖出/持仓三板块')
    p.add_argument('--sim-position-size', type=float, default=10000.0,
                   help='模拟每只股票固定金额 (默认10000)')
    a = p.parse_args()

    end_dt = None
    if a.end_date:
        end_dt = datetime.strptime(a.end_date, '%Y-%m-%d').date()

    bt = StartBreakoutBacktest(skip_sector_filter=a.skip_sector_filter)

    # ── 模拟模式 ──
    if a.sim:
        sim_exit = a.exit_mode if a.exit_mode != 'all' else 'fixed'
        sim_hold = a.hold if sim_exit == 'fixed' else 3
        df_buys, df_sells, df_holdings, df_history, sim_summary = bt.run_simulation(
            lookback=a.days, exit_mode=sim_exit, hold_days=sim_hold,
            end_date=end_dt, position_size=a.sim_position_size)

        # 导出 Excel
        out_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'data')
        suffix = f'_{sim_exit}'
        if a.skip_sector_filter: suffix += '_nosector'
        out = os.path.join(out_dir, f'start_breakout_simulation_{a.days}d{suffix}.xlsx')
        bt.export_simulation_xlsx(df_buys, df_sells, df_holdings, df_history,
                                   sim_summary, out)
        sys.exit(0)

    # ── 回测模式 (原有逻辑) ──
    df_fixed, df_dynamic, df_ma5, df_ma10, df_tp_ma5, ddf, total_days, ss = bt.run(
        lookback=a.days, exit_mode=a.exit_mode, hold_days=a.hold, end_date=end_dt)

    # 收集所有退出方式的交易
    all_modes = []
    if not df_fixed.empty:
        all_modes.append((f'固定持有{a.hold}天', df_fixed))
    if not df_dynamic.empty:
        all_modes.append(('动态止盈-5%/+10%', df_dynamic))
    if not df_ma5.empty:
        all_modes.append(('跌破MA5卖出', df_ma5))
    if not df_ma10.empty:
        all_modes.append(('跌破MA10卖出', df_ma10))
    if not df_tp_ma5.empty:
        all_modes.append(('+10%止盈+MA5跟踪', df_tp_ma5))

    # ── 对比报告 ──
    print(f'\n{"="*70}')
    print(f'  退出策略对比 (共{len(all_modes)}种)')
    print(f'{"="*70}')
    print(f'  {"策略":<20s} {"笔数":>6s} {"胜率":>7s} {"均收益":>8s} {"累计":>9s} {"盈亏比":>7s} {"中位数":>7s}')
    print(f'  {"-"*64}')

    for label, df_t in all_modes:
        s = bt._summarize(df_t)
        if not s: continue
        median_ret = df_t['net_ret'].median() if not df_t.empty else 0
        print(f'  {label:<20s} {s["trades"]:>6d} {s["win_rate"]:>6.1f}% '
              f'{s["avg_ret"]:>+7.2f}% {s["cum_ret"]:>+8.1f}% {s["pl_ratio"]:>6.2f} '
              f'{median_ret:>+6.2f}%')

    print(f'{"="*70}')
    print(f'  注: 累计收益 = 每笔独立1万求和/非真实资金曲线')
    print(f'  真实回报率: 100万本金/1%仓位(+3~5%), 5%仓位(+15~25%)')

    # ── Exit × ScoreLevel 交叉分析 ──
    if all_modes:
        print(f'\n{"="*90}')
        print(f'  Exit × ScoreLevel 交叉分析（退出方式 × V2评分层级）')
        print(f'{"="*90}')
        print(f'  {"退出方式":<16s} {"层级":<5s} {"笔数":>6s} {"胜率":>7s} {"均收益":>8s} {"中位数":>8s} {"复利CAGR":>9s}')
        print(f'  {"-"*65}')
        for label, df_t in all_modes:
            for tier in ['S','A','B','C','D']:
                sub = df_t[df_t['tier'] == tier]
                if len(sub) < 3: continue
                wr = sub['is_win'].mean() * 100
                avg = sub['net_ret'].mean()
                med = sub['net_ret'].median()
                equity = 1.0
                for r in sub['net_ret']:
                    equity *= (1 + r / 100)
                cagr = (equity - 1) * 100
                print(f'  {label:<16s} {tier:<5s} {len(sub):>6d} {wr:>6.1f}% {avg:>+7.2f}% {med:>+7.2f}% {cagr:>+8.1f}%')
        print(f'{"="*90}')

        # V2因子贡献分析（合并所有退出方式）
        combined = pd.concat([df_t for _, df_t in all_modes], ignore_index=True)
        if len(combined) > 50:
            print(f'\n  V2因子贡献 (大赚>10% vs 其他, {len(combined)}笔合并):')
            big_w = combined[combined['net_ret'] > 10]
            oth = combined[combined['net_ret'] <= 10]
            if len(big_w) >= 5 and len(oth) >= 5:
                for factor_key in ['信号强度', '距年高', '60日涨', '位置共振',
                                       '回踩幅度', '缩量', '止跌K线', '回踩天数']:
                    bw_sum = 0; ot_sum = 0
                    for _, row in big_w.iterrows():
                        try:
                            f = row['v2_factors']
                            if isinstance(f, str): f = eval(f)
                            if isinstance(f, dict): bw_sum += f.get(factor_key, 0)
                        except: pass
                    for _, row in oth.iterrows():
                        try:
                            f = row['v2_factors']
                            if isinstance(f, str): f = eval(f)
                            if isinstance(f, dict): ot_sum += f.get(factor_key, 0)
                        except: pass
                    bw_avg = bw_sum / len(big_w); ot_avg = ot_sum / len(oth)
                    diff = bw_avg - ot_avg
                    sig = ' **' if abs(diff) > 1 else ''
                    print(f'    {factor_key:<10s}:  大赚{bw_avg:4.1f}  普通{ot_avg:4.1f}  差{diff:+5.1f}{sig}')


    # ── 图表 ──
    out_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'data')
    for label, df_t in all_modes:
        if len(df_t) > 10:
            plot_charts(df_t, out_dir, label.replace('/', '_'))

    # ── Excel ──
    suffix = f'_{a.exit_mode}' if a.exit_mode != 'all' else ''
    if a.skip_sector_filter: suffix += '_nosector'
    out = os.path.join(out_dir, f'start_breakout_backtest_{a.days}d{suffix}.xlsx')
    summary_main = bt._summarize(df_fixed)
    export_xlsx(df_fixed, df_dynamic, ddf, summary_main,
                bt._summarize(df_dynamic) if not df_dynamic.empty else {}, ss, out)
