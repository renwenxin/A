"""竞价确认型三种选股回测 → 桌面xlsx

规则：
- 每种选股方式每日只选前3名
- T日开盘买入 → T+1日必须卖出（持仓1天）
- 止损-5% / 止盈+7%
- 清晰展示每日买入卖出标的
"""
import sys, os
from datetime import date, datetime, timedelta
from collections import defaultdict
import numpy as np

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from ashare_review.analysis.auction_confirm_backtest import (
    AuctionConfirmBacktest, RESULT_LABELS
)

# ============================================================
# 样式定义
# ============================================================
HEADER_FONT = Font(name='Microsoft YaHei', bold=True, size=11, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)

TITLE_FONT = Font(name='Microsoft YaHei', bold=True, size=14, color='1F3864')
SUBTITLE_FONT = Font(name='Microsoft YaHei', bold=True, size=12, color='2F5496')
SECTION_FONT = Font(name='Microsoft YaHei', bold=True, size=11, color='C55A11')

WIN_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')       # 绿色-盈利
LOSS_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')      # 红色-亏损
TIMEOUT_FILL = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')   # 黄色-超时
SKIPPED_FILL = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')   # 灰色-放弃
BUY_FILL = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')        # 蓝色-买入标记

ALL_KEYS = ['1进2', '竞价抢筹', '优化总筛选', '龙虎榜竞价']

METHOD_FILLS = {
    '1进2':         PatternFill(start_color='DAEEF3', end_color='DAEEF3', fill_type='solid'),  # 浅蓝
    '竞价抢筹':     PatternFill(start_color='E4DFEC', end_color='E4DFEC', fill_type='solid'),  # 浅紫
    '优化总筛选':   PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),  # 浅绿
    '龙虎榜竞价':   PatternFill(start_color='F8CBAD', end_color='F8CBAD', fill_type='solid'),  # 深橙
}

CENTER_ALIGN = Alignment(horizontal='center', vertical='center')
LEFT_ALIGN = Alignment(horizontal='left', vertical='center')
THIN_BORDER = Border(
    left=Side(style='thin', color='B0B0B0'),
    right=Side(style='thin', color='B0B0B0'),
    top=Side(style='thin', color='B0B0B0'),
    bottom=Side(style='thin', color='B0B0B0'),
)

NUM_FONT = Font(name='Consolas', size=10)
CN_FONT = Font(name='Microsoft YaHei', size=10)
BOLD_CN = Font(name='Microsoft YaHei', size=10, bold=True)


def style_header_row(ws, row, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def style_data_cell(ws, row, col, is_number=False, bold=False):
    cell = ws.cell(row=row, column=col)
    cell.font = BOLD_CN if bold else (NUM_FONT if is_number else CN_FONT)
    cell.alignment = CENTER_ALIGN if is_number else LEFT_ALIGN
    cell.border = THIN_BORDER


def auto_width(ws, min_width=8, max_width=42):
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = 0
        for cell in col_cells:
            if cell.value:
                val = str(cell.value)
                length = sum(2 if ord(c) > 127 else 1 for c in val)
                max_len = max(max_len, length)
        width = min(max(max_len + 3, min_width), max_width)
        ws.column_dimensions[col_letter].width = width


def result_fill(result_type):
    """根据交易结果返回填充色"""
    if result_type == 'win':
        return WIN_FILL
    elif result_type == 'loss':
        return LOSS_FILL
    elif result_type == 'timeout':
        return TIMEOUT_FILL
    else:
        return SKIPPED_FILL


def build_xlsx(result: dict, filepath: str):
    """生成Excel回测报告"""
    wb = Workbook()
    methods_data = result['methods']
    params = result['parameters']
    ranking = result['ranking']

    # ================================================================
    # Sheet 1: 每日操作明细（核心sheet — 买入+卖出对应展示）
    # ================================================================
    ws_daily = wb.active
    ws_daily.title = '每日操作明细'

    ws_daily.merge_cells('A1:R1')
    ws_daily.cell(row=1, column=1,
                  value=f'竞价确认型回测 — 每日买入/卖出明细（多因子精选 | Top3 | T日买→T+1卖 | 止损{params["stop_loss_pct"]:+.0f}% 止盈{params["take_profit_pct"]:+.0f}%）').font = TITLE_FONT
    ws_daily.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
    ws_daily.row_dimensions[1].height = 30

    # 列结构：按日期分组，每天显示三种方式的买入→卖出
    # A:日期 B:选股方式 C:排名 D:代码 E:名称 F:买入价 G:卖出价 H:结果 I:收益率% J:卖出日期 K:评分
    headers = ['买入日期', '选股方式', '排名', '代码', '名称', '买入价',
               '卖出价', '结果', '收益率(%)', '卖出日期', '评分', '止损价', '止盈价']
    for col, h in enumerate(headers, 1):
        ws_daily.cell(row=2, column=col, value=h)
    style_header_row(ws_daily, 2, len(headers))

    # 收集全部交易，按日期+方式+排名排序
    all_trades = []
    for key, stats in methods_data.items():
        for t in stats.get('all_trades', []):
            all_trades.append(t)

    # 按日期分组，组内按方式排序
    trades_by_date = defaultdict(list)
    for t in all_trades:
        trades_by_date[t['entry_date']].append(t)

    # 为每种方式每天的交易编号（前3名=1,2,3）
    for d, trades in trades_by_date.items():
        method_rank = defaultdict(int)
        for t in sorted(trades, key=lambda x: (x['method'], -x['score'])):
            method_rank[t['method']] += 1
            t['rank'] = method_rank[t['method']]

    # 按日期正序排列（方便看每日操作流程）
    sorted_dates = sorted(trades_by_date.keys())

    row = 3
    for d in sorted_dates:
        day_trades = trades_by_date[d]
        day_trades.sort(key=lambda x: (
            {'1进2': 0, '竞价抢筹': 1, '优化总筛选': 2, '龙虎榜竞价': 3}.get(x['method'], 99),
            x['rank']
        ))

        # 日期分隔行
        ws_daily.merge_cells(f'A{row}:M{row}')
        total_return = sum(t['return_pct'] for t in day_trades if t['result'] != 'skipped')
        win_count = sum(1 for t in day_trades if t['result'] == 'win')
        loss_count = sum(1 for t in day_trades if t['result'] == 'loss')
        skip_count = sum(1 for t in day_trades if t['result'] == 'skipped')
        valid_count = len(day_trades) - skip_count
        date_label = (f'▌ {d}  （{len(day_trades)}笔信号 | {valid_count}笔成交 | '
                      f'胜{win_count} 负{loss_count} 弃{skip_count} | '
                      f'当日收益 {total_return:+.1f}%）')
        ws_daily.cell(row=row, column=1, value=date_label).font = SECTION_FONT
        ws_daily.cell(row=row, column=1).fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        ws_daily.row_dimensions[row].height = 24
        row += 1

        # 每笔交易
        for t in day_trades:
            result_label = RESULT_LABELS.get(t['result'], t['result'])
            values = [
                t['entry_date'], t['method'], t['rank'],
                f"'{t['code']}", t['name'],
                t['entry_price'], t['exit_price'],
                result_label, t['return_pct'], t['exit_date'],
                t['score'], t.get('stop_loss', 0), t.get('target', 0),
            ]
            for col, val in enumerate(values, 1):
                ws_daily.cell(row=row, column=col, value=val)
                style_data_cell(ws_daily, row, col, is_number=col in (3, 6, 7, 9, 11, 12, 13))

            # 结果着色（H列和I列）
            rfill = result_fill(t['result'])
            ws_daily.cell(row=row, column=8).fill = rfill
            ws_daily.cell(row=row, column=9).fill = rfill

            # 方法色（B列）
            ws_daily.cell(row=row, column=2).fill = METHOD_FILLS.get(t['method'], PatternFill())

            # 买入价列加浅蓝背景
            ws_daily.cell(row=row, column=6).fill = BUY_FILL

            row += 1

        # 空行分隔
        row += 1

    auto_width(ws_daily)
    ws_daily.freeze_panes = 'A3'
    ws_daily.auto_filter.ref = f'A2:M{row}'

    # ================================================================
    # Sheet 2: 汇总对比
    # ================================================================
    ws_summary = wb.create_sheet('汇总对比')

    ws_summary.merge_cells('A1:H1')
    ws_summary.cell(row=1, column=1, value='竞价确认型四种选股方式 — 120日回测报告（优化版）').font = TITLE_FONT
    ws_summary.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
    ws_summary.row_dimensions[1].height = 30

    ws_summary.merge_cells('A2:H2')
    param_text = (f'规则：T日开盘买入 → T+1日强制卖出 | 每日每种方式≤3只（前3名）| '
                  f'止损{params["stop_loss_pct"]:+.0f}% 止盈{params["take_profit_pct"]:+.0f}% | '
                  f'回看{params["total_days"]}天 | {datetime.now().strftime("%Y-%m-%d %H:%M")}')
    ws_summary.cell(row=2, column=1, value=param_text).font = Font(name='Microsoft YaHei', size=9, color='888888')
    ws_summary.row_dimensions[2].height = 22

    # 排名
    ws_summary.merge_cells('A4:H4')
    ws_summary.cell(row=4, column=1, value='▎综合排名').font = SUBTITLE_FONT

    rank_headers = ['排名', '选股方式', '综合得分', '胜率(%)', '平均收益(%)', '盈亏比', '有效交易', '最大回撤(%)']
    for col, h in enumerate(rank_headers, 1):
        ws_summary.cell(row=5, column=col, value=h)
    style_header_row(ws_summary, 5, len(rank_headers))

    for i, r in enumerate(ranking):
        row = 6 + i
        stats = methods_data.get(r['method'], {})
        values = [i + 1, r['method'], r['composite_score'], r['win_rate'],
                  r['avg_return'], r['profit_factor'], r['valid_trades'],
                  stats.get('max_drawdown_pct', 0)]
        for col, val in enumerate(values, 1):
            ws_summary.cell(row=row, column=col, value=val)
            style_data_cell(ws_summary, row, col, is_number=col >= 3)
        if i == 0:
            gold = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
            for col in range(1, len(rank_headers) + 1):
                ws_summary.cell(row=row, column=col).fill = gold

    # 详细统计
    det = 6 + len(ranking) + 2
    ws_summary.merge_cells(f'A{det}:P{det}')
    ws_summary.cell(row=det, column=1, value='▎详细统计').font = SUBTITLE_FONT

    detail_h = ['选股方式', '信号总数', '有效交易', '高开放弃', '胜(止盈)', '负(止损)', '超时',
                '胜率(%)', '平均收益(%)', '平均盈利(%)', '平均亏损(%)',
                '总盈利(%)', '总亏损(%)', '盈亏比', '最大回撤(%)', '平均持仓(天)']
    for col, h in enumerate(detail_h, 1):
        ws_summary.cell(row=det + 1, column=col, value=h)
    style_header_row(ws_summary, det + 1, len(detail_h))

    for i, key in enumerate(ALL_KEYS):
        row = det + 2 + i
        s = methods_data.get(key, {})
        vals = [s.get('method', key), s.get('total_signals', 0), s.get('valid_trades', 0),
                s.get('skipped', 0), s.get('wins', 0), s.get('losses', 0), s.get('timeouts', 0),
                s.get('win_rate', 0), s.get('avg_return', 0), s.get('avg_win', 0),
                s.get('avg_loss', 0), s.get('total_profit', 0), s.get('total_loss', 0),
                s.get('profit_factor', 0), s.get('max_drawdown_pct', 0), s.get('avg_days_held', 0)]
        for col, val in enumerate(vals, 1):
            ws_summary.cell(row=row, column=col, value=val)
            style_data_cell(ws_summary, row, col, is_number=col >= 4)

    auto_width(ws_summary)
    ws_summary.freeze_panes = 'A6'

    # ================================================================
    # Sheet 3-5: 每种方式的独立明细
    # ================================================================
    for key in ALL_KEYS:
        stats = methods_data.get(key, {})
        if not stats:
            continue
        sname = key.replace('(', ' ').replace(')', '').strip()[:31]
        ws = wb.create_sheet(sname)

        ws.merge_cells('A1:M1')
        ws.cell(row=1, column=1, value=f'{stats["method"]} — 交易明细').font = TITLE_FONT
        ws.row_dimensions[1].height = 28

        stext = (f'信号{stats["total_signals"]} | 成交{stats["valid_trades"]} | '
                 f'弃{stats["skipped"]} | 胜{stats["wins"]} | 负{stats["losses"]} | '
                 f'超时{stats["timeouts"]} | 胜率{stats["win_rate"]:.1f}% | '
                 f'均收益{stats["avg_return"]:+.1f}% | 盈亏比{stats["profit_factor"]:.2f} | '
                 f'最大回撤{stats["max_drawdown_pct"]:.1f}%')
        ws.merge_cells('A2:M2')
        ws.cell(row=2, column=1, value=stext).font = Font(name='Microsoft YaHei', size=10, color='2F5496')
        ws.row_dimensions[2].height = 22

        mh = ['买入日期', '卖出日期', '代码', '名称', '买入价', '卖出价', '结果',
              '收益率(%)', '持仓天数', '评分', '止损价', '止盈价', '排名']
        for col, h in enumerate(mh, 1):
            ws.cell(row=3, column=col, value=h)
        style_header_row(ws, 3, len(mh))

        trades = stats.get('all_trades', [])
        # 按日期排序后给排名
        by_date = defaultdict(list)
        for t in trades:
            by_date[t['entry_date']].append(t)
        for d, tlist in by_date.items():
            tlist.sort(key=lambda x: -x['score'])
            for i, t in enumerate(tlist):
                t['rank'] = i + 1

        # 重新按日期排序
        trades.sort(key=lambda x: (x['entry_date'], -x['score']))

        for i, t in enumerate(trades):
            row = 4 + i
            rlabel = RESULT_LABELS.get(t['result'], t['result'])
            vals = [t['entry_date'], t['exit_date'], f"'{t['code']}", t['name'],
                    t['entry_price'], t['exit_price'], rlabel,
                    t['return_pct'], t['days_held'], t['score'],
                    t.get('stop_loss', 0), t.get('target', 0), t.get('rank', 0)]
            for col, val in enumerate(vals, 1):
                ws.cell(row=row, column=col, value=val)
                style_data_cell(ws, row, col, is_number=col >= 5)

            rfill = result_fill(t['result'])
            ws.cell(row=row, column=7).fill = rfill
            ws.cell(row=row, column=8).fill = rfill
            ws.cell(row=row, column=5).fill = BUY_FILL

        auto_width(ws)
        ws.freeze_panes = 'A4'
        ws.auto_filter.ref = f'A3:{get_column_letter(len(mh))}{3 + len(trades)}'

    # ================================================================
    # Sheet 6: 每日收益汇总
    # ================================================================
    ws6 = wb.create_sheet('每日收益汇总')

    ws6.merge_cells('A1:Q1')
    ws6.cell(row=1, column=1, value='竞价确认型 — 每日收益汇总（按买入日期）').font = TITLE_FONT
    ws6.row_dimensions[1].height = 28

    dh = ['日期', '1进2买入', '1进2卖出', '1进2胜', '1进2负', '1进2收益%',
          '竞价买入', '竞价卖出', '竞价胜', '竞价负', '竞价收益%',
          '优化买入', '优化卖出', '优化胜', '优化负', '优化收益%',
          '龙虎榜竞价买入', '龙虎榜竞价卖出', '龙虎榜竞价胜', '龙虎榜竞价负', '龙虎榜竞价收益%',
          '当日总收益(%)', '累计收益(%)']
    for col, h in enumerate(dh, 1):
        ws6.cell(row=2, column=col, value=h)
    style_header_row(ws6, 2, len(dh))

    daily = defaultdict(lambda: {
        '1进2': {'buy': 0, 'sell': 0, 'wins': 0, 'losses': 0, 'ret': 0.0},
        '竞价抢筹': {'buy': 0, 'sell': 0, 'wins': 0, 'losses': 0, 'ret': 0.0},
        '优化总筛选': {'buy': 0, 'sell': 0, 'wins': 0, 'losses': 0, 'ret': 0.0},
        '龙虎榜竞价': {'buy': 0, 'sell': 0, 'wins': 0, 'losses': 0, 'ret': 0.0},
    })

    for key, stats in methods_data.items():
        for t in stats.get('all_trades', []):
            d = t['entry_date']
            daily[d][key]['buy'] += 1
            if t['result'] != 'skipped':
                daily[d][key]['sell'] += 1
                daily[d][key]['ret'] += t['return_pct']
            if t['result'] == 'win':
                daily[d][key]['wins'] += 1
            elif t['result'] == 'loss':
                daily[d][key]['losses'] += 1

    cum_ret = 0.0
    for i, d in enumerate(sorted(daily.keys())):
        row = 3 + i
        ws6.cell(row=row, column=1, value=str(d))
        style_data_cell(ws6, row, 1)

        total_ret = 0.0
        col = 2
        for key in ALL_KEYS:
            info = daily[d][key]
            for v in [info['buy'], info['sell'], info['wins'], info['losses'], round(info['ret'], 1)]:
                ws6.cell(row=row, column=col, value=v)
                style_data_cell(ws6, row, col, is_number=True)
                col += 1
            total_ret += info['ret']

        cum_ret += total_ret
        ws6.cell(row=row, column=col, value=round(total_ret, 1))
        ws6.cell(row=row, column=col + 1, value=round(cum_ret, 1))
        style_data_cell(ws6, row, col, is_number=True)
        style_data_cell(ws6, row, col + 1, is_number=True)

        ret_cell = ws6.cell(row=row, column=col)
        if total_ret > 0:
            ret_cell.fill = WIN_FILL
        elif total_ret < 0:
            ret_cell.fill = LOSS_FILL

        # 累计收益着色
        cum_cell = ws6.cell(row=row, column=col + 1)
        if cum_ret > 0:
            cum_cell.fill = WIN_FILL
        elif cum_ret < 0:
            cum_cell.fill = LOSS_FILL

    auto_width(ws6)
    ws6.freeze_panes = 'A3'

    # 保存
    wb.save(filepath)
    print(f'\n[xlsx] 回测报告已生成 → {filepath}')
    return filepath


# ============================================================
# 主程序
# ============================================================
def main():
    desktop = os.path.join(os.path.expanduser('~'), 'Desktop')
    output_path = os.path.join(desktop, '竞价确认型_120日回测_优化版.xlsx')

    print('=' * 70)
    print('  竞价确认型四种选股方式 — 120日回测')
    print('  规则：T日开盘买 → T+1日强制卖 | 每日每方式≤3只')
    print('  止损-5% | 止盈+7% | 含优化总筛选+龙虎榜竞价')
    print('=' * 70)
    print()

    bt = AuctionConfirmBacktest()
    result = bt.run(
        lookback_days=120,
        stop_loss=-0.05,
        take_profit=0.07,
        max_hold=1,        # ← T+1强制卖出
        top_n=3,           # ← 每种方式只选前3名
        enable_lhb=True,   # ← 启用龙虎榜竞价抢筹
    )

    # 打印终端报告
    from ashare_review.analysis.auction_confirm_backtest import _print_comparison
    _print_comparison(result)

    # 生成xlsx
    build_xlsx(result, output_path)

    print(f'\n{"="*70}')
    print(f'  [OK] xlsx已保存到桌面: 竞价确认型_120日回测_四种策略.xlsx')
    print(f'  包含以下工作表:')
    print(f'    1. 每日操作明细 — 【核心】按日期展示买入→卖出标的')
    print(f'    2. 汇总对比 — 四种方式综合排名+详细统计')
    print(f'    3. 1进2 — 方式1独立明细')
    print(f'    4. 竞价抢筹 — 方式2独立明细')
    print(f'    5. 优化总筛选 — 方式3独立明细（多因子竞价精选）')
    print(f'    6. 龙虎榜竞价 — 方式4独立明细')
    print(f'    7. 每日收益汇总 — 按日期聚合的收益曲线')
    print(f'{"="*70}')


if __name__ == '__main__':
    main()
