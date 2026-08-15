"""V3 回测 — 市场情绪相关性分析"""
import json, sys, os
from datetime import datetime
from collections import defaultdict
import numpy as np
import pandas as pd

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
from ashare_review.data.tdx_reader import TdxReader

# ============================================================
# 1. Load backtest trades from xlsx
# ============================================================
import openpyxl
wb = openpyxl.load_workbook('ashare_review/analysis/v3_backtest_result.xlsx')
ws1 = wb['交易明细']
trades = []
for row in ws1.iter_rows(min_row=4, max_row=ws1.max_row, max_col=16, values_only=True):
    if row[0]:
        trades.append({
            'code': str(row[0]).zfill(6), 'buy_date': str(row[2]),
            'sell_date': str(row[4]), 'net_ret': float(row[7] or 0),
            'result': str(row[8]), 'days_held': int(row[9] or 0),
            'exit_reason': str(row[10]), 'score': int(row[12] or 0),
        })
print(f"1. Loaded {len(trades)} trades from backtest")

# ============================================================
# 2. Load ZT cache
# ============================================================
with open('ashare_review/data/zt_scan_cache.json', 'r') as f:
    zt_cache = json.load(f)
print(f"2. ZT cache: {len(zt_cache)} dates")

# ============================================================
# 3. Load SH index
# ============================================================
tdx = TdxReader()
sh = tdx.read_daily('999999', 'sh')
if sh is not None and len(sh) > 60:
    sh['date_str'] = sh['trade_date'].apply(lambda x: x.strftime('%Y%m%d') if hasattr(x, 'strftime') else str(x)[:8])
    sh['ma60'] = sh['close'].rolling(60).mean()
    sh['ma20'] = sh['close'].rolling(20).mean()
    # Volume in yi (亿)
    if 'amount' in sh.columns:
        sh['amt_yi'] = sh['amount'] / 1e8
    else:
        sh['amt_yi'] = sh['volume'] * sh['close'] / 1e8 / 10  # rough estimate
print(f"3. SH index: {len(sh)} rows")

# ============================================================
# 4. Build daily market data for backtest period
# ============================================================
market = {}
for ds_str in sorted(zt_cache.keys()):
    if not ('20250725' <= ds_str <= '20260725'):
        continue

    zt_list = zt_cache[ds_str]
    zt_count = len(zt_list)

    # Calc rank board height (连板高度) from ZT
    # For each stock in ZT, check if it was also ZT on previous days
    max_consecutive = 2
    if zt_list:
        codes_today = set(zt_list)  # zt_list is list of code strings
        # Check up to 10 previous days
        prev_dates = sorted(d for d in zt_cache if d < ds_str)
        for code in codes_today:
            consecutive = 1
            for pd_str in reversed(prev_dates[-10:]):
                pd_codes = set(zt_cache.get(pd_str, []))
                if code in pd_codes:
                    consecutive += 1
                else:
                    break
            max_consecutive = max(max_consecutive, consecutive)

    # SH index data
    sh_row = sh[sh['date_str'] == ds_str]
    if len(sh_row) > 0:
        sh_close = float(sh_row['close'].iloc[0])
        sh_chg = float(sh_row['close'].pct_change().iloc[0] * 100) if len(sh_row) > 1 else 0
        sh_amt = float(sh_row['amt_yi'].iloc[0])
        ma60 = float(sh_row['ma60'].iloc[0])
        above_ma60 = sh_close > ma60 if not pd.isna(ma60) else True
        # Index trend: > MA20 and MA20 sloping up
        ma20 = float(sh_row['ma20'].iloc[0]) if not pd.isna(sh_row['ma20'].iloc[0]) else sh_close
        above_ma20 = sh_close > ma20
    else:
        continue

    market[ds_str] = {
        'zt_count': zt_count,
        'max_board': max_consecutive,
        'sh_close': round(sh_close, 1),
        'sh_chg': round(sh_chg, 2),
        'sh_amt': round(sh_amt, 0),
        'above_ma60': above_ma60,
        'above_ma20': above_ma20,
    }

print(f"4. Market data: {len(market)} days computed")

# ============================================================
# 5. Merge trades with market data
# ============================================================
# Group by sell_date (the day we know the outcome)
trade_by_date = defaultdict(list)
for t in trades:
    ds = t['sell_date'].replace('-', '')
    trade_by_date[ds].append(t)

# Also group by buy_date
trade_by_buy_date = defaultdict(list)
for t in trades:
    ds = t['buy_date'].replace('-', '')
    trade_by_buy_date[ds].append(t)

# ============================================================
# 6. ANALYSIS
# ============================================================
print("\n" + "=" * 70)
print("  V3 回测 — 市场情绪相关性分析")
print("=" * 70)
print(f"  回测区间: 2025-07-26 ~ 2026-07-25")
print(f"  总交易: {len(trades)} 笔")
print(f"  有效市场数据: {len(market)} 天")

def analyze_bucket(label, buckets, get_metric, use_buy_date=False):
    """Generic bucket analysis"""
    print(f"\n--- {label} ---")
    for lo, hi in buckets:
        bucket = []
        date_map = trade_by_buy_date if use_buy_date else trade_by_date
        for ds, tlist in date_map.items():
            md = market.get(ds, {})
            val = get_metric(md)
            if (lo <= val < hi) or (hi == 99999 and val >= lo):
                bucket.extend(tlist)
        if bucket:
            n = len(bucket)
            wins = sum(1 for t in bucket if t['result'] == 'Win')
            avg = np.mean([t['net_ret'] for t in bucket])
            pnl = sum(t['net_ret'] for t in bucket)
            range_label = f"{lo}-{hi}" if hi < 99999 else f"{lo}+"
            print(f"  {range_label:>10}: {n:>4}笔  WR={wins/n*100:5.1f}%  avg={avg:+.2f}%  total={pnl:+.1f}%")

# --- A. ZT count ---
analyze_bucket("A. 涨停家数 vs 收益",
    [(0, 40), (40, 60), (60, 80), (80, 100), (100, 99999)],
    lambda md: md.get('zt_count', 0))

# --- B. Max board height ---
analyze_bucket("B. 连板高度 vs 收益",
    [(2, 3), (3, 4), (4, 6), (6, 8), (8, 99)],
    lambda md: md.get('max_board', 2))

# --- C. SH change ---
analyze_bucket("C. 上证涨跌幅 vs 收益",
    [(-10, -1.5), (-1.5, -0.5), (-0.5, 0.5), (0.5, 1.5), (1.5, 10)],
    lambda md: md.get('sh_chg', 0))

# --- D. SH volume ---
analyze_bucket("D. 上证成交额(亿) vs 收益",
    [(0, 2500), (2500, 3500), (3500, 4500), (4500, 6000), (6000, 99999)],
    lambda md: md.get('sh_amt', 0))

# --- E. MA60 ---
print("\n--- E. 上证 vs MA60 ---")
above = []; below = []
for ds, tlist in trade_by_date.items():
    md = market.get(ds, {})
    if md.get('above_ma60', True):
        above.extend(tlist)
    else:
        below.extend(tlist)
for label, tlist in [('上证>MA60', above), ('上证<MA60', below)]:
    if tlist:
        n = len(tlist); w = sum(1 for t in tlist if t['result']=='Win')
        print(f"  {label}: {n}笔 WR={w/n*100:.1f}% avg={np.mean([t['net_ret'] for t in tlist]):+.2f}%")

# --- F. ZT count TREND ---
print("\n--- F. 涨停家数变化方向 vs 收益 ---")
rising = []; falling = []; flat = []
sorted_dates = sorted(market.keys())
for i, ds in enumerate(sorted_dates):
    if ds not in trade_by_date:
        continue
    zt = market[ds].get('zt_count', 0)
    prev_zt = market[sorted_dates[i-1]].get('zt_count', zt) if i > 0 else zt
    tlist = trade_by_date[ds]
    if zt > prev_zt * 1.15:
        rising.extend(tlist)
    elif zt < prev_zt * 0.85:
        falling.extend(tlist)
    else:
        flat.extend(tlist)

for label, tlist in [('ZT急升(>+15%)', rising), ('ZT平稳(±15%)', flat), ('ZT急降(>-15%)', falling)]:
    if tlist:
        n = len(tlist); w = sum(1 for t in tlist if t['result']=='Win')
        print(f"  {label}: {n}笔 WR={w/n*100:.1f}% avg={np.mean([t['net_ret'] for t in tlist]):+.2f}%")

# --- G. Multi-factor combos ---
print("\n--- G. 多因子组合 ---")
combos = [
    ("ZT>=80 & SH>MA60 & SH上涨",
     lambda md: md.get('zt_count',0)>=80 and md.get('above_ma60',True) and md.get('sh_chg',0)>0),
    ("ZT>=80 & SH>MA60",
     lambda md: md.get('zt_count',0)>=80 and md.get('above_ma60',True)),
    ("ZT>=60 & SH>MA60",
     lambda md: md.get('zt_count',0)>=60 and md.get('above_ma60',True)),
    ("ZT<40 & SH<MA60",
     lambda md: md.get('zt_count',0)<40 and not md.get('above_ma60',True)),
    ("ZT<40",
     lambda md: md.get('zt_count',0)<40),
    ("SH成交>5000亿 & ZT>=60",
     lambda md: md.get('sh_amt',0)>=5000 and md.get('zt_count',0)>=60),
    ("连板>=5 & ZT>=60",
     lambda md: md.get('max_board',2)>=5 and md.get('zt_count',0)>=60),
]
for label, cond in combos:
    bucket = []
    for ds, tlist in trade_by_date.items():
        md = market.get(ds, {})
        if cond(md):
            bucket.extend(tlist)
    if bucket:
        n = len(bucket); w = sum(1 for t in bucket if t['result']=='Win')
        print(f"  {label}: {n:>4}笔 WR={w/n*100:5.1f}% avg={np.mean([t['net_ret'] for t in bucket]):+.2f}%")

# --- H. Daily win rate vs ZT scatter ---
print("\n--- H. 每日胜率 vs 涨停家数 (按天聚合) ---")
zt_bins = defaultdict(list)
for ds, tlist in trade_by_date.items():
    md = market.get(ds, {})
    zt = md.get('zt_count', 0)
    zt_bin = (zt // 20) * 20
    day_wr = sum(1 for t in tlist if t['result']=='Win') / max(len(tlist), 1) * 100
    day_avg = np.mean([t['net_ret'] for t in tlist])
    zt_bins[zt_bin].append((day_wr, day_avg, len(tlist)))

for zt_bin in sorted(zt_bins.keys()):
    data = zt_bins[zt_bin]
    avg_wr = np.mean([d[0] for d in data])
    avg_ret = np.mean([d[1] for d in data])
    total_t = sum(d[2] for d in data)
    print(f"  ZT~{zt_bin}: {len(data)}天 {total_t}笔, 日均胜率={avg_wr:.1f}%, 日均收益={avg_ret:+.2f}%")

# --- I. By buy-day market condition ---
print("\n--- I. 买入日市场环境 vs 收益 (按买入日分组) ---")
analyze_bucket("  买入日涨停家数 vs 收益",
    [(0, 40), (40, 60), (60, 80), (80, 100), (100, 99999)],
    lambda md: md.get('zt_count', 0), use_buy_date=True)

analyze_bucket("  买入日上证涨跌 vs 收益",
    [(-10, -1), (-1, 0), (0, 0.5), (0.5, 1), (1, 10)],
    lambda md: md.get('sh_chg', 0), use_buy_date=True)

# --- J. Hold days vs performance ---
print("\n--- J. 持有天数 vs 收益 ---")
for days in [1, 2, 3, 4, 5, 6, 7, 10]:
    bucket = [t for t in trades if t['days_held'] == days]
    if bucket:
        n = len(bucket); w = sum(1 for t in bucket if t['result']=='Win')
        print(f"  {days}天: {n:>4}笔 WR={w/n*100:5.1f}% avg={np.mean([t['net_ret'] for t in bucket]):+.2f}%")

print("\n" + "=" * 70)
print("  ANALYSIS COMPLETE")
print("=" * 70)
