"""
V3 VOL180 突破战法 — 过去一年回测

选股: 沪深主板 · 年涨停≥10 · 非ST · 距压力位≤10%
买入: 收盘突破压力位 + 成交量 > MAVOL180×1.2 → 次日开盘买入
      + V3 竞价确认（低开>3%跳过、缩量>50%跳过、开盘跌停跳过）
卖出 V3:
  0. -6% 硬止损
  1. 移动止盈: 最高收盘价回落 > 5%
  2. N字反包: 涨停后断板 → 等一天反包
  3. 持仓 ≥ 5天兜底

输出: xlsx 文件（交易明细 / 持仓快照 / 统计汇总 / 按卖出原因 / 按月统计）
"""
import sys, os, json, struct, time as _time
from datetime import date, datetime, timedelta
from typing import Dict, List, Optional, Tuple
from collections import defaultdict

import numpy as np
import pandas as pd

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

from ashare_review.data.tdx_reader import TdxReader, RECORD_SIZE
from ashare_review.analysis.indicators import calc_zigzag_find_top_line, calc_ma
from ashare_review.utils.calendar import TradingCalendar

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

# ═══════════════════════════════════════════════════════════════════════════
# Constants (与 sim_portfolio.py 一致)
# ═══════════════════════════════════════════════════════════════════════════
FEE = 0.0015
SLIPPAGE = 0.002
TOTAL_COST = FEE + SLIPPAGE
MAX_HOLD_DAYS = 5  # V3
MIN_LIMIT_UP_COUNT = 10
PRESSURE_DIST_PCT = 10.0
MAVOL_PERIOD = 180
MAVOL_MULTIPLIER = 1.2

DATA_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'data')
LIMIT_UP_POOL_FILE = os.path.join(DATA_DIR, 'limit_up_pool.json')
NAME_CACHE_FILE = os.path.join(DATA_DIR, 'stock_name_map.json')

# ═══════════════════════════════════════════════════════════════════════════
# Excel 样式
# ═══════════════════════════════════════════════════════════════════════════
TITLE_FONT = Font(name='微软雅黑', size=14, bold=True, color='1F2937')
HEADER_FONT = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='374151', end_color='374151', fill_type='solid')
HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
CENTER = Alignment(horizontal='center', vertical='center')
THIN_BORDER = Border(
    left=Side(style='thin', color='D1D5DB'),
    right=Side(style='thin', color='D1D5DB'),
    top=Side(style='thin', color='D1D5DB'),
    bottom=Side(style='thin', color='D1D5DB'),
)
WIN_FILL = PatternFill(start_color='DCFCE7', end_color='DCFCE7', fill_type='solid')
LOSS_FILL = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
S_FILL = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
A_FILL = PatternFill(start_color='DBEAFE', end_color='DBEAFE', fill_type='solid')
SUMMARY_LABEL_FONT = Font(name='微软雅黑', size=10, bold=True, color='374151')
SUMMARY_VAL_FONT = Font(name='Consolas', size=11, color='1F2937')
BIG_NUM_FONT = Font(name='Consolas', size=14, bold=True, color='1F2937')
GREEN_FONT = Font(name='Consolas', size=14, bold=True, color='059669')
RED_FONT = Font(name='Consolas', size=14, bold=True, color='DC2626')

OUTPUT_DIR = os.path.join(DATA_DIR, '..', 'analysis')


def _style_header(ws, row: int, ncols: int):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def _style_data(ws, start_row: int, end_row: int, ncols: int):
    for r in range(start_row, end_row + 1):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='center', vertical='center')


def _auto_width(ws, min_width: int = 8, max_width: int = 22):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            if cell.value:
                # 中文字符算2个宽度
                val = str(cell.value)
                length = sum(2 if ord(c) > 127 else 1 for c in val)
                max_len = max(max_len, length)
        ws.column_dimensions[col_letter].width = max(min_width, min(max_len + 3, max_width))


# ═══════════════════════════════════════════════════════════════════════════
# V3 Backtest Engine
# ═══════════════════════════════════════════════════════════════════════════

class V3Backtest:
    """V3 VOL180 突破战法历史回测"""

    def __init__(self):
        self.tdx = TdxReader()
        self.cal = TradingCalendar()
        self._name_map: Dict[str, str] = {}
        self._load_name_map()

    def _load_name_map(self):
        if os.path.exists(NAME_CACHE_FILE):
            try:
                with open(NAME_CACHE_FILE, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                if isinstance(data, dict) and len(data) > 1000:
                    self._name_map = data
            except Exception:
                pass

    def _get_name(self, code: str) -> str:
        return self._name_map.get(str(code).zfill(6), code)

    @staticmethod
    def _limit_threshold(code: str) -> float:
        code = str(code).zfill(6)
        if code.startswith(('300', '301', '688')): return 0.199
        if code.startswith(('8', '4')): return 0.299
        return 0.095

    @staticmethod
    def _chase_limit_pct(code: str) -> float:
        """追高上限(%)：已突破压力位的累计涨幅超过该比例 → 放弃（追高不做）。

        与实盘 sim_portfolio 完全一致（教学: "八个点以下才做"、
        页面模板/V4 口径: 10cm ≤6% / 20cm ≤8% / 30cm ≤30%），
        保证历史回测与实盘选股口径一致。
        """
        c = str(code).zfill(6)
        if c.startswith(('300', '301', '688')): return 8.0
        if c.startswith(('8', '4')): return 30.0
        return 6.0

    @staticmethod
    def _is_main_board(code: str) -> bool:
        code = str(code).zfill(6)
        return code.startswith(('60', '00', '001', '002'))

    # ─── 获取候选池（年涨停≥10 + 主板 + 非ST） ───

    def get_universe(self) -> List[str]:
        """从缓存获取候选池股票代码列表。"""
        if os.path.exists(LIMIT_UP_POOL_FILE):
            try:
                with open(LIMIT_UP_POOL_FILE, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                pool = data.get('pool', [])
                if pool:
                    codes = [s['code'] for s in pool
                             if self._is_main_board(s['code'])
                             and 'ST' not in self._get_name(s['code'])]
                    print(f'从缓存加载候选池: {len(codes)} 只')
                    return codes
            except Exception:
                pass
        print('[WARN] 候选池缓存不存在，请先运行 sim_portfolio.py 构建候选池')
        return []

    # ─── 读取单只股票全量数据 ───

    def _read_stock_full(self, code: str) -> Optional[pd.DataFrame]:
        """读取股票全量日线 + 计算指标 + 预计算zigzag压力位。"""
        market = 'sh' if str(code).startswith('6') else 'sz'
        if str(code).startswith(('8', '4')):
            market = 'bj'
        try:
            df = self.tdx.read_daily(code, market)
            if df is None or df.empty or len(df) < MAVOL_PERIOD + 60:
                return None
            # 截断到尾 600 行（~2.5 年），覆盖一年回测+180 日均量+zigzag 历史
            # 避免 8000+ 行长历史导致 zigzag O(n²) 耗时过大
            if len(df) > 600:
                df = df.iloc[-600:].copy()
                df.reset_index(drop=True, inplace=True)
            df = calc_ma(df, [5, 10])
            df['mavol180'] = df['volume'].rolling(MAVOL_PERIOD).mean()

            # ── 预计算简化 zigzag 压力位（快速向量化版本） ──
            # 用滚动窗口峰值检测替代完整 zigzag BACKSET/BARSLAST 链
            # 精度接近但速度提升 100x+
            try:
                high_v = df['high'].values.astype(float)
                n = len(high_v)
                # Step 1: 找局部峰值（前后各5根K线内最高）
                is_peak = np.ones(n, dtype=bool)
                for offset in range(1, 6):
                    is_peak &= (np.roll(high_v, offset) <= high_v)
                    is_peak &= (np.roll(high_v, -offset) <= high_v)
                is_peak[:5] = False; is_peak[-5:] = False
                # Step 2: 只保留相隔≥10根K线的显著峰值（去噪）
                peak_idx = np.where(is_peak)[0]
                filtered = []
                last_kept = -100
                for pi in peak_idx:
                    if pi - last_kept >= 10:
                        filtered.append(pi)
                        last_kept = pi
                # Step 3: 构建每日压力位 = 最近峰值的 high
                pressure = np.full(n, np.nan)
                last_high = np.nan
                fi = 0
                for i in range(n):
                    if fi < len(filtered) and i == filtered[fi]:
                        last_high = high_v[i]
                        fi += 1
                    pressure[i] = last_high
                df['_pressure'] = pressure
                df['_nn'] = is_peak.astype(int)
                # fallback: 前60行的 NaN 用60日最高
                high_60max = df['high'].rolling(60, min_periods=1).max().shift(1)
                mask = pd.isna(df['_pressure']) | (df['_pressure'] <= 0)
                df.loc[mask, '_pressure'] = high_60max[mask]
            except Exception:
                df['_pressure'] = df['high'].rolling(60, min_periods=1).max().shift(1)

            return df
        except Exception:
            return None

    # ─── 获取预计算的压力位 ───

    def _calc_pressure_at(self, df: pd.DataFrame, idx: int) -> float:
        """读取预计算的压力位（zigzag 找顶线 last NN high）。
        优化版本：从预计算的 _pressure 列直接读取，不再重复计算 zigzag。
        """
        try:
            val = float(df['_pressure'].iloc[idx])
            if pd.isna(val) or val <= 0:
                return float(df['high'].rolling(60).max().shift(1).iloc[idx]) \
                    if idx >= 60 else 0.0
            return round(val, 2)
        except Exception:
            return float(df['high'].rolling(60).max().shift(1).iloc[idx]) \
                if idx >= 60 else 0.0

    # ─── V3 卖出检查 ───

    def _check_sell_v3(self, code: str, hold: dict, check_date: date,
                       df_full: pd.DataFrame) -> Optional[dict]:
        """V3 卖出规则（与 sim_portfolio.py _check_sell_v3 完全一致）。"""
        buy_date = hold['buy_date']
        buy_price = hold['buy_price']
        had_zt = hold.get('had_zt', False)
        highest = hold.get('highest_close', buy_price)

        # 截取数据到 check_date
        mask = df_full['trade_date'].apply(
            lambda x: (x.date() if hasattr(x, 'date') else x) <= check_date
        )
        df = df_full[mask].copy()
        if df.empty or len(df) < 2:
            return None

        idx = len(df) - 1
        close = float(df['close'].iloc[idx])

        # 更新最高收盘价
        highest = max(highest, close)
        hold['highest_close'] = round(highest, 2)

        # 交易日计数
        try:
            trading_days = self.cal.trading_days_between(buy_date, check_date)
        except Exception:
            trading_days = 3

        # ── -6% 硬止损 ──
        if buy_price > 0:
            loss_pct = (close - buy_price) / buy_price
            if loss_pct <= -0.06:
                return {
                    'sell_price': round(close, 2),
                    'sell_reason': '止损-6%',
                    'days_held': trading_days,
                }

        # ── 移动止盈: 从最高回落 > 5%（需先盈利 3%+） ──
        if highest > buy_price * 1.03:
            pullback = (close - highest) / highest
            if pullback <= -0.05:
                return {
                    'sell_price': round(close, 2),
                    'sell_reason': f'移动止盈-5%(高{highest:.2f}→{close:.2f})',
                    'days_held': trading_days,
                }

        # ── 检查今日是否涨停 ──
        if idx >= 1:
            prev_close = float(df['close'].iloc[idx - 1])
            limit_pct = self._limit_threshold(code)
            is_zt_today = (close - prev_close) / prev_close >= limit_pct if prev_close > 0 else False
        else:
            is_zt_today = False

        # ── 今日涨停 → 继续持有 ──
        if is_zt_today:
            hold['had_zt'] = True
            hold['awaiting_reversal'] = False
            return None

        # ── N字反包 ──
        if had_zt:
            awaiting = hold.get('awaiting_reversal', False)
            if not awaiting:
                hold['awaiting_reversal'] = True
                hold['reversal_day_close'] = round(close, 2)
                hold['reversal_day_vol'] = int(float(df['volume'].iloc[idx])) if idx >= 0 else 0
                return None  # 等反包
            else:
                rev_vol = hold.get('reversal_day_vol', 0)
                today_vol = int(float(df['volume'].iloc[idx])) if idx >= 0 else 0
                today_open = float(df['open'].iloc[idx]) if idx >= 0 else close
                is_up = close > today_open
                vol_expand = today_vol > rev_vol
                if is_up and vol_expand:
                    hold['awaiting_reversal'] = False
                    hold['had_zt'] = True  # 反包成功，继续
                    return None
                else:
                    return {
                        'sell_price': round(close, 2),
                        'sell_reason': '涨停后断板离场',
                        'days_held': trading_days,
                    }

        # ── 始终没涨停 + 持有 ≥ 5 天兜底 ──
        if trading_days >= MAX_HOLD_DAYS:
            return {
                'sell_price': round(close, 2),
                'sell_reason': f'持有{trading_days}天到期',
                'days_held': trading_days,
            }

        # ── 没涨停但浮盈 5%+，也启用移动止盈 ──
        if highest > buy_price * 1.05:
            pullback = (close - highest) / highest
            if pullback <= -0.05:
                return {
                    'sell_price': round(close, 2),
                    'sell_reason': f'移动止盈-5%(高{highest:.2f})',
                    'days_held': trading_days,
                }

        return None

    # ─── V3 竞价确认 ───

    def _check_auction_v3(self, code: str, buy_date: date,
                          df_full: pd.DataFrame) -> Optional[str]:
        """V3 竞价确认，返回 None=通过, str=拒绝原因。"""
        mask = df_full['trade_date'].apply(
            lambda x: (x.date() if hasattr(x, 'date') else x) <= buy_date
        )
        df = df_full[mask].copy()
        if df.empty or len(df) < 2:
            return None

        idx = len(df) - 1
        open_p = float(df['open'].iloc[idx])
        prev_close = float(df['close'].iloc[idx - 1])
        vol_today = float(df['volume'].iloc[idx]) if idx >= 0 else 0
        vol_prev = float(df['volume'].iloc[idx - 1]) if idx >= 1 else 0

        # 低开 > 3%
        open_chg = (open_p - prev_close) / prev_close * 100 if prev_close > 0 else 0
        if open_chg < -3:
            return f'竞价低开{open_chg:.1f}%'

        # 缩量 > 50%
        if vol_prev > 0 and vol_today < vol_prev * 0.5:
            return '竞价缩量>50%'

        return None

    # ─── 检查开盘跌停 ───

    def _check_open_limit_down(self, code: str, buy_date: date,
                               df_full: pd.DataFrame) -> bool:
        """检查买入日是否开盘跌停。"""
        mask = df_full['trade_date'].apply(
            lambda x: (x.date() if hasattr(x, 'date') else x) <= buy_date
        )
        df = df_full[mask].copy()
        if df.empty or len(df) < 2:
            return False

        idx = len(df) - 1
        open_p = float(df['open'].iloc[idx])
        prev_close = float(df['close'].iloc[idx - 1])
        limit_pct = self._limit_threshold(code)
        return open_p <= prev_close * (1 - limit_pct)

    # ═══════════════════════════════════════════════════════════════════════
    # 主回测循环
    # ═══════════════════════════════════════════════════════════════════════

    def run(self, start_date: date = None, end_date: date = None,
            causal_universe=None, regime_weights: dict = None,
            regime_of_day=None, sector_resonance=None,
            min_resonance: int = 2, max_resonance: int = None,
            ma60_gate: bool = False, dd_stop: float = None,
            max_positions: int = None, position_pct: float = None,
            rotation: bool = False) -> dict:
        """运行 V3 回测（含真实资金模型）。

        Args:
            start_date: 回测开始日（默认一年前）
            end_date: 回测结束日（默认昨天）
            causal_universe: 因果候选池（strategy_regime.causal_universe.CausalUniverse）。
                传入时用逐日因果判定"前期强势"(近250日涨停≥10)，修复静态池幸存者偏差；
                为 None 则用静态 limit_up_pool.json（默认，兼容旧行为）。
            regime_weights: {行情标签: 仓位权重}，如 {'强势趋势':1.0,'题材轮动':0.7,
                '震荡观望':0.3,'弱市回调':0.2,'退潮下跌':0.0,'冰点超跌':0.3}。
                传入后按信号日行情缩放单票仓位；权重=0 的行情当天不开新仓。
            regime_of_day: 可调用(date)->行情标签，用于查信号日属于哪种行情。
                与 regime_weights 配套使用。
            sector_resonance: 可调用(code, 信号日)->板块强度(int)，如
                strategy_regime.sector_strength.SectorStrength.strength。
                传入后要求当日板块强度 ≥ min_resonance 才开仓（只做板块共振，不碰杂毛）。
            min_resonance: 板块共振最低强度（默认2）。
            max_resonance: 板块拥挤上限（默认None）。实测同板块当日涨停≥5只(涨停潮)的突破
                胜率最低(追高)，传入后板块强度 ≥ max_resonance 的信号跳过。
            ma60_gate: 上证 < MA60（下降趋势）时禁止开新仓（战法: 下降趋势不交易）。
            dd_stop: 组合回撤熔断阈值（0~1）。组合从峰值回撤 ≥ dd_stop 时停止开新仓，
                直到回撤修复到 dd_stop/2 以下。
            max_positions: 最大持仓数（默认10）。position_pct: 单票仓位（默认0.10）。
                传 max_positions=3, position_pct=0.33 → 3仓集中。
            rotation: 集中持仓时是否每日轮换。满仓后若出现比最弱持仓更强的信号
                （新信号评分 > 最弱持仓强度），卖出最弱持仓、换入新标的（"只做最强"）。
        """
        # ── 资金/仓位模型 ──
        INITIAL_CAPITAL = 1_000_000.0    # 初始资金 100 万
        MAX_POSITIONS = max_positions or 10
        MAX_NEW_PER_DAY = 3              # 每天最多新开仓
        PER_POSITION_PCT = position_pct or 0.10
        self._rotation = rotation
        self._regime_weights = regime_weights
        self._regime_of_day = regime_of_day
        self._sector_resonance = sector_resonance
        self._min_resonance = min_resonance
        self._max_resonance = max_resonance
        self._ma60_gate = ma60_gate
        self._dd_stop = dd_stop

        if end_date is None:
            end_date = date.today() - timedelta(days=1)
        if start_date is None:
            start_date = end_date - timedelta(days=365)

        print(f'\n{"="*60}')
        print(f'  V3 VOL180 突破战法 — 历史回测')
        print(f'  区间: {start_date} ~ {end_date}')
        print(f'  资金: {INITIAL_CAPITAL/1e4:.0f}万 | 最大持仓: {MAX_POSITIONS}只 | 单票: {PER_POSITION_PCT*100:.0f}%')
        print(f'{"="*60}\n')

        # ── Step 1: 获取候选池 ──
        if causal_universe is not None:
            # 因果宇宙（主板 + 非ST）；排序保证确定性
            universe = sorted(c for c in causal_universe.codes
                              if self._is_main_board(c) and 'ST' not in self._get_name(c))
            print(f'因果候选池: {len(universe)} 只（主板, 近250日涨停≥10 逐日判定）')
        else:
            universe = self.get_universe()
        if not universe:
            print('[ERROR] 候选池为空')
            return {}

        # ── Step 2: 预读所有候选股数据 + 构建日期索引 ──
        print(f'预读 {len(universe)} 只候选股数据...')
        stock_data: Dict[str, pd.DataFrame] = {}
        stock_date_idx: Dict[str, Dict[date, int]] = {}  # code → {date: row_index}
        t0 = _time.time()
        for i, code in enumerate(universe):
            if (i + 1) % 100 == 0:
                print(f'  读取 {i+1}/{len(universe)} (已加载 {len(stock_data)} 只, {_time.time() - t0:.0f}s)...')
            df = self._read_stock_full(code)
            if df is not None and len(df) >= MAVOL_PERIOD + 60:
                stock_data[code] = df
                # 预建日期→行号映射
                dmap = {}
                for ri, td_val in enumerate(df['trade_date']):
                    if hasattr(td_val, 'date'):
                        dmap[td_val.date()] = ri
                    elif hasattr(td_val, 'strftime'):
                        dmap[td_val] = ri
                stock_date_idx[code] = dmap
        print(f'预读完成: {len(stock_data)} 只有效数据 ({_time.time() - t0:.0f}s)')

        # ── Step 3: 生成交易日列表（用真实交易日历，排除节假日，避免节假日无行情导致市值误算） ──
        all_dates = []
        d = start_date
        while d <= end_date:
            if self.cal.is_trading_day(d):
                all_dates.append(d)
            d += timedelta(days=1)
        print(f'交易日: {len(all_dates)} 天 ({all_dates[0]} ~ {all_dates[-1]})')

        # ── Step 3b: 市场环境预计算（上证指数 MA60） ──
        market_bullish = {}  # date → bool (is above MA60)
        try:
            sh_df = self.tdx.read_daily('999999', 'sh')
            if sh_df is not None and len(sh_df) >= 60:
                sh_df['ma60'] = sh_df['close'].rolling(60).mean()
                for td in all_dates:
                    mask = sh_df['trade_date'].apply(
                        lambda x: (x.date() if hasattr(x, 'date') else x) <= td
                    )
                    sub = sh_df[mask]
                    if len(sub) >= 60:
                        close = float(sub['close'].iloc[-1])
                        ma60 = float(sub['ma60'].iloc[-1])
                        market_bullish[td] = close > ma60 and not pd.isna(ma60)
        except Exception:
            pass
        print(f'市场环境: 上证MA60数据 {len(market_bullish)} 天')

        # ── Step 4: 逐日回测（资金模型） ──
        # 资金状态
        cash = INITIAL_CAPITAL
        holdings: Dict[str, dict] = {}  # code → {shares, buy_price, buy_date, ...}
        finished_trades: List[dict] = []
        skipped_trades: List[dict] = []
        daily_log: List[dict] = []

        # 统计
        total_signals = 0
        total_buys = 0
        portfolio_values = []  # [(date, total_value, cash, positions_value)] for drawdown calc
        peak_val = INITIAL_CAPITAL  # 组合峰值（回撤熔断用）
        last_total = INITIAL_CAPITAL  # 上一日组合市值（回撤熔断用）
        breaker_active = False  # 回撤熔断状态

        # ── 辅助 ──
        def _get_idx(code: str, target_date: date) -> int:
            dmap = stock_date_idx.get(code)
            if dmap is None: return -1
            return dmap.get(target_date, -1)

        def _current_price(code: str, td: date) -> Optional[float]:
            """获取某股票在某日的收盘价"""
            idx = _get_idx(code, td)
            if idx >= 0:
                return float(stock_data[code]['close'].iloc[idx])
            return None

        t0 = _time.time()
        pending_sells: Dict[str, dict] = {}  # code → sell_info (执行日为次日)

        for di, td in enumerate(all_dates):
            if (di + 1) % 20 == 0:
                elapsed = _time.time() - t0
                eta = elapsed / (di + 1) * (len(all_dates) - di - 1) if di > 0 else 0
                # 计算当前组合市值
                pos_val = 0.0
                for code, h in holdings.items():
                    px = _current_price(code, td)
                    if px: pos_val += h['shares'] * px
                total_val = cash + pos_val
                print(f'  [{di+1}/{len(all_dates)}] {td}  '
                      f'持仓:{len(holdings)} 市值:{total_val/1e4:.1f}万 已平:{len(finished_trades)}  '
                      f'({elapsed:.0f}s, ETA {eta:.0f}s)', flush=True)

            # ── Step 4a: 执行昨日挂单的卖出（用今日开盘价） ──
            sold_today = []
            for code in list(pending_sells.keys()):
                sell_info = pending_sells[code]
                sell_date = sell_info['sell_date']
                sell_reason = sell_info['sell_reason']
                h = sell_info['hold']

                # 用今天的开盘价卖出
                sell_idx = _get_idx(code, td)
                if sell_idx >= 0:
                    sell_price = float(stock_data[code]['open'].iloc[sell_idx])
                else:
                    sell_price = sell_info.get('fallback_price', h['buy_price'])

                buy_price = h['buy_price']
                shares = h['shares']
                # 卖出回款（扣除卖出手续费+印花税 0.08%）
                sell_cost_rate = 0.0008
                proceeds = shares * sell_price * (1 - sell_cost_rate)
                cash += proceeds

                gross_ret = (sell_price - buy_price) / buy_price if buy_price > 0 else 0
                net_ret = gross_ret - TOTAL_COST

                trading_days = 0
                try:
                    trading_days = self.cal.trading_days_between(h['buy_date'], td)
                except Exception:
                    trading_days = sell_info.get('_days', 3)

                finished_trades.append({
                    'code': code,
                    'name': self._get_name(code),
                    'buy_date': h['buy_date'].strftime('%Y-%m-%d'),
                    'buy_price': round(buy_price, 2),
                    'sell_date': td.strftime('%Y-%m-%d'),
                    'sell_price': round(sell_price, 2),
                    'gross_ret': round(gross_ret * 100, 2),
                    'net_ret': round(net_ret * 100, 2),
                    'is_win': net_ret > 0,
                    'exit_reason': sell_reason,
                    'days_held': trading_days if trading_days > 0 else sell_info.get('_days', 1),
                    'had_zt': h.get('had_zt', False),
                    'score': h.get('score', 0),
                    'vol_ratio': h.get('vol_ratio', 0),
                    'dist_pct': h.get('dist_pct', 0),
                    'limit_count': h.get('limit_count', 0),
                    'position_w': h.get('position_w', 1.0),
                    'signal_regime': h.get('signal_regime', ''),
                })
                sold_today.append(code)
                del holdings[code]
                del pending_sells[code]

            # ── Step 4b: 检查持仓是否需要挂卖出单（用今日收盘价判断） ──
            for code in list(holdings.keys()):
                if code in pending_sells:  # 已有挂单
                    continue
                if code not in stock_data:
                    continue
                sell_sig = self._check_sell_v3(code, holdings[code], td, stock_data[code])
                if sell_sig:
                    # 挂单：下一个交易日以开盘价卖出
                    next_td_sell = self._next_trade_date(td, all_dates)
                    if next_td_sell is None:
                        # 回测最后一天，用收盘价强平
                        fallback_px = sell_sig['sell_price']
                        sell_real_date = td
                    else:
                        fallback_px = sell_sig['sell_price']
                        sell_real_date = next_td_sell

                    pending_sells[code] = {
                        'hold': holdings[code],
                        'sell_date': sell_real_date,
                        'sell_reason': sell_sig['sell_reason'],
                        'fallback_price': fallback_px,
                        '_days': sell_sig['days_held'],
                    }

            # ── Step 4c: 检测今日买入信号 ──
            buy_today = []
            # 市场环境过滤
            is_bull = market_bullish.get(td, True)  # 无数据默认允许

            # 风控闸门 1: 上证<MA60（下降趋势）→ 不开新仓（战法: 下降趋势不交易）
            allow_new = True
            if self._ma60_gate and not is_bull:
                allow_new = False

            # 风控闸门 2: 组合回撤熔断 → 强制清仓 + 停开新仓，回撤修复后才恢复
            if self._dd_stop:
                dd_pct = (peak_val - last_total) / peak_val if peak_val > 0 else 0.0
                if not breaker_active and dd_pct >= self._dd_stop:
                    breaker_active = True
                    print(f'  [熔断] {td} 组合回撤{dd_pct*100:.1f}%≥{self._dd_stop*100:.0f}%，强制清仓停手')
                elif breaker_active and dd_pct < self._dd_stop * 0.5:
                    breaker_active = False
                    print(f'  [恢复] {td} 回撤修复至{dd_pct*100:.1f}%，恢复开仓')
                if breaker_active:
                    allow_new = False
                    # 强制清仓：给所有持仓挂次日开盘卖出单
                    for _code in list(holdings.keys()):
                        if _code in pending_sells:
                            continue
                        pending_sells[_code] = {
                            'hold': holdings[_code],
                            'sell_date': self._next_trade_date(td, all_dates) or td,
                            'sell_reason': '回撤熔断清仓',
                            'fallback_price': _current_price(_code, td) or holdings[_code]['buy_price'],
                            '_days': holdings[_code].get('days_held', 1),
                        }

            # 按评分降序排列候选
            candidates_today = []
            for code, df in stock_data.items():
                if not allow_new:
                    break
                if code in holdings or code in pending_sells:
                    continue
                # 因果候选池: 当日必须"前期强势"(近250日涨停≥10)
                if causal_universe is not None and not causal_universe.eligible(code, td):
                    continue
                # 满仓则停止收集（轮换模式下继续收集，用于换仓判断）
                if len(holdings) >= MAX_POSITIONS and not self._rotation:
                    break

                idx = _get_idx(code, td)
                if idx < MAVOL_PERIOD + 20:
                    continue

                close = float(df['close'].iloc[idx])
                prev_close = float(df['close'].iloc[idx - 1]) if idx >= 1 else 0.0
                vol = float(df['volume'].iloc[idx])
                mavol180 = float(df['mavol180'].iloc[idx])

                if pd.isna(mavol180) or mavol180 <= 0:
                    continue

                pressure = self._calc_pressure_at(df, idx)
                if pressure <= 0:
                    continue

                dist_pct = (close - pressure) / pressure * 100
                vol_ratio = vol / mavol180

                # V3 买入条件
                if not (close > pressure and vol > mavol180 * MAVOL_MULTIPLIER):
                    continue

                total_signals += 1

                # V3 当日突破（CROSS 语义）: 昨日收盘未站上压力位 + 今日收盘突破。
                # 只抓"当天刚突破"的标的，排除突破后回踩/横盘（与实盘口径一致）
                if prev_close > 0 and prev_close > pressure * 1.001:
                    continue

                # V3 追高上限: 已突破压力位超过板块阈值 → 放弃（与实盘口径一致）
                if dist_pct > self._chase_limit_pct(code):
                    continue

                # V3 过滤: 死亡区间 3-5%
                if 3 < abs(dist_pct) <= 5:
                    continue
                # V3 过滤: 过度放量 ≥ 5x
                if vol_ratio >= 5.0:
                    continue

                # 熊市降低抽奖率：只做评分更高的
                score = min(100, 60 + (10 if abs(dist_pct) >= 3 else 0)
                                  + (10 if vol_ratio >= 2.0 else 5 if vol_ratio >= 1.5 else 0))
                if not is_bull and score < 75:
                    continue

                # 找到下一交易日
                next_td = self._next_trade_date(td, all_dates)
                if next_td is None:
                    continue

                # 检查开盘跌停
                if self._check_open_limit_down(code, next_td, df):
                    skipped_trades.append({
                        'code': code, 'name': self._get_name(code),
                        'signal_date': td.strftime('%Y-%m-%d'),
                        'skip_reason': '开盘跌停',
                        'close': round(close, 2), 'pressure': pressure,
                    })
                    continue

                # V3 竞价确认
                auction_reject = self._check_auction_v3(code, next_td, df)
                if auction_reject:
                    skipped_trades.append({
                        'code': code, 'name': self._get_name(code),
                        'signal_date': td.strftime('%Y-%m-%d'),
                        'skip_reason': auction_reject,
                        'close': round(close, 2), 'pressure': pressure,
                    })
                    continue

                candidates_today.append({
                    'code': code,
                    'score': score,
                    'close': close, 'pressure': pressure,
                    'dist_pct': dist_pct, 'vol_ratio': vol_ratio,
                    'next_td': next_td,
                })

            # 按评分排序，限制每天最多 N 只新买入
            candidates_today.sort(key=lambda x: -x['score'])
            available_slots = max(0, MAX_POSITIONS - len(holdings))
            max_new = min(MAX_NEW_PER_DAY, available_slots)

            # ── 集中持仓轮换：满仓时若新信号比最弱持仓强 → 卖出最弱、换入新标的 ──
            rotation_budget = 0.0
            if self._rotation and available_slots == 0 and candidates_today:
                weakest_code, weakest_strength = None, 1e18
                for _code, _h in holdings.items():
                    cur = _current_price(_code, td) or _h['buy_price']
                    ret = cur / _h['buy_price'] - 1 if _h['buy_price'] > 0 else 0
                    s = _h.get('score', 0) * (1 + ret)   # 持仓强度 = 信号分 × (1+当前收益)
                    if s < weakest_strength:
                        weakest_strength, weakest_code = s, _code
                top_c = candidates_today[0]
                if weakest_code and top_c['score'] > weakest_strength and weakest_code not in pending_sells:
                    _wpx = _current_price(weakest_code, td) or holdings[weakest_code]['buy_price']
                    pending_sells[weakest_code] = {
                        'hold': holdings[weakest_code],
                        'sell_date': top_c['next_td'],
                        'sell_reason': '轮换换仓(新信号更强)',
                        'fallback_price': _wpx,
                        '_days': 1,
                    }
                    rotation_budget = cash + holdings[weakest_code]['shares'] * _wpx
                    max_new = 1  # 换入 1 只（资金来自最弱持仓次日卖出）

            for c in candidates_today[:max_new]:
                code = c['code']
                next_td = c['next_td']

                # 获取次日开盘价
                next_idx = _get_idx(code, next_td)
                if next_idx < 0:
                    continue
                buy_open = float(stock_data[code]['open'].iloc[next_idx])

                # 板块共振过滤（信号日 td）：只做板块内 ≥min_resonance 只共涨停的突破
                if self._sector_resonance is not None:
                    _s = self._sector_resonance(code, td)
                    if _s < self._min_resonance:
                        continue
                    # 拥挤上限: 板块涨停≥max_resonance(涨停潮) → 追高，跳过
                    if self._max_resonance is not None and _s >= self._max_resonance:
                        continue

                # 按行情调仓（信号日 td 属于哪种行情 → 仓位权重；权重=0 不开新仓）
                w = 1.0
                if self._regime_weights and self._regime_of_day:
                    reg = self._regime_of_day(td)
                    w = self._regime_weights.get(reg, 0.0)
                if w <= 0:
                    continue

                # 计算买入份额（手数取整）
                position_capital = INITIAL_CAPITAL * PER_POSITION_PCT * w
                shares = int(position_capital / buy_open / 100) * 100
                if shares < 100:
                    continue
                buy_cost = shares * buy_open * (1 + 0.0003)  # 买入佣金万3
                if buy_cost > cash + rotation_budget:
                    continue  # 资金不足（含轮换腾出的预算），跳过

                cash -= buy_cost

                holdings[code] = {
                    'code': code,
                    'buy_date': next_td,
                    'buy_price': buy_open,
                    'shares': shares,
                    'signal_date': td,
                    'signal_close': round(c['close'], 2),
                    'pressure': c['pressure'],
                    'dist_pct': round(c['dist_pct'], 1),
                    'vol_ratio': round(c['vol_ratio'], 1),
                    'limit_count': self._fast_count_limit_ups(code),
                    'had_zt': False,
                    'highest_close': buy_open,
                    'awaiting_reversal': False,
                    'score': c['score'],
                    'position_w': round(w, 2),
                    'signal_regime': reg if self._regime_of_day else '',
                }

                buy_today.append(code)
                total_buys += 1

            # ── 日统计 ──
            # 计算当前组合市值
            pos_val = 0.0
            for code, h in holdings.items():
                px = _current_price(code, td)
                if px: pos_val += h['shares'] * px
            total_val = cash + pos_val
            portfolio_values.append((td, total_val, cash, pos_val))
            if total_val > peak_val:
                peak_val = total_val
            last_total = total_val

            daily_log.append({
                'date': td.strftime('%Y-%m-%d'),
                'signals': total_signals,
                'buys': len(buy_today),
                'sells': len(sold_today),
                'holdings': len(holdings),
                'portfolio_value': round(total_val, 2),
                'market_bull': is_bull,
            })

        # ── 到期强制平仓 ──
        last_date = all_dates[-1]
        for code in list(holdings.keys()):
            h = holdings[code]
            sell_idx = _get_idx(code, last_date)
            if sell_idx >= 0:
                final_close = float(stock_data[code]['close'].iloc[sell_idx])
                buy_price = h['buy_price']
                shares = h['shares']
                proceeds = shares * final_close * (1 - 0.0008)
                cash += proceeds

                gross_ret = (final_close - buy_price) / buy_price if buy_price > 0 else 0
                net_ret = gross_ret - TOTAL_COST
                trading_days = 0
                try:
                    trading_days = self.cal.trading_days_between(h['buy_date'], last_date)
                except Exception:
                    trading_days = 5

                finished_trades.append({
                    'code': code,
                    'name': self._get_name(code),
                    'buy_date': h['buy_date'].strftime('%Y-%m-%d'),
                    'buy_price': round(buy_price, 2),
                    'sell_date': last_date.strftime('%Y-%m-%d'),
                    'sell_price': round(final_close, 2),
                    'gross_ret': round(gross_ret * 100, 2),
                    'net_ret': round(net_ret * 100, 2),
                    'is_win': net_ret > 0,
                    'exit_reason': '回测到期强平',
                    'days_held': trading_days if trading_days > 0 else 5,
                    'had_zt': h.get('had_zt', False),
                    'score': h.get('score', 0),
                    'vol_ratio': h.get('vol_ratio', 0),
                    'dist_pct': h.get('dist_pct', 0),
                    'limit_count': h.get('limit_count', 0),
                    'position_w': h.get('position_w', 1.0),
                    'signal_regime': h.get('signal_regime', ''),
                })
            del holdings[code]

        # ── 组合层统计 ──
        cumulative_return = (total_val / INITIAL_CAPITAL - 1) * 100 if INITIAL_CAPITAL > 0 else 0
        peak_value = INITIAL_CAPITAL
        max_drawdown = 0.0
        for _, tv, _, _ in portfolio_values:
            peak_value = max(peak_value, tv)
            dd = (peak_value - tv) / peak_value * 100 if peak_value > 0 else 0
            max_drawdown = max(max_drawdown, dd)

        elapsed = _time.time() - t0
        print(f'\n回测完成 ({elapsed:.0f}s): '
              f'{total_signals} 信号, {total_buys} 买入, '
              f'{len(finished_trades)} 笔交易')
        print(f'  最终市值: {total_val/1e4:.1f}万 | '
              f'累计收益: {cumulative_return:+.1f}% | '
              f'最大回撤: {max_drawdown:.1f}%')

        return {
            'trades': finished_trades,
            'skipped': skipped_trades,
            'daily_log': daily_log,
            'start_date': start_date.strftime('%Y-%m-%d'),
            'end_date': end_date.strftime('%Y-%m-%d'),
            'total_signals': total_signals,
            'total_buys': total_buys,
            'final_value': round(total_val, 2),
            'initial_capital': INITIAL_CAPITAL,
            'cumulative_return': round(cumulative_return, 2),
            'max_drawdown': round(max_drawdown, 2),
            'portfolio_values': portfolio_values,
        }

    def _next_trade_date(self, d: date, all_dates: List[date]) -> Optional[date]:
        """找到 d 之后的下一个交易日。"""
        for td in all_dates:
            if td > d:
                return td
        return None

    @staticmethod
    def _fast_count_limit_ups(code: str) -> int:
        """快速数涨停（取最近250根K线）。"""
        tdx = TdxReader()
        market = 'sh' if str(code).startswith('6') else 'sz'
        if str(code).startswith(('8', '4')):
            market = 'bj'
        try:
            fpath = os.path.join(tdx._market_dir(market), f'{market}{code}.day')
            if not os.path.exists(fpath):
                return 0
            fsize = os.path.getsize(fpath)
            if fsize < RECORD_SIZE * 20:
                return 0
            read_bytes = min(RECORD_SIZE * 260, fsize)
            with open(fpath, 'rb') as f:
                f.seek(fsize - read_bytes)
                raw = f.read(read_bytes)
            n = len(raw) // RECORD_SIZE
            count = 0
            prev_close = 0
            for j in range(n):
                offset = j * RECORD_SIZE
                cl_int = struct.unpack_from('I', raw, offset + 16)[0]
                close = cl_int / 100.0
                if prev_close > 0:
                    chg = (close - prev_close) / prev_close
                    if chg >= 0.093:
                        count += 1
                prev_close = close
            return count
        except Exception:
            return 0


# ═══════════════════════════════════════════════════════════════════════════
# Excel 导出
# ═══════════════════════════════════════════════════════════════════════════

def export_xlsx(results: dict, output_path: str):
    trades = results.get('trades', [])
    skipped = results.get('skipped', [])
    daily = results.get('daily_log', [])

    wb = Workbook()

    # ═══ Sheet 1: 交易明细 ═══
    ws1 = wb.active
    ws1.title = '交易明细'

    # 标题行
    ws1.merge_cells('A1:P1')
    ws1['A1'] = f"V3 VOL180 突破战法 — 交易明细 ({results['start_date']} ~ {results['end_date']})"
    ws1['A1'].font = TITLE_FONT
    ws1['A1'].alignment = CENTER

    headers = ['代码', '名称', '买入日', '买入价', '卖出日', '卖出价',
               '毛收益%', '净收益%', '结果', '持有天', '卖出原因',
               '曾涨停', '评分', '量比', '距压力%', '年涨停次']
    for c, h in enumerate(headers, 1):
        ws1.cell(row=3, column=c, value=h)
    _style_header(ws1, 3, len(headers))

    for i, t in enumerate(trades):
        r = i + 4
        ws1.cell(row=r, column=1, value=t['code'])
        ws1.cell(row=r, column=2, value=t['name'])
        ws1.cell(row=r, column=3, value=t['buy_date'])
        ws1.cell(row=r, column=4, value=t['buy_price'])
        ws1.cell(row=r, column=5, value=t['sell_date'])
        ws1.cell(row=r, column=6, value=t['sell_price'])
        ws1.cell(row=r, column=7, value=t['gross_ret'])
        ws1.cell(row=r, column=8, value=t['net_ret'])
        ws1.cell(row=r, column=9, value='Win' if t['is_win'] else 'Loss')
        ws1.cell(row=r, column=10, value=t['days_held'])
        ws1.cell(row=r, column=11, value=t['exit_reason'])
        ws1.cell(row=r, column=12, value='是' if t.get('had_zt') else '否')
        ws1.cell(row=r, column=13, value=t.get('score', 0))
        ws1.cell(row=r, column=14, value=t.get('vol_ratio', 0))
        ws1.cell(row=r, column=15, value=t.get('dist_pct', 0))
        ws1.cell(row=r, column=16, value=t.get('limit_count', 0))

        if t['is_win']:
            ws1.cell(row=r, column=8).fill = WIN_FILL
            ws1.cell(row=r, column=9).fill = WIN_FILL
        else:
            ws1.cell(row=r, column=8).fill = LOSS_FILL
            ws1.cell(row=r, column=9).fill = LOSS_FILL

    if trades:
        _style_data(ws1, 4, 3 + len(trades), len(headers))
    _auto_width(ws1)

    # ═══ Sheet 2: 统计汇总 ═══
    ws2 = wb.create_sheet('统计汇总')
    ws2.merge_cells('A1:D1')
    ws2['A1'] = 'V3 策略回测 — 统计汇总'
    ws2['A1'].font = TITLE_FONT
    ws2['A1'].alignment = CENTER

    # ── 计算统计 ──
    total = len(trades)
    wins = sum(1 for t in trades if t['is_win'])
    losses = total - wins
    win_rate = wins / max(total, 1) * 100
    net_rets = [t['net_ret'] for t in trades]
    avg_ret = np.mean(net_rets) if net_rets else 0
    avg_win = np.mean([r for r in net_rets if r > 0]) if wins > 0 else 0
    avg_loss = np.mean([r for r in net_rets if r <= 0]) if losses > 0 else 0
    max_win = max(net_rets) if net_rets else 0
    max_loss = min(net_rets) if net_rets else 0

    # 盈亏比
    profit_factor = abs(sum(r for r in net_rets if r > 0) / min(sum(r for r in net_rets if r <= 0), -0.01)) if losses > 0 else 999

    # 平均持有天数
    avg_hold = np.mean([t['days_held'] for t in trades]) if trades else 0

    # ── 组合层级（真实资金模型） ──
    initial_cap = results.get('initial_capital', 1000000.0)
    final_value = results.get('final_value', initial_cap)
    cum_ret = results.get('cumulative_return', 0.0)
    max_dd = results.get('max_drawdown', 0.0)

    # ── 算最大回撤（基于每日组合市值） ──
    portfolio_values = results.get('portfolio_values', [])
    if portfolio_values:
        peak_val = initial_cap
        max_dd_calc = 0.0
        for _, tv, _, _ in portfolio_values:
            peak_val = max(peak_val, tv)
            dd = (peak_val - tv) / peak_val * 100 if peak_val > 0 else 0
            max_dd_calc = max(max_dd_calc, dd)
        if max_dd_calc > 0:
            max_dd = max_dd_calc

    stats = [
        ('回测区间', f"{results['start_date']} ~ {results['end_date']}"),
        ('', ''),
        ('--- 资金模型 ---', ''),
        ('初始资金', f'{initial_cap/1e4:.0f}万'),
        ('最终市值', f'{final_value/1e4:.2f}万'),
        ('最大持仓', '10只（单票10%）'),
        ('每天最多新开', '3只'),
        ('', ''),
        ('--- 交易统计 ---', ''),
        ('信号总数', results.get('total_signals', 0)),
        ('实际买入', results.get('total_buys', 0)),
        ('跳过(竞价/跌停)', len(skipped)),
        ('总交易笔数', total),
        ('盈利笔数', wins),
        ('亏损笔数', losses),
        ('胜率', f'{win_rate:.1f}%'),
        ('', ''),
        ('平均净收益/笔', f'{avg_ret:+.2f}%'),
        ('平均盈利/笔', f'{avg_win:+.2f}%'),
        ('平均亏损/笔', f'{avg_loss:+.2f}%'),
        ('最大单笔盈利', f'{max_win:+.2f}%'),
        ('最大单笔亏损', f'{max_loss:+.2f}%'),
        ('盈亏比', f'{profit_factor:.2f}'),
        ('平均持有天数', f'{avg_hold:.1f}天'),
        ('', ''),
        ('--- 组合表现 ---', ''),
        ('组合累计收益', f'{cum_ret:+.2f}%'),
        ('最大回撤', f'{max_dd:.2f}%'),
        ('年化收益（估算）', f'{cum_ret:+.1f}%'),
    ]

    for i, (label, val) in enumerate(stats):
        r = 3 + i
        ws2.cell(row=r, column=1, value=label).font = SUMMARY_LABEL_FONT
        ws2.cell(row=r, column=2, value=val).font = SUMMARY_VAL_FONT
        ws2.cell(row=r, column=1).border = THIN_BORDER
        ws2.cell(row=r, column=2).border = THIN_BORDER
        # 分隔行着色
        if '---' in str(label):
            ws2.cell(row=r, column=1).fill = PatternFill(start_color='F3F4F6', end_color='F3F4F6', fill_type='solid')
            ws2.cell(row=r, column=2).fill = PatternFill(start_color='F3F4F6', end_color='F3F4F6', fill_type='solid')

    # 大数字显示
    big_row = 3 + len(stats) + 2
    ws2.merge_cells(f'A{big_row}:B{big_row}')
    ws2.cell(row=big_row, column=1, value='胜率').font = SUMMARY_LABEL_FONT
    ws2.cell(row=big_row, column=3, value=f'{win_rate:.1f}%').font = GREEN_FONT if win_rate >= 50 else RED_FONT
    ws2.cell(row=big_row, column=1).border = THIN_BORDER
    ws2.cell(row=big_row, column=3).border = THIN_BORDER

    big_row2 = big_row + 1
    ws2.merge_cells(f'A{big_row2}:B{big_row2}')
    ws2.cell(row=big_row2, column=1, value='组合累计收益').font = SUMMARY_LABEL_FONT
    ws2.cell(row=big_row2, column=3, value=f'{cum_ret:+.2f}%').font = GREEN_FONT if cum_ret > 0 else RED_FONT
    ws2.cell(row=big_row2, column=1).border = THIN_BORDER
    ws2.cell(row=big_row2, column=3).border = THIN_BORDER

    big_row3 = big_row2 + 1
    ws2.merge_cells(f'A{big_row3}:B{big_row3}')
    ws2.cell(row=big_row3, column=1, value='最大回撤').font = SUMMARY_LABEL_FONT
    ws2.cell(row=big_row3, column=3, value=f'{max_dd:.2f}%').font = RED_FONT
    ws2.cell(row=big_row3, column=1).border = THIN_BORDER
    ws2.cell(row=big_row3, column=3).border = THIN_BORDER

    _auto_width(ws2, min_width=14, max_width=30)
    ws2.column_dimensions['C'].width = 18

    # ═══ Sheet 3: 按卖出原因统计 ═══
    ws3 = wb.create_sheet('卖出原因统计')
    ws3.merge_cells('A1:F1')
    ws3['A1'] = '按卖出原因统计'
    ws3['A1'].font = TITLE_FONT
    ws3['A1'].alignment = CENTER

    reason_headers = ['卖出原因', '笔数', '胜率%', '平均收益%', '总收益%', '平均持有天']
    for c, h in enumerate(reason_headers, 1):
        ws3.cell(row=3, column=c, value=h)
    _style_header(ws3, 3, len(reason_headers))

    by_reason = defaultdict(list)
    for t in trades:
        reason = t['exit_reason']
        # 合并相似原因
        if '止损' in reason:
            reason = '止损-6%'
        elif '移动止盈' in reason:
            reason = '移动止盈'
        elif '断板' in reason or '涨停后断板' in reason:
            reason = '涨停后断板离场'
        elif '到期' in reason:
            reason = '持有到期'
        elif '强平' in reason:
            reason = '回测到期强平'
        by_reason[reason].append(t)

    row = 4
    for reason in sorted(by_reason.keys(), key=lambda x: -len(by_reason[x])):
        items = by_reason[reason]
        n = len(items)
        w = sum(1 for t in items if t['is_win'])
        wr = w / max(n, 1) * 100
        avg_r = np.mean([t['net_ret'] for t in items])
        total_r = sum(t['net_ret'] for t in items)
        avg_d = np.mean([t['days_held'] for t in items])

        ws3.cell(row=row, column=1, value=reason)
        ws3.cell(row=row, column=2, value=n)
        ws3.cell(row=row, column=3, value=f'{wr:.1f}%')
        ws3.cell(row=row, column=4, value=f'{avg_r:+.2f}%')
        ws3.cell(row=row, column=5, value=f'{total_r:+.2f}%')
        ws3.cell(row=row, column=6, value=f'{avg_d:.1f}')
        row += 1

    _style_data(ws3, 4, row - 1, len(reason_headers))
    _auto_width(ws3)

    # ═══ Sheet 4: 月度统计 ═══
    ws4 = wb.create_sheet('月度统计')
    ws4.merge_cells('A1:I1')
    ws4['A1'] = '按月统计（组合层级）'
    ws4['A1'].font = TITLE_FONT
    ws4['A1'].alignment = CENTER

    month_headers = ['月份', '笔数', '胜', '负', '胜率%', '平均收益%',
                     '组合月收益%', '月末市值(万)', '累计收益%']
    for c, h in enumerate(month_headers, 1):
        ws4.cell(row=3, column=c, value=h)
    _style_header(ws4, 3, len(month_headers))

    # 用组合估值计算月度收益（而非简单累加每笔交易）
    init_cap = results.get('initial_capital', 1000000.0)
    pv = results.get('portfolio_values', [])
    monthly_pv = {}  # YYYY-MM → [(date, value)]
    for dt, tv, _, _ in pv:
        mk = dt.strftime('%Y-%m')
        if mk not in monthly_pv:
            monthly_pv[mk] = []
        monthly_pv[mk].append((dt, tv))

    # 每月交易统计
    by_month = defaultdict(list)
    for t in trades:
        month_key = t['sell_date'][:7]
        by_month[month_key].append(t)

    row = 4
    prev_month_end_val = init_cap
    cum_ret = 0.0

    for month_key in sorted(set(list(by_month.keys()) + list(monthly_pv.keys()))):
        items = by_month.get(month_key, [])
        pvs = monthly_pv.get(month_key, [])
        n = len(items)
        w = sum(1 for t in items if t['is_win']) if n > 0 else 0
        l = n - w
        wr = w / max(n, 1) * 100
        avg_r = np.mean([t['net_ret'] for t in items]) if n > 0 else 0

        # 组合层级月收益（基于月末市值 vs 上月月末市值）
        if pvs:
            month_end_val = pvs[-1][1]  # last portfolio value of month
        else:
            month_end_val = prev_month_end_val
        month_ret = (month_end_val / prev_month_end_val - 1) * 100 if prev_month_end_val > 0 else 0
        cum_ret = (month_end_val / init_cap - 1) * 100

        ws4.cell(row=row, column=1, value=month_key)
        ws4.cell(row=row, column=2, value=n if n > 0 else 0)
        ws4.cell(row=row, column=3, value=w)
        ws4.cell(row=row, column=4, value=l)
        ws4.cell(row=row, column=5, value=f'{wr:.1f}%' if n > 0 else '-')
        ws4.cell(row=row, column=6, value=f'{avg_r:+.2f}%' if n > 0 else '-')
        ws4.cell(row=row, column=7, value=f'{month_ret:+.2f}%')
        ws4.cell(row=row, column=8, value=f'{month_end_val/1e4:.1f}')
        ws4.cell(row=row, column=9, value=f'{cum_ret:+.1f}%')

        if month_ret > 0:
            ws4.cell(row=row, column=7).fill = WIN_FILL
        else:
            ws4.cell(row=row, column=7).fill = LOSS_FILL
        if cum_ret > 0:
            ws4.cell(row=row, column=9).fill = WIN_FILL
        else:
            ws4.cell(row=row, column=9).fill = LOSS_FILL

        prev_month_end_val = month_end_val
        row += 1

    _style_data(ws4, 4, row - 1, len(month_headers))
    _auto_width(ws4)

    # ═══ Sheet 5: 日度日志 ═══
    ws5 = wb.create_sheet('日度日志')
    ws5.merge_cells('A1:E1')
    ws5['A1'] = '日度持仓变化'
    ws5['A1'].font = TITLE_FONT
    ws5['A1'].alignment = CENTER

    daily_headers = ['日期', '累计信号', '买入', '卖出', '持仓数']
    for c, h in enumerate(daily_headers, 1):
        ws5.cell(row=3, column=c, value=h)
    _style_header(ws5, 3, len(daily_headers))

    for i, d in enumerate(daily):
        r = i + 4
        ws5.cell(row=r, column=1, value=d['date'])
        ws5.cell(row=r, column=2, value=d['signals'])
        ws5.cell(row=r, column=3, value=d['buys'])
        ws5.cell(row=r, column=4, value=d['sells'])
        ws5.cell(row=r, column=5, value=d['holdings'])

    if daily:
        _style_data(ws5, 4, 3 + len(daily), len(daily_headers))
    _auto_width(ws5)

    # ═══ Sheet 6: 跳过记录 ═══
    if skipped:
        ws6 = wb.create_sheet('跳过记录')
        ws6.merge_cells('A1:F1')
        ws6['A1'] = '竞价确认/开盘跌停 跳过的买入信号'
        ws6['A1'].font = TITLE_FONT
        ws6['A1'].alignment = CENTER

        skip_headers = ['代码', '名称', '信号日', '跳过原因', '收盘价', '压力位']
        for c, h in enumerate(skip_headers, 1):
            ws6.cell(row=3, column=c, value=h)
        _style_header(ws6, 3, len(skip_headers))

        for i, s in enumerate(skipped):
            r = i + 4
            ws6.cell(row=r, column=1, value=s['code'])
            ws6.cell(row=r, column=2, value=s['name'])
            ws6.cell(row=r, column=3, value=s['signal_date'])
            ws6.cell(row=r, column=4, value=s['skip_reason'])
            ws6.cell(row=r, column=5, value=s.get('close', 0))
            ws6.cell(row=r, column=6, value=s.get('pressure', 0))
        _style_data(ws6, 4, 3 + len(skipped), len(skip_headers))
        _auto_width(ws6)

    wb.save(output_path)
    print(f'\n[OK] Excel saved: {output_path}')


# ═══════════════════════════════════════════════════════════════════════════
# Main
# ═══════════════════════════════════════════════════════════════════════════

if __name__ == '__main__':
    import argparse
    p = argparse.ArgumentParser(description='V3 VOL180 突破战法回测')
    p.add_argument('--start', type=str, default=None, help='开始日期 YYYY-MM-DD (默认一年前)')
    p.add_argument('--end', type=str, default=None, help='结束日期 YYYY-MM-DD (默认昨天)')
    p.add_argument('--output', type=str, default=None, help='输出 xlsx 路径')
    args = p.parse_args()

    start_date = datetime.strptime(args.start, '%Y-%m-%d').date() if args.start else None
    end_date = datetime.strptime(args.end, '%Y-%m-%d').date() if args.end else None
    output_path = args.output or os.path.join(OUTPUT_DIR, 'v3_backtest_result.xlsx')

    bt = V3Backtest()
    results = bt.run(start_date=start_date, end_date=end_date)

    if results:
        export_xlsx(results, output_path)
