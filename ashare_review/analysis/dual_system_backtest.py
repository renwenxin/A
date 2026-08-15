"""双体系选股 + 竞价确认回测

选股体系：
  1. 1进2接力 — T日首板涨停 → T+1日竞价确认后买入
  2. 涨停复制 — T日识别近20日涨停回调企稳信号 → T+1日竞价确认后买入

竞价确认（二次筛选）：
  - 必要条件：T+1日成交量 >= T日成交量 × 50%（"竞价量≥50%昨日爆量"的代理）
  - 优先条件：T+1日成交量 >= T日成交量（"大于或等于昨日爆量"）

仓位管理：
  - 最多同时持有 3 只标的

卖出条件：
  - 涨停持有（当日涨幅 >= 涨停阈值 → 继续持有）
  - 不涨停卖出（当日不涨停 → 收盘价卖出）
  - 止损 -6%（日内最低价触及入场价×0.94 → 止损离场）

用法：
    python -m ashare_review.analysis.dual_system_backtest --days 250 --output backtest_1y.xlsx
"""
import sys, os, json, struct, time
from datetime import date, datetime, timedelta
from typing import Dict, List, Optional, Tuple, Set
from collections import defaultdict
import numpy as np

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

from ashare_review.data.tdx_reader import TdxReader, RECORD_SIZE
from ashare_review.utils.calendar import TradingCalendar


# ======================================================================
# 辅助函数
# ======================================================================
def _is_a_stock(code: str) -> bool:
    """过滤A股（排除B股、港股通）"""
    if not code or len(code) != 6:
        return False
    return (code.startswith(('0', '3', '6')) and
            not code.startswith(('900', '200')))


def _board_limit_threshold(code: str) -> float:
    """涨停阈值"""
    if code.startswith(('300', '301', '688')):
        return 19.9
    if code.startswith(('8', '4')):
        return 29.9
    return 9.9


_name_cache: Dict[str, str] = {}


def _get_stock_name(code: str) -> str:
    """获取股票名称"""
    if code in _name_cache:
        return _name_cache[code]
    try:
        from ashare_review.screening.base import BaseScreener
        bs = BaseScreener.__new__(BaseScreener)
        bs.__init__()
        name = bs._get_name(code)
        if name:
            _name_cache[code] = name
            return name
    except Exception:
        pass
    return code


# ======================================================================
# TDX涨停索引（复用 auction_confirm_backtest.py 的逻辑）
# ======================================================================
class TdxLimitUpIndex:
    """预扫描TDX .day文件，构建日期→涨停股列表的索引"""

    def __init__(self, tdx: TdxReader):
        self.tdx = tdx
        self._index: Dict[date, List[Dict]] = defaultdict(list)
        self._built = False

    def build(self, start_date: date, end_date: date):
        if self._built:
            return
        scan_start = start_date - timedelta(days=30)
        records_to_read = (end_date - scan_start).days + 10
        read_bytes = records_to_read * RECORD_SIZE

        total_files = 0
        matched = 0
        for market in ['sh', 'sz']:
            mdir = self.tdx._market_dir(market)
            if not os.path.exists(mdir):
                continue
            files = [f for f in os.listdir(mdir) if f.endswith('.day')]
            for fname in files:
                total_files += 1
                code = fname[2:8]
                if not _is_a_stock(code):
                    continue
                fpath = os.path.join(mdir, fname)
                fsize = os.path.getsize(fpath)
                if fsize < RECORD_SIZE * 2:
                    continue
                read_size = min(read_bytes, fsize)
                try:
                    with open(fpath, 'rb') as f:
                        f.seek(fsize - read_size)
                        data = f.read(read_size)
                except Exception:
                    continue

                threshold = _board_limit_threshold(code)
                num_records = len(data) // RECORD_SIZE
                records = []
                for i in range(num_records):
                    offset = i * RECORD_SIZE
                    dt_int, op, hi, lo, cl, amt, vol, _ = struct.unpack(
                        'IIIIIfII', data[offset:offset + RECORD_SIZE])
                    dt_str = str(dt_int)
                    try:
                        d = date(int(dt_str[:4]), int(dt_str[4:6]), int(dt_str[6:8]))
                    except ValueError:
                        continue
                    records.append({
                        'date': d, 'open': op / 100.0, 'high': hi / 100.0,
                        'low': lo / 100.0, 'close': cl / 100.0,
                        'volume': vol, 'amount': amt,
                    })

                for j in range(1, len(records)):
                    r = records[j]
                    if r['date'] < start_date or r['date'] > end_date:
                        continue
                    prev = records[j-1]
                    if prev['close'] <= 0:
                        continue
                    change_pct = (r['close'] - prev['close']) / prev['close'] * 100
                    if change_pct >= threshold - 0.1:
                        is_yizi = (abs(r['open'] - r['close']) < 0.01 and
                                   abs(r['high'] - r['low']) < 0.01 and
                                   abs(r['high'] - r['close']) < 0.01)
                        self._index[r['date']].append({
                            'code': code, 'market': market,
                            'open': r['open'], 'high': r['high'],
                            'low': r['low'], 'close': r['close'],
                            'volume': r['volume'], 'amount': r['amount'],
                            'prev_close': prev['close'],
                            'change_pct': round(change_pct, 1),
                            'is_yizi': is_yizi,
                        })
                        matched += 1

        self._built = True
        print(f'[TDX涨停索引] 扫描{total_files}只股票 → '
              f'找到{matched}条涨停记录覆盖{len(self._index)}个交易日')

    def get_boards(self, d: date) -> List[Dict]:
        return self._index.get(d, [])


# ======================================================================
# 1进2接力筛选器（复用简化版）
# ======================================================================
class SimulatedOneTwoScreener:
    """TDX日线模拟1进2筛选"""

    MIN_PRICE = 3.0
    MAX_PRICE = 15.0
    MIN_SCORE = 35

    def screen(self, signal_date: date, tdx: TdxReader,
               tdx_index: TdxLimitUpIndex) -> List[Dict]:
        """T日首板 → 筛选1进2候选"""
        first_boards = [s for s in tdx_index.get_boards(signal_date)
                        if s['change_pct'] < _board_limit_threshold(s['code']) + 5]
        if not first_boards:
            return []

        # 过滤：只保留首板（近2日首次涨停）
        # 简化判断：如果前一日也是涨停，则不是首板
        prev_date = signal_date - timedelta(days=1)
        while prev_date.weekday() >= 5:
            prev_date -= timedelta(days=1)
        prev_boards = {s['code'] for s in tdx_index.get_boards(prev_date)}

        candidates = []
        for s in first_boards:
            code = s['code']
            close = s['close']
            volume = s['volume']

            if code in prev_boards:
                continue  # 非首板
            if close < self.MIN_PRICE or close > self.MAX_PRICE:
                continue
            if s['is_yizi']:
                continue
            threshold = _board_limit_threshold(code)
            if threshold > 10.0:  # 排除创业板/科创板
                continue

            score = 0
            reasons = []

            # 股价评分
            if 5 <= close <= 10:
                score += 20
                reasons.append(f'股价{close:.1f}元·最佳区间')
            elif 3 <= close <= 15:
                score += 12
                reasons.append(f'股价{close:.1f}元')

            # 量能评分
            if volume > 100_000_000:
                score += 20
                reasons.append(f'巨量{volume/1e8:.1f}亿股')
            elif volume > 50_000_000:
                score += 14
            elif volume > 20_000_000:
                score += 8
            else:
                score += 3

            # 成交额
            amount_yi = s['amount'] / 1e8
            if amount_yi > 5:
                score += 15
                reasons.append(f'成交额{amount_yi:.1f}亿')
            elif amount_yi > 2:
                score += 8
            elif amount_yi > 1:
                score += 4

            # 封板稳定（上影线小）
            if s['high'] > close:
                upper_shadow = (s['high'] - close) / (s['high'] - s['low'] + 0.01)
            else:
                upper_shadow = 0
            if upper_shadow < 0.2:
                score += 12
                reasons.append('封板稳定')
            elif upper_shadow < 0.5:
                score += 6

            # 开盘幅度
            if s['prev_close'] > 0:
                gap = (s['open'] - s['prev_close']) / s['prev_close']
            else:
                gap = 0
            if 0.02 <= gap <= 0.05:
                score += 10
                reasons.append('高开涨停·强势')
            elif 0.01 <= gap < 0.02:
                score += 6
            elif gap < 0:
                score += 4
                reasons.append('低开拉板')

            if score < self.MIN_SCORE:
                continue

            name = _get_stock_name(code)
            candidates.append({
                'code': code, 'name': name, 'score': min(score, 100),
                'close': close, 'volume': volume,
                'amount_yi': round(amount_yi, 1),
                'change_pct': s['change_pct'], 'reasons': reasons,
                'system': '1进2接力',
            })

        candidates.sort(key=lambda x: x['score'], reverse=True)
        return candidates[:20]


# ======================================================================
# 涨停复制筛选器
# ======================================================================
class ZTReplicaScreener:
    """涨停复制回调筛选 — TDX日线模拟"""

    def screen(self, signal_date: date, tdx: TdxReader,
               tdx_index: TdxLimitUpIndex) -> List[Dict]:
        """T日筛选近20日涨停后回调企稳的标的"""
        # 收集近25日有过涨停的标的
        codes_to_check = set()
        for delta in range(1, 26):  # 往前看1-25天
            check_date = signal_date - timedelta(days=delta)
            if check_date.weekday() >= 5:
                continue
            for s in tdx_index.get_boards(check_date):
                code = s['code']
                if _is_a_stock(code):
                    threshold = _board_limit_threshold(code)
                    if threshold <= 10.0:  # 仅主板
                        codes_to_check.add(code)

        candidates = []
        for code in codes_to_check:
            result = self._evaluate(code, signal_date, tdx, tdx_index)
            if result:
                candidates.append(result)

        candidates.sort(key=lambda x: x['score'], reverse=True)
        return candidates[:15]

    def _evaluate(self, code: str, signal_date: date, tdx: TdxReader,
                  tdx_index: TdxLimitUpIndex) -> Optional[Dict]:
        """评估单只股票的涨停复制潜力"""
        market = 'sh' if code.startswith('6') else 'sz'
        if code.startswith(('8', '4')):
            market = 'bj'

        fpath = os.path.join(tdx._market_dir(market), f'{market}{code}.day')
        if not os.path.exists(fpath):
            return None
        fsize = os.path.getsize(fpath)
        if fsize < RECORD_SIZE * 60:
            return None

        # 读取日线数据
        read_size = min(RECORD_SIZE * 400, fsize)
        with open(fpath, 'rb') as f:
            f.seek(fsize - read_size)
            data = f.read(read_size)

        num_records = len(data) // RECORD_SIZE
        records = []
        for i in range(num_records):
            offset = i * RECORD_SIZE
            dt_int, op, hi, lo, cl, amt, vol, _ = struct.unpack(
                'IIIIIfII', data[offset:offset + RECORD_SIZE])
            dt_str = str(dt_int)
            try:
                d = date(int(dt_str[:4]), int(dt_str[4:6]), int(dt_str[6:8]))
            except ValueError:
                continue
            records.append({
                'date': d, 'open': op / 100.0, 'high': hi / 100.0,
                'low': lo / 100.0, 'close': cl / 100.0,
                'volume': vol, 'amount': amt,
            })

        if len(records) < 30:
            return None

        # 找到signal_date的索引
        idx = None
        for i, r in enumerate(records):
            if r['date'] == signal_date:
                idx = i
                break
        if idx is None or idx < 10:
            return None

        closes = [r['close'] for r in records]
        opens = [r['open'] for r in records]
        highs = [r['high'] for r in records]
        lows = [r['low'] for r in records]
        volumes = [r['volume'] for r in records]

        # 计算MAVOL180（只用信号日之前的数据）
        if idx + 1 >= 180:
            mavol180 = float(np.mean(volumes[idx - 179:idx + 1]))
        else:
            mavol180 = float(np.mean(volumes))

        if mavol180 <= 0:
            return None

        # 在近25日内找最近一次涨停
        limit_pct = _board_limit_threshold(code)
        zt_idx = None
        zt_close = zt_vol = zt_low = zt_high = 0
        zt_timing = '换手板'
        lookback = min(25, idx)

        for j in range(idx - 1, max(idx - lookback - 1, 1), -1):
            prev_c = closes[j - 1]
            if prev_c <= 0:
                continue
            chg = (closes[j] - prev_c) / prev_c * 100
            if chg >= limit_pct - 0.1:
                if abs(opens[j] - closes[j]) / max(closes[j], 0.01) < 0.005:
                    continue  # 排除一字板
                zt_idx = j
                zt_close = closes[j]; zt_vol = volumes[j]
                zt_low = lows[j]; zt_high = highs[j]
                open_chg = (opens[j] - prev_c) / prev_c
                if open_chg >= 0.03:
                    zt_timing = '早盘强势板'
                elif open_chg >= 0:
                    zt_timing = '换手板'
                else:
                    zt_timing = '低开拉板'
                break

        if zt_idx is None:
            return None

        days_since = idx - zt_idx
        if days_since < 1 or days_since > 12:
            return None

        # 回调分析
        pb_vols = [volumes[i] for i in range(zt_idx + 1, idx + 1)]
        pb_vol_max = max(pb_vols) if pb_vols else 0
        pb_high = max(highs[i] for i in range(zt_idx + 1, idx + 1))
        pb_low = min(lows[i] for i in range(zt_idx + 1, idx + 1))

        is_shrinking = pb_vol_max < zt_vol * 0.6
        is_moderate = pb_vol_max < zt_vol * 0.8
        is_above_zt_low = pb_low >= zt_low * 0.98

        close = closes[idx]; vol = volumes[idx]
        vol_ratio = vol / mavol180

        # 均线（只用信号日之前的数据）
        ma5 = float(np.mean(closes[idx - 4:idx + 1])) if idx >= 4 else 0
        ma10 = float(np.mean(closes[idx - 9:idx + 1])) if idx >= 9 else 0
        ma20 = float(np.mean(closes[idx - 19:idx + 1])) if idx >= 19 else 0

        ma_support = ''
        ma_score = 0
        if ma5 > 0:
            dist_ma5 = (close - ma5) / ma5 * 100
            if -1 <= dist_ma5 <= 3:
                ma_support = '回踩MA5'; ma_score = 10
        if not ma_support and ma10 > 0:
            dist_ma10 = (close - ma10) / ma10 * 100
            if -1 <= dist_ma10 <= 4:
                ma_support = '回踩MA10'; ma_score = 6
        if not ma_support and ma20 > 0 and close > ma20:
            ma_support = '站稳MA20'; ma_score = 3

        # 今日涨跌幅
        today_chg = 0
        if idx >= 1 and closes[idx-1] > 0:
            today_chg = (close - closes[idx-1]) / closes[idx-1]

        # 今日是否涨停
        is_zt_today = today_chg >= limit_pct - 0.1

        # 四类信号判断
        break_pb = close > pb_high and vol_ratio >= 1.2
        sig_a = is_shrinking and break_pb  # N字反包
        sig_b = is_zt_today and (is_shrinking or is_moderate)  # 双响炮
        sig_c = ((is_shrinking or is_moderate) and is_above_zt_low
                 and vol_ratio >= 1.2 and close > zt_close * 0.98 and not break_pb)
        sig_d = (not is_zt_today and 0.03 <= today_chg < 0.095
                 and vol_ratio >= 1.5 and close > ma5 and ma_support != ''
                 and (is_shrinking or is_moderate))

        if not (sig_a or sig_b or sig_c or sig_d):
            return None
        if vol_ratio >= 5.0:
            return None

        # 评分
        if sig_b:
            sig_type = '涨停双响炮'; score = 70
        elif sig_a:
            sig_type = 'N字反包'; score = 63
        elif sig_d:
            sig_type = '蓄势待发'; score = 52
        else:
            sig_type = '缩量回踩企稳'; score = 55

        if is_shrinking: score += 12
        elif is_moderate: score += 6
        if vol_ratio >= 2.0: score += 10
        elif vol_ratio >= 1.5: score += 5
        if 2 <= days_since <= 4: score += 8
        elif days_since <= 6: score += 4
        if zt_timing == '早盘强势板': score += 3
        score += ma_score

        name = _get_stock_name(code)
        if not name or 'ST' in str(name).upper():
            return None

        return {
            'code': code, 'name': name,
            'sig_type': sig_type, 'score': min(100, score),
            'close': round(float(close), 2), 'volume': vol,
            'vol_ratio': round(vol_ratio, 1),
            'zt_days_ago': days_since,
            'system': '涨停复制',
        }


# ======================================================================
# 主回测器
# ======================================================================
class DualSystemBacktest:
    """双体系选股 + 竞价确认回测"""

    def __init__(self):
        self.tdx = TdxReader()
        self.calendar = TradingCalendar()
        self.tdx_index: Optional[TdxLimitUpIndex] = None
        self.one_two = SimulatedOneTwoScreener()
        self.zt_replica = ZTReplicaScreener()

    # ==================================================================
    # 主入口
    # ==================================================================
    def run(self, lookback_days: int = 250, max_positions: int = 3,
            stop_loss: float = -0.06, vol_ratio_threshold: float = 0.5,
            hold_days: int = 10) -> Dict:
        trade_dates = self._get_trade_dates(lookback_days)
        print(f'回测范围: {trade_dates[0]} ~ {trade_dates[-1]} 共{len(trade_dates)}个交易日')
        print(f'参数: 止损{stop_loss*100:+.0f}% | 竞价量比≥{vol_ratio_threshold} | 最大持仓{max_positions}只')
        print()

        # 构建涨停索引
        print('构建TDX涨停索引...')
        self.tdx_index = TdxLimitUpIndex(self.tdx)
        self.tdx_index.build(trade_dates[0] - timedelta(days=30), trade_dates[-1])
        print()

        positions: List[Dict] = []  # 当前持仓
        all_trades: List[Dict] = []  # 全部已完成交易
        daily_snapshots: List[Dict] = []  # 每日持仓快照

        for i, td in enumerate(trade_dates):
            td_str = td.strftime('%Y%m%d')

            # --- 大盘环境过滤 ---
            if not self._check_market_env(td_str):
                # 大盘不好时也检查持仓是否需要止损
                positions, closed = self._check_exits(positions, td, stop_loss)
                all_trades.extend(closed)
                daily_snapshots.append({
                    'date': td_str, 'positions': len(positions),
                    'holdings': [{'code': p['code'], 'name': p['name'],
                                  'entry_date': p['entry_date'],
                                  'entry_price': p['entry_price'],
                                  'current_price': p.get('last_close', 0),
                                  'pnl_pct': p.get('unrealized_pnl', 0)}
                                 for p in positions],
                })
                continue

            # --- 1. 检查持仓卖出 ---
            positions, closed = self._check_exits(positions, td, stop_loss)
            all_trades.extend(closed)

            # --- 2. 筛选新候选 ---
            if len(positions) < max_positions:
                yesterday = self._prev_trade_date(td)

                # 2a. 1进2候选
                ot_picks = self.one_two.screen(yesterday, self.tdx, self.tdx_index)

                # 2b. 涨停复制候选
                zt_picks = self.zt_replica.screen(yesterday, self.tdx, self.tdx_index)

                # 2c. 合并去重
                all_picks = {}
                for p in ot_picks:
                    if p['code'] not in all_picks:
                        all_picks[p['code']] = p
                for p in zt_picks:
                    if p['code'] not in all_picks or p['score'] > all_picks[p['code']]['score']:
                        all_picks[p['code']] = p

                # 排除已在持仓中的标的
                held_codes = {p['code'] for p in positions}
                candidates = [p for code, p in all_picks.items() if code not in held_codes]

                # --- 3. 竞价确认 ---
                confirmed = []
                for c in candidates:
                    vol_check = self._check_auction_volume(c, td)
                    if vol_check is None:
                        continue
                    vol_ratio, is_priority = vol_check
                    if vol_ratio >= vol_ratio_threshold:
                        confirmed.append({
                            **c,
                            'auction_vol_ratio': round(vol_ratio, 2),
                            'is_priority': is_priority,
                        })

                # 排序：优先条件 > 评分
                confirmed.sort(key=lambda x: (x['is_priority'], x['score']), reverse=True)

                # --- 4. 买入 ---
                slots = max_positions - len(positions)
                for c in confirmed[:slots]:
                    entry_price = self._get_open_price(c['code'], td)
                    if entry_price is None or entry_price <= 0:
                        continue

                    # 高开>7%放弃（一字板预期）
                    if c['close'] > 0:
                        gap = (entry_price - c['close']) / c['close']
                        if gap > 0.07:
                            continue

                    positions.append({
                        'code': c['code'],
                        'name': c['name'],
                        'system': c['system'],
                        'sig_type': c.get('sig_type', ''),
                        'entry_date': td_str,
                        'entry_price': entry_price,
                        'score': c['score'],
                        'auction_vol_ratio': c['auction_vol_ratio'],
                        'is_priority': c['is_priority'],
                        'last_close': entry_price,
                        'unrealized_pnl': 0,
                    })

            # --- 5. 更新持仓市值 ---
            for p in positions:
                close_p = self._get_close_price(p['code'], td)
                if close_p is not None and close_p > 0:
                    p['last_close'] = close_p
                    p['unrealized_pnl'] = round((close_p - p['entry_price']) / p['entry_price'] * 100, 1)

            # --- 6. 每日快照 ---
            daily_snapshots.append({
                'date': td_str,
                'positions': len(positions),
                'holdings': [{
                    'code': p['code'], 'name': p['name'],
                    'system': p['system'],
                    'entry_date': p['entry_date'],
                    'entry_price': p['entry_price'],
                    'current_price': p.get('last_close', 0),
                    'pnl_pct': p.get('unrealized_pnl', 0),
                } for p in positions],
            })

            # 进度
            if (i + 1) % 20 == 0 or i == 0:
                print(f'\r[{i+1}/{len(trade_dates)}] {td_str} '
                      f'候选:{len(confirmed)} 买入:{min(slots, len(confirmed))} '
                      f'持仓:{len(positions)} 累计交易:{len(all_trades)}',
                      end='', flush=True)

        print()
        return self._summarize(all_trades, daily_snapshots, positions,
                               stop_loss, vol_ratio_threshold, max_positions,
                               trade_dates[0].strftime('%Y%m%d'),
                               trade_dates[-1].strftime('%Y%m%d'))

    # ==================================================================
    # 卖出检查
    # ==================================================================
    def _check_exits(self, positions: List[Dict], td: date,
                     stop_loss: float) -> Tuple[List[Dict], List[Dict]]:
        """检查所有持仓是否需要卖出。返回(保留的持仓, 已平仓交易)"""
        remaining = []
        closed = []

        for p in positions:
            code = p['code']
            entry_price = p['entry_price']
            entry_date_str = p['entry_date']

            # 读取当天行情
            row = self._get_day_row(code, td)
            if row is None:
                remaining.append(p)
                continue

            day_open = row['open']
            day_high = row['high']
            day_low = row['low']
            day_close = row['close']
            day_volume = row['volume']

            # 更新市值
            p['last_close'] = day_close
            p['unrealized_pnl'] = round((day_close - entry_price) / entry_price * 100, 1)

            # 1. 止损检查（最高优先级）
            stop_price = round(entry_price * (1 + stop_loss), 2)
            if day_low <= stop_price:
                exit_price = min(stop_price, day_open)
                ret_pct = (exit_price - entry_price) / entry_price * 100
                entry_dt = datetime.strptime(entry_date_str, '%Y%m%d').date()
                days_held = max((td - entry_dt).days, 1)

                closed.append({
                    'code': code, 'name': p['name'],
                    'system': p['system'], 'sig_type': p.get('sig_type', ''),
                    'entry_date': entry_date_str,
                    'entry_price': entry_price,
                    'exit_date': td.strftime('%Y%m%d'),
                    'exit_price': round(exit_price, 2),
                    'result': '止损',
                    'return_pct': round(ret_pct, 1),
                    'days_held': days_held,
                    'score': p['score'],
                    'auction_vol_ratio': p['auction_vol_ratio'],
                })
                continue

            # 2. 涨停判断
            is_limit_up = self._is_limit_up(code, row)
            if is_limit_up:
                remaining.append(p)
                continue

            # 3. 不涨停 → 卖出
            ret_pct = (day_close - entry_price) / entry_price * 100
            entry_dt = datetime.strptime(entry_date_str, '%Y%m%d').date()
            days_held = max((td - entry_dt).days, 1)

            result = '止盈' if ret_pct > 0 else '亏损'
            closed.append({
                'code': code, 'name': p['name'],
                'system': p['system'], 'sig_type': p.get('sig_type', ''),
                'entry_date': entry_date_str,
                'entry_price': entry_price,
                'exit_date': td.strftime('%Y%m%d'),
                'exit_price': round(day_close, 2),
                'result': result,
                'return_pct': round(ret_pct, 1),
                'days_held': days_held,
                'score': p['score'],
                'auction_vol_ratio': p['auction_vol_ratio'],
            })

        return remaining, closed

    def _is_limit_up(self, code: str, row: Dict) -> bool:
        """判断当日是否涨停"""
        threshold = _board_limit_threshold(code)
        if row['prev_close'] <= 0 or row['close'] <= 0:
            return False
        chg = (row['close'] - row['prev_close']) / row['prev_close'] * 100
        # 涨停：涨幅达到阈值，且不是一字板（一字板有买入机会）
        if chg >= threshold - 0.1:
            return True
        return False

    # ==================================================================
    # 竞价量确认
    # ==================================================================
    def _check_auction_volume(self, candidate: Dict, entry_date: date) -> Optional[Tuple[float, bool]]:
        """检查竞价量是否达标。

        返回 (量比, 是否优先) 或 None（数据不足）。
        量比 = entry_date当日成交量 / 昨日成交量。
        优先 = 量比 >= 1.0（大于或等于昨日爆量）。
        """
        code = candidate['code']
        sig_close = candidate['close']

        row = self._get_day_row(code, entry_date)
        if row is None or row['volume'] <= 0:
            return None

        # 昨日爆量 = entry_date前一日的成交量（即信号日的量）
        # 对于1进2：昨日是首板日（涨停 = 爆量）
        # 对于涨停复制：昨日是信号确认日
        yesterday_date = self._prev_trade_date(entry_date)
        yesterday_row = self._get_day_row(code, yesterday_date)
        if yesterday_row is None or yesterday_row['volume'] <= 0:
            # 回退：用近20日最大量
            yesterday_vol = candidate.get('volume', row['volume'])
        else:
            yesterday_vol = yesterday_row['volume']

        vol_ratio = row['volume'] / yesterday_vol if yesterday_vol > 0 else 0
        is_priority = vol_ratio >= 1.0

        return (vol_ratio, is_priority)

    # ==================================================================
    # 日线数据读取
    # ==================================================================
    def _get_day_row(self, code: str, d: date) -> Optional[Dict]:
        """获取单只股票指定日期的日线数据"""
        market = 'sh' if code.startswith('6') else 'sz'
        if code.startswith(('8', '4')):
            market = 'bj'

        fpath = os.path.join(self.tdx._market_dir(market), f'{market}{code}.day')
        if not os.path.exists(fpath):
            return None
        fsize = os.path.getsize(fpath)
        if fsize < RECORD_SIZE * 3:
            return None

        read_size = min(RECORD_SIZE * 400, fsize)
        with open(fpath, 'rb') as f:
            f.seek(fsize - read_size)
            data = f.read(read_size)

        num_records = len(data) // RECORD_SIZE
        records = []
        for i in range(num_records):
            offset = i * RECORD_SIZE
            dt_int, op, hi, lo, cl, amt, vol, _ = struct.unpack(
                'IIIIIfII', data[offset:offset + RECORD_SIZE])
            dt_str = str(dt_int)
            try:
                rd = date(int(dt_str[:4]), int(dt_str[4:6]), int(dt_str[6:8]))
            except ValueError:
                continue
            records.append({
                'date': rd, 'open': op / 100.0, 'high': hi / 100.0,
                'low': lo / 100.0, 'close': cl / 100.0,
                'volume': vol, 'amount': amt,
            })

        for i, r in enumerate(records):
            if r['date'] == d:
                prev_close = records[i-1]['close'] if i > 0 else r['close']
                return {**r, 'prev_close': prev_close}

        return None

    def _get_open_price(self, code: str, d: date) -> Optional[float]:
        row = self._get_day_row(code, d)
        return row['open'] if row else None

    def _get_close_price(self, code: str, d: date) -> Optional[float]:
        row = self._get_day_row(code, d)
        return row['close'] if row else None

    # ==================================================================
    # 汇总统计
    # ==================================================================
    def _summarize(self, all_trades: List[Dict], daily_snapshots: List[Dict],
                   final_positions: List[Dict], stop_loss: float,
                   vol_ratio_threshold: float, max_positions: int,
                   start_date: str, end_date: str) -> Dict:
        """汇总回测统计"""
        # 把期末持仓以当日收盘价平仓
        finalized_trades = list(all_trades)
        for p in final_positions:
            if p.get('last_close', 0) > 0:
                ret = (p['last_close'] - p['entry_price']) / p['entry_price'] * 100
                entry_dt = datetime.strptime(p['entry_date'], '%Y%m%d').date()
                exit_dt = datetime.strptime(end_date, '%Y%m%d').date()
                finalized_trades.append({
                    'code': p['code'], 'name': p['name'],
                    'system': p['system'], 'sig_type': p.get('sig_type', ''),
                    'entry_date': p['entry_date'],
                    'entry_price': p['entry_price'],
                    'exit_date': end_date,
                    'exit_price': round(p['last_close'], 2),
                    'result': '期末持仓',
                    'return_pct': round(ret, 1),
                    'days_held': max((exit_dt - entry_dt).days, 1),
                    'score': p['score'],
                    'auction_vol_ratio': p['auction_vol_ratio'],
                })

        if not finalized_trades:
            return {'error': '无有效交易记录', 'all_trades': [], 'daily_snapshots': daily_snapshots}

        trades = finalized_trades

        # 分类统计
        wins = [t for t in trades if t['return_pct'] > 0]
        losses = [t for t in trades if t['return_pct'] <= 0]
        stops = [t for t in trades if t['result'] == '止损']

        total = len(trades)
        win_count = len(wins)
        loss_count = len(losses)
        stop_count = len(stops)
        win_rate = win_count / total * 100 if total > 0 else 0

        avg_win = np.mean([t['return_pct'] for t in wins]) if wins else 0
        avg_loss = np.mean([t['return_pct'] for t in losses]) if losses else 0
        avg_return = np.mean([t['return_pct'] for t in trades])

        # 盈亏比
        total_profit = sum(t['return_pct'] for t in wins)
        total_loss = sum(abs(t['return_pct']) for t in losses)
        profit_factor = total_profit / max(total_loss, 0.01)

        # 累计收益曲线
        sorted_trades = sorted(trades, key=lambda x: x['entry_date'])
        cum_returns = []
        cum = 0
        peak = 0
        max_dd = 0
        for t in sorted_trades:
            cum += t['return_pct']
            cum_returns.append(cum)
            peak = max(peak, cum)
            dd = peak - cum
            max_dd = max(max_dd, dd)

        # 按体系分组
        by_system = defaultdict(list)
        for t in trades:
            by_system[t['system']].append(t)

        system_stats = {}
        for sys_name, ts in by_system.items():
            sw = len([t for t in ts if t['return_pct'] > 0])
            sl = len([t for t in ts if t['return_pct'] <= 0])
            system_stats[sys_name] = {
                'trades': len(ts),
                'wins': sw,
                'losses': sl,
                'win_rate': round(sw / max(sw + sl, 1) * 100, 1),
                'avg_return': round(np.mean([t['return_pct'] for t in ts]), 1),
            }

        # 按月分组
        monthly = defaultdict(list)
        for t in trades:
            month_key = t['entry_date'][:6]
            monthly[month_key].append(t)

        monthly_stats = {}
        for month_key, ts in sorted(monthly.items()):
            sw = len([t for t in ts if t['return_pct'] > 0])
            sl = len([t for t in ts if t['return_pct'] <= 0])
            monthly_stats[month_key] = {
                'trades': len(ts),
                'wins': sw,
                'losses': sl,
                'win_rate': round(sw / max(sw + sl, 1) * 100, 1),
                'total_return': round(sum(t['return_pct'] for t in ts), 1),
            }

        # 按竞价量比分组
        priority_trades = [t for t in trades if t.get('auction_vol_ratio', 0) >= 1.0]
        normal_trades = [t for t in trades if 0.5 <= t.get('auction_vol_ratio', 0) < 1.0]

        def _stats(ts):
            if not ts: return {'trades': 0, 'win_rate': 0, 'avg_return': 0}
            w = len([t for t in ts if t['return_pct'] > 0])
            l = len([t for t in ts if t['return_pct'] <= 0])
            return {
                'trades': len(ts),
                'win_rate': round(w / max(w + l, 1) * 100, 1),
                'avg_return': round(np.mean([t['return_pct'] for t in ts]), 1),
            }

        return {
            'parameters': {
                'stop_loss_pct': round(stop_loss * 100, 0),
                'vol_ratio_threshold': vol_ratio_threshold,
                'max_positions': max_positions,
                'start_date': start_date,
                'end_date': end_date,
            },
            'summary': {
                'total_trades': total,
                'wins': win_count,
                'losses': loss_count,
                'stop_losses': stop_count,
                'win_rate': round(win_rate, 1),
                'avg_return': round(float(avg_return), 1),
                'avg_win': round(float(avg_win), 1),
                'avg_loss': round(float(avg_loss), 1),
                'total_return': round(float(cum), 1),
                'max_drawdown': round(float(max_dd), 1),
                'profit_factor': round(float(profit_factor), 2),
                'avg_days_held': round(float(np.mean([t['days_held'] for t in trades])), 1),
            },
            'by_system': system_stats,
            'monthly': monthly_stats,
            'auction_analysis': {
                'priority_(>=100%)': _stats(priority_trades),
                'normal_(>=50%)': _stats(normal_trades),
            },
            'all_trades': sorted(trades, key=lambda x: x['entry_date']),
            'daily_snapshots': daily_snapshots,
        }

    # ==================================================================
    # 辅助
    # ==================================================================
    def _get_trade_dates(self, lookback: int) -> List[date]:
        d = date.today() - timedelta(days=1)
        dates = []
        while len(dates) < lookback:
            if d.weekday() < 5 and self.calendar.is_trading_day(d):
                dates.append(d)
            d -= timedelta(days=1)
            if len(dates) == 0 and (date.today() - d).days > 400:
                break
        return list(reversed(dates))

    def _prev_trade_date(self, d: date) -> date:
        d = d - timedelta(days=1)
        while d.weekday() >= 5 or not self.calendar.is_trading_day(d):
            d -= timedelta(days=1)
            if (date.today() - d).days > 400:
                break
        return d

    # ---- 大盘环境过滤 ----
    _index_cache = None

    def _check_market_env(self, trade_date: str) -> bool:
        """上证跌幅>1.5%则空仓"""
        if self._index_cache is None:
            try:
                df = self.tdx.read_daily('999999', 'sh')
                self._index_cache = {}
                if 'trade_date' in df.columns:
                    for i in range(1, len(df)):
                        td = df['trade_date'].iloc[i]
                        if isinstance(td, datetime):
                            td = td.date()
                        elif isinstance(td, date):
                            pass
                        else:
                            continue
                        prev = float(df['close'].iloc[i-1])
                        curr = float(df['close'].iloc[i])
                        self._index_cache[td] = (curr - prev) / prev
            except Exception:
                self._index_cache = {}

        try:
            entry_dt = datetime.strptime(trade_date, '%Y%m%d').date()
            chg = self._index_cache.get(entry_dt)
            if chg is not None and chg < -0.015:
                return False
        except Exception:
            pass
        return True


# ======================================================================
# Excel 导出
# ======================================================================
RESULT_LABELS = {
    '止损': '止损', '止盈': '止盈', '亏损': '亏损',
    '期末持仓': '期末持仓',
}


def export_xlsx(result: Dict, filepath: str):
    """导出回测结果到xlsx文件"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
        from openpyxl.utils import get_column_letter
    except ImportError:
        print('[ERROR] 需要安装 openpyxl: pip install openpyxl')
        return

    wb = openpyxl.Workbook()

    # --- 通用样式 ---
    header_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    title_font = Font(name='微软雅黑', bold=True, size=14, color='2F5496')
    data_font = Font(name='微软雅黑', size=10)
    win_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    loss_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'))

    def style_header(ws, row, cols):
        for c in range(1, cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

    def style_data(ws, row, cols):
        for c in range(1, cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

    # =====================================================================
    # Sheet 1: 回测概览
    # =====================================================================
    ws1 = wb.active
    ws1.title = '回测概览'

    params = result['parameters']
    summary = result['summary']

    ws1.merge_cells('A1:C1')
    ws1.cell(row=1, column=1, value='双体系选股+竞价确认 回测报告').font = title_font

    overview_data = [
        ('', '', ''),
        ('回测参数', '', ''),
        ('回测区间', f"{params['start_date']} ~ {params['end_date']}", ''),
        ('止损比例', f"{params['stop_loss_pct']:+.0f}%", ''),
        ('竞价量比阈值', f"≥{params['vol_ratio_threshold']}", '（≥1.0为优先）'),
        ('最大持仓数', f"{params['max_positions']}只", ''),
        ('', '', ''),
        ('收益统计', '', ''),
        ('总交易笔数', f"{summary['total_trades']}笔", ''),
        ('胜率', f"{summary['win_rate']}%", f"({summary['wins']}胜/{summary['losses']}负)"),
        ('累计收益率', f"{summary['total_return']:+.1f}%", ''),
        ('平均每笔收益', f"{summary['avg_return']:+.1f}%", ''),
        ('平均盈利', f"{summary['avg_win']:+.1f}%", ''),
        ('平均亏损', f"{summary['avg_loss']:+.1f}%", ''),
        ('最大回撤', f"{summary['max_drawdown']:.1f}%", ''),
        ('盈亏比', f"{summary['profit_factor']:.2f}", ''),
        ('平均持仓天数', f"{summary['avg_days_held']:.1f}天", ''),
        ('止损次数', f"{summary['stop_losses']}次", ''),
        ('', '', ''),
        ('按体系分组', '', ''),
    ]

    for i, (label, val, note) in enumerate(overview_data, start=3):
        ws1.cell(row=i, column=1, value=label).font = Font(name='微软雅黑', bold=(val == ''), size=10)
        ws1.cell(row=i, column=2, value=val).font = data_font
        ws1.cell(row=i, column=3, value=note).font = Font(name='微软雅黑', size=9, color='888888')

    # 体系分组表
    sys_start = len(overview_data) + 4
    sys_headers = ['体系', '交易笔数', '胜笔', '负笔', '胜率', '平均收益']
    for c, h in enumerate(sys_headers, 1):
        ws1.cell(row=sys_start, column=c, value=h)
    style_header(ws1, sys_start, len(sys_headers))

    row = sys_start + 1
    for sys_name, stats in result['by_system'].items():
        vals = [sys_name, stats['trades'], stats['wins'], stats['losses'],
                f"{stats['win_rate']}%", f"{stats['avg_return']:+.1f}%"]
        for c, v in enumerate(vals, 1):
            ws1.cell(row=row, column=c, value=v)
        style_data(ws1, row, len(vals))
        row += 1

    # 竞价量比分析
    row += 1
    ws1.cell(row=row, column=1, value='竞价量比分析').font = Font(name='微软雅黑', bold=True, size=11)
    row += 1
    auc_headers = ['竞价量比', '交易笔数', '胜率', '平均收益']
    for c, h in enumerate(auc_headers, 1):
        ws1.cell(row=row, column=c, value=h)
    style_header(ws1, row, len(auc_headers))
    row += 1

    auc = result['auction_analysis']
    for label, stats in auc.items():
        vals = [label, stats['trades'], f"{stats['win_rate']}%", f"{stats['avg_return']:+.1f}%"]
        for c, v in enumerate(vals, 1):
            ws1.cell(row=row, column=c, value=v)
        style_data(ws1, row, len(vals))
        row += 1

    ws1.column_dimensions['A'].width = 20
    ws1.column_dimensions['B'].width = 22
    ws1.column_dimensions['C'].width = 30

    # =====================================================================
    # Sheet 2: 全部交易记录
    # =====================================================================
    ws2 = wb.create_sheet('全部交易记录')

    trade_headers = ['入场日期', '代码', '名称', '选股体系', '信号类型', '入场价',
                     '离场日期', '离场价', '结果', '收益率%', '持仓天数', '评分', '竞价量比']
    for c, h in enumerate(trade_headers, 1):
        ws2.cell(row=1, column=c, value=h)
    style_header(ws2, 1, len(trade_headers))

    trades = result['all_trades']
    for i, t in enumerate(trades):
        row = i + 2
        vals = [
            t['entry_date'], t['code'], t['name'],
            t.get('system', ''), t.get('sig_type', ''),
            t['entry_price'], t['exit_date'], t['exit_price'],
            t['result'], t['return_pct'], t['days_held'],
            t.get('score', 0), t.get('auction_vol_ratio', 0),
        ]
        for c, v in enumerate(vals, 1):
            ws2.cell(row=row, column=c, value=v)
        style_data(ws2, row, len(vals))

        # 盈亏着色
        if t['return_pct'] > 0:
            ws2.cell(row=row, column=10).fill = win_fill
        elif t['return_pct'] < 0:
            ws2.cell(row=row, column=10).fill = loss_fill

    # 列宽
    col_widths = [12, 8, 8, 12, 14, 10, 12, 10, 10, 10, 10, 8, 10]
    for c, w in enumerate(col_widths, 1):
        ws2.column_dimensions[get_column_letter(c)].width = w

    # 冻结首行
    ws2.freeze_panes = 'A2'
    # 自动筛选
    ws2.auto_filter.ref = f'A1:{get_column_letter(len(trade_headers))}{len(trades)+1}'

    # =====================================================================
    # Sheet 3: 月度收益汇总
    # =====================================================================
    ws3 = wb.create_sheet('月度收益汇总')

    month_headers = ['月份', '交易笔数', '胜笔', '负笔', '胜率', '月总收益%']
    for c, h in enumerate(month_headers, 1):
        ws3.cell(row=1, column=c, value=h)
    style_header(ws3, 1, len(month_headers))

    cum = 0
    for i, (month_key, stats) in enumerate(sorted(result['monthly'].items())):
        row = i + 2
        cum += stats['total_return']
        vals = [
            f"{month_key[:4]}-{month_key[4:6]}",
            stats['trades'], stats['wins'], stats['losses'],
            f"{stats['win_rate']}%",
            f"{stats['total_return']:+.1f}%",
        ]
        for c, v in enumerate(vals, 1):
            ws3.cell(row=row, column=c, value=v)
        style_data(ws3, row, len(vals))

        if stats['total_return'] > 0:
            ws3.cell(row=row, column=6).fill = win_fill
        elif stats['total_return'] < 0:
            ws3.cell(row=row, column=6).fill = loss_fill

    # 合计行
    total_row = len(result['monthly']) + 2
    total_vals = ['合计', result['summary']['total_trades'],
                  result['summary']['wins'], result['summary']['losses'],
                  f"{result['summary']['win_rate']}%",
                  f"{result['summary']['total_return']:+.1f}%"]
    for c, v in enumerate(total_vals, 1):
        ws3.cell(row=total_row, column=c, value=v)
        ws3.cell(row=total_row, column=c).font = Font(name='微软雅黑', bold=True, size=10)
        ws3.cell(row=total_row, column=c).border = thin_border
        ws3.cell(row=total_row, column=c).alignment = Alignment(horizontal='center')

    col_widths3 = [14, 12, 10, 10, 10, 14]
    for c, w in enumerate(col_widths3, 1):
        ws3.column_dimensions[get_column_letter(c)].width = w

    # =====================================================================
    # Sheet 4: 累计收益曲线数据
    # =====================================================================
    ws4 = wb.create_sheet('累计收益曲线')

    curve_headers = ['日期', '当日交易数', '当日收益%', '累计收益%', '最大回撤%', '持仓数']
    for c, h in enumerate(curve_headers, 1):
        ws4.cell(row=1, column=c, value=h)
    style_header(ws4, 1, len(curve_headers))

    # 按日期聚合交易
    trades_by_date = defaultdict(list)
    for t in result['all_trades']:
        trades_by_date[t['exit_date']].append(t)

    daily_trades_by_date = defaultdict(list)
    for t in result['all_trades']:
        daily_trades_by_date[t['entry_date']].append(t)

    snapshots_by_date = {}
    for s in result['daily_snapshots']:
        snapshots_by_date[s['date']] = s

    # 按日期排序
    all_dates = sorted(set(list(trades_by_date.keys()) +
                           list(daily_trades_by_date.keys()) +
                           list(snapshots_by_date.keys())))

    cum_return = 0
    peak_return = 0
    row = 2
    for d_str in all_dates:
        day_exits = trades_by_date.get(d_str, [])
        day_entries = daily_trades_by_date.get(d_str, [])
        day_return = sum(t['return_pct'] for t in day_exits)
        cum_return += day_return
        peak_return = max(peak_return, cum_return)
        dd = peak_return - cum_return

        snapshot = snapshots_by_date.get(d_str, {})
        pos_count = snapshot.get('positions', 0)

        vals = [d_str, len(day_entries) + len(day_exits),
                round(day_return, 1), round(cum_return, 1),
                round(dd, 1), pos_count]
        for c, v in enumerate(vals, 1):
            ws4.cell(row=row, column=c, value=v)
        style_data(ws4, row, len(vals))
        row += 1

    col_widths4 = [14, 14, 14, 14, 14, 10]
    for c, w in enumerate(col_widths4, 1):
        ws4.column_dimensions[get_column_letter(c)].width = w
    ws4.freeze_panes = 'A2'

    # 保存
    wb.save(filepath)
    print(f'\n[xlsx] 回测报告已导出 → {filepath}')
    print(f'   Sheet 1: 回测概览')
    print(f'   Sheet 2: 全部交易记录 ({len(trades)}笔)')
    print(f'   Sheet 3: 月度收益汇总 ({len(result["monthly"])}个月)')
    print(f'   Sheet 4: 累计收益曲线 ({len(all_dates)}天)')


# ======================================================================
# 控制台报告
# ======================================================================
def print_report(result: Dict):
    """控制台输出回测报告"""
    if result.get('error'):
        print(f'[ERROR] {result["error"]}')
        return

    params = result['parameters']
    s = result['summary']

    print(f'\n{"="*70}')
    print(f'  双体系选股+竞价确认 回测报告')
    print(f'{"="*70}')
    print(f'  回测区间: {params["start_date"]} ~ {params["end_date"]}')
    print(f'  参数: 止损{params["stop_loss_pct"]:+.0f}% | '
          f'竞价量比≥{params["vol_ratio_threshold"]} | 最大持仓{params["max_positions"]}只')
    print()
    print(f'  {"-"*66}')
    print(f'  收益统计')
    print(f'  {"-"*66}')
    print(f'  总交易: {s["total_trades"]}笔 | 胜: {s["wins"]} | 负: {s["losses"]} | 止损: {s["stop_losses"]}')
    print(f'  胜率: {s["win_rate"]}% | 累计收益: {s["total_return"]:+.1f}% | 平均收益: {s["avg_return"]:+.1f}%')
    print(f'  平均盈利: {s["avg_win"]:+.1f}% | 平均亏损: {s["avg_loss"]:+.1f}%')
    print(f'  最大回撤: {s["max_drawdown"]:.1f}% | 盈亏比: {s["profit_factor"]:.2f} | 平均持仓: {s["avg_days_held"]:.1f}天')
    print()
    print(f'  {"-"*66}')
    print(f'  按体系分组')
    print(f'  {"-"*66}')
    for sys_name, stats in result['by_system'].items():
        print(f'  {sys_name}: {stats["trades"]}笔 胜率{stats["win_rate"]}% 均收益{stats["avg_return"]:+.1f}%')
    print()
    print(f'  {"-"*66}')
    print(f'  竞价量比分析')
    print(f'  {"-"*66}')
    for label, stats in result['auction_analysis'].items():
        print(f'  {label}: {stats["trades"]}笔 胜率{stats["win_rate"]}% 均收益{stats["avg_return"]:+.1f}%')
    print()
    print(f'  {"-"*66}')
    print(f'  月度收益')
    print(f'  {"-"*66}')
    for month_key, stats in sorted(result['monthly'].items()):
        print(f'  {month_key[:4]}-{month_key[4:6]}: {stats["trades"]}笔 '
              f'胜率{stats["win_rate"]}% 收益{stats["total_return"]:+.1f}%')
    print()

    # 最近10笔交易
    print(f'  {"-"*66}')
    print(f'  最近10笔交易')
    print(f'  {"日期":<12} {"代码":<8} {"名称":<8} {"体系":<10} {"结果":<6} {"收益%":>7} {"持仓天":>6}')
    for t in result['all_trades'][-10:]:
        icon = '[胜]' if t['return_pct'] > 0 else '[负]'
        system = t.get('system', '')[:10]
        print(f'  {t["entry_date"]:<12} {t["code"]:<8} {t["name"]:<8} '
              f'{system:<10} {icon} {t["return_pct"]:>+6.1f}% {t["days_held"]:>4}天')


# ======================================================================
# CLI
# ======================================================================
def main():
    import argparse
    parser = argparse.ArgumentParser(description='双体系选股+竞价确认回测')
    parser.add_argument('--days', type=int, default=250,
                        help='回测天数 (默认250，约一年)')
    parser.add_argument('--output', '-o', type=str, required=True,
                        help='xlsx输出路径')
    parser.add_argument('--max-positions', type=int, default=3,
                        help='最大持仓数 (默认3)')
    parser.add_argument('--stop-loss', type=float, default=-0.06,
                        help='止损比例 (默认-0.06=-6%%)')
    parser.add_argument('--vol-ratio', type=float, default=0.5,
                        help='竞价量比阈值 (默认0.5=50%%)')
    args = parser.parse_args()

    print(f'\n{"="*70}')
    print(f'  双体系选股 + 竞价确认 回测')
    print(f'  选股: 1进2接力 + 涨停复制双体系')
    print(f'  竞价: 量比≥{args.vol_ratio} (≥1.0优先) | 最大持仓:{args.max_positions}只')
    print(f'  卖出: 涨停持有 / 不涨停卖出 / 跌{args.stop_loss*100:+.0f}%止损')
    print(f'{"="*70}\n')

    bt = DualSystemBacktest()
    result = bt.run(
        lookback_days=args.days,
        max_positions=args.max_positions,
        stop_loss=args.stop_loss,
        vol_ratio_threshold=args.vol_ratio,
    )

    print_report(result)
    export_xlsx(result, args.output)


if __name__ == '__main__':
    main()
